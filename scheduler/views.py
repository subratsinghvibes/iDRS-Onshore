from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.utils import timezone
from django.db import models
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, parser_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd
import csv
import io
import json
import math
from datetime import datetime, date, timedelta
from decimal import Decimal
import logging
import hmac
import hashlib
import time
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import CompanyCode, UserProfile, Rig, Well, Schedule, Assignment, UnassignedWell, StagedWell, DrillingBenchmark, RigBuildingNorm, RigBuildingAdjustment
from .serializers import (
    RigSerializer, WellSerializer, ScheduleSerializer, AssignmentSerializer,
    UnassignedWellSerializer, ScheduleCreateSerializer, GanttDataSerializer,
    AssignmentUpdateSerializer, BulkDataUploadSerializer, ScheduleStatsSerializer
)
from .optimization import DrillingScheduler
from .well_rejection_analyzer import WellRejectionAnalyzer

logger = logging.getLogger(__name__)


def get_user_location(user):
    """
    Get the location for a user. Returns None if user can access all locations.
    
    Args:
        user: Django User object
    
    Returns:
        CompanyCode object or None if user has access to all locations
    """
    if not user or not user.is_authenticated:
        return None
    
    # Superusers and staff can see all locations
    if user.is_superuser or user.is_staff:
        return None
    
    # Check if user has a profile
    try:
        profile = user.profile
        if profile.can_view_all_locations:
            return None
        return profile.location
    except UserProfile.DoesNotExist:
        # Create profile if it doesn't exist
        UserProfile.objects.create(user=user)
        return None


def get_user_accessible_locations(user):
    """
    Get all locations a user can access.
    
    Args:
        user: Django User object
    
    Returns:
        QuerySet of CompanyCode objects
    """
    if not user or not user.is_authenticated:
        return CompanyCode.objects.none()
    
    # Superusers can see all locations
    if user.is_superuser:
        return CompanyCode.objects.filter(is_active=True)
    
    # Check if user has a profile
    try:
        profile = user.profile
        return profile.get_accessible_locations()
    except UserProfile.DoesNotExist:
        UserProfile.objects.create(user=user)
        return CompanyCode.objects.none()


# =============================================================================
# ROLE-BASED PERMISSION HELPERS
# =============================================================================

def get_authorized_user(user):
    """
    Get the AuthorizedUser for a user.
    
    Args:
        user: Django User object
    
    Returns:
        AuthorizedUser object or None if no role assigned
    """
    if not user or not user.is_authenticated:
        return None
    
    try:
        from .models import AuthorizedUser
        # Try to get AuthorizedUser by linked user
        return AuthorizedUser.objects.get(user=user)
    except:
        # Fallback: try by CPF number (username)
        try:
            from .models import AuthorizedUser
            return AuthorizedUser.objects.get(cpf_no=user.username)
        except:
            return None


def user_can_view_all_locations(user):
    """
    Check if user can view all locations (admin or L1 role).
    
    Args:
        user: Django User object
    
    Returns:
        Boolean
    """
    if not user or not user.is_authenticated:
        return False
    
    # Superusers always have full access
    if user.is_superuser:
        return True
    
    # Staff members (L1) can view all locations
    if user.is_staff:
        return True
    
    # Check UserProfile setting (primary source)
    if hasattr(user, 'profile') and user.profile:
        return user.profile.can_view_all_locations
    
    # Fallback: Check role from AuthorizedUser (legacy)
    authorized_user = get_authorized_user(user)
    if authorized_user:
        return authorized_user.can_view_all_locations()
    
    # Default to False for regular users
    return False


def user_is_admin(user):
    """
    Check if user has admin role (full access).
    
    Args:
        user: Django User object
    
    Returns:
        Boolean
    """
    if not user or not user.is_authenticated:
        return False
    
    # Superusers are always admins
    if user.is_superuser:
        return True
    
    # Check role from AuthorizedUser
    authorized_user = get_authorized_user(user)
    if authorized_user:
        return authorized_user.is_admin_role()
    
    # Default to old behavior - staff are admins
    return user.is_staff


def get_user_assigned_location(user):
    """
    Get the assigned location for a user with 'user' role.
    
    Args:
        user: Django User object
    
    Returns:
        Location string or None
    """
    if not user or not user.is_authenticated:
        return None
    
    # First try to get location from UserProfile (primary source)
    if hasattr(user, 'profile') and user.profile and user.profile.location:
        # Return the location name from the CompanyCode object
        return user.profile.location.location
    
    # Fallback: try to get from AuthorizedUser table (legacy)
    authorized_user = get_authorized_user(user)
    if authorized_user and authorized_user.role == 'user':
        return authorized_user.assigned_location
    
    return None


def create_child_schedule(parent_schedule, branch_type="reschedule", custom_suffix=None,
                          input_wells_count=None, input_rigs_count=None, time_limit_seconds=None):
    """
    Create a properly named child schedule with version management
    
    Args:
        parent_schedule: The parent Schedule instance
        branch_type: Type of branching (reschedule, add_well, delete_well, etc.)
        custom_suffix: Optional custom suffix instead of auto-generated version
        input_wells_count: Number of wells for optimization
        input_rigs_count: Number of rigs for optimization
        time_limit_seconds: Solver time limit in seconds
    
    Returns:
        New Schedule instance with proper naming and relationships
    """
    # Get the base name from the root schedule
    base_name = parent_schedule.get_base_name()
    
    # Get next version number
    next_version = parent_schedule.get_next_version_number()
    
    # Create display name
    if custom_suffix:
        new_name = f"{base_name} {custom_suffix}"
    else:
        new_name = f"{base_name} v{next_version}"
    
    # Create the child schedule
    child_schedule = Schedule.objects.create(
        name=new_name,
        financial_year=parent_schedule.financial_year,
        parent_schedule=parent_schedule,
        branch_type=branch_type,
        version_number=next_version,
        status='RUNNING',
        created_by=parent_schedule.created_by,
        input_wells_count=input_wells_count,
        input_rigs_count=input_rigs_count,
        time_limit_seconds=time_limit_seconds,
    )
    
    logger.info(f"Created child schedule: '{new_name}' (v{next_version}) from parent '{parent_schedule.name}'")
    return child_schedule


@login_required
def index(request):
    """Main dashboard view"""
    return render(request, 'scheduler/dashboard.html')


@login_required
def dashboard(request):
    """Dashboard page view"""
    return render(request, 'scheduler/dashboard.html')


@login_required
def about(request):
    """About page view"""
    return render(request, 'scheduler/about.html')


@login_required
def test_appsense(request):
    """AppSense diagnostic page view"""
    return render(request, 'test_appsense.html')


@login_required
def get_appsense_url(request):
    """Generate signed AppSense URL on server-side (no HTTPS required)"""
    from .models import ExternalAppSetting
    
    try:
        # Get settings from database
        setting = ExternalAppSetting.get_setting()
        external_url = setting.url
        secret_key = setting.secret_key
        
        # Generate signature on server
        source_app = 'iDRS'
        username = request.user.username
        timestamp = int(time.time())
        
        # Build message for HMAC (pipe-separated format)
        message = f"{source_app}|{username}|{timestamp}"
        
        # Generate HMAC-SHA256 signature
        signature = hmac.new(
            secret_key.encode('utf-8'),
            message.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        # Build URL with parameters
        params = {
            'source_app': source_app,
            'user': username,
            'timestamp': timestamp,
            'sig': signature
        }
        
        full_url = f"{external_url}/report/?{urlencode(params)}"
        
        return JsonResponse({
            'success': True,
            'url': full_url
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def product_showcase(request):
    """Product showcase page for marketing/PR purposes"""
    return render(request, 'scheduler/product_showcase.html')


@login_required
def data_management(request):
    """Data management page view with role-based access control"""
    # Build a mapping from company_code to location for normalization
    from scheduler.models import CompanyCode
    company_code_to_location = {}
    for cc in CompanyCode.objects.all():
        if cc.company_code and cc.location:
            company_code_to_location[cc.company_code.upper()] = cc.location
            # Also map location to itself for cases where asset_id IS the location
            company_code_to_location[cc.location.upper()] = cc.location
    
    # Get unique asset_ids from both rigs and wells that actually have data
    rig_asset_ids = Rig.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    well_asset_ids = Well.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    # Normalize asset_ids to location values using CompanyCode mapping
    normalized_locations = set()
    for asset_id in list(rig_asset_ids) + list(well_asset_ids):
        if asset_id:
            # Try to map to location via CompanyCode
            normalized = company_code_to_location.get(asset_id.upper(), asset_id)
            normalized_locations.add(normalized)
    
    active_locations = sorted(normalized_locations)
    
    # Get user's role-based permissions
    user_location = None
    user_location_code = None
    can_view_all = user_can_view_all_locations(request.user)
    is_admin = user_is_admin(request.user)
    
    # If user has 'user' role, get their assigned location
    if not can_view_all:
        assigned_location = get_user_assigned_location(request.user)
        if assigned_location:
            user_location = assigned_location
            user_location_code = assigned_location
    
    context = {
        'asset_ids': active_locations,  # Normalized location values
        'user_location': user_location,
        'user_location_code': user_location_code,
        'can_view_all_locations': can_view_all,
        'is_admin': is_admin,
    }
    return render(request, 'scheduler/data_management.html', context)


@login_required
@staff_member_required
def user_management(request):
    """User management page - standalone page for managing users and roles"""
    is_admin = user_is_admin(request.user)
    context = {
        'is_admin': is_admin,
    }
    return render(request, 'scheduler/user_management.html', context)


@login_required
@staff_member_required
def company_codes(request):
    """Company codes management page - standalone page for managing company codes"""
    is_admin = user_is_admin(request.user)
    context = {
        'is_admin': is_admin,
    }
    return render(request, 'scheduler/company_codes.html', context)


@login_required
def view_all_rigs(request):
    """View all rigs in a comprehensive table format"""
    # Get user's role-based permissions
    can_view_all = user_can_view_all_locations(request.user)
    user_location = None
    
    if not can_view_all:
        assigned_location = get_user_assigned_location(request.user)
        if assigned_location:
            user_location = assigned_location
    
    context = {
        'can_view_all_locations': can_view_all,
        'user_location': user_location,
    }
    return render(request, 'scheduler/view_all_rigs.html', context)


@login_required
def view_all_wells(request):
    """View all wells in a comprehensive table format"""
    # Get user's role-based permissions
    can_view_all = user_can_view_all_locations(request.user)
    user_location = None
    
    if not can_view_all:
        assigned_location = get_user_assigned_location(request.user)
        if assigned_location:
            user_location = assigned_location
    
    context = {
        'can_view_all_locations': can_view_all,
        'user_location': user_location,
    }
    return render(request, 'scheduler/view_all_wells.html', context)


@login_required
def scheduling(request):
    """Scheduling page view with role-based access control"""
    # Build a mapping from company_code to location for normalization
    from scheduler.models import CompanyCode
    company_code_to_location = {}
    for cc in CompanyCode.objects.all():
        if cc.company_code and cc.location:
            company_code_to_location[cc.company_code.upper()] = cc.location
            # Also map location to itself for cases where asset_id IS the location
            company_code_to_location[cc.location.upper()] = cc.location
    
    # Get unique asset_ids from both rigs and wells that actually have data
    rig_asset_ids = Rig.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    well_asset_ids = Well.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    # Normalize asset_ids to location values using CompanyCode mapping
    normalized_locations = set()
    for asset_id in list(rig_asset_ids) + list(well_asset_ids):
        if asset_id:
            # Try to map to location via CompanyCode
            normalized = company_code_to_location.get(asset_id.upper(), asset_id)
            normalized_locations.add(normalized)
    
    active_locations = sorted(normalized_locations)
    
    # Import financial year functions
    from .models import get_financial_year_choices, get_current_financial_year
    
    # Get user's role-based permissions
    user_location = None
    user_location_code = None
    can_view_all = user_can_view_all_locations(request.user)
    is_admin = user_is_admin(request.user)
    
    # If user has 'user' role, get their assigned location
    if not can_view_all:
        assigned_location = get_user_assigned_location(request.user)
        if assigned_location:
            user_location = assigned_location
            user_location_code = assigned_location
    
    context = {
        'asset_ids': active_locations,  # Only show locations with actual data
        'financial_year_choices': get_financial_year_choices(),
        'current_financial_year': get_current_financial_year(),
        'user_location': user_location,
        'user_location_code': user_location_code,
        'can_view_all_locations': can_view_all,
        'is_admin': is_admin,
    }
    return render(request, 'scheduler/scheduling.html', context)


@login_required
def schedules_list(request):
    """Schedules management page view"""
    # Build a mapping from company_code to location for normalization
    from scheduler.models import CompanyCode
    company_code_to_location = {}
    for cc in CompanyCode.objects.all():
        if cc.company_code and cc.location:
            company_code_to_location[cc.company_code.upper()] = cc.location
            # Also map location to itself for cases where asset_id IS the location
            company_code_to_location[cc.location.upper()] = cc.location
    
    # Get unique asset_ids from both rigs and wells that actually have data
    rig_asset_ids = Rig.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    well_asset_ids = Well.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    # Normalize asset_ids to location values using CompanyCode mapping
    normalized_locations = set()
    for asset_id in list(rig_asset_ids) + list(well_asset_ids):
        if asset_id:
            # Try to map to location via CompanyCode
            normalized = company_code_to_location.get(asset_id.upper(), asset_id)
            normalized_locations.add(normalized)
    
    active_locations = sorted(normalized_locations)
    
    # Get user's role-based permissions
    user_location = None
    can_view_all = user_can_view_all_locations(request.user)
    
    # If user has 'user' role, get their assigned location
    if not can_view_all:
        assigned_location = get_user_assigned_location(request.user)
        if assigned_location:
            user_location = assigned_location
    
    context = {
        'asset_ids': active_locations,
        'user_location': user_location,
        'can_view_all_locations': can_view_all,
    }
    return render(request, 'scheduler/schedules.html', context)


@login_required
def schedule_maps(request):
    """Generate and display movement maps for a schedule"""
    schedule_id = request.GET.get('schedule_id')

    
    if not schedule_id:
        # Get location filter if provided
        location_filter = request.GET.get('location', None)
        
        # Get available schedules for user to choose from
        available_schedules = Schedule.objects.all().order_by('-created_at')
        
        # Filter schedules by location if specified - with proper company code mapping
        if location_filter:
            # Build mapping from location to all associated company codes
            location_to_codes = {}
            for cc in CompanyCode.objects.all():
                if cc.company_code and cc.location:
                    if cc.location not in location_to_codes:
                        location_to_codes[cc.location] = set()
                    location_to_codes[cc.location].add(cc.company_code)
                    location_to_codes[cc.location].add(cc.location)
            
            # Get all company codes associated with the selected location
            matching_codes = set()
            if location_filter in location_to_codes:
                matching_codes = location_to_codes[location_filter]
            else:
                for loc, codes in location_to_codes.items():
                    if loc.upper() == location_filter.upper():
                        matching_codes = codes
                        break
            
            if not matching_codes:
                matching_codes = {location_filter}
            
            # Filter schedules that have assignments with wells from any of the matching codes
            from django.db.models import Q
            code_filter = Q()
            for code in matching_codes:
                code_filter |= Q(assignments__well__asset_id__iexact=code)
            available_schedules = available_schedules.filter(code_filter).distinct()
        
        # Apply slice after filtering
        available_schedules = available_schedules[:10]
        
        # Get asset IDs for location filter with normalization
        company_code_to_location = {}
        for cc in CompanyCode.objects.all():
            if cc.company_code and cc.location:
                company_code_to_location[cc.company_code.upper()] = cc.location
                company_code_to_location[cc.location.upper()] = cc.location
        
        rig_asset_ids = Rig.objects.exclude(
            asset_id__isnull=True
        ).exclude(
            asset_id__exact=''
        ).values_list('asset_id', flat=True).distinct()
        
        well_asset_ids = Well.objects.exclude(
            asset_id__isnull=True
        ).exclude(
            asset_id__exact=''
        ).values_list('asset_id', flat=True).distinct()
        
        normalized_locations = set()
        for asset_id in list(rig_asset_ids) + list(well_asset_ids):
            if asset_id:
                normalized = company_code_to_location.get(asset_id.upper(), asset_id)
                normalized_locations.add(normalized)
        
        asset_ids = sorted(normalized_locations)
        
        return render(request, 'scheduler/schedule_maps.html', {
            'error': 'No schedule ID provided',
            'available_schedules': available_schedules,
            'show_schedule_selector': True,
            'asset_ids': asset_ids,
            'selected_location': location_filter,
        })
    
    try:
        # Get location filter if provided
        location_filter = request.GET.get('location', None)
        
        # Get the schedule
        schedule = Schedule.objects.get(id=schedule_id)
        assignments = Assignment.objects.filter(schedule=schedule)
        
        # Filter assignments by location if specified
        if location_filter:
            assignments = assignments.filter(well__asset_id=location_filter)
        
        if not assignments.exists():
            error_message = f'No assignments found for this schedule'
            if location_filter:
                error_message += f' in location {location_filter}'
            # Get asset IDs for location filter
            asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
            asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
            
            return render(request, 'scheduler/schedule_maps.html', {
                'error': error_message,
                'schedule': schedule,
                'selected_location': location_filter,
                'asset_ids': asset_ids,
            })
        
        # Prepare data for the scheduler
        rigs_data = []
        wells_data = []
        
        # Get unique rigs and wells from assignments
        for assignment in assignments:
            rig = assignment.rig
            well = assignment.well
            
            # Add rig data
            rig_data = {
                'name': rig.name,
                'rig_type': rig.rig_type,
                'start_date': rig.start_date,
                'end_date': rig.end_date,
                'rig_capacity_hp': rig.rig_capacity_hp,
                'daily_cost_inr': rig.daily_cost_inr,
                'drilling_capacity_m': rig.drilling_capacity_m,
                'mobilization_time_days': rig.mobilization_time_days,
                'maintenance_schedule': rig.maintenance_schedule,
                'crew_availability': rig.crew_availability,
                'hpht_suitability': rig.hpht_suitability,
                'ilm_cost_fixed': rig.ilm_cost_fixed,
                'ilm_cost_per_km': rig.ilm_cost_per_km,
                'ilm_cost_cluster': rig.ilm_cost_cluster,
                'bop_stack': rig.bop_stack,
                'tds_availability': rig.tds_availability
            }
            
            # Add well data
            well_data = {
                'name': well.name,
                'sn': well.sn,
                'asset_id': well.asset_id,
                'well_type': well.well_type,
                'well_profile': well.well_profile,
                'depth': well.depth,
                'rig_capacity_required_hp': well.rig_capacity_required_hp,
                'drl_days': well.drl_days,
                'pt_days': well.pt_days,
                'duration': well.duration,
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'rtd': well.rtd,
                'bop_stack': well.bop_stack,
                'tds_requirement': well.tds_requirement,
                'footprint': well.footprint,
                'preferred_rig': well.preferred_rig,
                'expected_potential': well.expected_potential,
                'priority': well.priority
            }
            
            # Check if already added to avoid duplicates
            if not any(r['name'] == rig.name for r in rigs_data):
                rigs_data.append(rig_data)
            if not any(w['name'] == well.name for w in wells_data):
                wells_data.append(well_data)
        
        # Create scheduler instance with existing assignments
        from .optimization import DrillingScheduler
        scheduler = DrillingScheduler(rigs_data, wells_data)
        
        # Manually set results based on existing assignments
        scheduler.results = {
            'assignments': [],
            'unassigned_wells': [],
            'total_drilling_cost': 0,
            'total_ilm_cost': 0,
            'unassigned_wells_count': 0,
            'solver_status': 'OPTIMAL'
        }
        
        # Convert assignments to scheduler format
        for assignment in assignments:
            assignment_data = {
                'rig': assignment.rig.name,
                'well': assignment.well.name,
                'well_start_date': assignment.well_start_date,
                'well_end_date': assignment.well_end_date,
                'rig_start_date': assignment.rig.start_date,
                'rig_end_date': assignment.rig.end_date,
                'rtd': assignment.well.rtd,
                'priority': assignment.well.priority,
                'duration': assignment.well.duration,
                'depth': assignment.well.depth,
                'required_hp': assignment.well.rig_capacity_required_hp,
                'drilling_cost': float(assignment.drilling_cost or 0),
                'ilm_cost': float(assignment.ilm_cost or 0),
                'latitude': float(assignment.well.latitude),
                'longitude': float(assignment.well.longitude),
                'sequence_order': getattr(assignment, 'sequence_order', 1)
            }
            scheduler.results['assignments'].append(assignment_data)
        
        # Generate maps
        gantt_chart_html = scheduler.generate_gantt_chart()
        
        # Build assignment data for Leaflet map
        import json as json_module
        map_assignments = []
        for a in scheduler.results['assignments']:
            map_assignments.append({
                'rig': a['rig'],
                'well': a['well'],
                'latitude': float(a['latitude']),
                'longitude': float(a['longitude']),
                'well_start_date': str(a['well_start_date']),
                'well_end_date': str(a['well_end_date']),
                'duration': a['duration'],
                'depth': a['depth'],
                'required_hp': a['required_hp'],
                'rtd': str(a.get('rtd', '')),
                'sequence_order': a.get('sequence_order', 1),
            })
        
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        context = {
            'schedule': schedule,
            'map_assignments_json': json_module.dumps(map_assignments),
            'gantt_chart_html': gantt_chart_html,
            'assignments_count': len(assignments),
            'asset_ids': asset_ids,
            'selected_location': location_filter,
        }
        
        return render(request, 'scheduler/schedule_maps.html', context)
        
    except Schedule.DoesNotExist:
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        return render(request, 'scheduler/schedule_maps.html', {
            'error': 'Schedule not found',
            'asset_ids': asset_ids,
        })
    except Exception as e:
        logger.error(f"Error generating schedule maps: {str(e)}")
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        return render(request, 'scheduler/schedule_maps.html', {
            'error': f'Error generating maps: {str(e)}',
            'asset_ids': asset_ids,
        })


@login_required
@staff_member_required
def er_diagram(request):
    """ER Diagram and database schema visualization page - Admin only"""
    from django.apps import apps
    from django.db import connection
    
    # Get ALL models from ALL apps (not just scheduler app)
    all_models = []
    for app_config in apps.get_app_configs():
        all_models.extend(app_config.get_models())
    
    # Sort models by app name and model name for better organization
    app_models = sorted(all_models, key=lambda model: (model._meta.app_label, model.__name__))
    
    # Build model information
    models_info = {}
    apps_info = {}
    
    for model in app_models:
        model_name = model.__name__
        app_label = model._meta.app_label
        
        # Initialize app info if not exists
        if app_label not in apps_info:
            apps_info[app_label] = {
                'name': app_label,
                'models': [],
                'display_name': app_label.replace('_', ' ').title()
            }
        
        model_info = {
            'name': model_name,
            'app_label': app_label,
            'table_name': model._meta.db_table,
            'fields': [],
            'relationships': [],
            'doc_string': model.__doc__ or '',
            'verbose_name': str(model._meta.verbose_name),
            'verbose_name_plural': str(model._meta.verbose_name_plural),
        }
        
        # Add to app's models list
        apps_info[app_label]['models'].append(model_name)
        
        # Get field information
        for field in model._meta.get_fields():
            field_info = {
                'name': field.name,
                'type': field.__class__.__name__,
                'null': getattr(field, 'null', False),
                'blank': getattr(field, 'blank', False),
                'unique': getattr(field, 'unique', False),
                'primary_key': getattr(field, 'primary_key', False),
                'max_length': getattr(field, 'max_length', None),
                'help_text': getattr(field, 'help_text', ''),
                'choices': getattr(field, 'choices', None),
            }
            
            # Handle relationship fields
            if hasattr(field, 'related_model') and field.related_model:
                field_info['related_model'] = field.related_model.__name__
                field_info['related_table'] = field.related_model._meta.db_table
                
                # Add to relationships
                relationship = {
                    'from_model': model_name,
                    'to_model': field.related_model.__name__,
                    'field_name': field.name,
                    'relationship_type': field.__class__.__name__,
                }
                model_info['relationships'].append(relationship)
            
            model_info['fields'].append(field_info)
        
        models_info[model_name] = model_info
    
    # Get database statistics
    with connection.cursor() as cursor:
        stats = {}
        for model in app_models:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {model._meta.db_table}")
                result = cursor.fetchone()
                if result:
                    stats[model.__name__] = result[0]
                else:
                    stats[model.__name__] = 0
            except Exception as e:
                stats[model.__name__] = f"Error: {str(e)}"
    
    context = {
        'models_info': models_info,
        'apps_info': apps_info,
        'stats': stats,
        'total_models': len(app_models),
        'total_apps': len(apps_info),
    }
    
    return render(request, 'scheduler/er_diagram.html', context)


@login_required
def interactive_gantt(request):
    """Interactive Gantt chart view"""
    # Get asset IDs for location filter with normalization
    from scheduler.models import CompanyCode
    company_code_to_location = {}
    for cc in CompanyCode.objects.all():
        if cc.company_code and cc.location:
            company_code_to_location[cc.company_code.upper()] = cc.location
            company_code_to_location[cc.location.upper()] = cc.location
    
    rig_asset_ids = Rig.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    well_asset_ids = Well.objects.exclude(
        asset_id__isnull=True
    ).exclude(
        asset_id__exact=''
    ).values_list('asset_id', flat=True).distinct()
    
    normalized_locations = set()
    for asset_id in list(rig_asset_ids) + list(well_asset_ids):
        if asset_id:
            normalized = company_code_to_location.get(asset_id.upper(), asset_id)
            normalized_locations.add(normalized)
    
    asset_ids = sorted(normalized_locations)
    
    context = {
        'asset_ids': asset_ids,
    }
    return render(request, 'scheduler/interactive_gantt.html', context)


@login_required
def schedule_detail(request, schedule_id):
    """Schedule detail view showing comprehensive schedule information"""
    schedule = get_object_or_404(Schedule, id=schedule_id)
    
    # Get all assignments for this schedule - using all_objects to include deleted items
    assignments = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well').order_by('rig__name', 'sequence_order')
    
    # Get unassigned wells
    unassigned_wells = UnassignedWell.objects.filter(schedule=schedule).select_related('well').order_by('well__sn')
    
    # Get input data - use all_objects to include any deleted wells/rigs that are in this schedule
    # This ensures historical schedules display correctly even if wells/rigs were deleted
    all_wells = Well.all_objects.all().order_by('sn')
    all_rigs = Rig.all_objects.all().order_by('name')
    
    # Identify assigned wells and rigs
    assigned_well_ids = assignments.values_list('well_id', flat=True)
    assigned_rig_ids = assignments.values_list('rig_id', flat=True)
    
    # Group assignments by rig for better display
    assignments_by_rig = {}
    for assignment in assignments:
        rig_name = assignment.rig.name
        if rig_name not in assignments_by_rig:
            assignments_by_rig[rig_name] = []
        assignments_by_rig[rig_name].append(assignment)
    
    # Calculate utilization statistics
    total_wells = all_wells.count()
    assigned_wells_count = len(assigned_well_ids)
    unassigned_wells_count = unassigned_wells.count()
    
    # For rig statistics, use only active (non-deleted) rigs
    active_rigs = Rig.objects.filter(is_deleted=False)
    total_rigs = active_rigs.count()
    utilized_rigs_count = len(set(assigned_rig_ids))
    
    # Get unused rigs with reasons - only show active (non-deleted) rigs
    unused_rigs = active_rigs.exclude(id__in=assigned_rig_ids).order_by('name')
    unused_rigs_count = unused_rigs.count()
    
    # Calculate costs
    total_drilling_cost = assignments.aggregate(
        total=models.Sum('drilling_cost')
    )['total'] or Decimal('0.00')
    
    total_ilm_cost = assignments.aggregate(
        total=models.Sum('ilm_cost')
    )['total'] or Decimal('0.00')
    
    total_ilm_days = assignments.aggregate(
        total=models.Sum('ilm_days')
    )['total'] or Decimal('0.0')
    
    context = {
        'schedule': schedule,
        'assignments': assignments,
        'assignments_by_rig': assignments_by_rig,
        'unassigned_wells': unassigned_wells,
        'all_wells': all_wells,
        'all_rigs': all_rigs,
        'assigned_well_ids': assigned_well_ids,
        'assigned_rig_ids': assigned_rig_ids,
        'unused_rigs': unused_rigs,
        'stats': {
            'total_wells': total_wells,
            'assigned_wells_count': assigned_wells_count,
            'unassigned_wells_count': unassigned_wells_count,
            'total_rigs': total_rigs,
            'utilized_rigs_count': utilized_rigs_count,
            'unused_rigs_count': unused_rigs_count,
            'total_drilling_cost': total_drilling_cost,
            'total_ilm_cost': total_ilm_cost,
            'total_ilm_days': total_ilm_days,
            'total_cost': total_drilling_cost + total_ilm_cost,
        }
    }
    
    return render(request, 'scheduler/schedule_detail.html', context)


@login_required
def schedule_comparison(request):
    """Schedule comparison view for analyzing multiple schedules"""
    schedule_ids = request.GET.getlist('schedules')
    
    if not schedule_ids:
        # If no schedules specified, show empty comparison page
        return render(request, 'scheduler/schedule_comparison.html', {
            'schedules': [],
            'error_message': 'No schedules selected for comparison. Please select schedules from the scheduling page.'
        })
    
    if len(schedule_ids) > 3:
        # Limit to 3 schedules for better display
        schedule_ids = schedule_ids[:3]
    
    schedules_data = []
    
    for schedule_id in schedule_ids:
        try:
            schedule = Schedule.objects.get(id=schedule_id)
            
            # Get assignments and statistics for this schedule
            assignments = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
            unassigned_wells = UnassignedWell.objects.filter(schedule=schedule).select_related('well')
            
            # Calculate statistics
            total_drilling_cost = assignments.aggregate(
                total=models.Sum('drilling_cost')
            )['total'] or Decimal('0.00')
            
            total_ilm_cost = assignments.aggregate(
                total=models.Sum('ilm_cost')
            )['total'] or Decimal('0.00')
            
            total_cost = total_drilling_cost + total_ilm_cost
            
            # Get rig utilization - count unique rigs, not wells
            utilized_rigs = assignments.values_list('rig__name', flat=True).distinct()
            utilized_rigs_count = len(set(utilized_rigs))
            
            # Get detailed rig information
            rig_details = []
            for rig_name in set(utilized_rigs):
                rig_assignments = assignments.filter(rig__name=rig_name)
                rig_details.append({
                    'name': rig_name,
                    'wells_count': rig_assignments.count(),
                    'total_cost': rig_assignments.aggregate(
                        total=models.Sum('drilling_cost') + models.Sum('ilm_cost')
                    )['total'] or Decimal('0.00')
                })
            
            # Group assignments by rig
            assignments_by_rig = {}
            for assignment in assignments:
                rig_name = assignment.rig.name
                if rig_name not in assignments_by_rig:
                    assignments_by_rig[rig_name] = []
                assignments_by_rig[rig_name].append(assignment)
            
            # Calculate efficiency metrics
            assigned_wells_count = assignments.count()
            unassigned_wells_count = unassigned_wells.count()
            total_wells_count = assigned_wells_count + unassigned_wells_count
            success_rate = (assigned_wells_count / total_wells_count * 100) if total_wells_count > 0 else 0
            cost_per_well = (total_cost / assigned_wells_count) if assigned_wells_count > 0 else 0
            wells_per_rig = (assigned_wells_count / utilized_rigs_count) if utilized_rigs_count > 0 else 0
            
            schedule_data = {
                'schedule': schedule,
                'assignments': assignments,
                'assignments_by_rig': assignments_by_rig,
                'unassigned_wells': unassigned_wells,
                'rig_details': rig_details,
                'stats': {
                    'assigned_wells_count': assigned_wells_count,
                    'unassigned_wells_count': unassigned_wells_count,
                    'total_wells_count': total_wells_count,
                    'utilized_rigs_count': utilized_rigs_count,
                    'utilized_rigs': list(set(utilized_rigs)),
                    'total_drilling_cost': total_drilling_cost,
                    'total_ilm_cost': total_ilm_cost,
                    'total_cost': total_cost,
                    'success_rate': success_rate,
                    'cost_per_well': cost_per_well,
                    'wells_per_rig': wells_per_rig,
                }
            }
            
            schedules_data.append(schedule_data)
            
        except Schedule.DoesNotExist:
            continue
    
    if not schedules_data:
        return render(request, 'scheduler/schedule_comparison.html', {
            'schedules': [],
            'error_message': 'No valid schedules found for comparison.'
        })
    
    context = {
        'schedules': schedules_data,
        'comparison_count': len(schedules_data)
    }
    
    return render(request, 'scheduler/schedule_comparison.html', context)


class RigViewSet(viewsets.ModelViewSet):
    """ViewSet for managing drilling rigs"""
    queryset = Rig.objects.all()  # Default: only active (non-deleted) rigs
    serializer_class = RigSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        # Check if we need to include deleted records (e.g., for admin data management)
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        queryset = Rig.all_objects.all() if include_deleted else Rig.objects.all()
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                # Filter by location FK OR by asset_id matching location code (for legacy data)
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(location=user_location) | Q(asset_id__iexact=user_location.code)
                )
        
        rig_type = self.request.query_params.get('rig_type', None)
        available_now = self.request.query_params.get('available_now', None)
        asset_id = self.request.query_params.get('asset_id', None)
        location_id = self.request.query_params.get('location_id', None)
        
        if rig_type:
            queryset = queryset.filter(rig_type=rig_type)
        
        if available_now == 'true':
            today = timezone.now().date()
            queryset = queryset.filter(
                start_date__lte=today,
                end_date__gte=today
            )
        
        if asset_id:
            # Normalize asset_id filter: look up CompanyCode to find matching location/company_code
            # This handles cases where filter value is 'Cambay' but data has 'CBY' or 'CAMBAY'
            from django.db.models import Q
            try:
                # Try to find CompanyCode where location or company_code matches (case-insensitive)
                company_code = CompanyCode.objects.filter(
                    Q(location__iexact=asset_id) | Q(company_code__iexact=asset_id)
                ).first()
                if company_code:
                    # Filter by matching either location name or company code
                    queryset = queryset.filter(
                        Q(asset_id__iexact=company_code.location) | 
                        Q(asset_id__iexact=company_code.company_code) |
                        Q(location=company_code)
                    )
                else:
                    # Fallback to case-insensitive exact match
                    queryset = queryset.filter(asset_id__iexact=asset_id)
            except Exception:
                # Fallback to case-insensitive exact match
                queryset = queryset.filter(asset_id__iexact=asset_id)
        
        # Additional location filter from query params (within user's accessible locations)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        return queryset

    def destroy(self, request, *args, **kwargs):
        """Override hard delete: perform soft-delete to avoid ProtectedError and preserve history"""
        try:
            rig = self.get_object()
            # If already deleted, return idempotent success
            if getattr(rig, 'is_deleted', False):
                return Response({'message': 'Rig already deleted'}, status=status.HTTP_200_OK)
            rig.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Rig soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete rig: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def soft_delete(self, request, pk=None):
        """Soft delete a rig without affecting existing schedules"""
        try:
            rig = self.get_object()
            if rig.is_deleted:
                return Response({'message': 'Rig already deleted'}, status=status.HTTP_200_OK)
            rig.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Rig soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete rig: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def hard_delete(self, request, pk=None):
        """Permanently delete a rig from the database"""
        try:
            rig = self.get_object()
            rig_name = rig.name
            # Perform actual database deletion
            rig.delete()
            return Response({'message': f'Rig "{rig_name}" permanently deleted'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to permanently delete rig: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def bulk_soft_delete(self, request):
        """Soft delete multiple rigs (recoverable)"""
        try:
            rig_ids = request.data.get('rig_ids', [])
            if not rig_ids:
                return Response({'error': 'No rig IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            rigs = Rig.objects.filter(id__in=rig_ids)
            count = rigs.count()
            user = request.user if request.user.is_authenticated else None
            for rig in rigs:
                rig.soft_delete(user=user)
            
            return Response({
                'message': f'Soft-deleted {count} rig(s). They can be recovered later.',
                'deleted_count': count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to soft delete rigs: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def bulk_hard_delete(self, request):
        """Permanently delete multiple rigs and all related records"""
        try:
            rig_ids = request.data.get('rig_ids', [])
            
            if not rig_ids:
                return Response({'error': 'No rig IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            from scheduler.models import (
                Schedule, Assignment, ScheduleRig, ScheduleWell, UnassignedWell,
                ExecutionSchedule, ExecutionRig, ExecutionWell, ExecutionLog, ExecutionScenario
            )
            
            # Always cascade - delete ALL related records first
            
            # 1. Delete ExecutionWell records referencing these rigs
            ExecutionWell.objects.filter(rig_id__in=rig_ids).delete()
            
            # 2. Delete ExecutionRig records referencing these rigs
            ExecutionRig.objects.filter(rig_id__in=rig_ids).delete()
            
            # 3. Get schedules that reference these rigs
            schedule_ids = list(ScheduleRig.objects.filter(rig_id__in=rig_ids).values_list('schedule_id', flat=True).distinct())
            schedule_count = len(schedule_ids)
            assignment_count = Assignment.objects.filter(rig_id__in=rig_ids).count()
            
            # 4. Delete assignments referencing these rigs
            Assignment.objects.filter(rig_id__in=rig_ids).delete()
            
            # 5. Delete ScheduleRig records
            ScheduleRig.objects.filter(rig_id__in=rig_ids).delete()
            
            # 6. Delete schedules and all their related records (including execution chain)
            for schedule_id in schedule_ids:
                # Delete execution chain for this schedule
                exec_schedule_ids = list(ExecutionSchedule.objects.filter(source_schedule_id=schedule_id).values_list('id', flat=True))
                for exec_id in exec_schedule_ids:
                    ExecutionLog.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionScenario.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionWell.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionRig.objects.filter(execution_schedule_id=exec_id).delete()
                ExecutionSchedule.objects.filter(source_schedule_id=schedule_id).delete()
                
                # Delete schedule's own records
                Assignment.objects.filter(schedule_id=schedule_id).delete()
                ScheduleWell.objects.filter(schedule_id=schedule_id).delete()
                ScheduleRig.objects.filter(schedule_id=schedule_id).delete()
                UnassignedWell.objects.filter(schedule_id=schedule_id).delete()
                Schedule.objects.filter(id=schedule_id).delete()
            
            # 7. Now delete the rigs (including soft-deleted ones)
            # Note: delete() returns (total_rows_deleted, {model_label: count})
            # We use the breakdown dict to get only the Rig count, not cascade-deleted records
            delete_result = Rig.all_objects.filter(id__in=rig_ids).delete()
            deleted_count = delete_result[1].get('scheduler.Rig', 0)
            
            return Response({
                'message': f'Permanently deleted {deleted_count} rig(s), {schedule_count} schedule(s), and {assignment_count} assignment(s)',
                'deleted_count': deleted_count,
                'deleted_schedules': schedule_count,
                'deleted_assignments': assignment_count
            }, status=status.HTTP_200_OK)
                
        except Exception as e:
            logger.error(f"Error in bulk_hard_delete rigs: {str(e)}")
            return Response({'error': f'Failed to bulk delete rigs: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        """Bulk upload rigs from CSV file"""
        serializer = BulkDataUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        rigs_file = serializer.validated_data.get('rigs_file')
        if not rigs_file:
            return Response(
                {'error': 'No rigs file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Delegate to the robust handler
        return handle_rigs_upload(rigs_file)
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export rigs data as CSV"""
        try:
            queryset = self.get_queryset()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="rigs_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = [
                'ID', 'Name', 'Asset ID', 'Rig Type', 'Start Date', 'End Date',
                'Rig Capacity (HP)', 'Daily Cost (INR)', 'Drilling Capacity (m)',
                'Mobilization Time (Days)', 'Maintenance Schedule', 'Crew Availability',
                'HPHT Suitability', 'ILM Cost Fixed', 'ILM Cost per km',
                'ILM Cost Cluster', 'BOP Stack', 'TDS Availability', 'Created At'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for rig in queryset:
                writer.writerow([
                    str(rig.id),
                    rig.name,
                    rig.asset_id or '',
                    rig.rig_type,
                    rig.start_date.strftime('%Y-%m-%d') if rig.start_date else '',
                    rig.end_date.strftime('%Y-%m-%d') if rig.end_date else '',
                    rig.rig_capacity_hp,
                    rig.daily_cost_inr,
                    rig.drilling_capacity_m,
                    rig.mobilization_time_days or '',
                    rig.maintenance_schedule or '',
                    rig.crew_availability or '',
                    rig.hpht_suitability or '',
                    rig.ilm_cost_fixed or '',
                    rig.ilm_cost_per_km or '',
                    rig.ilm_cost_cluster or '',
                    rig.bop_stack or '',
                    rig.tds_availability or '',
                    rig.created_at.strftime('%Y-%m-%d %H:%M:%S') if rig.created_at else ''
                ])
            
            return response
            
        except Exception as e:
            return JsonResponse(
                {'error': f'Failed to export CSV: {str(e)}'}, 
                status=500
            )


class WellViewSet(viewsets.ModelViewSet):
    """ViewSet for managing wells"""
    queryset = Well.objects.all()  # Default: only active (non-deleted) wells
    serializer_class = WellSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        # Check if we need to include deleted records (e.g., for admin data management)
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        queryset = Well.all_objects.all() if include_deleted else Well.objects.all()
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                # Filter by location FK OR by asset_id matching location code (for legacy data)
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(location=user_location) | Q(asset_id__iexact=user_location.code)
                )
        
        asset_id = self.request.query_params.get('asset_id', None)
        well_type = self.request.query_params.get('well_type', None)
        priority = self.request.query_params.get('priority', None)
        location_id = self.request.query_params.get('location_id', None)
        
        if asset_id:
            # Normalize asset_id filter: look up CompanyCode to find matching location/company_code
            # This handles cases where filter value is 'Cambay' but data has 'CBY' or 'CAMBAY'
            from django.db.models import Q
            try:
                # Try to find CompanyCode where location or company_code matches (case-insensitive)
                company_code = CompanyCode.objects.filter(
                    Q(location__iexact=asset_id) | Q(company_code__iexact=asset_id)
                ).first()
                if company_code:
                    # Filter by matching either location name or company code
                    queryset = queryset.filter(
                        Q(asset_id__iexact=company_code.location) | 
                        Q(asset_id__iexact=company_code.company_code) |
                        Q(location=company_code)
                    )
                else:
                    # Fallback to case-insensitive exact match
                    queryset = queryset.filter(asset_id__iexact=asset_id)
            except Exception:
                # Fallback to case-insensitive exact match
                queryset = queryset.filter(asset_id__iexact=asset_id)
                
        if well_type:
            queryset = queryset.filter(well_type=well_type)
        if priority:
            queryset = queryset.filter(priority=priority)
        
        # Additional location filter from query params (within user's accessible locations)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        return queryset

    def destroy(self, request, *args, **kwargs):
        """Override hard delete: perform soft-delete to avoid ProtectedError and preserve history"""
        try:
            well = self.get_object()
            # If already deleted, return idempotent success
            if getattr(well, 'is_deleted', False):
                return Response({'message': 'Well already deleted'}, status=status.HTTP_200_OK)
            well.soft_delete(user=request.user if request.user.is_authenticated else None)
            # Clean up staged wells that were imported from this well
            StagedWell.objects.filter(imported_well=well).update(status='PENDING', imported_well=None, imported_at=None)
            return Response({'message': 'Well soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete well: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def soft_delete(self, request, pk=None):
        """Soft delete a well without affecting existing schedules"""
        try:
            well = self.get_object()
            if well.is_deleted:
                return Response({'message': 'Well already deleted'}, status=status.HTTP_200_OK)
            well.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Well soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete well: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def hard_delete(self, request, pk=None):
        """Permanently delete a well from the database"""
        try:
            well = self.get_object()
            well_name = well.name
            # Perform actual database deletion
            well.delete()
            return Response({'message': f'Well "{well_name}" permanently deleted'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to permanently delete well: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def bulk_soft_delete(self, request):
        """Soft delete multiple wells (recoverable)"""
        try:
            well_ids = request.data.get('well_ids', [])
            if not well_ids:
                return Response({'error': 'No well IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            wells = Well.objects.filter(id__in=well_ids)
            count = wells.count()
            user = request.user if request.user.is_authenticated else None
            for well in wells:
                well.soft_delete(user=user)
            
            # Reset staged wells that imported these wells back to PENDING
            StagedWell.objects.filter(imported_well_id__in=well_ids).update(
                status='PENDING', imported_well=None, imported_at=None
            )
            
            return Response({
                'message': f'Soft-deleted {count} well(s). They can be recovered later.',
                'deleted_count': count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to soft delete wells: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def bulk_hard_delete(self, request):
        """Permanently delete multiple wells and all related records"""
        try:
            well_ids = request.data.get('well_ids', [])
            
            if not well_ids:
                return Response({'error': 'No well IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
            
            from scheduler.models import (
                Assignment, ScheduleWell, UnassignedWell, Schedule, ScheduleRig,
                ExecutionSchedule, ExecutionRig, ExecutionWell, ExecutionLog, ExecutionScenario
            )
            
            # Always cascade - delete ALL related records first
            
            # 1. Delete ExecutionWell records referencing these wells
            ExecutionWell.objects.filter(well_id__in=well_ids).delete()
            
            # 2. Get schedules that reference these wells
            schedule_ids = list(ScheduleWell.objects.filter(well_id__in=well_ids).values_list('schedule_id', flat=True).distinct())
            schedule_count = len(schedule_ids)
            assignment_count = Assignment.objects.filter(well_id__in=well_ids).count()
            
            # 3. Delete assignments referencing these wells
            Assignment.objects.filter(well_id__in=well_ids).delete()
            
            # 4. Delete UnassignedWell records referencing these wells  
            UnassignedWell.objects.filter(well_id__in=well_ids).delete()
            
            # 5. Delete ScheduleWell records
            ScheduleWell.objects.filter(well_id__in=well_ids).delete()
            
            # 6. Delete schedules and all their related records (including execution chain)
            for schedule_id in schedule_ids:
                # Delete execution chain for this schedule
                exec_schedule_ids = list(ExecutionSchedule.objects.filter(source_schedule_id=schedule_id).values_list('id', flat=True))
                for exec_id in exec_schedule_ids:
                    ExecutionLog.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionScenario.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionWell.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionRig.objects.filter(execution_schedule_id=exec_id).delete()
                ExecutionSchedule.objects.filter(source_schedule_id=schedule_id).delete()
                
                # Delete schedule's own records
                Assignment.objects.filter(schedule_id=schedule_id).delete()
                ScheduleWell.objects.filter(schedule_id=schedule_id).delete()
                ScheduleRig.objects.filter(schedule_id=schedule_id).delete()
                UnassignedWell.objects.filter(schedule_id=schedule_id).delete()
                Schedule.objects.filter(id=schedule_id).delete()
            
            # 7. Clean up staged wells that reference these wells
            #    - Reset imported staged wells so they can be re-uploaded
            #    - Or delete them if the source well is being permanently removed
            staged_cleaned = StagedWell.objects.filter(imported_well_id__in=well_ids).count()
            StagedWell.objects.filter(imported_well_id__in=well_ids).delete()
            
            # 8. Also delete staged wells that match by name (for wells imported without FK link)
            well_names = list(Well.all_objects.filter(id__in=well_ids).values_list('name', flat=True))
            if well_names:
                name_matched = StagedWell.objects.filter(name__in=well_names, status='IMPORTED').count()
                StagedWell.objects.filter(name__in=well_names, status='IMPORTED').delete()
                staged_cleaned += name_matched
            
            # 9. Now delete the wells (including soft-deleted ones)
            # Note: delete() returns (total_rows_deleted, {model_label: count})
            # We use the breakdown dict to get only the Well count, not cascade-deleted records
            delete_result = Well.all_objects.filter(id__in=well_ids).delete()
            deleted_count = delete_result[1].get('scheduler.Well', 0)
            
            return Response({
                'message': f'Permanently deleted {deleted_count} well(s), {schedule_count} schedule(s), and {assignment_count} assignment(s)',
                'deleted_count': deleted_count,
                'deleted_schedules': schedule_count,
                'deleted_assignments': assignment_count,
                'staged_wells_cleaned': staged_cleaned
            }, status=status.HTTP_200_OK)
                
        except Exception as e:
            logger.error(f"Error in bulk_hard_delete wells: {str(e)}")
            return Response({'error': f'Failed to bulk delete wells: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        """Bulk upload wells from CSV file"""
        serializer = BulkDataUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        wells_file = serializer.validated_data.get('wells_file')
        if not wells_file:
            return Response(
                {'error': 'No wells file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Delegate to the robust handler
        return handle_wells_upload(wells_file)
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export wells data as CSV"""
        try:
            queryset = self.get_queryset()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="wells_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = [
                'ID', 'SN', 'Asset ID', 'Name', 'Well Type', 'Well Profile', 'Depth',
                'Rig Capacity Required (HP)', 'DRL Days', 'PT Days', 'Duration',
                'Latitude', 'Longitude', 'RTD', 'BOP Stack', 'TDS Requirement',
                'Footprint', 'Preferred Rig', 'Expected Potential', 'Priority', 'Created At'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for well in queryset:
                writer.writerow([
                    str(well.id),
                    well.sn or '',
                    well.asset_id or '',
                    well.name,
                    well.well_type,
                    well.well_profile or '',
                    well.depth or '',
                    well.rig_capacity_required_hp or '',
                    well.drl_days or '',
                    well.pt_days or '',
                    well.duration or '',
                    well.latitude or '',
                    well.longitude or '',
                    well.rtd.strftime('%Y-%m-%d') if well.rtd else '',
                    well.bop_stack or '',
                    well.tds_requirement or '',
                    well.footprint or '',
                    well.preferred_rig or '',
                    well.expected_potential or '',
                    well.priority or '',
                    well.created_at.strftime('%Y-%m-%d %H:%M:%S') if well.created_at else ''
                ])
            
            return response
            
        except Exception as e:
            return JsonResponse(
                {'error': f'Failed to export CSV: {str(e)}'}, 
                status=500
            )


#new code piece
def run_full_optimization(rig_queryset=None, well_queryset=None, base_start_date=None, time_limit_seconds=60, financial_year=None):
    """
    Runs the full optimizer and returns the result dictionary.
    
    Args:
        rig_queryset: QuerySet of rigs to include (default: all rigs)
        well_queryset: QuerySet of wells to include (default: all wells)
        base_start_date: Base date for schedule (day 0)
        time_limit_seconds: Solver time limit
        financial_year: Financial year string (e.g., "2024-2025") to constrain well start dates
    """
    from .models import parse_financial_year
    
    rigs = rig_queryset.order_by('name') if rig_queryset is not None else Rig.objects.all().order_by('name')
    wells = well_queryset.order_by('name') if well_queryset is not None else Well.objects.all().order_by('name')
    rigs_data = [rig.to_dict() for rig in rigs]
    wells_data = [well.to_dict() for well in wells]
    
    # Parse FY constraints if provided
    fy_start_date = None
    fy_end_date = None
    if financial_year:
        try:
            fy_start_date, fy_end_date = parse_financial_year(financial_year)
            logger.info(f"run_full_optimization using FY constraints: {fy_start_date} to {fy_end_date}")
        except ValueError as e:
            logger.warning(f"Could not parse financial year: {e}")
    
    scheduler = DrillingScheduler(rigs_data, wells_data, base_start_date, fy_start_date=fy_start_date, fy_end_date=fy_end_date)
    result = scheduler.solve(time_limit_seconds=time_limit_seconds, deterministic=True)
    return result
#-- new code piece
class ScheduleViewSet(viewsets.ModelViewSet):
    """ViewSet for managing schedules"""
    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            from .serializers import ScheduleListSerializer
            return ScheduleListSerializer
        return self.serializer_class

    def get_queryset(self):
        queryset = Schedule.objects.all().order_by('-created_at')
        # Optimise FK lookups that serializers need
        queryset = queryset.select_related('location', 'created_by', 'parent_schedule')
        # Annotate assignment count so the list serializer can use it
        if self.action == 'list':
            from django.db.models import Count
            queryset = queryset.annotate(assignments_count=Count('assignments'))
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                queryset = queryset.filter(location=user_location)
        
        # Check if request has query_params (DRF Request) or GET (Django Request)
        if hasattr(self.request, 'query_params'):
            status_filter = self.request.query_params.get('status', None)
            asset_id_filter = self.request.query_params.get('asset_id', None)
            location_id = self.request.query_params.get('location_id', None)
        else:
            status_filter = self.request.GET.get('status', None)
            asset_id_filter = self.request.GET.get('asset_id', None)
            location_id = self.request.GET.get('location_id', None)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter schedules by asset_id (location) - with company code mapping
        if asset_id_filter:
            # Build a mapping from location to all associated company codes
            location_to_codes = {}
            for cc in CompanyCode.objects.all():
                if cc.company_code and cc.location:
                    if cc.location not in location_to_codes:
                        location_to_codes[cc.location] = set()
                    location_to_codes[cc.location].add(cc.company_code)
                    location_to_codes[cc.location].add(cc.location)  # Include location itself
            
            # Get all company codes associated with the selected location
            matching_codes = set()
            # First check if asset_id_filter is itself a location name
            if asset_id_filter in location_to_codes:
                matching_codes = location_to_codes[asset_id_filter]
            else:
                # Try to find by case-insensitive match
                for loc, codes in location_to_codes.items():
                    if loc.upper() == asset_id_filter.upper():
                        matching_codes = codes
                        break
                    # Also check if the filter matches any company code
                    for code in codes:
                        if code.upper() == asset_id_filter.upper():
                            matching_codes = codes
                            break
                    if matching_codes:
                        break
            
            # If no mapping found, fall back to direct match
            if not matching_codes:
                matching_codes = {asset_id_filter}
            
            # Filter schedules that have assignments with wells from any of the matching codes
            from django.db.models import Q
            code_filter = Q()
            for code in matching_codes:
                code_filter |= Q(assignments__well__asset_id__iexact=code)
            queryset = queryset.filter(code_filter).distinct()
        
        # Additional location filter from query params (within user's accessible locations)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        return queryset
    
    def destroy(self, request, *args, **kwargs):
        """Delete a schedule and all related records (cascade through Execution models)"""
        try:
            schedule = self.get_object()
            schedule_name = schedule.name
            schedule_id = schedule.id
            
            from scheduler.models import (
                Assignment, ScheduleRig, ScheduleWell, UnassignedWell,
                ExecutionSchedule, ExecutionRig, ExecutionWell, ExecutionLog, ExecutionScenario
            )
            
            # Delete execution chain first (PROTECT FKs)
            exec_schedule_ids = list(
                ExecutionSchedule.objects.filter(source_schedule_id=schedule_id)
                .values_list('id', flat=True)
            )
            for exec_id in exec_schedule_ids:
                ExecutionLog.objects.filter(execution_schedule_id=exec_id).delete()
                ExecutionScenario.objects.filter(execution_schedule_id=exec_id).delete()
                ExecutionWell.objects.filter(execution_schedule_id=exec_id).delete()
                ExecutionRig.objects.filter(execution_schedule_id=exec_id).delete()
            ExecutionSchedule.objects.filter(source_schedule_id=schedule_id).delete()
            
            # Delete schedule's own related records
            Assignment.objects.filter(schedule_id=schedule_id).delete()
            ScheduleWell.objects.filter(schedule_id=schedule_id).delete()
            ScheduleRig.objects.filter(schedule_id=schedule_id).delete()
            UnassignedWell.objects.filter(schedule_id=schedule_id).delete()
            
            # Delete child schedules (branches)
            child_ids = list(Schedule.objects.filter(parent_schedule_id=schedule_id).values_list('id', flat=True))
            for child_id in child_ids:
                # Recursively clean child execution chains
                child_exec_ids = list(
                    ExecutionSchedule.objects.filter(source_schedule_id=child_id)
                    .values_list('id', flat=True)
                )
                for exec_id in child_exec_ids:
                    ExecutionLog.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionScenario.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionWell.objects.filter(execution_schedule_id=exec_id).delete()
                    ExecutionRig.objects.filter(execution_schedule_id=exec_id).delete()
                ExecutionSchedule.objects.filter(source_schedule_id=child_id).delete()
                Assignment.objects.filter(schedule_id=child_id).delete()
                ScheduleWell.objects.filter(schedule_id=child_id).delete()
                ScheduleRig.objects.filter(schedule_id=child_id).delete()
                UnassignedWell.objects.filter(schedule_id=child_id).delete()
            Schedule.objects.filter(parent_schedule_id=schedule_id).delete()
            
            # Finally delete the schedule itself
            schedule.delete()
            
            return Response({'message': f'Schedule "{schedule_name}" deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error deleting schedule: {str(e)}")
            return Response({'error': f'Failed to delete schedule: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def create_schedule(self, request):
        """Create and run a new schedule optimization"""
        serializer = ScheduleCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        
        schedule = None
        try:
            # --- Phase 1: Create schedule record (committed immediately so it's visible) ---
            schedule_location = None
            if request.user and request.user.is_authenticated:
                user_location = get_user_location(request.user)
                if user_location:
                    schedule_location = user_location
            
            schedule = Schedule.objects.create(
                name=validated_data['name'],
                financial_year=validated_data['financial_year'],
                location=schedule_location,
                status='RUNNING',
                created_by=request.user if request.user.is_authenticated else None,
                input_wells_count=len(validated_data['well_ids']),
                input_rigs_count=len(validated_data['rig_ids']),
                time_limit_seconds=validated_data.get('time_limit_seconds', 600),
            )
            
            # Get rigs and wells (ordered for deterministic scheduling)
            rigs = Rig.objects.filter(id__in=validated_data['rig_ids']).order_by('name')
            wells = Well.objects.filter(id__in=validated_data['well_ids']).order_by('name')
            
            # If no location set yet, try to get from first well
            if not schedule_location and wells.exists():
                first_well = wells.first()
                if first_well.location:
                    schedule.location = first_well.location
                    schedule.save()
            
            if not rigs.exists():
                schedule.status = 'FAILED'
                schedule.solver_status = 'NO_RIGS'
                schedule.save()
                return Response(
                    {'error': 'No valid rigs found'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not wells.exists():
                schedule.status = 'FAILED'
                schedule.solver_status = 'NO_WELLS'
                schedule.save()
                return Response(
                    {'error': 'No valid wells found'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # --- Phase 2: Run optimization (outside transaction so schedule stays visible) ---
            try:
                rigs_data = list(rigs.values())
                wells_data = list(wells.values())
                
                from .models import parse_financial_year
                fy_start_date = None
                fy_end_date = None
                try:
                    fy_start_date, fy_end_date = parse_financial_year(validated_data['financial_year'])
                    logger.info(f"Schedule FY constraints: {fy_start_date} to {fy_end_date} (wells must START within this period)")
                except ValueError as e:
                    logger.warning(f"Could not parse financial year: {e}. Running without FY constraints.")
                
                scheduler = DrillingScheduler(
                    rigs_data, 
                    wells_data,
                    fy_start_date=fy_start_date,
                    fy_end_date=fy_end_date
                )
                results = scheduler.solve(time_limit_seconds=validated_data['time_limit_seconds'], deterministic=True)
                
                if results and results.get('status') in ['OPTIMAL', 'FEASIBLE']:
                    # --- Phase 3: Save results atomically ---
                    with transaction.atomic():
                        schedule.status = 'COMPLETED'
                        schedule.completed_at = timezone.now()
                        schedule.total_drilling_cost = Decimal(str(results.get('total_drilling_cost', 0)))
                        schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
                        schedule.project_end_date = results.get('project_end_date')
                        schedule.unassigned_wells_count = results.get('unassigned_wells_count', 0)
                        schedule.solver_status = results.get('solver_status')
                        schedule.solve_time_seconds = results.get('solve_time_seconds')
                        schedule.optimality_gap_percent = results.get('optimality_gap_percent')
                        schedule.schedule_hash = results.get('schedule_hash')
                        schedule.save()
                        
                        from .models import ScheduleRig, ScheduleWell
                        
                        schedule_rigs = []
                        for rig in rigs:
                            schedule_rigs.append(ScheduleRig(schedule=schedule, rig=rig))
                        ScheduleRig.objects.bulk_create(schedule_rigs, ignore_conflicts=True)
                        
                        schedule_wells = []
                        for well in wells:
                            schedule_wells.append(ScheduleWell(schedule=schedule, well=well))
                        ScheduleWell.objects.bulk_create(schedule_wells, ignore_conflicts=True)
                        
                        # Create assignments with proper sequence order
                        assignments_by_rig = {}
                        for assignment_data in results.get('assignments', []):
                            rig_name = assignment_data['rig']
                            if rig_name not in assignments_by_rig:
                                assignments_by_rig[rig_name] = []
                            assignments_by_rig[rig_name].append(assignment_data)
                        
                        for rig_name, rig_assignments in assignments_by_rig.items():
                            rig_assignments.sort(key=lambda x: x['well_start_date'])
                            for i, assignment_data in enumerate(rig_assignments, 1):
                                assignment_data['calculated_sequence_order'] = i
                        
                        for assignment_data in results.get('assignments', []):
                            try:
                                rig = rigs.get(name=assignment_data['rig'])
                                well = wells.get(name=assignment_data['well'])
                            except (Rig.DoesNotExist, Well.DoesNotExist) as e:
                                logger.warning(f"Assignment creation failed - {str(e)}: rig={assignment_data.get('rig')}, well={assignment_data.get('well')}")
                                continue
                                
                            Assignment.objects.create(
                                schedule=schedule,
                                rig=rig,
                                well=well,
                                well_start_date=assignment_data['well_start_date'],
                                well_end_date=assignment_data['well_end_date'],
                                rtd_check=assignment_data.get('rtd_check', 'OK'),
                                well_start_check=assignment_data.get('well_start_check', 'OK'),
                                well_end_check=assignment_data.get('well_end_check', 'OK'),
                                depth_check=assignment_data.get('depth_check', 'OK'),
                                hp_check=assignment_data.get('hp_check', 'OK'),
                                bop_check=assignment_data.get('bop_check', 'OK'),
                                tds_check=assignment_data.get('tds_check', 'OK'),
                                rig_type_check=assignment_data.get('rig_type_check', 'OK'),
                                drilling_cost=Decimal(str(assignment_data.get('drilling_cost_inr', assignment_data.get('drilling_cost', 0)))),
                                ilm_cost=Decimal(str(assignment_data.get('ilm_cost', 0))),
                                ilm_days=Decimal(str(assignment_data.get('ilm_days', 0))),
                                sequence_order=assignment_data.get('calculated_sequence_order', 1)
                            )
                        
                        # Create unassigned wells with detailed rejection analysis
                        assigned_well_names = [assignment['well'] for assignment in results.get('assignments', [])]
                        
                        wells_data = []
                        rigs_data = []
                        
                        for well in wells:
                            wells_data.append({
                                'name': well.name,
                                'depth': well.depth,
                                'duration': well.duration,
                                'rtd': well.rtd,
                                'rig_capacity_required_hp': well.rig_capacity_required_hp,
                                'bop_stack': well.bop_stack,
                                'tds_requirement': well.tds_requirement,
                                'priority': well.priority,
                                'latitude': float(well.latitude) if well.latitude else 0.0,
                                'longitude': float(well.longitude) if well.longitude else 0.0,
                                'footprint': well.footprint,
                            })
                        
                        for rig in rigs:
                            rigs_data.append({
                                'name': rig.name,
                                'start_date': rig.start_date,
                                'end_date': rig.end_date,
                                'rig_capacity_hp': rig.rig_capacity_hp,
                                'drilling_capacity_m': rig.drilling_capacity_m,
                                'bop_stack': rig.bop_stack,
                                'tds_availability': rig.tds_availability,
                                'daily_cost_inr': float(rig.daily_cost_inr),
                                'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                                'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                                'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                                'rig_type': rig.rig_type,
                            })
                        
                        wells_df = pd.DataFrame(wells_data)
                        rigs_df = pd.DataFrame(rigs_data)
                        
                        analyzer = WellRejectionAnalyzer(wells_df, rigs_df, timezone.now().date())
                        
                        for well_name in results.get('unassigned_wells', []):
                            try:
                                well = wells.get(name=well_name)
                            except Well.DoesNotExist:
                                logger.warning(f"Unassigned well creation failed - Well not found: {well_name}")
                                continue
                            
                            detailed_reason = analyzer.analyze_well_rejection(well_name, assigned_well_names)
                            
                            UnassignedWell.objects.create(
                                schedule=schedule,
                                well=well,
                                reason=detailed_reason
                            )
                    
                    return Response(ScheduleSerializer(schedule).data)
                
                else:
                    schedule.status = 'FAILED'
                    schedule.solver_status = results.get('status', 'NO_SOLUTION') if results else 'NO_SOLUTION'
                    schedule.save()
                    
                    return Response(
                        {'error': 'No feasible solution found. This can happen when rig availability '
                                  'windows are too short for the selected wells, required equipment '
                                  'exceeds rig capabilities, or Financial Year constraints are too '
                                  'restrictive. Try adding more rigs or reducing the number of wells.'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                    
            except Exception as e:
                if schedule:
                    schedule.status = 'FAILED'
                    schedule.solver_status = str(e)[:20]
                    schedule.save()
                logger.error(f"Optimization failed: {str(e)}", exc_info=True)
                user_msg = self._friendly_error_message(str(e))
                return Response(
                    {'error': user_msg}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
                    
        except Exception as e:
            if schedule:
                schedule.status = 'FAILED'
                schedule.save()
            logger.error(f"Schedule creation failed: {str(e)}", exc_info=True)
            user_msg = self._friendly_error_message(str(e))
            return Response(
                {'error': user_msg}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @staticmethod
    def _friendly_error_message(raw: str) -> str:
        """Translate raw Python exception text into a user-readable message."""
        lowered = raw.lower()
        if 'nattype' in lowered or 'nat' in lowered and 'date' in lowered:
            return ('One or more rigs/wells have missing or invalid dates. '
                    'Please check that all Rig start/end dates and Well RTD '
                    'dates are valid and not empty.')
        if 'unsupported operand' in lowered and 'date' in lowered:
            return ('A date calculation failed due to invalid date values in '
                    'the selected rigs or wells. Please verify all dates are '
                    'correct in Data Management.')
        if 'keyerror' in lowered or 'key error' in lowered:
            return ('Required data is missing for one or more rigs/wells. '
                    'Please ensure all required fields are filled in Data Management.')
        if 'division by zero' in lowered or 'zerodivision' in lowered:
            return ('A calculation error occurred (division by zero). '
                    'Please check that well durations and rig capacities are not zero.')
        if 'timeout' in lowered or 'timed out' in lowered:
            return 'The optimization timed out. Try increasing the time limit or reducing the problem size.'
        if 'memory' in lowered:
            return ('The server ran out of memory. Try optimizing fewer wells/rigs '
                    'or contact your administrator.')
        # Generic fallback — still useful, hides internal stack trace
        return (f'An unexpected error occurred during optimization. '
                f'Please try again or contact support if the problem persists. '
                f'(Technical detail: {raw[:200]})')
    
    @action(detail=True, methods=['post'])
    def cancel_schedule(self, request, pk=None):
        """Cancel a RUNNING schedule optimization.
        
        The actual OR-Tools solver runs synchronously in the create_schedule
        request, so the HTTP connection is still open.  When the frontend
        calls AbortController.abort(), Django receives a *new* request
        hitting this endpoint.  We mark the schedule as CANCELLED so the
        frontend knows it was intentionally stopped.
        """
        schedule = self.get_object()
        if schedule.status == 'RUNNING':
            schedule.status = 'CANCELLED'
            schedule.save()
            return Response({'status': 'cancelled', 'message': 'Optimization cancelled by user.'})
        return Response(
            {'error': f'Schedule is not running (current status: {schedule.status})'},
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=True, methods=['post'])
    def reschedule_with_actuals(self, request, pk=None):
        """Re-optimize an existing schedule using actual start/end dates.

        Expected payload:
        {
          "actuals": [
             {"well": "Well-1", "rig": "Rig-1", "actual_start_date": "2025-05-10", "actual_end_date": "2025-06-05"},
             {"assignment_id": "<uuid>", "actual_start_date": "2025-07-01"}
          ],
          "time_limit_seconds": 60
        }
        - If rig is omitted, it will be inferred from the current schedule's assignment for that well/assignment.
        - If only one of start/end is provided, only that boundary is pinned.
        - A new Schedule will be created to store results (original remains unchanged).
        """
        schedule = self.get_object()

        payload = request.data if isinstance(request.data, dict) else {}
        actuals = payload.get('actuals', []) or []
        time_limit_seconds = int(payload.get('time_limit_seconds', 60))

        if not isinstance(actuals, list) or len(actuals) == 0:
            return Response({'error': 'Provide a non-empty "actuals" list'}, status=status.HTTP_400_BAD_REQUEST)

        # Build current rigs/wells dataset from the ORIGINAL SCOPE of the schedule
        from .models import ScheduleRig, ScheduleWell
        
        # Get the originally selected rigs and wells for this schedule
        schedule_rigs = ScheduleRig.objects.filter(schedule=schedule).select_related('rig')
        schedule_wells = ScheduleWell.objects.filter(schedule=schedule).select_related('well')
        
        if not schedule_rigs.exists() or not schedule_wells.exists():
            # Fallback to assignment-based detection if no scope tracking available
            assignments_qs = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
            if not assignments_qs.exists():
                return Response({'error': 'No assignments or scope found for the provided schedule'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Use rigs and wells from assignments as fallback
            rigs = sorted([a.rig for a in assignments_qs], key=lambda r: r.name)
            wells = sorted([a.well for a in assignments_qs], key=lambda w: w.name)
            rigs_seen = set([r.name for r in rigs])
            wells_seen = set([w.name for w in wells])
        else:
            # Use the originally selected scope
            rigs = sorted([sr.rig for sr in schedule_rigs], key=lambda r: r.name)
            wells = sorted([sw.well for sw in schedule_wells], key=lambda w: w.name)
            rigs_seen = set([r.name for r in rigs])
            wells_seen = set([w.name for w in wells])

        rigs_data: list[dict] = []
        wells_data: list[dict] = []

        # Build the data structures for rigs and wells
        for rig in rigs:
            if rig.name not in [r['name'] for r in rigs_data]:  # Avoid duplicates
                rigs_data.append({
                    'name': rig.name,
                    'rig_type': rig.rig_type,
                    'start_date': rig.start_date,
                    'end_date': rig.end_date,
                    'rig_capacity_hp': rig.rig_capacity_hp,
                    'daily_cost_inr': float(rig.daily_cost_inr),
                    'drilling_capacity_m': rig.drilling_capacity_m,
                    'mobilization_time_days': rig.mobilization_time_days,
                    'maintenance_schedule': rig.maintenance_schedule,
                    'crew_availability': rig.crew_availability,
                    'hpht_suitability': rig.hpht_suitability,
                    'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                    'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                    'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                    'bop_stack': rig.bop_stack,
                    'tds_availability': rig.tds_availability,
                })
        
        for well in wells:
            if well.name not in [w['name'] for w in wells_data]:  # Avoid duplicates
                wells_data.append({
                    'name': well.name,
                    'sn': well.sn,
                    'asset_id': well.asset_id,
                    'well_type': well.well_type,
                    'well_profile': well.well_profile,
                    'depth': well.depth,
                    'rig_capacity_required_hp': well.rig_capacity_required_hp,
                    'drl_days': well.drl_days,
                    'pt_days': well.pt_days,
                    'duration': well.duration,
                    'latitude': float(well.latitude),
                    'longitude': float(well.longitude),
                    'rtd': well.rtd,
                    'bop_stack': well.bop_stack,
                    'tds_requirement': well.tds_requirement,
                    'footprint': well.footprint,
                    'preferred_rig': well.preferred_rig,
                    'expected_potential': well.expected_potential,
                    'priority': well.priority,
                })
        
        # Build mapping from existing assignments
        assignments_qs = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
        well_to_rig: dict[str, str] = {}
        assignment_map: dict[str, dict] = {}
        
        for a in assignments_qs:
            well_to_rig[a.well.name] = a.rig.name
            assignment_map[str(a.id)] = {
                'well': a.well.name,
                'rig': a.rig.name,
            }

        # STEP 1: Build comprehensive actuals list from EXISTING assignments with actual dates
        # This ensures ALL locked dates are preserved with absolute priority
        normalized_actuals: list[dict] = []
        
        # Add ALL existing actual dates from original schedule as FIXED constraints
        for assignment in assignments_qs:
            if assignment.actual_start_date or assignment.actual_end_date:
                normalized_actuals.append({
                    'well': assignment.well.name,
                    'rig': assignment.rig.name,
                    'actual_start_date': assignment.actual_start_date,
                    'actual_end_date': assignment.actual_end_date,
                })
        
        # STEP 2: Process new actuals provided in the API call (these override existing if same well-rig)
        actuals_map = {}  # Track well-rig combinations to avoid duplicates
        for existing in normalized_actuals:
            key = (existing['well'], existing['rig'])
            actuals_map[key] = existing
            
        for rec in actuals:
            if not isinstance(rec, dict):
                continue
            well = rec.get('well')
            rig = rec.get('rig')
            assignment_id = rec.get('assignment_id')
            
            # Infer well/rig from assignment_id if not provided
            if not well and assignment_id and str(assignment_id) in assignment_map:
                well = assignment_map[str(assignment_id)]['well']
                rig = rig or assignment_map[str(assignment_id)]['rig']
            if well and not rig:
                # infer rig from current schedule
                rig = well_to_rig.get(well)
            if not well or not rig:
                # skip records we cannot map to current schedule
                continue
            
            # Filter out None values and only include records with actual dates
            actual_start_date = rec.get('actual_start_date')
            actual_end_date = rec.get('actual_end_date')
            
            if actual_start_date or actual_end_date:
                key = (well, rig)
                # Override existing or add new actual dates
                actuals_map[key] = {
                    'well': well,
                    'rig': rig,
                    'actual_start_date': actual_start_date,
                    'actual_end_date': actual_end_date,
                }

        # Convert back to list with only valid actual dates (sorted for determinism)
        normalized_actuals = []
        for key in sorted(actuals_map.keys()):
            actual_data = actuals_map[key]
            # Only include if at least one actual date is provided and not None
            if actual_data.get('actual_start_date') or actual_data.get('actual_end_date'):
                normalized_actuals.append(actual_data)

        logger.info(f"Fixed actuals for optimization: {len(normalized_actuals)} well-rig combinations with locked dates")
        
        # Parse Financial Year constraints from the schedule
        from .models import parse_financial_year
        fy_start_date = None
        fy_end_date = None
        if schedule.financial_year:
            try:
                fy_start_date, fy_end_date = parse_financial_year(schedule.financial_year)
                logger.info(f"Re-optimize using FY constraints: {fy_start_date} to {fy_end_date}")
            except ValueError as e:
                logger.warning(f"Could not parse schedule financial year: {e}")
        
        if not normalized_actuals:
            logger.warning("No actual dates found in existing assignments or provided actuals - running normal optimization")
            # If no actual dates to lock, run normal optimization instead
            scheduler = DrillingScheduler(rigs_data, wells_data, fy_start_date=fy_start_date, fy_end_date=fy_end_date)
            results = scheduler.solve(time_limit_seconds=time_limit_seconds, deterministic=True)
        else:
            # Run rescheduling optimizer with locked actual dates as FIXED constraints
            scheduler = DrillingScheduler(rigs_data, wells_data, fy_start_date=fy_start_date, fy_end_date=fy_end_date)
            results = scheduler.solve_with_actuals(normalized_actuals, time_limit_seconds=time_limit_seconds, deterministic=True)

        if not results or results.get('status') not in ['OPTIMAL', 'FEASIBLE']:
            # Provide detailed analysis when solution is infeasible
            failure_analysis = scheduler.analyze_infeasible_solution(normalized_actuals)
            solver_status = results.get('status') if results else 'NO_RESULT'
            
            return Response({
                'error': f"No feasible solution: {solver_status}",
                'error_type': 'INFEASIBLE_SOLUTION',
                'detailed_analysis': failure_analysis,
                'solver_status': solver_status,
                'actuals_provided': len(normalized_actuals)
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create a properly named child schedule
        new_schedule = create_child_schedule(
            schedule, branch_type="reschedule",
            input_wells_count=len(wells_data),
            input_rigs_count=len(rigs_data),
            time_limit_seconds=time_limit_seconds,
        )
        
        # Update with optimization results
        new_schedule.status = 'COMPLETED'
        new_schedule.completed_at = timezone.now()
        new_schedule.total_drilling_cost = Decimal(str(results.get('total_drilling_cost', 0)))
        new_schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
        new_schedule.project_end_date = results.get('project_end_date')
        new_schedule.unassigned_wells_count = len(results.get('unassigned_wells', []))
        new_schedule.solver_status = results.get('solver_status')
        new_schedule.solve_time_seconds = results.get('solve_time_seconds')
        new_schedule.save()
        
        # Copy the scope tracking from the original schedule
        # Create ScheduleRig entries for the new schedule
        schedule_rigs = []
        for rig in rigs:
            schedule_rigs.append(ScheduleRig(schedule=new_schedule, rig=rig))
        ScheduleRig.objects.bulk_create(schedule_rigs, ignore_conflicts=True)
        
        # Create ScheduleWell entries for the new schedule
        schedule_wells = []
        for well in wells:
            schedule_wells.append(ScheduleWell(schedule=new_schedule, well=well))
        ScheduleWell.objects.bulk_create(schedule_wells, ignore_conflicts=True)

        # Create assignments with proper sequence order
        assignments_by_rig: dict[str, list] = {}
        for ad in results.get('assignments', []):
            assignments_by_rig.setdefault(ad['rig'], []).append(ad)
        for rig_name, arr in assignments_by_rig.items():
            arr.sort(key=lambda x: x['well_start_date'])
            for i, ad in enumerate(arr, 1):
                ad['calculated_sequence_order'] = i

        # Persist assignments
        # Build quick lookup
        rigs_by_name = {r.name: r for r in Rig.objects.filter(name__in=rigs_seen).order_by('name')}
        wells_by_name = {w.name: w for w in Well.objects.filter(name__in=wells_seen).order_by('name')}
        
        # Create a lookup for actual dates from the provided actuals list AND existing assignments
        actuals_lookup = {}
        
        # First, get existing actual dates from the original schedule
        for assignment in assignments_qs:
            if assignment.actual_start_date or assignment.actual_end_date:
                key = (assignment.well.name, assignment.rig.name)
                actuals_lookup[key] = {
                    'actual_start_date': assignment.actual_start_date,
                    'actual_end_date': assignment.actual_end_date
                }
        
        # Then, override/add with provided actuals (new actuals take precedence)
        for actual in normalized_actuals:
            key = (actual['well'], actual['rig'])
            actuals_lookup[key] = {
                'actual_start_date': actual.get('actual_start_date'),
                'actual_end_date': actual.get('actual_end_date')
            }

        # Build a lookup for original planned dates from the parent schedule
        original_planned_lookup = {}
        for assignment in assignments_qs:
            key = (assignment.well.name, assignment.rig.name)
            # If the assignment already has original planned dates, use those
            # Otherwise, use the current well_start/end_date as the original
            original_planned_lookup[key] = {
                'original_planned_start': assignment.original_planned_start or assignment.well_start_date,
                'original_planned_end': assignment.original_planned_end or assignment.well_end_date
            }
        
        for ad in results.get('assignments', []):
            rig_obj = rigs_by_name.get(ad['rig'])
            well_obj = wells_by_name.get(ad['well'])
            if not rig_obj or not well_obj:
                continue
            
            # Check if this well-rig combination has actual dates
            actual_dates = actuals_lookup.get((ad['well'], ad['rig']), {})
            
            # Get original planned dates (preserve them from parent schedule)
            original_planned = original_planned_lookup.get((ad['well'], ad['rig']), {})
            
            Assignment.objects.create(
                schedule=new_schedule,
                rig=rig_obj,
                well=well_obj,
                well_start_date=ad['well_start_date'],
                well_end_date=ad['well_end_date'],
                original_planned_start=original_planned.get('original_planned_start'),
                original_planned_end=original_planned.get('original_planned_end'),
                actual_start_date=actual_dates.get('actual_start_date'),
                actual_end_date=actual_dates.get('actual_end_date'),
                rtd_check=ad.get('rtd_check', 'OK'),
                well_start_check=ad.get('well_start_check', 'OK'),
                well_end_check=ad.get('well_end_check', 'OK'),
                depth_check=ad.get('depth_check', 'OK'),
                hp_check=ad.get('hp_check', 'OK'),
                bop_check=ad.get('bop_check', 'OK'),
                tds_check=ad.get('tds_check', 'OK'),
                rig_type_check=ad.get('rig_type_check', 'OK'),
                drilling_cost=Decimal(str(ad.get('drilling_cost_inr', ad.get('drilling_cost', 0)))),
                ilm_cost=Decimal(str(ad.get('ilm_cost', 0))),
                ilm_days=Decimal(str(ad.get('ilm_days', 0))),
                sequence_order=ad.get('calculated_sequence_order', 1),
            )

        # Save unassigned wells info
        assigned_well_names = [a['well'] for a in results.get('assignments', [])]
        wells_df = pd.DataFrame([{
            'name': w.name,
            'depth': w.depth,
            'duration': w.duration,
            'rtd': w.rtd,
            'rig_capacity_required_hp': w.rig_capacity_required_hp,
            'bop_stack': w.bop_stack,
            'tds_requirement': w.tds_requirement,
            'priority': w.priority,
            'latitude': float(w.latitude),
            'longitude': float(w.longitude),
            'footprint': w.footprint,
        } for w in Well.objects.filter(name__in=wells_seen).order_by('name')])
        rigs_df = pd.DataFrame([{
            'name': r.name,
            'start_date': r.start_date,
            'end_date': r.end_date,
            'rig_capacity_hp': r.rig_capacity_hp,
            'drilling_capacity_m': r.drilling_capacity_m,
            'bop_stack': r.bop_stack,
            'tds_availability': r.tds_availability,
            'daily_cost_inr': float(r.daily_cost_inr),
            'ilm_cost_fixed': float(r.ilm_cost_fixed),
            'ilm_cost_per_km': float(r.ilm_cost_per_km),
            'ilm_cost_cluster': float(r.ilm_cost_cluster),
            'rig_type': r.rig_type,
        } for r in Rig.objects.filter(name__in=rigs_seen).order_by('name')])

        analyzer = WellRejectionAnalyzer(wells_df, rigs_df, timezone.now().date())
        for well_name in results.get('unassigned_wells', []):
            try:
                well_obj = wells_by_name.get(well_name) or Well.objects.get(name=well_name)
            except Well.DoesNotExist:
                continue
            UnassignedWell.objects.create(
                schedule=new_schedule,
                well=well_obj,
                reason=analyzer.analyze_well_rejection(well_name, assigned_well_names)
            )

        return Response(ScheduleSerializer(new_schedule).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def gantt_data(self, request, pk=None):
        """Get Gantt chart data for a schedule"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        assignments = Assignment.objects.filter(schedule=schedule).order_by('rig__name', 'sequence_order')
        
        tasks = []
        rigs = set()
        wells = set()
        
        for assignment in assignments:
            task = {
                'id': str(assignment.id),
                'name': assignment.well.name,
                'rig': assignment.rig.name,
                'start': assignment.well_start_date.isoformat(),
                'end': assignment.well_end_date.isoformat(),
                'priority': assignment.well.priority,
                'well_type': assignment.well.well_type,
                'depth': assignment.well.depth,
                'asset': assignment.well.asset_id,
                'latitude': float(assignment.well.latitude),
                'longitude': float(assignment.well.longitude),
                'drilling_cost': float(assignment.drilling_cost),
                'ilm_cost': float(assignment.ilm_cost),
                'sequence_order': assignment.sequence_order,
                'original_planned_start': assignment.original_planned_start.isoformat() if assignment.original_planned_start else None,
                'original_planned_end': assignment.original_planned_end.isoformat() if assignment.original_planned_end else None,
                'actual_start_date': assignment.actual_start_date.isoformat() if assignment.actual_start_date else None,
                'actual_end_date': assignment.actual_end_date.isoformat() if assignment.actual_end_date else None,
                'has_actuals': bool(assignment.actual_start_date or assignment.actual_end_date),
                'is_well_deleted': assignment.well.is_deleted,
                'is_rig_deleted': assignment.rig.is_deleted,
                'checks': {
                    'rtd': assignment.rtd_check,
                    'hp': assignment.hp_check,
                    'depth': assignment.depth_check,
                    'bop': assignment.bop_check,
                    'tds': assignment.tds_check,
                    'rig_type': assignment.rig_type_check,
                }
            }
            tasks.append(task)
            rigs.add(assignment.rig.name)
            wells.add(assignment.well.name)
        
        # Calculate date range
        if tasks:
            start_dates = [task['start'] for task in tasks]
            end_dates = [task['end'] for task in tasks]
            date_range = {
                'start': min(start_dates),
                'end': max(end_dates)
            }
        else:
            date_range = {'start': None, 'end': None}
        
        data = {
            'schedule_id': str(schedule.id),
            'tasks': tasks,
            'rigs': sorted(list(rigs)),
            'wells': sorted(list(wells)),
            'date_range': date_range
        }
        
        serializer = GanttDataSerializer(data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reschedule_task(self, request, pk=None):
        """Reschedule a specific task and re-optimize the entire schedule"""
        schedule = self.get_object()
        



        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task_id = request.data.get('task_id')
        new_start_date = request.data.get('new_start_date')
        
        if not task_id or not new_start_date:
            return Response(
                {'error': 'task_id and new_start_date are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            from decimal import Decimal
            
            # Parse the new start date
            if isinstance(new_start_date, str):
                new_start_date = datetime.strptime(new_start_date, '%Y-%m-%d').date()
            
            # Get the assignment to reschedule
            assignment = Assignment.objects.get(id=task_id, schedule=schedule)
            
            # Create a properly named child schedule for the rescheduling
            new_schedule = create_child_schedule(schedule, branch_type="reschedule",
                                                  time_limit_seconds=1800)
            
            # Get wells and rigs from the ORIGINAL SCOPE of the schedule
            from .models import ScheduleRig, ScheduleWell
            
            # Get the originally selected rigs and wells for this schedule
            schedule_rigs = ScheduleRig.objects.filter(schedule=schedule).select_related('rig')
            schedule_wells = ScheduleWell.objects.filter(schedule=schedule).select_related('well')
            
            if not schedule_rigs.exists() or not schedule_wells.exists():
                # Fallback to assignment-based detection if no scope tracking available
                original_assignments = Assignment.objects.filter(schedule=schedule)
                wells_in_schedule = sorted([a.well for a in original_assignments], key=lambda w: w.name)
                rigs_in_schedule = sorted(set([a.rig for a in original_assignments]), key=lambda r: r.name)
            else:
                # Use the originally selected scope
                rigs_in_schedule = sorted([sr.rig for sr in schedule_rigs], key=lambda r: r.name)
                wells_in_schedule = sorted([sw.well for sw in schedule_wells], key=lambda w: w.name)
            
            # Collect ALL existing actual dates as FIXED constraints  
            original_assignments = Assignment.objects.filter(schedule=schedule)
            actuals_for_optimizer = []
            
            # Add ALL existing locked dates as fixed constraints
            for orig_assignment in original_assignments:
                if orig_assignment.actual_start_date or orig_assignment.actual_end_date:
                    actuals_for_optimizer.append({
                        'well': orig_assignment.well.name,
                        'rig': orig_assignment.rig.name,
                        'actual_start_date': orig_assignment.actual_start_date,
                        'actual_end_date': orig_assignment.actual_end_date,
                    })
            
            # Add the NEW reschedule constraint (this overrides any existing actual dates for this well-rig combo)
            reschedule_actual = {
                'well': assignment.well.name,
                'rig': assignment.rig.name,
                'actual_start_date': new_start_date,
                'actual_end_date': None  # Let optimizer calculate end date based on duration
            }
            
            # Remove any existing actual for this well-rig combination and add the new one
            actuals_for_optimizer = [a for a in actuals_for_optimizer 
                                   if not (a['well'] == assignment.well.name and a['rig'] == assignment.rig.name)]
            actuals_for_optimizer.append(reschedule_actual)
            # Sort actuals for deterministic constraint application
            actuals_for_optimizer.sort(key=lambda a: (a['well'], a['rig']))
            
            logger.info(f"Rescheduling with {len(actuals_for_optimizer)} fixed actual constraints")
            
            # Convert rigs and wells to data format for optimizer
            rigs_data = []
            wells_data = []
            
            for rig in rigs_in_schedule:
                rigs_data.append({
                    'name': rig.name,
                    'rig_type': rig.rig_type,
                    'start_date': rig.start_date,
                    'end_date': rig.end_date,
                    'rig_capacity_hp': rig.rig_capacity_hp,
                    'daily_cost_inr': float(rig.daily_cost_inr),
                    'drilling_capacity_m': rig.drilling_capacity_m,
                    'mobilization_time_days': rig.mobilization_time_days,
                    'maintenance_schedule': rig.maintenance_schedule,
                    'crew_availability': rig.crew_availability,
                    'hpht_suitability': rig.hpht_suitability,
                    'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                    'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                    'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                    'bop_stack': rig.bop_stack,
                    'tds_availability': rig.tds_availability,
                })

            for well in wells_in_schedule:
                wells_data.append({
                    'name': well.name,
                    'sn': well.sn,
                    'asset_id': well.asset_id,
                    'well_type': well.well_type,
                    'well_profile': well.well_profile,
                    'depth': well.depth,
                    'rig_capacity_required_hp': well.rig_capacity_required_hp,
                    'drl_days': well.drl_days,
                    'pt_days': well.pt_days,
                    'duration': well.duration,
                    'latitude': float(well.latitude),
                    'longitude': float(well.longitude),
                    'rtd': well.rtd,
                    'bop_stack': well.bop_stack,
                    'tds_requirement': well.tds_requirement,
                    'footprint': well.footprint,
                    'preferred_rig': well.preferred_rig,
                    'expected_potential': well.expected_potential,
                    'priority': well.priority,
                })
            
            # Run optimization with locked actual dates as FIXED constraints
            from .optimization import DrillingScheduler
            
            # Parse Financial Year constraints from the schedule
            from .models import parse_financial_year
            fy_start_date = None
            fy_end_date = None
            if schedule.financial_year:
                try:
                    fy_start_date, fy_end_date = parse_financial_year(schedule.financial_year)
                    logger.info(f"Reschedule using FY constraints: {fy_start_date} to {fy_end_date}")
                except ValueError as e:
                    logger.warning(f"Could not parse schedule financial year: {e}")
            
            scheduler = DrillingScheduler(rigs_data, wells_data, fy_start_date=fy_start_date, fy_end_date=fy_end_date)
            
            if actuals_for_optimizer:
                results = scheduler.solve_with_actuals(actuals_for_optimizer, time_limit_seconds=1800, deterministic=True)
            else:
                results = scheduler.solve(time_limit_seconds=1800, deterministic=True)
                
            success = results and results.get('status') in ['OPTIMAL', 'FEASIBLE']
            
            if success:
                # Update schedule with results
                new_schedule.status = 'COMPLETED'
                new_schedule.completed_at = timezone.now()
                new_schedule.total_drilling_cost = Decimal(str(results.get('total_drilling_cost', 0)))
                new_schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
                new_schedule.project_end_date = results.get('project_end_date')
                new_schedule.unassigned_wells_count = len(results.get('unassigned_wells', []))
                new_schedule.solver_status = results.get('solver_status')
                new_schedule.solve_time_seconds = results.get('solve_time_seconds')
                new_schedule.save()
                
                # Copy scope tracking from original schedule
                from .models import ScheduleRig, ScheduleWell
                
                # Create ScheduleRig entries for the new schedule
                schedule_rigs = []
                for rig in rigs_in_schedule:
                    schedule_rigs.append(ScheduleRig(schedule=new_schedule, rig=rig))
                ScheduleRig.objects.bulk_create(schedule_rigs, ignore_conflicts=True)
                
                # Create ScheduleWell entries for the new schedule  
                schedule_wells = []
                for well in wells_in_schedule:
                    schedule_wells.append(ScheduleWell(schedule=new_schedule, well=well))
                ScheduleWell.objects.bulk_create(schedule_wells, ignore_conflicts=True)
                
                # Create assignments with sequence order and preserve actual dates
                assignments_by_rig = {}
                for ad in results.get('assignments', []):
                    assignments_by_rig.setdefault(ad['rig'], []).append(ad)
                for rig_name, arr in assignments_by_rig.items():
                    arr.sort(key=lambda x: x['well_start_date'])
                    for i, ad in enumerate(arr, 1):
                        ad['calculated_sequence_order'] = i
                
                # Create assignment lookup for actual dates
                actuals_lookup = {}
                for actual in actuals_for_optimizer:
                    key = (actual['well'], actual['rig'])
                    actuals_lookup[key] = {
                        'actual_start_date': actual.get('actual_start_date'),
                        'actual_end_date': actual.get('actual_end_date')
                    }
                
                # Create assignments
                for ad in results.get('assignments', []):
                    try:
                        rig = next(r for r in rigs_in_schedule if r.name == ad['rig'])
                        well = next(w for w in wells_in_schedule if w.name == ad['well'])
                        
                        # Get actual dates for this well-rig combination
                        actual_dates = actuals_lookup.get((ad['well'], ad['rig']), {})
                        
                        Assignment.objects.create(
                            schedule=new_schedule,
                            rig=rig,
                            well=well,
                            well_start_date=ad['well_start_date'],
                            well_end_date=ad['well_end_date'],
                            actual_start_date=actual_dates.get('actual_start_date'),
                            actual_end_date=actual_dates.get('actual_end_date'),
                            rtd_check=ad.get('rtd_check', 'OK'),
                            well_start_check=ad.get('well_start_check', 'OK'),
                            well_end_check=ad.get('well_end_check', 'OK'),
                            depth_check=ad.get('depth_check', 'OK'),
                            hp_check=ad.get('hp_check', 'OK'),
                            bop_check=ad.get('bop_check', 'OK'),
                            tds_check=ad.get('tds_check', 'OK'),
                            rig_type_check=ad.get('rig_type_check', 'OK'),
                            drilling_cost=Decimal(str(ad.get('drilling_cost_inr', ad.get('drilling_cost', 0)))),
                            ilm_cost=Decimal(str(ad.get('ilm_cost', 0))),
                            ilm_days=Decimal(str(ad.get('ilm_days', 0))),
                            sequence_order=ad.get('calculated_sequence_order', 1),
                        )
                    except (StopIteration, Exception) as e:
                        logger.warning(f"Could not create assignment for {ad.get('well')} -> {ad.get('rig')}: {e}")
                        continue
                
                # Return the new schedule data
                return Response({
                    'success': True,
                    'new_schedule_id': str(new_schedule.id),
                    'message': 'Schedule rescheduled successfully with locked dates preserved'
                })
            else:
                new_schedule.status = 'FAILED'
                new_schedule.save()
                return Response(
                    {'error': 'Rescheduling optimization failed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Assignment.DoesNotExist:
            return Response(
                {'error': 'Assignment not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Rescheduling error: {e}")
            return Response(
                {'error': f'Rescheduling failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def debug_assignments(self, request, pk=None):
        """Debug endpoint to see assignment data structure"""
        schedule = self.get_object()
        assignments = Assignment.objects.filter(schedule=schedule)
        
        debug_data = []
        for assignment in assignments:
            debug_data.append({
                'id': str(assignment.id),
                'id_type': type(assignment.id).__name__,
                'well_name': assignment.well.name,
                'rig_name': assignment.rig.name,
                'sequence_order': assignment.sequence_order,
            })
        
        return Response({
            'schedule_id': str(schedule.id),
            'assignment_count': len(debug_data),
            'assignments': debug_data
        })
    
    @action(detail=True, methods=['post'])
    def delete_assignment(self, request, pk=None):
        """Delete an assignment and re-optimize with remaining wells/rigs from the original schedule scope"""
        schedule = self.get_object()
        
        logger.info(f"Delete assignment request received for schedule: {schedule.id}")
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request data: {request.data}")
        logger.info(f"Content type: {request.content_type}")
        
        if schedule.status != 'COMPLETED':
            logger.warning(f"Schedule {schedule.id} is not completed, status: {schedule.status}")
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task_id = request.data.get('task_id')
        logger.info(f"Task ID to delete: {task_id}")
        
        if not task_id:
            logger.error("No task_id provided in request")
            return Response(
                {'error': 'task_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            
            # Get the assignment to delete
            logger.info(f"Looking for assignment with ID: {task_id} in schedule: {schedule.id}")
            assignment_to_delete = Assignment.objects.get(id=task_id, schedule=schedule)
            logger.info(f"Found assignment to delete: {assignment_to_delete.well.name} on {assignment_to_delete.rig.name}")
            
            # Create a properly named child schedule for deleting a well
            new_schedule = create_child_schedule(
                schedule, 
                branch_type="delete_well", 
                custom_suffix=f"- {assignment_to_delete.well.name}"
            )
            logger.info(f"Created new schedule: {new_schedule.id}")
            
            # REVISED APPROACH: Use ORIGINAL SCHEDULE SCOPE (wells and rigs that were selected for the parent schedule)
            # Get the well that was deleted
            deleted_well = assignment_to_delete.well
            logger.info(f"Deleted well: {deleted_well.name}")
            
            # Get wells and rigs from the ORIGINAL SCOPE of the schedule
            from .models import ScheduleRig, ScheduleWell
            
            # Get the originally selected rigs and wells for this schedule
            schedule_rigs = ScheduleRig.objects.filter(schedule=schedule).select_related('rig')
            schedule_wells = ScheduleWell.objects.filter(schedule=schedule).select_related('well')
            
            if not schedule_rigs.exists() or not schedule_wells.exists():
                # Fallback to assignment-based detection if no scope tracking available
                logger.info("No ScheduleRig/ScheduleWell records found, using assignment-based detection")
                original_assignments = Assignment.objects.filter(schedule=schedule)
                wells_in_schedule = [a.well for a in original_assignments]
                rigs_in_schedule = sorted(set([a.rig for a in original_assignments]), key=lambda r: r.name)
                
                # Also include unassigned wells from the original schedule
                from .models import UnassignedWell
                unassigned_wells = UnassignedWell.objects.filter(schedule=schedule).select_related('well')
                unassigned_well_objects = [uw.well for uw in unassigned_wells]
                
                # Combine assigned and unassigned wells
                all_wells_in_scope = set(wells_in_schedule + unassigned_well_objects)
            else:
                # Use the originally selected scope
                logger.info(f"Found {schedule_rigs.count()} rigs and {schedule_wells.count()} wells in original scope")
                rigs_in_schedule = [sr.rig for sr in schedule_rigs]
                all_wells_in_scope = set([sw.well for sw in schedule_wells])
            
            # Remove the deleted well from the scope
            all_wells_in_scope.discard(deleted_well)
            
            # Convert to queryset for optimization
            well_ids = [w.id for w in all_wells_in_scope]
            rig_ids = [r.id for r in rigs_in_schedule]
            
            all_wells_except_deleted = Well.objects.filter(id__in=well_ids).order_by('name')
            all_rigs = Rig.objects.filter(id__in=rig_ids).order_by('name')
            
            logger.info(f"Re-optimizing with {all_wells_except_deleted.count()} wells (from original scope minus deleted) and {all_rigs.count()} rigs (from original scope)")
            
            # Let the optimizer determine the base date automatically based on rig availability
            # This matches the behavior of normal optimization runs
            base_start_date = None  # Let DrillingScheduler use earliest rig start date
            logger.info(f"Using automatic base start date (will be determined by optimizer based on rig availability)")
            
            # Run optimization with fresh database state (all wells except deleted + all rigs)
            success = self._run_optimization_with_constraints(
                new_schedule, all_rigs, all_wells_except_deleted, {}, base_start_date, original_schedule=schedule
            )
            
            if success:
                # Refresh the schedule data and return it for frontend update
                new_schedule.refresh_from_db()
                serializer = ScheduleSerializer(new_schedule)
                return Response({
                    'success': True,
                    'schedule': serializer.data,
                    'new_schedule_id': str(new_schedule.id),
                    'deleted_well_id': str(deleted_well.id),
                    'deleted_well_name': deleted_well.name,
                    'message': f'Assignment for {assignment_to_delete.well.name} deleted and remaining wells from original scope re-optimized'
                })
            else:
                new_schedule.status = 'FAILED'
                new_schedule.save()
                return Response(
                    {'error': 'Re-optimization after deletion failed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Assignment.DoesNotExist:
            logger.error(f"Assignment with ID {task_id} not found in schedule {schedule.id}")
            return Response(
                {'error': 'Assignment not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError as e:
            logger.error(f"Invalid task_id format: {task_id}, error: {e}")
            return Response(
                {'error': f'Invalid task ID format: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Deletion error: {e}")
            logger.error(f"Error type: {type(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Deletion failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def add_assignment(self, request, pk=None):
        """Add an unscheduled well to the current schedule and re-optimize"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        well_id = request.data.get('well_id')
        
        if not well_id:
            return Response(
                {'error': 'well_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            
            # Get the well to add
            well_to_add = Well.objects.get(id=well_id)
            
            # Check if the well is already scheduled in this schedule
            existing_assignment = Assignment.objects.filter(schedule=schedule, well=well_to_add).first()
            if existing_assignment:
                return Response(
                    {'error': f'Well {well_to_add.name} is already scheduled in this schedule'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create a properly named child schedule for adding a well
            new_schedule = create_child_schedule(
                schedule, 
                branch_type="add_well", 
                custom_suffix=f"+ {well_to_add.name}"
            )
            
            # Copy scope tracking from original schedule and add the new well
            from .models import ScheduleRig, ScheduleWell
            
            # Copy original schedule's rigs
            original_schedule_rigs = ScheduleRig.objects.filter(schedule=schedule)
            for orig_rig in original_schedule_rigs:
                ScheduleRig.objects.create(schedule=new_schedule, rig=orig_rig.rig)
                
            # Copy original schedule's wells and add the new one
            original_schedule_wells = ScheduleWell.objects.filter(schedule=schedule)
            for orig_well in original_schedule_wells:
                ScheduleWell.objects.create(schedule=new_schedule, well=orig_well.well)
            # Add the new well
            ScheduleWell.objects.create(schedule=new_schedule, well=well_to_add)
            
            # Get all wells from current schedule plus the new well
            original_assignments = Assignment.objects.filter(schedule=schedule)
            
            # Extract locked wells (those with actual dates) to preserve their positions
            locked_wells_data = []
            for assignment in original_assignments:
                if assignment.actual_start_date or assignment.actual_end_date:
                    locked_wells_data.append({
                        'well': assignment.well.name,
                        'rig': assignment.rig.name,
                        'actual_start_date': assignment.actual_start_date.isoformat() if assignment.actual_start_date else None,
                        'actual_end_date': assignment.actual_end_date.isoformat() if assignment.actual_end_date else None,
                    })
            
            logger.info(f"Found {len(locked_wells_data)} locked wells with actual dates that will be preserved")
            for locked in locked_wells_data:
                logger.info(f"  - Locked: {locked['well']} on {locked['rig']} ({locked['actual_start_date']} to {locked['actual_end_date']})")
            
            # Get well IDs from current assignments plus the new well
            well_ids_current = [a.well.id for a in original_assignments]
            well_ids_current.append(well_to_add.id)
            rig_ids_in_schedule = list(set([a.rig.id for a in original_assignments]))
            
            # Create QuerySets for all wells (current + new) and rigs (ordered for determinism)
            wells_in_schedule = Well.objects.filter(id__in=well_ids_current).order_by('name')
            rigs_in_schedule = Rig.objects.filter(id__in=rig_ids_in_schedule).order_by('name')
            
            logger.info(f"Re-optimizing with {wells_in_schedule.count()} wells (added {well_to_add.name}) and {rigs_in_schedule.count()} rigs")
            
            # Use the earliest start date from the original schedule
            if original_assignments.exists():
                earliest_start_date = min(a.well_start_date for a in original_assignments)
                base_start_date = earliest_start_date
                logger.info(f"Using base start date from original schedule: {base_start_date}")
            else:
                # Let the optimizer determine base date from rig availability
                base_start_date = None
                logger.info(f"Using automatic base start date determination")
            
            # Run optimization with the added well and locked constraints
            success = self._run_optimization_with_constraints(
                new_schedule, rigs_in_schedule, wells_in_schedule, locked_wells_data, base_start_date, original_schedule=schedule
            )
            
            if success:
                return Response({
                    'success': True,
                    'new_schedule_id': str(new_schedule.id),
                    'message': f'Well {well_to_add.name} added and schedule re-optimized'
                })
            else:
                new_schedule.status = 'FAILED'
                new_schedule.save()
                return Response(
                    {'error': 'Re-optimization after adding well failed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Well.DoesNotExist:
            return Response(
                {'error': 'Well not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Add assignment error: {e}")
            return Response(
                {'error': f'Adding well failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _run_optimization_with_constraints(self, schedule, rigs, wells, fixed_assignments=None, base_start_date=None, original_schedule=None):
        """Run optimization with optional fixed assignments and custom base start date"""
        try:
            from .optimization import DrillingScheduler
            from datetime import date, datetime
            from decimal import Decimal
            
            # Use provided base_start_date or let optimizer determine it automatically
            # The DrillingScheduler will automatically use earliest rig start date when None is passed
            
            logger.info(f"Starting optimization with {wells.count() if hasattr(wells, 'count') else len(wells)} wells and {rigs.count() if hasattr(rigs, 'count') else len(rigs)} rigs")
            if base_start_date:
                logger.info(f"Using provided base start date: {base_start_date}")
            else:
                logger.info(f"Using automatic base start date determination from rig availability")
            
            # Convert QuerySets to dictionaries for the optimizer (ordered for determinism)
            rigs_data = list(rigs.order_by('name').values()) if hasattr(rigs, 'order_by') else rigs
            wells_data = list(wells.order_by('name').values()) if hasattr(wells, 'order_by') else wells
            
            # DEBUGGING: Log exact wells being optimized
            well_names = [w['name'] for w in wells_data]
            logger.info(f"Wells being optimized: {sorted(well_names)}")
            
            # DEBUGGING: Check for any wells that might have problematic constraints
            problematic_wells = []
            for well in wells_data:
                if not well.get('rtd') or not well.get('duration') or well.get('duration', 0) <= 0:
                    problematic_wells.append(f"{well['name']}: rtd={well.get('rtd')}, duration={well.get('duration')}")
            
            if problematic_wells:
                logger.warning(f"Wells with potential constraint issues: {problematic_wells}")
            
            # Parse Financial Year constraints from the schedule
            from .models import parse_financial_year
            fy_start_date = None
            fy_end_date = None
            if schedule.financial_year:
                try:
                    fy_start_date, fy_end_date = parse_financial_year(schedule.financial_year)
                    logger.info(f"_run_optimization_with_constraints using FY constraints: {fy_start_date} to {fy_end_date}")
                except ValueError as e:
                    logger.warning(f"Could not parse schedule financial year: {e}")
            
            # Initialize scheduler with the correct base start date and FY constraints
            scheduler = DrillingScheduler(rigs_data, wells_data, base_start_date=base_start_date, fy_start_date=fy_start_date, fy_end_date=fy_end_date)
            
            # solve() and solve_with_actuals() run the full pipeline internally
            # (preprocess → setup_variables → add_constraints → add_ilm → set_objective → decision strategy)
            # No need to call them separately.
            
            # If we have locked wells (fixed_assignments), use solve_with_actuals to respect them
            if fixed_assignments and len(fixed_assignments) > 0:
                logger.info(f"Running optimization WITH {len(fixed_assignments)} locked wells as fixed constraints")
                results = scheduler.solve_with_actuals(fixed_assignments, time_limit_seconds=1800, deterministic=True)
            else:
                # Use standard solver for unconstrained optimization
                logger.info("Running standard optimization with 1800s time limit (30 minutes)")
                results = scheduler.solve(time_limit_seconds=1800, deterministic=True)
            
            if results and results.get('status') in ['OPTIMAL', 'FEASIBLE']:
                logger.info(f"Optimization successful: {len(results['assignments'])} assignments created")
                
                # Clear any existing assignments and unscheduled wells for this schedule to prevent duplicates
                Assignment.objects.filter(schedule=schedule).delete()
                UnassignedWell.objects.filter(schedule=schedule).delete()
                
                # Validate assignments for overlaps before creating them
                assignment_data_list = results['assignments']
                rig_schedules = {}  # Track each rig's schedule to prevent overlaps
                
                # Sort assignments by rig and start date to check for overlaps
                assignment_data_list.sort(key=lambda x: (x['rig'], x['well_start_date']))
                
                valid_assignments = []
                for assignment_data in assignment_data_list:
                    rig_name = assignment_data['rig']
                    start_date = assignment_data['well_start_date']
                    end_date = assignment_data['well_end_date']
                    
                    # Check for overlaps with existing assignments for this rig
                    has_overlap = False
                    if rig_name in rig_schedules:
                        for existing_start, existing_end in rig_schedules[rig_name]:
                            # Check if dates overlap
                            if not (end_date <= existing_start or start_date >= existing_end):
                                logger.warning(f"Overlap detected for rig {rig_name}: {assignment_data['well']} ({start_date}-{end_date}) overlaps with existing assignment ({existing_start}-{existing_end})")
                                has_overlap = True
                                break
                    
                    if not has_overlap:
                        valid_assignments.append(assignment_data)
                        # Track this rig's schedule
                        if rig_name not in rig_schedules:
                            rig_schedules[rig_name] = []
                        rig_schedules[rig_name].append((start_date, end_date))
                
                logger.info(f"Creating {len(valid_assignments)} valid assignments (removed {len(assignment_data_list) - len(valid_assignments)} overlapping assignments)")
                
                # Calculate proper sequence order for valid assignments
                assignments_by_rig = {}
                for assignment_data in valid_assignments:
                    rig_name = assignment_data['rig']
                    if rig_name not in assignments_by_rig:
                        assignments_by_rig[rig_name] = []
                    assignments_by_rig[rig_name].append(assignment_data)
                
                # Sort each rig's assignments by start date and assign sequence numbers
                for rig_name, rig_assignments in assignments_by_rig.items():
                    # Sort by start date
                    rig_assignments.sort(key=lambda x: x['well_start_date'])
                    # Assign sequence numbers starting from 1
                    for i, assignment_data in enumerate(rig_assignments, 1):
                        assignment_data['calculated_sequence_order'] = i
                
                # Prepare actual dates and original planned dates lookup if original schedule is provided
                actual_dates_lookup = {}
                original_planned_lookup = {}
                if original_schedule:
                    original_assignments = Assignment.objects.filter(schedule=original_schedule)
                    for orig_assignment in original_assignments:
                        key = (orig_assignment.well.name, orig_assignment.rig.name)
                        
                        # Preserve actual dates if they exist
                        if orig_assignment.actual_start_date or orig_assignment.actual_end_date:
                            actual_dates_lookup[key] = {
                                'actual_start_date': orig_assignment.actual_start_date,
                                'actual_end_date': orig_assignment.actual_end_date
                            }
                        
                        # Preserve original planned dates if they exist, otherwise use current planned dates
                        original_planned_lookup[key] = {
                            'original_planned_start': orig_assignment.original_planned_start or orig_assignment.well_start_date,
                            'original_planned_end': orig_assignment.original_planned_end or orig_assignment.well_end_date
                        }
                
                # Create assignments for the new schedule with proper sequence order
                for assignment_data in valid_assignments:
                    well = Well.objects.get(name=assignment_data['well'])
                    rig = Rig.objects.get(name=assignment_data['rig'])
                    
                    # Get actual dates if they exist for this well-rig combination
                    actual_dates = actual_dates_lookup.get((well.name, rig.name), {})
                    
                    # Get original planned dates if they exist
                    original_planned = original_planned_lookup.get((well.name, rig.name), {})
                    
                    Assignment.objects.create(
                        schedule=schedule,
                        rig=rig,
                        well=well,
                        well_start_date=assignment_data['well_start_date'],
                        well_end_date=assignment_data['well_end_date'],
                        original_planned_start=original_planned.get('original_planned_start'),
                        original_planned_end=original_planned.get('original_planned_end'),
                        actual_start_date=actual_dates.get('actual_start_date'),
                        actual_end_date=actual_dates.get('actual_end_date'),
                        rtd_check=assignment_data.get('rtd_check', 'OK'),
                        well_start_check=assignment_data.get('well_start_check', 'OK'),
                        well_end_check=assignment_data.get('well_end_check', 'OK'),
                        depth_check=assignment_data.get('depth_check', 'OK'),
                        hp_check=assignment_data.get('hp_check', 'OK'),
                        bop_check=assignment_data.get('bop_check', 'OK'),
                        tds_check=assignment_data.get('tds_check', 'OK'),
                        rig_type_check=assignment_data.get('rig_type_check', 'OK'),
                        drilling_cost=Decimal(str(assignment_data.get('drilling_cost_inr', assignment_data.get('drilling_cost', 0)))),
                        ilm_cost=Decimal(str(assignment_data.get('ilm_cost', 0))),
                        ilm_days=Decimal(str(assignment_data.get('ilm_days', 0))),
                        sequence_order=assignment_data.get('calculated_sequence_order', 1)
                    )
                
                # Create unassigned wells with detailed rejection analysis
                assigned_well_names = [assignment['well'] for assignment in valid_assignments]
                
                # Get all wells and rigs for analysis
                all_wells = Well.objects.all()
                all_rigs = Rig.objects.all()
                
                # Prepare data for analyzer
                wells_data = []
                rigs_data = []
                
                for well in all_wells:
                    wells_data.append({
                        'name': well.name,
                        'depth': well.depth,
                        'duration': well.duration,
                        'rtd': well.rtd,
                        'rig_capacity_required_hp': well.rig_capacity_required_hp,
                        'bop_stack': well.bop_stack,
                        'tds_requirement': well.tds_requirement,
                        'priority': well.priority,
                        'latitude': float(well.latitude) if well.latitude else 0.0,
                        'longitude': float(well.longitude) if well.longitude else 0.0,
                        'footprint': well.footprint,
                    })
                
                for rig in all_rigs:
                    rigs_data.append({
                        'name': rig.name,
                        'start_date': rig.start_date,
                        'end_date': rig.end_date,
                        'rig_capacity_hp': rig.rig_capacity_hp,
                        'drilling_capacity_m': rig.drilling_capacity_m,
                        'bop_stack': rig.bop_stack,
                        'tds_availability': rig.tds_availability,
                        'daily_cost_inr': float(rig.daily_cost_inr),
                        'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                        'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                        'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                        'rig_type': rig.rig_type,
                    })
                
                # Convert to DataFrames for analysis
                wells_df = pd.DataFrame(wells_data)
                rigs_df = pd.DataFrame(rigs_data)
                
                # Initialize analyzer
                analyzer = WellRejectionAnalyzer(wells_df, rigs_df, timezone.now().date())
                
                for unassigned_well_name in results.get('unassigned_wells', []):
                    # Handle both simple string format and dict format
                    if isinstance(unassigned_well_name, dict):
                        well_name = unassigned_well_name.get('well', unassigned_well_name.get('name'))
                        # If reason is already provided in dict, use it, otherwise analyze
                        existing_reason = unassigned_well_name.get('reason', '')
                        if existing_reason and existing_reason != 'Not assigned':
                            reason = existing_reason
                        else:
                            reason = analyzer.analyze_well_rejection(well_name, assigned_well_names)
                    else:
                        well_name = str(unassigned_well_name)
                        reason = analyzer.analyze_well_rejection(well_name, assigned_well_names)
                    
                    well = Well.objects.get(name=well_name)
                    UnassignedWell.objects.create(
                        schedule=schedule,
                        well=well,
                        reason=reason
                    )
                
                # Update schedule
                schedule.status = 'COMPLETED'
                schedule.total_drilling_cost = Decimal(str(results['total_drilling_cost']))
                schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
                schedule.project_end_date = results.get('project_end_date')
                schedule.unassigned_wells_count = results.get('unassigned_wells_count', 0)
                schedule.solver_status = results.get('solver_status', 'UNKNOWN')
                schedule.solve_time_seconds = results.get('solve_time_seconds', 0)
                schedule.completed_at = datetime.now()
                schedule.save()
                
                return True
            else:
                return False
                
        except Exception as e:
            logger.error(f"Optimization error: {e}")
            return False
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Get schedule statistics"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        assignments = Assignment.objects.filter(schedule=schedule)
        unassigned = UnassignedWell.objects.filter(schedule=schedule)
        
        # Calculate statistics
        total_drilling_cost_cr = float(schedule.total_drilling_cost or 0) / 10000000
        total_ilm_cost_cr = float(schedule.total_ilm_cost or 0) / 10000000
        total_cost_cr = total_drilling_cost_cr + total_ilm_cost_cr
        
        # Project duration
        if assignments.exists():
            project_start = assignments.aggregate(min_start=models.Min('well_start_date'))['min_start']
            project_end = assignments.aggregate(max_end=models.Max('well_end_date'))['max_end']
            project_duration_days = (project_end - project_start).days + 1 if project_start and project_end else 0
        else:
            project_duration_days = 0
        
        # Priority breakdown
        priority_counts = {}
        for priority in ['HIGH', 'MEDIUM', 'LOW']:
            assigned = assignments.filter(well__priority=priority).count()
            unassigned_count = unassigned.filter(well__priority=priority).count()
            priority_counts[priority] = {
                'assigned': assigned,
                'unassigned': unassigned_count,
                'total': assigned + unassigned_count
            }
        
        # Rig utilization
        rig_utilizations = []
        for rig in Rig.objects.filter(assignments__schedule=schedule).distinct():
            rig_assignments = assignments.filter(rig=rig)
            total_assigned_days = sum(a.well.duration for a in rig_assignments)
            total_available_days = rig.duration_days
            utilization_pct = (total_assigned_days / total_available_days * 100) if total_available_days > 0 else 0
            
            rig_utilizations.append({
                'rig_name': rig.name,
                'total_available_days': total_available_days,
                'total_assigned_days': total_assigned_days,
                'utilization_percentage': round(utilization_pct, 2),
                'wells_assigned': rig_assignments.count(),
                'idle_days': max(0, total_available_days - total_assigned_days),
                'drilling_cost': float(sum(a.drilling_cost for a in rig_assignments)),
                'ilm_cost': float(sum(a.ilm_cost for a in rig_assignments))
            })
        
        stats_data = {
            'total_assignments': assignments.count(),
            'total_unassigned': unassigned.count(),
            'total_drilling_cost_cr': total_drilling_cost_cr,
            'total_ilm_cost_cr': total_ilm_cost_cr,
            'total_cost_cr': total_cost_cr,
            'project_duration_days': project_duration_days,
            'rig_utilization': rig_utilizations,
            'priority_breakdown': priority_counts,
            'solver_status': schedule.solver_status,
            'optimality_gap_percent': schedule.optimality_gap_percent,
            'schedule_hash': schedule.schedule_hash,
        }
        
        serializer = ScheduleStatsSerializer(stats_data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def unscheduled_resources(self, request, pk=None):
        """Get wells and rigs that were selected for the schedule but not assigned/used"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get wells that were SELECTED for this schedule (input)
        selected_well_ids = schedule.selected_wells.values_list('well_id', flat=True)
        
        # Get wells that were ASSIGNED in this schedule (output)
        assigned_well_ids = Assignment.objects.filter(schedule=schedule).values_list('well_id', flat=True)
        
        # Get unassigned wells WITH rejection reasons
        unassigned_wells_qs = schedule.unassigned_wells.select_related('well').order_by('well__name')
        
        # Get rigs that were SELECTED for this schedule (input)
        selected_rig_ids = schedule.selected_rigs.values_list('rig_id', flat=True)
        
        # Get rigs that were USED in this schedule (output)
        used_rig_ids = Assignment.objects.filter(schedule=schedule).values_list('rig_id', flat=True).distinct()
        
        # Get unused rigs (selected but not used)
        unused_rig_ids = set(selected_rig_ids) - set(used_rig_ids)
        unused_rigs = Rig.objects.filter(id__in=unused_rig_ids).order_by('name')
        
        # Serialize unassigned wells with rejection reasons
        unscheduled_wells_data = []
        for unassigned in unassigned_wells_qs:
            well = unassigned.well
            unscheduled_wells_data.append({
                'id': str(well.id),
                'name': well.name,
                'asset_id': well.asset_id,
                'priority': well.priority,
                'depth': well.depth,
                'duration': well.duration,
                'well_type': well.well_type,
                'rtd': well.rtd.strftime('%Y-%m-%d'),
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'reason': unassigned.reason  # Include rejection reason
            })
        
        # Serialize unused rigs with reason
        unused_rigs_data = []
        for rig in unused_rigs:
            unused_rigs_data.append({
                'id': str(rig.id),
                'name': rig.name,
                'rig_type': rig.rig_type,
                'capacity_hp': rig.rig_capacity_hp,
                'daily_cost': float(rig.daily_cost_inr),
                'drilling_capacity': rig.drilling_capacity_m,
                'start_date': rig.start_date.strftime('%Y-%m-%d'),
                'end_date': rig.end_date.strftime('%Y-%m-%d'),
                'duration_days': rig.duration_days,
                'reason': 'Not selected by optimization algorithm (higher cost or lower efficiency)'  # Default reason for rigs
            })
        
        return Response({
            'unscheduled_wells': unscheduled_wells_data,
            'unused_rigs': unused_rigs_data,
            'summary': {
                'total_wells_selected': len(selected_well_ids),
                'wells_assigned': len(assigned_well_ids),
                'unscheduled_wells_count': len(unscheduled_wells_data),
                'total_rigs_selected': len(selected_rig_ids),
                'rigs_used': len(set(used_rig_ids)),
                'unused_rigs_count': len(unused_rigs_data)
            }
        })
    
    @action(detail=True, methods=['get'])
    def rig_statistics(self, request, pk=None):
        """Get detailed rig statistics including rig-months, well type split, and meterage"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        assignments = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
        
        # Group assignments by rig
        rig_stats = {}
        for assignment in assignments:
            rig_name = assignment.rig.name
            if rig_name not in rig_stats:
                rig_stats[rig_name] = {
                    'rig_name': rig_name,
                    'rig_type': assignment.rig.rig_type,
                    'total_days': 0,
                    'dev_wells': 0,
                    'exp_wells': 0,
                    'dev_days': 0,
                    'exp_days': 0,
                    'dev_meterage': 0,
                    'exp_meterage': 0,
                }
            
            # Calculate operating days for this assignment
            if assignment.well_start_date and assignment.well_end_date:
                days = (assignment.well_end_date - assignment.well_start_date).days + 1
            else:
                days = assignment.well.duration
            
            rig_stats[rig_name]['total_days'] += days
            
            # Classify by well type
            is_exploration = assignment.well.well_type.upper() in ['EXP', 'EXPLORATION']
            
            if is_exploration:
                rig_stats[rig_name]['exp_wells'] += 1
                rig_stats[rig_name]['exp_days'] += days
                rig_stats[rig_name]['exp_meterage'] += assignment.well.depth
            else:
                rig_stats[rig_name]['dev_wells'] += 1
                rig_stats[rig_name]['dev_days'] += days
                rig_stats[rig_name]['dev_meterage'] += assignment.well.depth
        
        # Calculate rig-months and format data
        rig_statistics = []
        for rig_name, stats in rig_stats.items():
            # Calculate rig-months (total_days / 30.41)
            total_rig_months = stats['total_days'] / 30.41
            dev_rig_months = stats['dev_days'] / 30.41
            exp_rig_months = stats['exp_days'] / 30.41
            
            rig_statistics.append({
                'rig_name': stats['rig_name'],
                'rig_type': stats['rig_type'],
                'total_rig_months': round(total_rig_months, 2),
                'total_days': stats['total_days'],
                'development': {
                    'well_count': stats['dev_wells'],
                    'rig_months': round(dev_rig_months, 2),
                    'days': stats['dev_days'],
                    'meterage': stats['dev_meterage']
                },
                'exploration': {
                    'well_count': stats['exp_wells'],
                    'rig_months': round(exp_rig_months, 2),
                    'days': stats['exp_days'],
                    'meterage': stats['exp_meterage']
                },
                'total_meterage': stats['dev_meterage'] + stats['exp_meterage']
            })
        
        # Sort by rig name
        rig_statistics.sort(key=lambda x: x['rig_name'])
        
        return Response({
            'rig_statistics': rig_statistics,
            'summary': {
                'total_rigs': len(rig_statistics),
                'total_rig_months': round(sum(r['total_rig_months'] for r in rig_statistics), 2),
                'total_meterage': sum(r['total_meterage'] for r in rig_statistics)
            }
        })
    
    @action(detail=True, methods=['get'])
    def allocated_resources(self, request, pk=None):
        """Get wells and rigs allocated to the selected schedule"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get wells assigned to this schedule
        assigned_wells = Assignment.objects.filter(schedule=schedule).select_related('well').order_by('well__name')
        
        # Get rigs used in this schedule
        used_rigs = Assignment.objects.filter(schedule=schedule).select_related('rig').values_list('rig', flat=True).distinct()
        allocated_rigs = Rig.objects.filter(id__in=used_rigs).order_by('name')
        
        # Serialize allocated wells
        allocated_wells_data = []
        for assignment in assigned_wells:
            well = assignment.well
            allocated_wells_data.append({
                'id': str(well.id),
                'name': well.name,
                'asset_id': well.asset_id,
                'priority': well.priority,
                'depth': well.depth,
                'duration': well.duration,
                'well_type': well.well_type,
                'rtd': well.rtd.strftime('%Y-%m-%d'),
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'assigned_start_date': assignment.well_start_date.strftime('%Y-%m-%d'),
                'assigned_end_date': assignment.well_end_date.strftime('%Y-%m-%d'),
                'assigned_rig': assignment.rig.name
            })
        
        # Serialize allocated rigs
        allocated_rigs_data = []
        for rig in allocated_rigs:
            allocated_rigs_data.append({
                'id': str(rig.id),
                'name': rig.name,
                'rig_type': rig.rig_type,
                'capacity_hp': rig.rig_capacity_hp,
                'daily_cost': float(rig.daily_cost_inr),
                'drilling_capacity': rig.drilling_capacity_m,
                'start_date': rig.start_date.strftime('%Y-%m-%d'),
                'end_date': rig.end_date.strftime('%Y-%m-%d'),
                'duration_days': rig.duration_days
            })
        
        return Response({
            'allocated_wells': allocated_wells_data,
            'allocated_rigs': allocated_rigs_data,
            'summary': {
                'total_wells_in_schedule': len(allocated_wells_data),
                'total_rigs_in_schedule': len(allocated_rigs_data),
                'schedule_name': schedule.name,
                'schedule_status': schedule.status
            }
        })
    
    @action(detail=True, methods=['post'])
    def update_assignment(self, request, pk=None):
        """Update assignment dates (for drag-and-drop functionality)"""
        schedule = self.get_object()
        serializer = AssignmentUpdateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        
        try:
            assignment = Assignment.objects.get(
                id=validated_data['assignment_id'],
                schedule=schedule
            )
            
            # Update start date and calculate end date
            new_start_date = validated_data['new_start_date']
            new_end_date = new_start_date + timedelta(days=assignment.well.duration - 1)
            
            assignment.well_start_date = new_start_date
            assignment.well_end_date = new_end_date
            
            # Update rig if provided
            if 'new_rig_id' in validated_data:
                new_rig = get_object_or_404(Rig, id=validated_data['new_rig_id'])
                assignment.rig = new_rig
            
            # Re-validate constraints
            checks = self._validate_assignment(assignment)
            assignment.rtd_check = checks.get('rtd_check', 'OK')
            assignment.well_start_check = checks.get('well_start_check', 'OK')
            assignment.well_end_check = checks.get('well_end_check', 'OK')
            assignment.depth_check = checks.get('depth_check', 'OK')
            assignment.hp_check = checks.get('hp_check', 'OK')
            assignment.bop_check = checks.get('bop_check', 'OK')
            assignment.tds_check = checks.get('tds_check', 'OK')
            assignment.rig_type_check = checks.get('rig_type_check', 'OK')
            
            assignment.save()
            
            return Response(AssignmentSerializer(assignment).data)
            
        except Assignment.DoesNotExist:
            return Response(
                {'error': 'Assignment not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Update failed: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export schedules data as CSV"""
        try:
            queryset = self.get_queryset()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="schedules_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = [
                'ID', 'Name', 'Status', 'Total Drilling Cost', 'Total ILM Cost',
                'Project End Date', 'Unassigned Wells Count', 'Solver Status',
                'Solve Time (seconds)', 'Created At', 'Updated At', 'Completed At'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for schedule in queryset:
                writer.writerow([
                    str(schedule.id),
                    schedule.name,
                    schedule.status,
                    schedule.total_drilling_cost or 0,
                    schedule.total_ilm_cost or 0,
                    schedule.project_end_date.strftime('%Y-%m-%d') if schedule.project_end_date else '',
                    schedule.unassigned_wells_count or 0,
                    schedule.solver_status or '',
                    schedule.solve_time_seconds or 0,
                    schedule.created_at.strftime('%Y-%m-%d %H:%M:%S') if schedule.created_at else '',
                    schedule.updated_at.strftime('%Y-%m-%d %H:%M:%S') if schedule.updated_at else '',
                    schedule.completed_at.strftime('%Y-%m-%d %H:%M:%S') if schedule.completed_at else ''
                ])
            
            return response
            
        except Exception as e:
            return JsonResponse(
                {'error': f'Failed to export CSV: {str(e)}'}, 
                status=500
            )
    
    def _validate_assignment(self, assignment):
        """Validate an assignment against constraints"""
        well = assignment.well
        rig = assignment.rig
        well_start_date = assignment.well_start_date
        well_end_date = assignment.well_end_date
        
        checks = {}
        
        # RTD check
        checks['rtd_check'] = 'OK' if well_start_date >= well.rtd else 'NOK'
        
        # Well start date check
        checks['well_start_check'] = 'OK' if (well_start_date - well.rtd).days >= 0 else 'NOK'
        
        # Well end date check
        checks['well_end_check'] = 'OK' if well_end_date <= rig.end_date else 'NOK'
        
        # Depth check
        checks['depth_check'] = 'OK' if rig.drilling_capacity_m >= well.depth else 'NOK'
        
        # HP check
        checks['hp_check'] = 'OK' if rig.rig_capacity_hp >= well.rig_capacity_required_hp else 'NOK'
        
        # BOP check
        checks['bop_check'] = 'OK' if rig.bop_stack >= well.bop_stack else 'NOK'
        
        # TDS check
        if well.tds_requirement == 'Y':
            checks['tds_check'] = 'OK' if rig.tds_availability == 'Y' else 'NOK'
        else:
            checks['tds_check'] = 'OK'
        
        # Rig type check
        checks['rig_type_check'] = 'OK' if well.footprint == rig.rig_type else 'NOK'
        
        return checks


class AssignmentViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing assignments"""
    queryset = Assignment.objects.all()
    serializer_class = AssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = Assignment.objects.all()
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                # Filter assignments by schedule location or well location
                queryset = queryset.filter(
                    models.Q(schedule__location=user_location) |
                    models.Q(well__location=user_location)
                )
        
        # Check if request has query_params (DRF Request) or GET (Django Request)
        if hasattr(self.request, 'query_params'):
            schedule_id = self.request.query_params.get('schedule_id', None)
            rig_id = self.request.query_params.get('rig_id', None)
        else:
            schedule_id = self.request.GET.get('schedule_id', None)
            rig_id = self.request.GET.get('rig_id', None)
        
        if schedule_id:
            queryset = queryset.filter(schedule_id=schedule_id)
        if rig_id:
            queryset = queryset.filter(rig_id=rig_id)
        
        return queryset.order_by('rig__name', 'sequence_order')


@login_required
@csrf_exempt
def export_schedule_csv(request, schedule_id):
    """Export schedule to CSV format"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        assignments = Assignment.objects.filter(schedule=schedule).order_by('rig__name', 'sequence_order')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="schedule_{schedule.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Rig', 'Well', 'RTD', 'RTD Check', 'Well Start Date', 'Well Start Date Check (+ve)',
            'Well End Date', 'Well End Date Check (-ve)', 'Rig Start Date', 'Rig End Date',
            'Well Priority', 'Duration (days)', 'required_depth', 'rig depth capacity',
            'depth check', 'required_hp', 'rig hp capacity', 'HP check', 'required_bop',
            'rig bop capacity', 'BOP check', 'required_tds', 'rig tds availability',
            'TDS check', 'required_rig_type', 'rig type', 'Rig Type check',
            'Latitude', 'Longitude'
        ])
        
        # Write data
        for assignment in assignments:
            writer.writerow([
                assignment.rig.name,
                assignment.well.name,
                assignment.well.rtd,
                assignment.rtd_check,
                assignment.well_start_date,
                assignment.well_start_check,
                assignment.well_end_date,
                assignment.well_end_check,
                assignment.rig.start_date,
                assignment.rig.end_date,
                assignment.well.priority,
                assignment.well.duration,
                assignment.well.depth,
                assignment.rig.drilling_capacity_m,
                assignment.depth_check,
                assignment.well.rig_capacity_required_hp,
                assignment.rig.rig_capacity_hp,
                assignment.hp_check,
                assignment.well.bop_stack,
                assignment.rig.bop_stack,
                assignment.bop_check,
                assignment.well.tds_requirement,
                assignment.rig.tds_availability,
                assignment.tds_check,
                assignment.well.footprint,
                assignment.rig.rig_type,
                assignment.rig_type_check,
                assignment.well.latitude,
                assignment.well.longitude
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def export_schedule_excel(request, schedule_id):
    """Export schedule to Excel format with multiple sheets and formatting"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        assignments = Assignment.objects.filter(schedule=schedule).order_by('rig__name', 'sequence_order')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet if it exists
        if wb.active:
            wb.remove(wb.active)
        
        # Create schedule overview sheet
        ws_overview = wb.create_sheet("Schedule Overview")
        
        # Add title and basic info
        ws_overview['A1'] = f'Schedule: {schedule.name}'
        ws_overview['A1'].font = Font(size=16, bold=True)
        ws_overview['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_overview['A3'] = f'Total Wells: {assignments.count()}'
        ws_overview['A4'] = f'Total Rigs: {assignments.values("rig").distinct().count()}'
        
        # Create detailed assignments sheet
        ws_details = wb.create_sheet("Detailed Assignments")
        
        # Headers
        headers = [
            'Rig', 'Well', 'Asset ID', 'Well Type', 'Priority', 'RTD', 'RTD Check', 
            'Well Start Date', 'Well Start Check', 'Well End Date', 'Well End Check',
            'Rig Start Date', 'Rig End Date', 'Duration (days)', 'Depth (m)', 
            'Required HP', 'Rig HP', 'HP Check', 'Required BOP', 'Rig BOP', 'BOP Check',
            'Required TDS', 'Rig TDS', 'TDS Check', 'Required Type', 'Rig Type', 'Type Check',
            'Latitude', 'Longitude', 'Sequence Order'
        ]
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_num, assignment in enumerate(assignments, 2):
            data = [
                assignment.rig.name,
                assignment.well.name,
                assignment.well.asset_id,
                assignment.well.well_type,
                assignment.well.priority,
                assignment.well.rtd,
                assignment.rtd_check,
                assignment.well_start_date,
                assignment.well_start_check,
                assignment.well_end_date,
                assignment.well_end_check,
                assignment.rig.start_date,
                assignment.rig.end_date,
                assignment.well.duration,
                assignment.well.depth,
                assignment.well.rig_capacity_required_hp,
                assignment.rig.rig_capacity_hp,
                assignment.hp_check,
                assignment.well.bop_stack,
                assignment.rig.bop_stack,
                assignment.bop_check,
                assignment.well.tds_requirement,
                assignment.rig.tds_availability,
                assignment.tds_check,
                assignment.well.footprint,
                assignment.rig.rig_type,
                assignment.rig_type_check,
                assignment.well.latitude,
                assignment.well.longitude,
                assignment.sequence_order
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_details.cell(row=row_num, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Color code based on checks
                if col in [7, 9, 11, 18, 21, 24, 27]:  # Check columns
                    if value == 'PASS':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == 'FAIL':
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Create rig summary sheet
        ws_rigs = wb.create_sheet("Rig Summary")
        
        # Get rig summary data
        rig_data = []
        for rig in assignments.values('rig__name', 'rig__rig_type', 'rig__rig_capacity_hp', 'rig__start_date', 'rig__end_date').distinct():
            rig_assignments = assignments.filter(rig__name=rig['rig__name'])
            rig_data.append([
                rig['rig__name'],
                rig['rig__rig_type'],
                rig['rig__rig_capacity_hp'],
                rig['rig__start_date'],
                rig['rig__end_date'],
                rig_assignments.count(),
                sum([a.well.duration for a in rig_assignments])
            ])
        
        # Rig headers
        rig_headers = ['Rig Name', 'Type', 'Capacity (HP)', 'Start Date', 'End Date', 'Wells Assigned', 'Total Duration (days)']
        
        for col, header in enumerate(rig_headers, 1):
            cell = ws_rigs.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, rig_row in enumerate(rig_data, 2):
            for col, value in enumerate(rig_row, 1):
                ws_rigs.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for ws in [ws_overview, ws_details, ws_rigs]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="schedule_{schedule.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_unified(request):
    """Unified bulk upload endpoint for both rigs and wells"""
    try:
        data_type = request.data.get('data_type')
        file_obj = request.data.get('file')
        
        if not data_type:
            return Response(
                {'error': 'data_type parameter is required (rigs or wells)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not file_obj:
            return Response(
                {'error': 'file parameter is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if data_type == 'rigs':
            return handle_rigs_upload(file_obj)
        elif data_type == 'wells':
            return handle_wells_upload(file_obj)
        else:
            return Response(
                {'error': 'data_type must be either "rigs" or "wells"'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
    except Exception as e:
        return Response(
            {'error': f'Upload failed: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


###############################################################################
# ─── CSV Upload Helper Utilities ─────────────────────────────────────────────
###############################################################################

def _read_upload_file_to_dataframe(file_obj):
    """
    Robustly read an uploaded file (CSV or Excel) into a pandas DataFrame.

    Handles:
      - Multiple encodings (UTF-8 BOM, CP1252, Latin-1, UTF-16)
      - Excel files (.xlsx, .xls)
      - Tab-separated / semicolon-separated files
      - BOM characters
      - Completely empty files
      - Files with only headers
      - Binary / corrupt files

    Returns (df, error_message).  If df is None, error_message explains why.
    """
    try:
        raw_bytes = file_obj.read()
    except Exception as e:
        return None, f"Could not read the uploaded file: {str(e)}"

    if not raw_bytes or len(raw_bytes.strip()) == 0:
        return None, "The uploaded file is empty."

    filename = getattr(file_obj, 'name', '') or ''

    # ── Try Excel first if extension hints at it ──
    if filename.lower().endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(io.BytesIO(raw_bytes), engine='openpyxl')
            if df.empty:
                return None, "The uploaded Excel file contains no data rows."
            return df, None
        except Exception as e:
            return None, f"Failed to read Excel file: {str(e)}. Try saving as CSV (UTF-8) instead."

    # ── CSV with multiple encoding attempts ──
    encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1', 'utf-16', 'iso-8859-1']
    separators = [',', ';', '\t', '|']
    df = None
    last_error = None

    for encoding in encodings:
        for sep in separators:
            try:
                candidate = pd.read_csv(
                    io.BytesIO(raw_bytes),
                    encoding=encoding,
                    sep=sep,
                    skipinitialspace=True,
                    on_bad_lines='warn',
                    dtype=str,            # read everything as string first
                )
                # A valid parse should have >1 column (unless single-column CSV)
                if candidate is not None and len(candidate.columns) >= 1:
                    # If only 1 column and the header contains the separator, wrong sep
                    if len(candidate.columns) == 1 and any(s in str(candidate.columns[0]) for s in [',', ';', '\t', '|'] if s != sep):
                        continue
                    df = candidate
                    break
            except UnicodeDecodeError:
                continue
            except pd.errors.ParserError as e:
                last_error = str(e)
                continue
            except Exception as e:
                last_error = str(e)
                continue
        if df is not None:
            break

    # ── Fallback: try treating as Excel even without extension ──
    if df is None:
        try:
            df = pd.read_excel(io.BytesIO(raw_bytes), engine='openpyxl')
        except Exception:
            pass

    if df is None:
        hint = f" Last parse error: {last_error}" if last_error else ""
        return None, (
            f"Could not parse the uploaded file.{hint} "
            "Please ensure it is a valid CSV (UTF-8) or Excel (.xlsx) file."
        )

    if df.empty or len(df) == 0:
        return None, "The file was parsed but contains no data rows (only headers?)."

    # ── Clean up column names ──
    # Strip whitespace, BOM chars, and invisible characters from column names
    df.columns = [str(c).strip().strip('\ufeff').strip('\u200b') for c in df.columns]

    # Drop completely empty rows and columns
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')

    if df.empty:
        return None, "All rows in the file are empty after cleanup."

    return df, None


def _normalize_column_name(col):
    """Normalize a column name for fuzzy matching: lower, strip, remove special chars."""
    import re
    return re.sub(r'[^a-z0-9]', '', str(col).lower().strip())


def _build_column_map(df_columns, mapping_dict):
    """
    Build a mapping from actual CSV columns to model field names.
    Uses exact match first, then normalized (fuzzy) match.

    mapping_dict: {csv_header_variant: model_field, ...}
    Returns: dict to pass to df.rename(columns=...)
    """
    # Build normalized lookup: normalized_key -> model_field
    norm_lookup = {}
    for csv_name, field_name in mapping_dict.items():
        norm_lookup[_normalize_column_name(csv_name)] = field_name

    rename_map = {}
    for col in df_columns:
        # Exact match
        if col in mapping_dict:
            rename_map[col] = mapping_dict[col]
        else:
            # Fuzzy match
            norm = _normalize_column_name(col)
            if norm in norm_lookup:
                rename_map[col] = norm_lookup[norm]
    return rename_map


def _safe_int(value, field_name=''):
    """Safely convert a value to int, handling commas, spaces, floats, suffixes like 5M/10K, etc."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    s = str(value).strip().replace(',', '').replace(' ', '')
    if not s or s.upper() in ('NAN', 'NONE', 'NIL', 'NA', 'N/A', '-', ''):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        pass
    # Handle suffixed values like "5M" (extract leading digits)
    import re
    m = re.match(r'^([\d.]+)\s*([KkMmBb]?)$', s)
    if m:
        try:
            num = float(m.group(1))
            suffix = m.group(2).upper()
            # Don't multiply — in drilling context "5M" means "5 million PSI class" etc.
            # Just extract the numeric part
            return int(num)
        except (ValueError, TypeError):
            pass
    # Also try stripping all non-digit chars as last resort
    digits_only = re.sub(r'[^\d]', '', s)
    if digits_only:
        try:
            return int(digits_only)
        except (ValueError, TypeError):
            pass
    raise ValueError(f"Cannot convert '{value}' to integer for field '{field_name}'")


def _safe_decimal(value, field_name=''):
    """Safely convert a value to Decimal, handling commas, currency symbols, etc."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    s = str(value).strip().replace(',', '').replace(' ', '').replace('₹', '').replace('$', '').replace('INR', '')
    if not s or s.upper() in ('NAN', 'NONE', 'NIL', 'NA', 'N/A', '-', ''):
        return None
    try:
        return Decimal(s)
    except Exception:
        raise ValueError(f"Cannot convert '{value}' to decimal number for field '{field_name}'")


def _safe_date(value, field_name=''):
    """
    Safely parse a date from many formats.
    Handles: DD-MM-YYYY, DD/MM/YY, YYYY-MM-DD, Excel serial numbers, etc.
    Returns a date object or None.
    Dates beyond 31-12-2099 are capped to 31-12-2099.
    """
    # Maximum allowed year – prevents year-9999 values that break pandas/optimizer
    _MAX_YEAR = 2099
    _MAX_DATE = date(2099, 12, 31)

    def _cap(d):
        """Cap a date to _MAX_DATE if it exceeds the allowed year."""
        if d and d.year > _MAX_YEAR:
            logger.info(f"Date {d} for '{field_name}' exceeds year {_MAX_YEAR}, capped to {_MAX_DATE}")
            return _MAX_DATE
        return d

    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, date):
        return _cap(value)
    s = str(value).strip()
    if not s or s.upper() in ('NAN', 'NONE', 'NIL', 'NA', 'N/A', '-', 'NAT', ''):
        return None

    # ── Excel serial number detection ──
    # Excel dates are stored as number of days since 1899-12-30.
    # Typical range: 1 (1900-01-01) to ~2958465 (9999-12-31)
    # We detect if the string is purely numeric (int or float) within valid Excel range.
    try:
        numeric_val = float(s)
        # Only treat as Excel serial if it's a positive integer-like value
        # and doesn't look like a year (avoid interpreting "2024" as Excel serial)
        if numeric_val > 365 and numeric_val == int(numeric_val):
            serial = int(numeric_val)
            # Valid Excel serial range: roughly 366 (1901-01-01) to 2958465 (9999-12-31)
            if 366 <= serial <= 2958465:
                from datetime import timedelta
                excel_epoch = datetime(1899, 12, 30)
                result = (excel_epoch + timedelta(days=serial)).date()
                logger.debug(f"Converted Excel serial {serial} to date {result} for field '{field_name}'")
                return _cap(result)
    except (ValueError, TypeError, OverflowError):
        pass

    # ── Try explicit formats first (most common in Indian drilling data) ──
    formats = [
        '%d-%m-%Y', '%d/%m/%Y', '%d-%m-%y', '%d/%m/%y',
        '%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y', '%m/%d/%Y',
        '%d %b %Y', '%d %B %Y', '%b %d, %Y', '%B %d, %Y',
        '%d-%b-%Y', '%d-%b-%y', '%d/%b/%Y',
        '%Y%m%d',
    ]
    for fmt in formats:
        try:
            # Use datetime.strptime first (handles year 9999, no nanosecond limits)
            return _cap(datetime.strptime(s, fmt).date())
        except (ValueError, TypeError):
            continue

    # Last resort: let pandas guess (but catch out-of-bounds for far-future dates)
    try:
        return _cap(pd.to_datetime(s, dayfirst=True).date())
    except Exception:
        # Final fallback with strptime for remaining formats
        try:
            return _cap(datetime.strptime(s, '%d/%m/%Y').date())
        except (ValueError, TypeError):
            pass
        raise ValueError(f"Cannot parse date '{value}' for field '{field_name}'. "
                         f"Use DD-MM-YYYY or YYYY-MM-DD format.")


def _safe_string(value, max_length=None, clean_non_ascii=False):
    """Safely convert value to a cleaned string or None.
    Optionally strip non-printable / corrupt chars."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    s = str(value).strip()
    if not s or s.upper() in ('NAN', 'NONE', 'NIL', 'NA', 'N/A'):
        return None
    # Remove BOM, zero-width spaces, and other invisible chars
    s = s.replace('\ufeff', '').replace('\u200b', '').replace('\u00a0', ' ')
    if clean_non_ascii:
        # Replace common mojibake / replacement chars with empty string
        import re
        s = re.sub(r'[\ufffd\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)
    s = s.strip()
    if not s:
        return None
    if max_length:
        s = s[:max_length]
    return s


def _yes_no(value, default='N'):
    """Convert various yes/no representations to 'Y' or 'N'."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return default
    s = str(value).strip().upper()
    if s in ('Y', 'YES', '1', 'TRUE', 'T'):
        return 'Y'
    if s in ('N', 'NO', '0', 'FALSE', 'F'):
        return 'N'
    if not s or s in ('NAN', 'NONE', 'NIL', 'NA', 'N/A', '-', ''):
        return default
    return default


###############################################################################
# ─── Robust Rig Upload ──────────────────────────────────────────────────────
###############################################################################

def handle_rigs_upload(file_obj):
    """Handle rigs CSV/Excel upload with comprehensive error handling."""

    # ── Step 1: Read file into DataFrame ──
    df, file_error = _read_upload_file_to_dataframe(file_obj)
    if df is None:
        return Response({'error': file_error}, status=status.HTTP_400_BAD_REQUEST)

    try:
        logger.info(f"Rig upload: parsed {len(df)} rows, columns: {list(df.columns)}")

        # ── Step 2: Map columns ──
        rig_column_mapping = {
            'Rig': 'name', 'Name': 'name', 'Rig Name': 'name', 'rig': 'name', 'name': 'name', 'RIG': 'name',
            'Asset ID': 'asset_id', 'Asset_ID': 'asset_id', 'AssetID': 'asset_id', 'asset_id': 'asset_id', 'ASSET ID': 'asset_id',
            'Rig Type': 'rig_type', 'rig_type': 'rig_type', 'Type': 'rig_type', 'RIG TYPE': 'rig_type',
            'Start Date': 'start_date', 'start_date': 'start_date', 'StartDate': 'start_date', 'START DATE': 'start_date',
            'End Date': 'end_date', 'end_date': 'end_date', 'EndDate': 'end_date', 'END DATE': 'end_date',
            'Rig Capacity (HP)': 'rig_capacity_hp', 'Rig Capacity HP': 'rig_capacity_hp', 'rig_capacity_hp': 'rig_capacity_hp',
            'Capacity HP': 'rig_capacity_hp', 'HP': 'rig_capacity_hp', 'RIG CAPACITY (HP)': 'rig_capacity_hp',
            'Daily Cost (INR)': 'daily_cost_inr', 'Daily Cost INR': 'daily_cost_inr', 'daily_cost_inr': 'daily_cost_inr',
            'Daily Cost': 'daily_cost_inr', 'Cost': 'daily_cost_inr', 'DAILY COST (INR)': 'daily_cost_inr',
            'Drilling Capacity (m)': 'drilling_capacity_m', 'Drilling Capacity m': 'drilling_capacity_m',
            'drilling_capacity_m': 'drilling_capacity_m', 'Drilling Capacity': 'drilling_capacity_m', 'DRILLING CAPACITY (M)': 'drilling_capacity_m',
            'Mobilization Time (Days)': 'mobilization_time_days', 'Mobilization Time Days': 'mobilization_time_days',
            'mobilization_time_days': 'mobilization_time_days', 'Mob Time': 'mobilization_time_days', 'MOBILIZATION TIME (DAYS)': 'mobilization_time_days',
            'Maintenance Schedule': 'maintenance_schedule', 'maintenance_schedule': 'maintenance_schedule', 'MAINTENANCE SCHEDULE': 'maintenance_schedule',
            'Crew Availability': 'crew_availability', 'crew_availability': 'crew_availability', 'CREW AVAILABILITY': 'crew_availability',
            'HPHT Suitability': 'hpht_suitability', 'hpht_suitability': 'hpht_suitability', 'HPHT': 'hpht_suitability', 'HPHT SUITABILITY': 'hpht_suitability',
            'ILM COST FIXED': 'ilm_cost_fixed', 'ILM Cost Fixed': 'ilm_cost_fixed', 'ilm_cost_fixed': 'ilm_cost_fixed', 'ILM Fixed': 'ilm_cost_fixed',
            'ILM COST per km': 'ilm_cost_per_km', 'ILM Cost per km': 'ilm_cost_per_km', 'ilm_cost_per_km': 'ilm_cost_per_km',
            'ILM per km': 'ilm_cost_per_km', 'ILM COST PER KM': 'ilm_cost_per_km',
            'ILM COST CLUSTER': 'ilm_cost_cluster', 'ILM Cost Cluster': 'ilm_cost_cluster', 'ilm_cost_cluster': 'ilm_cost_cluster',
            'ILM Cluster': 'ilm_cost_cluster', 'ILM COST CLUSTER': 'ilm_cost_cluster',
            'BOP Stack': 'bop_stack', 'bop_stack': 'bop_stack', 'BOP': 'bop_stack', 'BOP STACK': 'bop_stack',
            'TDS Availability': 'tds_availability', 'tds_availability': 'tds_availability', 'TDS': 'tds_availability', 'TDS AVAILABILITY': 'tds_availability',
        }

        rename_map = _build_column_map(df.columns, rig_column_mapping)
        df = df.rename(columns=rename_map)

        # ── Step 3: Validate required column exists ──
        if 'name' not in df.columns:
            available = ', '.join(list(df.columns)[:20])
            return Response(
                {'error': (
                    'Required column "Rig" or "Name" not found in upload. '
                    f'Columns detected: [{available}]. '
                    'Please ensure your file has a column named "Rig" or "Name".'
                )},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Step 4: Define valid model fields to avoid passing junk columns ──
        valid_rig_fields = {
            'name', 'asset_id', 'rig_type', 'start_date', 'end_date',
            'rig_capacity_hp', 'daily_cost_inr', 'drilling_capacity_m',
            'mobilization_time_days', 'maintenance_schedule', 'crew_availability',
            'hpht_suitability', 'ilm_cost_fixed', 'ilm_cost_per_km',
            'ilm_cost_cluster', 'bop_stack', 'tds_availability', 'location',
        }

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for idx, (index, row) in enumerate(df.iterrows()):
            row_num = idx + 2  # +2 because header is row 1, data starts at row 2
            row_errors = []

            try:
                raw = row.to_dict()

                # ── Parse name (required) ──
                rig_name = _safe_string(raw.get('name'), max_length=50, clean_non_ascii=True)
                if not rig_name:
                    errors.append(f"Row {row_num}: Rig name is empty or missing. Skipped.")
                    continue

                rig_data = {'name': rig_name}

                # ── Parse asset_id ──
                rig_data['asset_id'] = _safe_string(raw.get('asset_id'), max_length=50)

                # ── Parse rig_type ──
                rt = _safe_string(raw.get('rig_type'), max_length=10)
                if rt:
                    rt_upper = rt.upper().strip()
                    if rt_upper in ('MOBILE', 'MOB', 'M'):
                        rig_data['rig_type'] = 'Mobile'
                    elif rt_upper in ('FIXED', 'FIX', 'F'):
                        rig_data['rig_type'] = 'Fixed'
                    else:
                        rig_data['rig_type'] = rt  # Pass through; validator will catch
                else:
                    row_errors.append("Rig Type is missing")
                    rig_data['rig_type'] = 'Mobile'  # default to continue

                # ── Parse dates ──
                try:
                    rig_data['start_date'] = _safe_date(raw.get('start_date'), 'Start Date')
                    if rig_data['start_date'] is None:
                        row_errors.append("Start Date is missing or unparseable")
                except ValueError as e:
                    row_errors.append(str(e))
                    rig_data['start_date'] = None

                try:
                    rig_data['end_date'] = _safe_date(raw.get('end_date'), 'End Date')
                    if rig_data['end_date'] is None:
                        row_errors.append("End Date is missing or unparseable")
                except ValueError as e:
                    row_errors.append(str(e))
                    rig_data['end_date'] = None

                # Validate date range
                if rig_data.get('start_date') and rig_data.get('end_date'):
                    if rig_data['start_date'] > rig_data['end_date']:
                        row_errors.append(f"Start Date ({rig_data['start_date']}) is after End Date ({rig_data['end_date']})")

                # ── Parse numeric fields ──
                numeric_fields = [
                    ('rig_capacity_hp', 'Rig Capacity (HP)', True),
                    ('daily_cost_inr', 'Daily Cost (INR)', True),
                    ('drilling_capacity_m', 'Drilling Capacity (m)', True),
                    ('ilm_cost_fixed', 'ILM Cost Fixed', False),
                    ('ilm_cost_per_km', 'ILM Cost per km', False),
                    ('ilm_cost_cluster', 'ILM Cost Cluster', False),
                    ('bop_stack', 'BOP Stack', False),
                ]
                for field, label, required in numeric_fields:
                    raw_val = raw.get(field)
                    try:
                        if field in ('daily_cost_inr', 'ilm_cost_fixed', 'ilm_cost_per_km', 'ilm_cost_cluster'):
                            parsed = _safe_decimal(raw_val, label)
                            if parsed is None and required:
                                row_errors.append(f"{label} is missing")
                                parsed = Decimal('0')
                            elif parsed is None:
                                parsed = Decimal('0')
                            rig_data[field] = parsed
                        else:
                            parsed = _safe_int(raw_val, label)
                            if parsed is None and required:
                                row_errors.append(f"{label} is missing")
                                parsed = 0
                            elif parsed is None:
                                parsed = 0
                            rig_data[field] = parsed
                    except ValueError as e:
                        row_errors.append(str(e))
                        rig_data[field] = Decimal('0') if 'cost' in field else 0

                # ── Parse string/choice fields ──
                rig_data['mobilization_time_days'] = _safe_string(raw.get('mobilization_time_days'), max_length=10)
                rig_data['maintenance_schedule'] = _safe_string(raw.get('maintenance_schedule'), max_length=100)

                crew = _safe_string(raw.get('crew_availability'))
                if crew:
                    crew_upper = crew.upper().strip()
                    if crew_upper in ('OK', 'AVAILABLE', 'YES', 'Y', '1'):
                        rig_data['crew_availability'] = 'OK'
                    elif crew_upper in ('NOT_OK', 'NOT OK', 'NOTOK', 'UNAVAILABLE', 'NO', 'N', '0'):
                        rig_data['crew_availability'] = 'NOT_OK'
                    else:
                        rig_data['crew_availability'] = 'OK'
                else:
                    rig_data['crew_availability'] = 'OK'

                rig_data['hpht_suitability'] = _yes_no(raw.get('hpht_suitability'), default='N')
                rig_data['tds_availability'] = _yes_no(raw.get('tds_availability'), default='Y')

                # ── Auto-set location FK from asset_id ──
                if rig_data.get('asset_id'):
                    try:
                        from django.db.models import Q
                        loc = CompanyCode.objects.filter(
                            Q(location__iexact=rig_data['asset_id']) |
                            Q(company_code__iexact=rig_data['asset_id'])
                        ).first()
                        if loc:
                            rig_data['location'] = loc
                    except Exception:
                        pass

                # ── Remove any keys not in valid fields ──
                rig_data = {k: v for k, v in rig_data.items() if k in valid_rig_fields}

                # ── If there are blocking errors (missing required dates), skip row ──
                if rig_data.get('start_date') is None or rig_data.get('end_date') is None:
                    all_issues = '; '.join(row_errors) if row_errors else 'Missing required date fields'
                    errors.append(f"Row {row_num} (Rig: {rig_name}): {all_issues}. Skipped.")
                    continue

                # ── Create or update rig ──
                existing_rig = Rig.all_objects.filter(name=rig_data['name']).first()

                if existing_rig:
                    data_changed = False
                    for key, value in rig_data.items():
                        if key in ('name', 'location'):
                            continue
                        existing_value = getattr(existing_rig, key, None)
                        try:
                            if isinstance(value, (Decimal, float)) and existing_value is not None:
                                if abs(float(value) - float(existing_value)) > 0.01:
                                    data_changed = True
                                    break
                            elif isinstance(value, date) and isinstance(existing_value, date):
                                if value != existing_value:
                                    data_changed = True
                                    break
                            elif str(existing_value or '').strip() != str(value or '').strip():
                                data_changed = True
                                break
                        except Exception:
                            if str(existing_value) != str(value):
                                data_changed = True
                                break

                    if data_changed:
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        rig_data['name'] = f"{rig_name}_v{timestamp}"
                        Rig.objects.create(**rig_data)
                        created_count += 1
                    else:
                        if existing_rig.is_deleted:
                            existing_rig.is_deleted = False
                            existing_rig.deleted_at = None
                            existing_rig.deleted_by = None
                            existing_rig.save()
                            updated_count += 1
                        else:
                            skipped_count += 1
                else:
                    Rig.objects.create(**rig_data)
                    created_count += 1

                # Append non-fatal warnings
                if row_errors:
                    errors.append(f"Row {row_num} (Rig: {rig_name}): Imported with warnings — {'; '.join(row_errors)}")

            except Exception as e:
                rig_label = _safe_string(raw.get('name')) or f'row {row_num}'
                errors.append(f"Row {row_num} (Rig: {rig_label}): {str(e)}")
                logger.exception(f"Error processing rig row {row_num}")
                continue

        failed_count = len([e for e in errors if 'Skipped' in e or 'ERROR' in e.upper()])
        warning_count = len(errors) - failed_count

        # Build clear summary message
        parts = []
        if created_count:
            parts.append(f'{created_count} created')
        if updated_count:
            parts.append(f'{updated_count} revived')
        if skipped_count:
            parts.append(f'{skipped_count} unchanged')
        if failed_count:
            parts.append(f'{failed_count} failed')
        summary = ', '.join(parts) if parts else 'No rigs processed'

        return Response({
            'message': f'Rigs upload: {summary} (out of {len(df)} rows)',
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': errors,
            'error_count': len(errors),
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception("Rig upload failed at top level")
        return Response(
            {'error': f'Failed to process rigs file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


###############################################################################
# ─── Robust Well Upload ─────────────────────────────────────────────────────
###############################################################################

def handle_wells_upload(file_obj):
    """Handle wells CSV/Excel upload with comprehensive error handling."""

    # ── Step 1: Read file into DataFrame ──
    df, file_error = _read_upload_file_to_dataframe(file_obj)
    if df is None:
        return Response({'error': file_error}, status=status.HTTP_400_BAD_REQUEST)

    try:
        logger.info(f"Well upload: parsed {len(df)} rows, columns: {list(df.columns)}")

        # ── Step 2: Map columns ──
        well_column_mapping = {
            'SN': 'sn', 'sn': 'sn', 'S.No': 'sn', 'S.N.': 'sn', 'Serial': 'sn', 'Sl No': 'sn', 'SN.': 'sn',
            'Asset_ID': 'asset_id', 'Asset ID': 'asset_id', 'AssetID': 'asset_id', 'asset_id': 'asset_id',
            'ASSET ID': 'asset_id', 'Asset': 'asset_id', 'Location': 'asset_id',
            'Well': 'name', 'Name': 'name', 'Well Name': 'name', 'name': 'name', 'WELL': 'name', 'Well_Name': 'name',
            'Type of well': 'well_type', 'Well Type': 'well_type', 'well_type': 'well_type', 'WellType': 'well_type',
            'TYPE OF WELL': 'well_type', 'Type': 'well_type', 'WELL TYPE': 'well_type',
            'Well Profile': 'well_profile', 'well_profile': 'well_profile', 'WellProfile': 'well_profile',
            'Profile': 'well_profile', 'WELL PROFILE': 'well_profile',
            'Depth': 'depth', 'depth': 'depth', 'DEPTH': 'depth', 'Depth (m)': 'depth', 'Depth(m)': 'depth',
            'Rig Capacity Required (HP)': 'rig_capacity_required_hp', 'Rig Capacity Required HP': 'rig_capacity_required_hp',
            'rig_capacity_required_hp': 'rig_capacity_required_hp', 'RIG CAPACITY REQUIRED (HP)': 'rig_capacity_required_hp',
            'Capacity Required HP': 'rig_capacity_required_hp', 'HP Required': 'rig_capacity_required_hp',
            'Rig Capacity Required': 'rig_capacity_required_hp', 'Required HP': 'rig_capacity_required_hp',
            'DRL_DAYS': 'drl_days', 'DRL Days': 'drl_days', 'drl_days': 'drl_days', 'DRL DAYS': 'drl_days',
            'Drilling Days': 'drl_days', 'DrlDays': 'drl_days',
            'PT_DAYS': 'pt_days', 'PT Days': 'pt_days', 'pt_days': 'pt_days', 'PT DAYS': 'pt_days',
            'Post Test Days': 'pt_days', 'PtDays': 'pt_days', 'Testing Days': 'pt_days',
            'Duration': 'duration', 'duration': 'duration', 'DURATION': 'duration', 'Total Duration': 'duration',
            'Latitude': 'latitude', 'latitude': 'latitude', 'LATITUDE': 'latitude', 'Lat': 'latitude',
            'Longitude': 'longitude', 'longitude': 'longitude', 'LONGITUDE': 'longitude', 'Long': 'longitude', 'Lng': 'longitude',
            'RTD': 'rtd', 'rtd': 'rtd', 'Ready To Drill': 'rtd', 'READY TO DRILL': 'rtd', 'Ready Date': 'rtd',
            'BOP Stack': 'bop_stack', 'bop_stack': 'bop_stack', 'BOP': 'bop_stack', 'BOP STACK': 'bop_stack',
            'TDS Requirement': 'tds_requirement', 'tds_requirement': 'tds_requirement', 'TDS': 'tds_requirement',
            'TDS REQUIREMENT': 'tds_requirement', 'TDS Req': 'tds_requirement',
            'Footprint': 'footprint', 'footprint': 'footprint', 'FOOTPRINT': 'footprint', 'Rig Footprint': 'footprint',
            'Preferred Rig': 'preferred_rig', 'preferred_rig': 'preferred_rig', 'PREFERRED RIG': 'preferred_rig',
            'Preferred_Rig': 'preferred_rig', 'Pref Rig': 'preferred_rig',
            'Expected_Potential': 'expected_potential', 'Expected Potential': 'expected_potential',
            'expected_potential': 'expected_potential', 'EXPECTED POTENTIAL': 'expected_potential', 'Potential': 'expected_potential',
            'Priority': 'priority', 'priority': 'priority', 'PRIORITY': 'priority',
        }

        rename_map = _build_column_map(df.columns, well_column_mapping)
        df = df.rename(columns=rename_map)

        # ── Step 3: Validate required columns ──
        if 'name' not in df.columns:
            available = ', '.join(list(df.columns)[:20])
            return Response(
                {'error': (
                    'Required column "Well" or "Name" not found in upload. '
                    f'Columns detected: [{available}]. '
                    'Please ensure your file has a column named "Well" or "Name".'
                )},
                status=status.HTTP_400_BAD_REQUEST
            )

        if 'asset_id' not in df.columns:
            available = ', '.join(list(df.columns)[:20])
            return Response(
                {'error': (
                    'Required column "Asset_ID" or "Asset ID" not found in upload. '
                    f'Columns detected: [{available}]. '
                    'Please ensure your file has a column named "Asset_ID" or "Asset ID".'
                )},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Step 4: Valid model fields ──
        valid_well_fields = {
            'sn', 'asset_id', 'name', 'well_type', 'well_profile', 'depth',
            'rig_capacity_required_hp', 'drl_days', 'pt_days', 'duration',
            'latitude', 'longitude', 'rtd', 'bop_stack', 'tds_requirement',
            'footprint', 'preferred_rig', 'expected_potential', 'priority', 'location',
        }

        # Get next available SN
        from django.db.models import Max
        max_sn = Well.all_objects.aggregate(Max('sn'))['sn__max'] or 0
        next_sn = max_sn + 1

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for idx, (index, row) in enumerate(df.iterrows()):
            row_num = idx + 2
            row_errors = []

            try:
                raw = row.to_dict()

                # ── Parse name (required) ──
                well_name = _safe_string(raw.get('name'), max_length=50)
                if not well_name:
                    errors.append(f"Row {row_num}: Well name is empty or missing. Skipped.")
                    continue

                # ── Parse asset_id (required) ──
                asset_id = _safe_string(raw.get('asset_id'), max_length=50)
                if not asset_id:
                    errors.append(f"Row {row_num} (Well: {well_name}): Asset ID is empty or missing. Skipped.")
                    continue

                well_data = {'name': well_name, 'asset_id': asset_id}

                # ── Parse well_type ──
                wt = _safe_string(raw.get('well_type'), max_length=3)
                if wt:
                    wt_upper = wt.upper().strip()
                    if wt_upper in ('EXP', 'EXPLORATION', 'EXPL', 'E'):
                        well_data['well_type'] = 'EXP'
                    elif wt_upper in ('DEV', 'DEVELOPMENT', 'D'):
                        well_data['well_type'] = 'Dev'
                    else:
                        well_data['well_type'] = wt[:3]
                        row_errors.append(f"Unrecognized well type '{wt}', used as-is")
                else:
                    row_errors.append("Well Type is missing, defaulting to 'Dev'")
                    well_data['well_type'] = 'Dev'

                # ── Parse well_profile ──
                wp = _safe_string(raw.get('well_profile'), max_length=2)
                if wp:
                    wp_upper = wp.upper().strip()
                    if wp_upper in ('DI', 'DIR', 'DIRECTIONAL'):
                        well_data['well_profile'] = 'DI'
                    elif wp_upper in ('VE', 'VERT', 'VERTICAL', 'V'):
                        well_data['well_profile'] = 'VE'
                    elif wp_upper in ('SD', 'SIDE', 'SIDETRACK', 'S'):
                        well_data['well_profile'] = 'SD'
                    else:
                        well_data['well_profile'] = wp[:2]
                        row_errors.append(f"Unrecognized well profile '{wp}', used as-is")
                else:
                    row_errors.append("Well Profile is missing, defaulting to 'VE'")
                    well_data['well_profile'] = 'VE'

                # ── Parse integer fields ──
                int_fields = [
                    ('depth', 'Depth', True),
                    ('rig_capacity_required_hp', 'Rig Capacity Required (HP)', True),
                    ('drl_days', 'DRL Days', True),
                    ('pt_days', 'PT Days', True),
                    ('duration', 'Duration', True),
                    ('bop_stack', 'BOP Stack', False),
                ]
                for field, label, required in int_fields:
                    try:
                        parsed = _safe_int(raw.get(field), label)
                        if parsed is None and required:
                            row_errors.append(f"{label} is missing")
                            parsed = 0
                        elif parsed is None:
                            parsed = 0
                        well_data[field] = parsed
                    except ValueError as e:
                        row_errors.append(str(e))
                        well_data[field] = 0

                # Auto-calculate duration if missing but drl_days + pt_days available
                if well_data.get('duration', 0) == 0 and well_data.get('drl_days', 0) > 0:
                    well_data['duration'] = well_data.get('drl_days', 0) + well_data.get('pt_days', 0)
                    if well_data['duration'] > 0:
                        row_errors.append(f"Duration auto-calculated as DRL Days + PT Days = {well_data['duration']}")

                # ── Parse decimal fields (latitude, longitude) ──
                for field, label in [('latitude', 'Latitude'), ('longitude', 'Longitude')]:
                    try:
                        parsed = _safe_decimal(raw.get(field), label)
                        if parsed is None:
                            row_errors.append(f"{label} is missing, defaulting to 0")
                            parsed = Decimal('0')
                        well_data[field] = parsed
                    except ValueError as e:
                        row_errors.append(str(e))
                        well_data[field] = Decimal('0')

                # ── Parse RTD date ──
                try:
                    well_data['rtd'] = _safe_date(raw.get('rtd'), 'RTD')
                    if well_data['rtd'] is None:
                        row_errors.append("RTD (Ready To Drill) date is missing")
                except ValueError as e:
                    row_errors.append(str(e))
                    well_data['rtd'] = None

                # ── Parse choice/string fields ──
                well_data['tds_requirement'] = _yes_no(raw.get('tds_requirement'), default='Y')

                fp = _safe_string(raw.get('footprint'), max_length=10)
                if fp:
                    fp_upper = fp.upper().strip()
                    if fp_upper in ('MOBILE', 'MOB', 'M'):
                        well_data['footprint'] = 'Mobile'
                    elif fp_upper in ('FIXED', 'FIX', 'F'):
                        well_data['footprint'] = 'Fixed'
                    else:
                        well_data['footprint'] = fp
                        row_errors.append(f"Unrecognized footprint '{fp}', used as-is")
                else:
                    well_data['footprint'] = 'Mobile'
                    row_errors.append("Footprint is missing, defaulting to 'Mobile'")

                well_data['preferred_rig'] = _safe_string(raw.get('preferred_rig'), max_length=50)
                well_data['expected_potential'] = _safe_string(raw.get('expected_potential'), max_length=20)

                pri = _safe_string(raw.get('priority'))
                if pri:
                    pri_upper = pri.upper().strip()
                    if pri_upper in ('HIGH', 'H', '1'):
                        well_data['priority'] = 'HIGH'
                    elif pri_upper in ('MEDIUM', 'MED', 'M', '2'):
                        well_data['priority'] = 'MEDIUM'
                    elif pri_upper in ('LOW', 'L', '3'):
                        well_data['priority'] = 'LOW'
                    else:
                        well_data['priority'] = 'MEDIUM'
                        row_errors.append(f"Unrecognized priority '{pri}', defaulting to MEDIUM")
                else:
                    well_data['priority'] = 'MEDIUM'

                # ── Auto-set location FK from asset_id ──
                if well_data.get('asset_id'):
                    try:
                        from django.db.models import Q
                        loc = CompanyCode.objects.filter(
                            Q(company_code__iexact=well_data['asset_id']) |
                            Q(location__iexact=well_data['asset_id'])
                        ).first()
                        if loc:
                            well_data['location'] = loc
                    except Exception:
                        pass

                # ── Remove any keys not in valid fields ──
                well_data = {k: v for k, v in well_data.items() if k in valid_well_fields}

                # ── If RTD is missing, skip row (it's critical for scheduling) ──
                if well_data.get('rtd') is None:
                    all_issues = '; '.join(row_errors) if row_errors else 'Missing RTD date'
                    errors.append(f"Row {row_num} (Well: {well_name}): {all_issues}. Skipped.")
                    continue

                # ── Create or update well ──
                import re as _re
                all_matching_wells = Well.all_objects.filter(
                    asset_id=asset_id, name=well_name
                ) | Well.all_objects.filter(
                    asset_id=asset_id, name__startswith=f"{well_name}_v"
                )
                most_recent_well = all_matching_wells.order_by('-id').first()

                if most_recent_well:
                    data_changed = False
                    changed_fields = []
                    for key, value in well_data.items():
                        if key in ('name', 'asset_id', 'sn', 'location'):
                            continue
                        existing_value = getattr(most_recent_well, key, None)
                        try:
                            # Both None
                            if (value is None or (isinstance(value, float) and math.isnan(value))) and existing_value is None:
                                continue
                            if (value is None or (isinstance(value, float) and math.isnan(value))) and existing_value is not None:
                                data_changed = True
                                changed_fields.append(f"{key}: {existing_value} -> None")
                                continue
                            if existing_value is None:
                                data_changed = True
                                changed_fields.append(f"{key}: None -> {value}")
                                continue
                            # Date comparison
                            if isinstance(existing_value, date):
                                new_date = value if isinstance(value, date) else _safe_date(value, key)
                                if existing_value != new_date:
                                    data_changed = True
                                    changed_fields.append(f"{key}: {existing_value} -> {new_date}")
                                continue
                            # Numeric comparison
                            if isinstance(existing_value, (int, float, Decimal)):
                                if abs(float(existing_value) - float(value)) > 0.01:
                                    data_changed = True
                                    changed_fields.append(f"{key}: {existing_value} -> {value}")
                                continue
                            # String comparison
                            if str(existing_value).strip() != str(value).strip():
                                data_changed = True
                                changed_fields.append(f"{key}: '{existing_value}' -> '{value}'")
                        except Exception:
                            if str(existing_value) != str(value):
                                data_changed = True
                                changed_fields.append(f"{key}: {existing_value} -> {value}")

                    if data_changed:
                        logger.info(f"Creating versioned well for {well_name}: {', '.join(changed_fields)}")
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        versioned_data = well_data.copy()
                        versioned_data['name'] = f"{well_name}_v{timestamp}"
                        versioned_data['sn'] = next_sn
                        next_sn += 1
                        well = Well.objects.create(**versioned_data)
                        created_count += 1
                    else:
                        base_well = Well.all_objects.filter(name=well_name, asset_id=asset_id).first()
                        if base_well and base_well.is_deleted:
                            base_well.is_deleted = False
                            base_well.deleted_at = None
                            base_well.deleted_by = None
                            base_well.save()
                            updated_count += 1
                        else:
                            skipped_count += 1
                else:
                    # New well
                    if 'sn' not in well_data or well_data.get('sn') is None:
                        well_data['sn'] = next_sn
                        next_sn += 1
                    else:
                        # Validate provided SN isn't a duplicate
                        try:
                            sn_val = int(float(str(well_data['sn'])))
                            if Well.all_objects.filter(sn=sn_val).exists():
                                well_data['sn'] = next_sn
                                next_sn += 1
                                row_errors.append(f"SN {sn_val} already exists, auto-assigned new SN {well_data['sn']}")
                            else:
                                well_data['sn'] = sn_val
                        except (ValueError, TypeError):
                            well_data['sn'] = next_sn
                            next_sn += 1

                    well = Well.objects.create(**well_data)
                    created_count += 1

                    # ILM well pair distances are calculated automatically
                    # via the post_save signal (well_saved_trigger_ilm) using on_commit

                # Append non-fatal warnings
                if row_errors:
                    errors.append(f"Row {row_num} (Well: {well_name}): Imported with warnings — {'; '.join(row_errors)}")

            except Exception as e:
                well_label = _safe_string(raw.get('name')) if 'raw' in dir() else f'row {row_num}'
                errors.append(f"Row {row_num} (Well: {well_label}): {str(e)}")
                logger.exception(f"Error processing well row {row_num}")
                continue

        failed_count = len([e for e in errors if 'Skipped' in e or 'ERROR' in e.upper()])
        warning_count = len(errors) - failed_count

        # Build clear summary message
        parts = []
        if created_count:
            parts.append(f'{created_count} created')
        if updated_count:
            parts.append(f'{updated_count} revived')
        if skipped_count:
            parts.append(f'{skipped_count} unchanged')
        if failed_count:
            parts.append(f'{failed_count} failed')
        summary = ', '.join(parts) if parts else 'No wells processed'

        return Response({
            'message': f'Wells upload: {summary} (out of {len(df)} rows)',
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': errors,
            'error_count': len(errors),
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.exception("Well upload failed at top level")
        return Response(
            {'error': f'Failed to process wells file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@login_required
def wells_by_location(request):
    """API endpoint to get wells filtered by asset_id/location"""
    asset_id = request.GET.get('asset_id')
    
    try:
        if asset_id:
            wells = Well.objects.filter(asset_id=asset_id)
        else:
            wells = Well.objects.all()
            
        wells_data = []
        for well in wells:
            wells_data.append({
                'id': str(well.id),
                'name': well.name,
                'sn': well.sn,
                'asset_id': well.asset_id,
                'well_type': well.well_type,
                'well_profile': well.well_profile,
                'depth': well.depth,
                'rig_capacity_required_hp': well.rig_capacity_required_hp,
                'drl_days': well.drl_days,
                'pt_days': well.pt_days,
                'duration': well.duration,
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'rtd': well.rtd.isoformat() if well.rtd else None,
                'bop_stack': well.bop_stack,
                'tds_requirement': well.tds_requirement,
                'footprint': well.footprint,
                'preferred_rig': well.preferred_rig,
                'expected_potential': float(well.expected_potential) if well.expected_potential else None,
                'priority': well.priority
            })
            
        return JsonResponse({
            'success': True,
            'wells': wells_data,
            'count': len(wells_data),
            'asset_id': asset_id
        })
        
    except Exception as e:
        logger.error(f"Error fetching wells by location: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@login_required
def rigs_by_location(request):
    """API endpoint to get rigs filtered by asset_id/location"""
    asset_id = request.GET.get('asset_id')
    
    try:
        if asset_id:
            rigs = Rig.objects.filter(asset_id=asset_id)
        else:
            rigs = Rig.objects.all()
            
        rigs_data = []
        for rig in rigs:
            rigs_data.append({
                'id': str(rig.id),
                'name': rig.name,
                'asset_id': rig.asset_id,
                'rig_type': rig.rig_type,
                'start_date': rig.start_date.isoformat() if rig.start_date else None,
                'end_date': rig.end_date.isoformat() if rig.end_date else None,
                'rig_capacity_hp': rig.rig_capacity_hp,
                'daily_cost_inr': float(rig.daily_cost_inr) if rig.daily_cost_inr else None,
                'drilling_capacity_m': rig.drilling_capacity_m,
                'mobilization_time_days': rig.mobilization_time_days,
                'maintenance_schedule': rig.maintenance_schedule,
                'crew_availability': rig.crew_availability,
                'hpht_suitability': rig.hpht_suitability,
                'ilm_cost_fixed': float(rig.ilm_cost_fixed) if rig.ilm_cost_fixed else None,
                'ilm_cost_per_km': float(rig.ilm_cost_per_km) if rig.ilm_cost_per_km else None,
                'ilm_cost_cluster': float(rig.ilm_cost_cluster) if rig.ilm_cost_cluster else None,
                'bop_stack': rig.bop_stack,
                'tds_availability': rig.tds_availability
            })
            
        return JsonResponse({
            'success': True,
            'rigs': rigs_data,
            'count': len(rigs_data),
            'asset_id': asset_id
        })
        
    except Exception as e:
        logger.error(f"Error fetching rigs by location: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@login_required
def get_user_location_info(request):
    """
    Get the current user's location information and accessible locations.
    
    Returns:
        - user_location: The location assigned to the user (or null if admin)
        - can_view_all: Boolean indicating if user can view all locations
        - accessible_locations: List of locations the user can access
    """
    try:
        user = request.user
        user_location = get_user_location(user)
        accessible_locations = get_user_accessible_locations(user)
        
        # Format the response
        response_data = {
            'user': {
                'username': user.username,
                'email': user.email,
                'is_superuser': user.is_superuser,
                'is_staff': user.is_staff,
            },
            'user_location': None,
            'can_view_all': user.is_superuser,
            'accessible_locations': []
        }
        
        # Check user's profile
        if hasattr(user, 'profile') and user.profile:
            response_data['can_view_all'] = user.profile.can_view_all_locations or user.is_superuser
            if user.profile.location:
                response_data['user_location'] = {
                    'id': str(user.profile.location.id),
                    'code': user.profile.location.code,
                    'name': user.profile.location.name,
                }
        
        # Format accessible locations
        for location in accessible_locations:
            response_data['accessible_locations'].append({
                'id': str(location.id),
                'code': location.code,
                'name': location.name,
                'rig_count': location.rigs.count(),
                'well_count': location.wells.count(),
            })
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Error getting user location info: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@login_required
def get_all_locations(request):
    """
    Get all active locations. Available to all authenticated users.
    """
    try:
        locations = CompanyCode.objects.filter(is_active=True).order_by('location', 'company_code')
        
        locations_data = []
        for location in locations:
            locations_data.append({
                'id': str(location.id),
                'code': location.code,
                'name': location.name,
                'description': location.description,
                'rig_count': location.rigs.count(),
                'well_count': location.wells.count(),
            })
        
        return Response({
            'locations': locations_data,
            'count': len(locations_data)
        })
        
    except Exception as e:
        logger.error(f"Error getting all locations: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
def well_upload(request):
    """
    Render the well CSV upload page
    """
    return render(request, 'scheduler/well_upload.html')


@login_required
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def well_csv_upload(request):
    """
    Handle CSV file upload for bulk well creation or update from location release data.
    Wells are saved to the StagedWell table for field completion before final import.
    
    For existing wells: Only updates fields that are currently null/empty.
    For new wells: Creates new records.
    
    CSV Column Mappings:
    - Company Code -> Asset_ID
    - WellId -> Well name
    - Category -> Well type (D=Dev, E=EXP)
    - Targ._Depth -> Depth
    - CORD:_LAT_&_LONG -> Latitude / Longitude (parsed from DMS format)
    - CVIL_WRK_END_DATE -> RTD
    - (priority) -> Priority
    - Fieild_Name / Field_Name -> field_name
    """
    try:
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        csv_file = request.FILES['file']
        
        # Validate file extension
        if not csv_file.name.endswith('.csv'):
            return Response({'error': 'File must be a CSV'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Read CSV file
        try:
            raw_bytes = csv_file.read()
            decoded_file = None
            for encoding in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                try:
                    decoded_file = raw_bytes.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if decoded_file is None:
                return Response({'error': 'Error reading CSV file: unable to decode file with any supported encoding (utf-8, latin-1, cp1252)'}, status=status.HTTP_400_BAD_REQUEST)
            io_string = io.StringIO(decoded_file)
            df = pd.read_csv(io_string)
        except Exception as e:
            logger.error(f"Error reading CSV file: {str(e)}")
            return Response({'error': f'Error reading CSV file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get user location for data filtering/assignment
        user_location = get_user_location(request.user)
        
        # Process rows
        created_staged_wells = []
        updated_staged_wells = []
        errors = []
        skipped_rows = []
        
        for idx, row in df.iterrows():
            try:
                # Skip rows without essential data
                if pd.isna(row.get('WellId')) or str(row.get('WellId')).strip() == '':
                    skipped_rows.append(f"Row {idx + 2}: Missing WellId")
                    continue
                
                # Extract and map fields
                well_name = str(row['WellId']).strip()
                
                # Company Code -> Asset_ID
                asset_id = str(row.get('Company Code', '')).strip() if pd.notna(row.get('Company Code')) else ''
                
                # Try to find matching CompanyCode by company_code field to get location
                location_from_asset = None
                if asset_id:
                    try:
                        company_code_obj = CompanyCode.objects.filter(company_code=asset_id).first()
                        if company_code_obj:
                            location_from_asset = company_code_obj
                        else:
                            logger.warning(f"Row {idx + 2} ({well_name}): Asset ID '{asset_id}' not found in Company Codes table")
                    except Exception as cc_error:
                        logger.warning(f"Row {idx + 2} ({well_name}): Error looking up company code - {str(cc_error)}")
                
                # Determine location: prefer location from asset_id match, fallback to user_location
                well_location = location_from_asset if location_from_asset else user_location
                
                # Category -> Well Type (D=Dev, E=EXP)
                category = str(row.get('Category', '')).strip().upper() if pd.notna(row.get('Category')) else ''
                well_type = 'Dev' if category == 'D' else 'EXP' if category == 'E' else 'Dev'
                
                # Targ._Depth -> Depth
                try:
                    depth = int(float(row.get('Targ._Depth', 0))) if pd.notna(row.get('Targ._Depth')) else 0
                    if depth <= 0:
                        skipped_rows.append(f"Row {idx + 2} ({well_name}): Invalid depth")
                        continue
                except (ValueError, TypeError):
                    skipped_rows.append(f"Row {idx + 2} ({well_name}): Invalid depth value")
                    continue
                
                # Parse coordinates from DMS format: "22 21' 05.7958"N, 72 30' 18.9544"E"
                # Wrap in try-except to catch regex pattern errors
                coord_str = str(row.get('CORD:_LAT_&_LONG', '')).strip() if pd.notna(row.get('CORD:_LAT_&_LONG')) else ''
                latitude, longitude = None, None
                
                if coord_str and coord_str.lower() != 'nan':
                    try:
                        latitude, longitude = parse_dms_coordinates(coord_str)
                        # If parsing fails, log it but don't reject the well
                        if latitude is None or longitude is None:
                            logger.warning(f"Row {idx + 2} ({well_name}): Could not parse coordinates '{coord_str}', will need manual entry")
                    except Exception as coord_error:
                        logger.warning(f"Row {idx + 2} ({well_name}): Coordinate parsing error - {str(coord_error)}")
                        # Continue without coordinates
                
                # CVIL_WRK_END_DATE -> RTD
                rtd_value = row.get('CVIL_WRK_END_DATE', None)
                rtd = None
                try:
                    rtd = parse_date_field(rtd_value)
                except Exception as date_error:
                    logger.warning(f"Row {idx + 2} ({well_name}): Date parsing error - {str(date_error)}")
                
                # Keep RTD as None if not provided (blank RTD is allowed)
                
                # (priority) -> Priority
                priority_value = str(row.get('(priority)', '')).strip().upper() if pd.notna(row.get('(priority)')) else ''
                if priority_value in ['HIGH', 'MEDIUM', 'LOW']:
                    priority = priority_value
                else:
                    priority = 'MEDIUM'  # Default
                
                # Extract Field_Name (note: CSV has "Fieild_Name" typo)
                field_name = str(row.get('Fieild_Name', '')).strip() if pd.notna(row.get('Fieild_Name')) else ''
                if not field_name:  # Try standard spelling too
                    field_name = str(row.get('Field_Name', '')).strip() if pd.notna(row.get('Field_Name')) else ''
                # Convert to title case for consistency
                if field_name:
                    field_name = field_name.title()
                
                # Check if well already exists in staging
                existing_well = StagedWell.objects.filter(name=well_name).first()
                
                if existing_well:
                    # UPDATE MODE: Only update fields that are currently null or empty
                    updated = False
                    
                    # Update location: prefer location from asset_id, then user_location, only if currently empty
                    if not existing_well.location and well_location:
                        existing_well.location = well_location
                        updated = True
                    
                    # Update asset_id only if empty
                    if not existing_well.asset_id and asset_id:
                        existing_well.asset_id = asset_id
                        updated = True
                    
                    # Update field_name only if empty
                    if not existing_well.field_name and field_name:
                        existing_well.field_name = field_name
                        updated = True
                    
                    # Update well_type only if empty
                    if not existing_well.well_type and well_type:
                        existing_well.well_type = well_type
                        updated = True
                    
                    # Update depth only if 0 or null
                    if (not existing_well.depth or existing_well.depth == 0) and depth and depth > 0:
                        existing_well.depth = depth
                        updated = True
                    
                    # Update latitude only if null
                    if existing_well.latitude is None and latitude is not None:
                        existing_well.latitude = latitude
                        updated = True
                    
                    # Update longitude only if null
                    if existing_well.longitude is None and longitude is not None:
                        existing_well.longitude = longitude
                        updated = True
                    
                    # Update rtd only if null (don't overwrite user-set dates)
                    if existing_well.rtd is None and rtd:
                        existing_well.rtd = rtd
                        updated = True
                    
                    # Update priority only if empty
                    if not existing_well.priority and priority:
                        existing_well.priority = priority
                        updated = True
                    
                    if updated:
                        with transaction.atomic():
                            existing_well.save()
                        updated_staged_wells.append(well_name)
                    else:
                        skipped_rows.append(f"Row {idx + 2} ({well_name}): Well exists and all fields already populated")
                    
                    continue
                
                # Check if well already exists in final Wells table
                if Well.objects.filter(name=well_name).exists():
                    skipped_rows.append(f"Row {idx + 2} ({well_name}): Well already exists in wells database")
                    continue
                
                # CREATE MODE: Create new staged well
                with transaction.atomic():
                    staged_well = StagedWell.objects.create(
                        location=well_location,
                        asset_id=asset_id,
                        name=well_name,
                        field_name=field_name,
                        well_type=well_type,
                        depth=depth,
                        latitude=latitude,
                        longitude=longitude,
                        rtd=rtd,
                        priority=priority,
                        uploaded_by=request.user,
                        status='PENDING'
                    )
                    created_staged_wells.append(staged_well.name)
                
            except Exception as e:
                logger.error(f"Error processing row {idx + 2}: {str(e)}")
                errors.append(f"Row {idx + 2}: {str(e)}")
        
        # Prepare response
        message_parts = []
        if created_staged_wells:
            message_parts.append(f'{len(created_staged_wells)} new wells uploaded')
        if updated_staged_wells:
            message_parts.append(f'{len(updated_staged_wells)} wells updated with missing fields')
        
        response_data = {
            'success': True,
            'message': '. '.join(message_parts) if message_parts else 'No changes made',
            'created_count': len(created_staged_wells),
            'created_wells': created_staged_wells,
            'updated_count': len(updated_staged_wells),
            'updated_wells': updated_staged_wells,
            'skipped_count': len(skipped_rows),
            'skipped_rows': skipped_rows,
            'error_count': len(errors),
            'errors': errors
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error in well CSV upload: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def parse_dms_coordinates(coord_str):
    """
    Parse DMS (Degrees Minutes Seconds) coordinates to decimal degrees
    Supports multiple formats:
    - "22° 21' 05.7958"N, 72° 30' 18.9544"E"
    - "22 21' 05.7958"N, 72 30' 18.9544"E"
    - "22 13' 16.3886"N, 73 01 20.3441"E" (no seconds quotes)
    - "22#13'50.0560"N 73#05'54.8945"E" (# instead of degree symbol)
    - "22°14'03.8580"N 73°05'54.7142"E" (no comma)
    - "23 21 06.5050 N & 72 18 49.0381 E" (space-separated with &)
    - "23 23 09.988 72 17 30.624" (space-separated, no direction)
    - Degrees and minutes only (no seconds)
    
    Returns:
        tuple: (latitude, longitude) in decimal degrees, or (None, None) if parsing fails
    """
    import re
    
    try:
        # Clean the string
        coord_str = coord_str.strip()
        
        if not coord_str:
            return None, None
        
        # Replace common variations to normalize
        coord_str = coord_str.replace('#', '°')  # Handle # as degree symbol
        coord_str = coord_str.replace('"', ' ')  # Replace closing quotes with space
        coord_str = coord_str.replace('″', ' ')  # Unicode double prime
        coord_str = coord_str.replace('′', "'")  # Unicode single prime to apostrophe
        coord_str = coord_str.replace('&', ',')  # Replace & with comma for easier parsing
        coord_str = coord_str.replace(';', ',')  # Replace semicolon with comma for easier parsing
        
        # Pattern 0: Compact DDMMSS.sss format (6 integer digits per coordinate)
        # Handles: "232625.72, 722917.98"  "233855.338783, 721447.852957"  "234004.7082, 721436.9850"
        # The integer part is always DDMMSS (2 deg + 2 min + 2 sec digits)
        pattern_compact = r'^\s*(\d{6}(?:\.\d*)?)\s*,\s*(\d{6}(?:\.\d*)?)\s*$'
        match_compact = re.match(pattern_compact, coord_str.strip())
        if match_compact:
            def _parse_compact(s):
                int_part, _, dec_part = s.partition('.')
                if len(int_part) < 6:
                    return None
                deg = int(int_part[:-4])
                min_ = int(int_part[-4:-2])
                sec = float(int_part[-2:] + ('.' + dec_part if dec_part else ''))
                return deg + min_ / 60.0 + sec / 3600.0
            lat_val = _parse_compact(match_compact.group(1))
            lon_val = _parse_compact(match_compact.group(2))
            if lat_val is not None and lon_val is not None:
                return round(lat_val, 6), round(lon_val, 6)
        
        # Pattern 1: Space-separated numbers without symbols (most common in current dataset)
        # Handles: "23 21 06.5050 N , 72 18 49.0381 E" or "23 21 06.5050 72 18 49.0381"
        # This pattern looks for 3 numbers (deg min sec) optionally followed by direction, repeated twice
        pattern_space = r'([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s*([NSEW])?\s*[,\s]\s*([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)\s*([NSEW])?'
        match = re.search(pattern_space, coord_str, re.IGNORECASE)
        
        if match:
            lat_deg = float(match.group(1))
            lat_min = float(match.group(2))
            lat_sec = float(match.group(3))
            lat_dir = match.group(4).upper() if match.group(4) else 'N'  # Default to N if not specified
            
            lon_deg = float(match.group(5))
            lon_min = float(match.group(6))
            lon_sec = float(match.group(7))
            lon_dir = match.group(8).upper() if match.group(8) else 'E'  # Default to E if not specified
            
            latitude = lat_deg + (lat_min / 60.0) + (lat_sec / 3600.0)
            if lat_dir == 'S':
                latitude = -latitude
            
            longitude = lon_deg + (lon_min / 60.0) + (lon_sec / 3600.0)
            if lon_dir == 'W':
                longitude = -longitude
            
            # Round to 6 decimal places for database storage
            return round(latitude, 6), round(longitude, 6)
        
        # Try multiple patterns in order of complexity
        
        # Pattern 2: Full DMS with seconds (degrees minutes seconds direction)
        # Handles: 22° 21' 05.7958"N or 22 21' 05.7958 N or 22°21'05.7958N
        pattern_full = r"([0-9.]+)\s*[°\s]\s*([0-9.]+)\s*['\']\s*([0-9.]+)\s*[\"′″\']?\s*([NSEW])"
        matches = re.findall(pattern_full, coord_str, re.IGNORECASE)
        
        if len(matches) >= 2:
            # Parse latitude (first match)
            lat_deg = float(matches[0][0])
            lat_min = float(matches[0][1])
            lat_sec = float(matches[0][2])
            lat_dir = matches[0][3].upper()
            
            latitude = lat_deg + (lat_min / 60.0) + (lat_sec / 3600.0)
            if lat_dir == 'S':
                latitude = -latitude
            
            # Parse longitude (second match)
            lon_deg = float(matches[1][0])
            lon_min = float(matches[1][1])
            lon_sec = float(matches[1][2])
            lon_dir = matches[1][3].upper()
            
            longitude = lon_deg + (lon_min / 60.0) + (lon_sec / 3600.0)
            if lon_dir == 'W':
                longitude = -longitude
            
            # Round to 6 decimal places for database storage
            return round(latitude, 6), round(longitude, 6)
        
        # Pattern 2b: Full DMS with seconds but NO direction marker
        # Handles: "23°26'07.7788  72°28'36.0839"  "23°26' 34.3614  72°21' 06.2119"
        pattern_nodirection = r"([0-9.]+)\s*°\s*([0-9.]+)\s*['\u2019]\s*([0-9.]+)"
        matches_nd = re.findall(pattern_nodirection, coord_str, re.IGNORECASE)
        if len(matches_nd) >= 2:
            lat_deg = float(matches_nd[0][0])
            lat_min = float(matches_nd[0][1])
            lat_sec = float(matches_nd[0][2])
            lon_deg = float(matches_nd[1][0])
            lon_min = float(matches_nd[1][1])
            lon_sec = float(matches_nd[1][2])
            latitude = lat_deg + (lat_min / 60.0) + (lat_sec / 3600.0)
            longitude = lon_deg + (lon_min / 60.0) + (lon_sec / 3600.0)
            return round(latitude, 6), round(longitude, 6)
        
        # Pattern 3: Degrees and minutes only (no seconds)
        # Handles: 22° 21'N or 22 21 N
        pattern_dm = r"([0-9.]+)\s*[°\s]\s*([0-9.]+)\s*['\']\s*([NSEW])"
        matches = re.findall(pattern_dm, coord_str, re.IGNORECASE)
        
        if len(matches) >= 2:
            # Parse latitude (first match)
            lat_deg = float(matches[0][0])
            lat_min = float(matches[0][1])
            lat_dir = matches[0][2].upper()
            
            latitude = lat_deg + (lat_min / 60.0)
            if lat_dir == 'S':
                latitude = -latitude
            
            # Parse longitude (second match)
            lon_deg = float(matches[1][0])
            lon_min = float(matches[1][1])
            lon_dir = matches[1][2].upper()
            
            longitude = lon_deg + (lon_min / 60.0)
            if lon_dir == 'W':
                longitude = -longitude
            
            # Round to 6 decimal places for database storage
            return round(latitude, 6), round(longitude, 6)
        
        # Pattern 4: Try splitting by comma and parsing each part separately
        # Handles cases where there might be formatting issues
        if ',' in coord_str:
            parts = coord_str.split(',')
            if len(parts) == 2:
                # Extract all numbers from each part
                lat_numbers = re.findall(r'[0-9.]+', parts[0])
                lon_numbers = re.findall(r'[0-9.]+', parts[1])
                
                # Extract direction
                lat_dir = 'N' if 'N' in parts[0].upper() else ('S' if 'S' in parts[0].upper() else 'N')
                lon_dir = 'E' if 'E' in parts[1].upper() else ('W' if 'W' in parts[1].upper() else 'E')
                
                if len(lat_numbers) >= 2 and len(lon_numbers) >= 2:
                    # At minimum we have degrees and minutes
                    lat_deg = float(lat_numbers[0])
                    lat_min = float(lat_numbers[1])
                    lat_sec = float(lat_numbers[2]) if len(lat_numbers) > 2 else 0.0
                    
                    lon_deg = float(lon_numbers[0])
                    lon_min = float(lon_numbers[1])
                    lon_sec = float(lon_numbers[2]) if len(lon_numbers) > 2 else 0.0
                    
                    latitude = lat_deg + (lat_min / 60.0) + (lat_sec / 3600.0)
                    if lat_dir == 'S':
                        latitude = -latitude
                    
                    longitude = lon_deg + (lon_min / 60.0) + (lon_sec / 3600.0)
                    if lon_dir == 'W':
                        longitude = -longitude
                    
                    # Round to 6 decimal places for database storage
                    return round(latitude, 6), round(longitude, 6)
        
        # If all patterns fail, return None
        logger.warning(f"Could not parse coordinates '{coord_str}' with any known pattern")
        return None, None
        
    except Exception as e:
        logger.error(f"Error parsing coordinates '{coord_str}': {str(e)}")
        return None, None


def parse_date_field(date_value):
    """
    Parse various date formats from CSV to Python date object
    Handles: YYYYMMDD, standard date formats, and 0 values
    
    Returns:
        date object or None if parsing fails
    """
    try:
        if pd.isna(date_value) or date_value == 0 or date_value == '0':
            return None
        
        # Convert to string and clean
        date_str = str(date_value).strip()
        
        if not date_str or date_str == '0':
            return None
        
        # Try YYYYMMDD format (e.g., 20250217)
        if len(date_str) == 8 and date_str.isdigit():
            year = int(date_str[0:4])
            month = int(date_str[4:6])
            day = int(date_str[6:8])
            return date(year, month, day)
        
        # Try standard date parsing
        parsed_date = pd.to_datetime(date_str, errors='coerce')
        if pd.notna(parsed_date):
            return parsed_date.date()
        
        return None
        
    except Exception as e:
        logger.error(f"Error parsing date '{date_value}': {str(e)}")
        return None


# =============================================================================
# STAGED WELL MANAGEMENT VIEWS
# =============================================================================

@login_required
def staged_wells_management(request):
    """
    Render the staged wells management page where users can complete additional fields
    """
    return render(request, 'scheduler/staged_wells.html')


@login_required
@api_view(['GET'])
def get_staged_wells(request):
    """
    Get all staged wells for the current user's location
    """
    try:
        user_location = get_user_location(request.user)
        
        # Filter by location if user has location restriction
        if user_location:
            staged_wells = StagedWell.objects.filter(location=user_location)
        else:
            # User can see all locations
            staged_wells = StagedWell.objects.all()
        
        # Filter by status if provided
        status_filter = request.GET.get('status')
        if status_filter:
            staged_wells = staged_wells.filter(status=status_filter)
        
        # Order by most recent first
        staged_wells = staged_wells.order_by('-uploaded_at')
        
        from .serializers import StagedWellSerializer
        serializer = StagedWellSerializer(staged_wells, many=True)
        
        # Calculate processed count (IMPORTED staged wells)
        if user_location:
            processed_count = StagedWell.objects.filter(location=user_location, status='IMPORTED').count()
        else:
            processed_count = StagedWell.objects.filter(status='IMPORTED').count()
        
        return Response({
            'success': True,
            'count': len(serializer.data),
            'staged_wells': serializer.data,
            'processed_count': processed_count
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching staged wells: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['GET'])
def get_staged_well_detail(request, staged_well_id):
    """
    Get details of a specific staged well
    """
    try:
        staged_well = get_object_or_404(StagedWell, id=staged_well_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and staged_well.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        from .serializers import StagedWellSerializer
        serializer = StagedWellSerializer(staged_well)
        
        return Response({
            'success': True,
            'staged_well': serializer.data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching staged well detail: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT', 'POST'])
def bulk_update_staged_wells(request):
    """
    Bulk update multiple staged wells at once.
    Expects JSON body: { "updates": [ { "id": "uuid", "fields": { ... } }, ... ] }
    Returns summary of successes and failures.
    """
    try:
        updates = request.data.get('updates', [])
        if not updates:
            return Response({'error': 'No updates provided'}, status=status.HTTP_400_BAD_REQUEST)

        user_location = get_user_location(request.user)
        from .serializers import StagedWellUpdateSerializer, StagedWellSerializer

        success_count = 0
        error_count = 0
        errors = []

        for item in updates:
            well_id = item.get('id')
            fields = item.get('fields', {})
            if not well_id or not fields:
                error_count += 1
                errors.append(f'Missing id or fields')
                continue

            try:
                staged_well = StagedWell.objects.get(id=well_id)

                # Check location access
                if user_location and staged_well.location != user_location:
                    error_count += 1
                    errors.append(f'{staged_well.name}: Access denied')
                    continue

                if staged_well.status == 'IMPORTED':
                    error_count += 1
                    errors.append(f'{staged_well.name}: Already imported')
                    continue

                serializer = StagedWellUpdateSerializer(staged_well, data=fields, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    # Update status if all required fields are filled
                    if staged_well.is_ready_to_import:
                        staged_well.status = 'COMPLETED'
                        staged_well.completed_by = request.user
                        staged_well.completed_at = timezone.now()
                        staged_well.save()
                    success_count += 1
                else:
                    error_count += 1
                    errors.append(f'{well_id}: {serializer.errors}')

            except StagedWell.DoesNotExist:
                error_count += 1
                errors.append(f'{well_id}: Not found')
            except Exception as e:
                error_count += 1
                errors.append(f'{well_id}: {str(e)}')

        return Response({
            'success': True,
            'updated_count': success_count,
            'error_count': error_count,
            'errors': errors[:20],  # Limit error details
            'message': f'Updated {success_count} wells, {error_count} errors'
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Error in bulk update staged wells: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT', 'PATCH'])
def update_staged_well(request, staged_well_id):
    """
    Update additional fields for a staged well
    """
    try:
        staged_well = get_object_or_404(StagedWell, id=staged_well_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and staged_well.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Don't allow updates if already imported
        if staged_well.status == 'IMPORTED':
            return Response({'error': 'Cannot update imported well'}, status=status.HTTP_400_BAD_REQUEST)
        
        from .serializers import StagedWellUpdateSerializer
        serializer = StagedWellUpdateSerializer(staged_well, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            
            # Update status if all required fields are filled
            if staged_well.is_ready_to_import:
                staged_well.status = 'COMPLETED'
                staged_well.completed_by = request.user
                staged_well.completed_at = timezone.now()
                staged_well.save()
            
            from .serializers import StagedWellSerializer
            response_serializer = StagedWellSerializer(staged_well)
            
            return Response({
                'success': True,
                'message': 'Staged well updated successfully',
                'staged_well': response_serializer.data
            }, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        logger.error(f"Error updating staged well: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def finalize_staged_well(request, staged_well_id):
    """
    Finalize a staged well by moving it to the Well table
    """
    try:
        staged_well = get_object_or_404(StagedWell, id=staged_well_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and staged_well.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if already imported
        if staged_well.status == 'IMPORTED':
            return Response({'error': 'Well already imported'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if all required fields are filled
        if not staged_well.is_ready_to_import:
            return Response({
                'error': 'Cannot finalize well with missing fields',
                'missing_fields': staged_well.missing_fields
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if well already exists with same name
        if Well.objects.filter(name=staged_well.name).exists():
            return Response({'error': f'Well with name {staged_well.name} already exists'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get next serial number
        max_sn = Well.objects.aggregate(models.Max('sn'))['sn__max'] or 0
        next_sn = max_sn + 1
        
        # Create the well
        with transaction.atomic():
            well = Well.objects.create(
                sn=next_sn,
                location=staged_well.location,
                asset_id=staged_well.asset_id,
                name=staged_well.name,
                well_type=staged_well.well_type,
                well_profile=staged_well.well_profile,
                depth=staged_well.depth,
                rig_capacity_required_hp=staged_well.rig_capacity_required_hp,
                drl_days=staged_well.drl_days,
                pt_days=staged_well.pt_days,
                duration=staged_well.duration,
                latitude=staged_well.latitude,
                longitude=staged_well.longitude,
                rtd=staged_well.rtd,
                bop_stack=staged_well.bop_stack,
                tds_requirement=staged_well.tds_requirement,
                footprint=staged_well.footprint,
                preferred_rig=staged_well.preferred_rig or '',
                expected_potential=staged_well.expected_potential or '',
                priority=staged_well.priority
            )
            
            # ILM well pair distances are calculated automatically
            # via the post_save signal (well_saved_trigger_ilm) using on_commit
            
            # Update staged well status
            staged_well.status = 'IMPORTED'
            staged_well.imported_at = timezone.now()
            staged_well.imported_well = well
            staged_well.save()
        
        return Response({
            'success': True,
            'message': f'Well {well.name} successfully imported',
            'well_id': well.id,
            'well_sn': well.sn
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error finalizing staged well: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def finalize_all_staged_wells(request):
    """
    Finalize all ready staged wells by moving them to the Well table
    """
    try:
        user_location = get_user_location(request.user)
        
        # Get all completed staged wells for user's location
        if user_location:
            staged_wells = StagedWell.objects.filter(
                location=user_location,
                status='COMPLETED'
            )
        else:
            staged_wells = StagedWell.objects.filter(status='COMPLETED')
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        # Use a single transaction so all on_commit ILM callbacks fire
        # only after the entire batch, preventing SQLite lock contention
        with transaction.atomic():
            for staged_well in staged_wells:
                try:
                    # Check if well already exists
                    if Well.objects.filter(name=staged_well.name).exists():
                        skipped_count += 1
                        errors.append(f'{staged_well.name}: Well already exists')
                        continue
                    
                    # Get next serial number
                    max_sn = Well.objects.aggregate(models.Max('sn'))['sn__max'] or 0
                    next_sn = max_sn + 1
                    
                    # Create the well
                    well = Well.objects.create(
                        sn=next_sn,
                        location=staged_well.location,
                        asset_id=staged_well.asset_id,
                        name=staged_well.name,
                        well_type=staged_well.well_type,
                        well_profile=staged_well.well_profile,
                        depth=staged_well.depth,
                        rig_capacity_required_hp=staged_well.rig_capacity_required_hp,
                        drl_days=staged_well.drl_days,
                        pt_days=staged_well.pt_days,
                        duration=staged_well.duration,
                        latitude=staged_well.latitude,
                        longitude=staged_well.longitude,
                        rtd=staged_well.rtd,
                        bop_stack=staged_well.bop_stack,
                        tds_requirement=staged_well.tds_requirement,
                        footprint=staged_well.footprint,
                        preferred_rig=staged_well.preferred_rig or '',
                        expected_potential=staged_well.expected_potential or '',
                        priority=staged_well.priority
                    )
                    
                    # Update staged well status
                    staged_well.status = 'IMPORTED'
                    staged_well.imported_at = timezone.now()
                    staged_well.imported_well = well
                    staged_well.save()
                    
                    imported_count += 1
                    
                except Exception as e:
                    logger.error(f"Error importing staged well {staged_well.name}: {str(e)}")
                    errors.append(f'{staged_well.name}: {str(e)}')
                    skipped_count += 1
        
        return Response({
            'success': True,
            'imported_count': imported_count,
            'skipped_count': skipped_count,
            'errors': errors
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error finalizing all staged wells: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def bulk_set_rtd_staged_wells(request):
    """
    Set RTD date for multiple staged wells at once
    """
    try:
        data = request.data
        well_ids = data.get('well_ids', [])
        rtd_date = data.get('rtd')
        
        if not well_ids:
            return Response({'error': 'No well IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not rtd_date:
            return Response({'error': 'No RTD date provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse the date
        from datetime import datetime
        rtd_parsed = datetime.strptime(rtd_date, '%Y-%m-%d').date()
        
        user_location = get_user_location(request.user)
        
        # Get staged wells by IDs and location
        if user_location:
            staged_wells = StagedWell.objects.filter(
                id__in=well_ids,
                location=user_location
            ).exclude(status='IMPORTED')
        else:
            staged_wells = StagedWell.objects.filter(
                id__in=well_ids
            ).exclude(status='IMPORTED')
        
        updated_count = staged_wells.update(rtd=rtd_parsed)
        
        return Response({
            'success': True,
            'updated_count': updated_count
        }, status=status.HTTP_200_OK)
        
    except ValueError as e:
        logger.error(f"Error parsing RTD date: {str(e)}")
        return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f"Error bulk setting RTD for staged wells: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_staged_well(request, staged_well_id):
    """
    Delete a staged well (only if not yet imported)
    """
    try:
        staged_well = get_object_or_404(StagedWell, id=staged_well_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and staged_well.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Don't allow deletion if already imported
        if staged_well.status == 'IMPORTED':
            return Response({'error': 'Cannot delete imported well'}, status=status.HTTP_400_BAD_REQUEST)
        
        well_name = staged_well.name
        staged_well.delete()
        
        return Response({
            'success': True,
            'message': f'Staged well {well_name} deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting staged well: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def bulk_delete_staged_wells(request):
    """
    Delete all staged wells matching the current filter (or all if no filter).
    Accepts optional query params:
        - location: filter by location
        - include_imported: if 'true', also delete imported staged wells
    """
    try:
        from .models import CompanyCode
        from django.db.models import Q
        user_location = get_user_location(request.user)
        location_filter = request.query_params.get('location', None)
        include_imported = request.query_params.get('include_imported', 'false').lower() == 'true'
        
        if include_imported:
            qs = StagedWell.objects.all()
        else:
            qs = StagedWell.objects.exclude(status='IMPORTED')
        
        if user_location:
            # user_location is already a CompanyCode object
            qs = qs.filter(location=user_location)
        elif location_filter:
            # Resolve location name string to CompanyCode object
            location_obj = CompanyCode.objects.filter(
                Q(location__iexact=location_filter) |
                Q(company_code__iexact=location_filter) |
                Q(name__iexact=location_filter)
            ).first()
            if location_obj:
                qs = qs.filter(location=location_obj)
            else:
                return Response({'error': f'Location "{location_filter}" not found'}, status=status.HTTP_400_BAD_REQUEST)
        
        count = qs.count()
        qs.delete()
        
        return Response({
            'success': True,
            'message': f'{count} staged well(s) deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error bulk deleting staged wells: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# BENCHMARK MANAGEMENT VIEWS
# =============================================================================

@login_required
def benchmark_management(request):
    """
    Page to manage drilling benchmarks
    """
    return render(request, 'scheduler/benchmark_management.html')


@login_required
@api_view(['GET'])
def get_benchmarks(request):
    """
    Get all drilling benchmarks - filtered by user location if not admin
    """
    try:
        benchmarks = DrillingBenchmark.objects.all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            benchmarks = benchmarks.filter(location=user_location)
        
        benchmarks = benchmarks.order_by('location', 'pool', 'field')
        
        data = []
        for benchmark in benchmarks:
            data.append({
                'id': str(benchmark.id),
                'location': benchmark.location.location if benchmark.location else '',
                'pool': benchmark.pool,
                'well_category': benchmark.well_category,
                'well_depth_start': benchmark.well_depth_start,
                'well_depth_end': benchmark.well_depth_end,
                'field': benchmark.field,
                'drilling_depth': benchmark.drilling_depth,
                'benchmark_days': float(benchmark.benchmark_days),
                'loc_spec_factor': benchmark.loc_spec_factor,
                'created_at': benchmark.created_at.isoformat(),
                'updated_at': benchmark.updated_at.isoformat(),
            })
        
        return Response({'data': data}, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching benchmarks: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_benchmark(request):
    """
    Create a new drilling benchmark
    """
    try:
        from .models import CompanyCode
        data = request.data
        
        # Get CompanyCode from location string
        location_obj = None
        if data.get('location'):
            location_obj = CompanyCode.objects.filter(
                models.Q(location=data['location']) | 
                models.Q(name=data['location']) |
                models.Q(company_code=data['location'])
            ).first()
        
        # Normalize field name to title case for consistency
        field_value = data.get('field', '')
        if field_value:
            field_value = field_value.strip().title()
        
        benchmark = DrillingBenchmark.objects.create(
            location=location_obj,
            pool=data['pool'],
            well_category=data['well_category'],
            well_depth_start=data['well_depth_start'],
            well_depth_end=data['well_depth_end'],
            field=field_value,
            drilling_depth=data['drilling_depth'],
            benchmark_days=data['benchmark_days'],
            loc_spec_factor=data.get('loc_spec_factor', 'Main Pool')
        )
        
        return Response({
            'success': True,
            'message': 'Benchmark created successfully',
            'benchmark': {
                'id': str(benchmark.id),
                'location': benchmark.location.location if benchmark.location else '',
                'pool': benchmark.pool,
                'well_category': benchmark.well_category,
                'well_depth_start': benchmark.well_depth_start,
                'well_depth_end': benchmark.well_depth_end,
                'field': benchmark.field,
                'drilling_depth': benchmark.drilling_depth,
                'benchmark_days': float(benchmark.benchmark_days),
                'loc_spec_factor': benchmark.loc_spec_factor,
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating benchmark: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_benchmark(request, benchmark_id):
    """
    Update an existing drilling benchmark
    """
    try:
        from .models import CompanyCode
        benchmark = get_object_or_404(DrillingBenchmark, id=benchmark_id)
        data = request.data
        
        # Handle location update
        if 'location' in data:
            if data['location']:
                location_obj = CompanyCode.objects.filter(
                    models.Q(location=data['location']) | 
                    models.Q(name=data['location']) |
                    models.Q(company_code=data['location'])
                ).first()
                benchmark.location = location_obj
            else:
                benchmark.location = None
        
        benchmark.pool = data.get('pool', benchmark.pool)
        benchmark.well_category = data.get('well_category', benchmark.well_category)
        benchmark.well_depth_start = data.get('well_depth_start', benchmark.well_depth_start)
        benchmark.well_depth_end = data.get('well_depth_end', benchmark.well_depth_end)
        # Normalize field name to title case for consistency
        field_val = data.get('field', benchmark.field)
        if field_val:
            field_val = field_val.strip().title()
        benchmark.field = field_val
        benchmark.drilling_depth = data.get('drilling_depth', benchmark.drilling_depth)
        benchmark.benchmark_days = data.get('benchmark_days', benchmark.benchmark_days)
        benchmark.loc_spec_factor = data.get('loc_spec_factor', benchmark.loc_spec_factor)
        benchmark.save()
        
        return Response({
            'success': True,
            'message': 'Benchmark updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating benchmark: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_benchmark(request, benchmark_id):
    """
    Delete a drilling benchmark
    """
    try:
        benchmark = get_object_or_404(DrillingBenchmark, id=benchmark_id)
        field_name = benchmark.field
        benchmark.delete()
        
        return Response({
            'success': True,
            'message': f'Benchmark for {field_name} deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting benchmark: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# RIG BUILDING NORMS MANAGEMENT VIEWS
# =============================================================================

@login_required
def rig_norms_management(request):
    """
    Page to manage rig building norms
    """
    return render(request, 'scheduler/rig_norms_management.html')


@login_required
@api_view(['GET'])
def get_rig_norms(request):
    """
    Get all rig building norms - filtered by location parameter or user location if not admin
    """
    try:
        norms = RigBuildingNorm.objects.all()
        
        # Check for explicit location parameter in query string
        location_param = request.GET.get('location', '').strip()
        
        if location_param:
            # Filter by explicit location parameter
            from .models import CompanyCode
            location_obj = CompanyCode.objects.filter(
                models.Q(location__iexact=location_param) | 
                models.Q(name__iexact=location_param) |
                models.Q(company_code__iexact=location_param)
            ).first()
            if location_obj:
                norms = norms.filter(location=location_obj)
        else:
            # Apply location-based filtering for non-admin users
            user_location = get_user_location(request.user)
            if user_location:
                norms = norms.filter(location=user_location)
        
        norms = norms.order_by('location', 'rig_name')
        
        data = []
        for norm in norms:
            data.append({
                'id': str(norm.id),
                'location': norm.location.location if norm.location else '',
                'rig_name': norm.rig_name,
                'days': norm.days,
                'top_drive': norm.top_drive,
                'rig_type': norm.rig_type,
                'created_at': norm.created_at.isoformat(),
                'updated_at': norm.updated_at.isoformat(),
            })
        
        return Response({'data': data}, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching rig norms: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_rig_norm(request):
    """
    Create a new rig building norm
    """
    try:
        from .models import CompanyCode
        data = request.data
        
        # Get CompanyCode from location string
        location_obj = None
        if data.get('location'):
            location_obj = CompanyCode.objects.filter(
                models.Q(location=data['location']) | 
                models.Q(name=data['location']) |
                models.Q(company_code=data['location'])
            ).first()
        
        # Check if norm with same location and rig name already exists
        existing = RigBuildingNorm.objects.filter(
            location=location_obj,
            rig_name=data['rig_name']
        ).first()
        
        if existing:
            return Response({
                'error': 'Rig norm with this location and name already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        norm = RigBuildingNorm.objects.create(
            location=location_obj,
            rig_name=data['rig_name'],
            days=data['days'],
            top_drive=data.get('top_drive', False),
            rig_type=data.get('rig_type', 'Fixed')
        )
        
        return Response({
            'success': True,
            'message': 'Rig norm created successfully',
            'norm': {
                'id': str(norm.id),
                'location': norm.location.location if norm.location else '',
                'rig_name': norm.rig_name,
                'days': norm.days,
                'top_drive': norm.top_drive,
                'rig_type': norm.rig_type,
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating rig norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_rig_norm(request, norm_id):
    """
    Update an existing rig building norm
    """
    try:
        from .models import CompanyCode
        norm = get_object_or_404(RigBuildingNorm, id=norm_id)
        data = request.data
        
        # Handle location update
        if 'location' in data:
            if data['location']:
                location_obj = CompanyCode.objects.filter(
                    models.Q(location=data['location']) | 
                    models.Q(name=data['location']) |
                    models.Q(company_code=data['location'])
                ).first()
                norm.location = location_obj
            else:
                norm.location = None
        
        norm.rig_name = data.get('rig_name', norm.rig_name)
        norm.days = data.get('days', norm.days)
        norm.top_drive = data.get('top_drive', norm.top_drive)
        norm.rig_type = data.get('rig_type', norm.rig_type)
        norm.save()
        
        return Response({
            'success': True,
            'message': 'Rig norm updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating rig norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_rig_norm(request, norm_id):
    """
    Delete a rig building norm
    """
    try:
        norm = get_object_or_404(RigBuildingNorm, id=norm_id)
        rig_name = norm.rig_name
        norm.delete()
        
        return Response({
            'success': True,
            'message': f'Rig norm for {rig_name} deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting rig norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def map_rigs_to_norms(request):
    """
    Automatically map rigs to rig building norms using partial text matching
    Returns proposed mappings for user confirmation
    Filters by location parameter if provided
    """
    try:
        from difflib import SequenceMatcher
        from .models import CompanyCode
        
        # Get location filter from request body
        location_param = request.data.get('location', '').strip()
        location_obj = None
        
        if location_param:
            # Find the location object
            location_obj = CompanyCode.objects.filter(
                models.Q(location__iexact=location_param) | 
                models.Q(name__iexact=location_param) |
                models.Q(company_code__iexact=location_param)
            ).first()
        
        # Get all active rigs and norms
        # Note: Rig has SoftDeleteMixin (is_deleted field), but RigBuildingNorm does not
        rigs = Rig.objects.filter(is_deleted=False).select_related('location')
        norms = RigBuildingNorm.objects.all().select_related('location')
        
        # Apply location filter if specified
        if location_obj:
            rigs = rigs.filter(location=location_obj)
            norms = norms.filter(location=location_obj)
        
        proposed_mappings = []
        
        for rig in rigs:
            best_match = None
            best_score = 0
            
            for norm in norms:
                # Calculate similarity score
                score = 0
                
                # Primary matching: Compare rig names using partial matching
                rig_name_lower = rig.name.lower()
                norm_name_lower = norm.rig_name.lower()
                
                # Check if one contains the other
                if norm_name_lower in rig_name_lower or rig_name_lower in norm_name_lower:
                    score += 0.6
                
                # Use SequenceMatcher for similarity
                similarity = SequenceMatcher(None, rig_name_lower, norm_name_lower).ratio()
                score += similarity * 0.4
                
                # Boost score if locations match
                if rig.location and norm.location and rig.location.id == norm.location.id:
                    score += 0.2
                
                # Boost score if rig types match
                if rig.rig_type and norm.rig_type:
                    if rig.rig_type.lower() == norm.rig_type.lower():
                        score += 0.1
                
                # Update best match if this is better
                if score > best_score and score > 0.3:  # Minimum threshold of 30%
                    best_score = score
                    best_match = norm
            
            # If we found a good match, add to proposed mappings
            if best_match:
                proposed_mappings.append({
                    'rig_id': str(rig.id),
                    'rig_name': rig.name,
                    'rig_location': rig.location.location if rig.location else 'N/A',
                    'rig_type': rig.rig_type,
                    'rig_tds': rig.tds_availability,
                    'norm_id': str(best_match.id),
                    'norm_rig_name': best_match.rig_name,
                    'norm_location': best_match.location.location if best_match.location else 'N/A',
                    'norm_days': best_match.days,
                    'confidence': round(best_score * 100, 1)
                })
        
        # Sort by confidence score (highest first)
        proposed_mappings.sort(key=lambda x: x['confidence'], reverse=True)
        
        return Response({
            'success': True,
            'mappings': proposed_mappings,
            'total_rigs': rigs.count(),
            'mapped_count': len(proposed_mappings)
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error mapping rigs to norms: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def save_rig_norm_mappings(request):
    """
    Save the confirmed rig-to-norm mappings
    Updates the Rig model with the matched norm reference
    """
    try:
        mappings = request.data.get('mappings', [])
        
        if not mappings:
            return Response({
                'success': False,
                'message': 'No mappings provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        saved_count = 0
        errors = []
        
        for mapping in mappings:
            try:
                rig = Rig.objects.get(id=mapping['rig_id'])
                norm = RigBuildingNorm.objects.get(id=mapping['norm_id'])
                
                # Store the norm information in a JSON field or create a relationship
                # For now, we'll add a note to the rig indicating the mapping
                # You can extend this to add a ForeignKey field to Rig model later
                
                # Check if Rig model has a rig_building_norm field
                if hasattr(rig, 'rig_building_norm'):
                    rig.rig_building_norm = norm
                    rig.save()
                    saved_count += 1
                else:
                    # If no direct field exists, we can store it in notes or create the field
                    errors.append(f"Rig {rig.name}: No rig_building_norm field available")
                    
            except Rig.DoesNotExist:
                errors.append(f"Rig with ID {mapping['rig_id']} not found")
            except RigBuildingNorm.DoesNotExist:
                errors.append(f"Rig norm with ID {mapping['norm_id']} not found")
            except Exception as e:
                errors.append(f"Error saving mapping: {str(e)}")
        
        response_data = {
            'success': True,
            'saved_count': saved_count,
            'total_mappings': len(mappings)
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] = f'Saved {saved_count} of {len(mappings)} mappings with some errors'
        else:
            response_data['message'] = f'Successfully saved {saved_count} mappings'
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error saving rig norm mappings: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# RIG BUILDING ADJUSTMENT RULES API
# =============================================================================

@login_required
@api_view(['GET'])
def get_rig_adjustments(request):
    """
    Get all rig building adjustment rules - filtered by user location if not admin
    """
    try:
        adjustments = RigBuildingAdjustment.objects.all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            adjustments = adjustments.filter(location=user_location)
        
        adjustments = adjustments.order_by('location', 'category', 'priority', 'condition')
        
        data = []
        for adj in adjustments:
            data.append({
                'id': str(adj.id),
                'location': adj.location.location if adj.location else None,
                'condition': adj.condition,
                'category': adj.category,
                'category_display': adj.get_category_display(),
                'adjustment_type': adj.adjustment_type,
                'adjustment_type_display': adj.get_adjustment_type_display(),
                'adjustment_value': float(adj.adjustment_value) if adj.adjustment_value else None,
                'adjustment_display': adj.adjustment_display,
                'unit': adj.unit,
                'min_distance': float(adj.min_distance) if adj.min_distance else None,
                'max_distance': float(adj.max_distance) if adj.max_distance else None,
                'applies_to_rig_type': adj.applies_to_rig_type,
                'max_depth': float(adj.max_depth) if adj.max_depth else None,
                'notes': adj.notes,
                'is_active': adj.is_active,
                'priority': adj.priority,
                'created_at': adj.created_at.isoformat(),
                'updated_at': adj.updated_at.isoformat(),
            })
        
        return Response({'data': data}, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching rig adjustments: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_rig_adjustment(request):
    """
    Create a new rig building adjustment rule
    """
    try:
        from .models import CompanyCode
        data = request.data
        
        # Handle location
        location_obj = None
        if data.get('location'):
            location_obj = CompanyCode.objects.filter(
                models.Q(location=data['location']) |
                models.Q(name=data['location']) |
                models.Q(company_code=data['location'])
            ).first()
        
        adjustment = RigBuildingAdjustment.objects.create(
            location=location_obj,
            condition=data['condition'],
            category=data.get('category', 'other'),
            adjustment_type=data.get('adjustment_type', 'add'),
            adjustment_value=data.get('adjustment_value'),
            adjustment_display=data['adjustment_display'],
            unit=data.get('unit'),
            min_distance=data.get('min_distance'),
            max_distance=data.get('max_distance'),
            applies_to_rig_type=data.get('applies_to_rig_type'),
            max_depth=data.get('max_depth'),
            notes=data.get('notes'),
            is_active=data.get('is_active', True),
            priority=data.get('priority', 0)
        )
        
        return Response({
            'success': True,
            'message': 'Adjustment rule created successfully',
            'adjustment': {
                'id': str(adjustment.id),
                'condition': adjustment.condition,
                'adjustment_display': adjustment.adjustment_display,
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating rig adjustment: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_rig_adjustment(request, adjustment_id):
    """
    Update an existing rig building adjustment rule
    """
    try:
        from .models import CompanyCode
        adjustment = get_object_or_404(RigBuildingAdjustment, id=adjustment_id)
        data = request.data
        
        # Handle location update
        if 'location' in data:
            if data['location']:
                location_obj = CompanyCode.objects.filter(
                    models.Q(location=data['location']) |
                    models.Q(name=data['location']) |
                    models.Q(company_code=data['location'])
                ).first()
                adjustment.location = location_obj
            else:
                adjustment.location = None
        
        adjustment.condition = data.get('condition', adjustment.condition)
        adjustment.category = data.get('category', adjustment.category)
        adjustment.adjustment_type = data.get('adjustment_type', adjustment.adjustment_type)
        adjustment.adjustment_value = data.get('adjustment_value', adjustment.adjustment_value)
        adjustment.adjustment_display = data.get('adjustment_display', adjustment.adjustment_display)
        adjustment.unit = data.get('unit', adjustment.unit)
        adjustment.min_distance = data.get('min_distance', adjustment.min_distance)
        adjustment.max_distance = data.get('max_distance', adjustment.max_distance)
        adjustment.applies_to_rig_type = data.get('applies_to_rig_type', adjustment.applies_to_rig_type)
        adjustment.max_depth = data.get('max_depth', adjustment.max_depth)
        adjustment.notes = data.get('notes', adjustment.notes)
        adjustment.is_active = data.get('is_active', adjustment.is_active)
        adjustment.priority = data.get('priority', adjustment.priority)
        adjustment.save()
        
        return Response({
            'success': True,
            'message': 'Adjustment rule updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating rig adjustment: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_rig_adjustment(request, adjustment_id):
    """
    Delete a rig building adjustment rule
    """
    try:
        adjustment = get_object_or_404(RigBuildingAdjustment, id=adjustment_id)
        condition = adjustment.condition
        adjustment.delete()
        
        return Response({
            'success': True,
            'message': f'Adjustment rule deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting rig adjustment: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# AUTO-CALCULATION HELPER FUNCTIONS
# =============================================================================

def calculate_drilling_days(field_name, well_profile, depth, location_identifier=None):
    """
    Calculate DRL_DAYS based on benchmark data and daily drilling rates.
    
    Logic:
    1. Find matching benchmark where field matches and well depth is in range
    2. Get benchmark_days for drilling_depth from benchmark
    3. Calculate difference between benchmark's drilling_depth and actual well depth
    4. Find daily drilling rate for that depth range
    5. Adjust benchmark_days based on depth difference and drilling rate
    
    Example: 
    - Well VDEF at 1765m in VADATAL field
    - Benchmark: (Cambay, VD 3, Directional, 1725-2100, Vadatal, 1870m, 19.5 days)
    - Daily Rate: (CAMBAY, 1501-2000, Vadatal, 50 m/day)
    - Depth diff: 1870 - 1765 = 105m (well is shallower)
    - Time adjustment: 105 / 50 = 2.1 days
    - DRL Days: 19.5 - 2.1 = 17.4 days
    """
    from .models import DailyDrillingRate
    
    if not field_name or not depth:
        return None
    
    try:
        depth = int(depth)
    except (ValueError, TypeError):
        return None
    
    # Resolve location identifier (may be CompanyCode id or location string)
    location_obj = None
    if location_identifier:
        try:
            location_obj = CompanyCode.objects.filter(id=location_identifier).first()
        except Exception:
            location_obj = None

        if not location_obj:
            try:
                location_obj = CompanyCode.objects.filter(location__iexact=str(location_identifier)).first()
            except Exception:
                location_obj = None

    # Map well profile code to well_category
    profile_mapping = {
        'DI': 'Directional',
        'VE': 'Vertical', 
        'SD': 'Sidetrack',
        'Directional': 'Directional',
        'Vertical': 'Vertical',
        'Sidetrack': 'Sidetrack',
    }
    well_category = profile_mapping.get(well_profile, 'Directional')
    
    # Step 1: Find matching benchmark where:
    # - field matches (case-insensitive)
    # - well depth falls within well_depth_start and well_depth_end range
    # - well_category matches (optional, try with category first, then without)
    # If we have a location, prefer benchmarks for that location first
    benchmark = None
    if location_obj:
        benchmark = DrillingBenchmark.objects.filter(
            location=location_obj,
            field__iexact=field_name,
            well_category=well_category,
            well_depth_start__lte=depth,
            well_depth_end__gte=depth
        ).first()

    # If no location-specific benchmark, try with category only
    if not benchmark:
        benchmark = DrillingBenchmark.objects.filter(
            field__iexact=field_name,
            well_category=well_category,
            well_depth_start__lte=depth,
            well_depth_end__gte=depth
        ).first()
    
    # If still no match (either category mismatch or no category in DB), try without well_category
    if not benchmark:
        # If location provided, try location + depth only
        if location_obj:
            benchmark = DrillingBenchmark.objects.filter(
                location=location_obj,
                field__iexact=field_name,
                well_depth_start__lte=depth,
                well_depth_end__gte=depth
            ).first()

        if not benchmark:
            benchmark = DrillingBenchmark.objects.filter(
                field__iexact=field_name,
                well_depth_start__lte=depth,
                well_depth_end__gte=depth
            ).first()
    
    if not benchmark:
        logger.debug(f"No benchmark found for field={field_name}, depth={depth}, category={well_category}")
        return None
    
    benchmark_days = float(benchmark.benchmark_days)
    benchmark_drilling_depth = benchmark.drilling_depth
    
    logger.info(f"Found benchmark: field={benchmark.field}, depth_range={benchmark.well_depth_start}-{benchmark.well_depth_end}, "
                f"drilling_depth={benchmark_drilling_depth}, benchmark_days={benchmark_days}")
    
    # Step 2: Calculate depth difference
    depth_difference = benchmark_drilling_depth - depth
    # Positive means well is shallower than benchmark, negative means deeper
    
    if depth_difference == 0:
        # Well depth matches benchmark drilling depth exactly
        return benchmark_days
    
    # Step 3: Find daily drilling rate for this depth range and location/field
    # Get location from benchmark
    location = benchmark.location
    
    # Find matching daily drilling rate
    # Try to match by location + field + depth range
    daily_rate = None
    
    if location:
        daily_rate = DailyDrillingRate.objects.filter(
            location=location,
            field__iexact=field_name,
            depth_start__lte=depth,
            depth_end__gte=depth
        ).first()
        
        # If not found with field, try just location + depth range
        if not daily_rate:
            daily_rate = DailyDrillingRate.objects.filter(
                location=location,
                depth_start__lte=depth,
                depth_end__gte=depth
            ).first()
    
    # If still not found, try matching by field only
    if not daily_rate:
        daily_rate = DailyDrillingRate.objects.filter(
            field__iexact=field_name,
            depth_start__lte=depth,
            depth_end__gte=depth
        ).first()
    
    if not daily_rate:
        # No daily rate found - return benchmark days as-is
        logger.warning(f"No daily drilling rate found for depth={depth}, returning benchmark_days as-is")
        return benchmark_days
    
    per_day_depth = float(daily_rate.per_day_depth)
    
    if per_day_depth <= 0:
        logger.warning(f"Invalid per_day_depth={per_day_depth}, returning benchmark_days as-is")
        return benchmark_days
    
    logger.info(f"Found daily rate: location={daily_rate.location}, field={daily_rate.field}, "
                f"depth_range={daily_rate.depth_start}-{daily_rate.depth_end}, rate={per_day_depth} m/day")
    
    # Step 4: Calculate time adjustment
    # depth_difference / per_day_depth = days to add or subtract
    time_adjustment = abs(depth_difference) / per_day_depth
    
    # Step 5: Calculate final DRL days
    if depth_difference > 0:
        # Well is shallower than benchmark drilling depth -> subtract days
        drl_days = benchmark_days - time_adjustment
    else:
        # Well is deeper than benchmark drilling depth -> add days
        drl_days = benchmark_days + time_adjustment
    
    # Ensure we don't return negative days
    drl_days = max(drl_days, 1.0)
    
    logger.info(f"Calculated DRL days: benchmark={benchmark_days}, depth_diff={depth_difference}, "
                f"adjustment={time_adjustment:.2f}, final_drl_days={drl_days:.2f}")
    
    return math.ceil(drl_days)


def calculate_pt_days(location_identifier=None, depth=None, well_type=None):
    """
    Calculate PT_DAYS by looking up `CompletionTestingNorm` table.

    Matching priority:
    1. location (CompanyCode) + depth in [well_depth_start, well_depth_end] + well_type
    2. depth + well_type (regardless of location)
    3. depth only (regardless of well_type)

    Returns: float rounded to 2 decimals, or default 7.0 if no norm found.
    """
    from .models import CompletionTestingNorm, CompanyCode

    DEFAULT_PT_DAYS = 7.0

    # Validate depth
    try:
        depth_val = int(depth)
    except (TypeError, ValueError):
        return DEFAULT_PT_DAYS

    # Try to resolve location identifier to CompanyCode instance
    location_obj = None
    if location_identifier:
        # Try UUID/ID lookup
        try:
            location_obj = CompanyCode.objects.filter(id=location_identifier).first()
        except Exception:
            location_obj = None

        # If not found by ID, try matching by location string (case-insensitive)
        if not location_obj:
            try:
                location_obj = CompanyCode.objects.filter(location__iexact=str(location_identifier)).first()
            except Exception:
                location_obj = None

    # 1. location + depth + well_type
    if location_obj and well_type:
        norm = CompletionTestingNorm.objects.filter(
            location=location_obj,
            well_depth_start__lte=depth_val,
            well_depth_end__gte=depth_val,
            well_type__iexact=well_type
        ).first()
        if norm:
            return round(float(norm.days), 2)

    # 2. depth + well_type (any location)
    if well_type:
        norm = CompletionTestingNorm.objects.filter(
            well_depth_start__lte=depth_val,
            well_depth_end__gte=depth_val,
            well_type__iexact=well_type
        ).first()
        if norm:
            return round(float(norm.days), 2)

    # 3. depth only (any well_type)
    norm = CompletionTestingNorm.objects.filter(
        well_depth_start__lte=depth_val,
        well_depth_end__gte=depth_val
    ).first()
    if norm:
        return round(float(norm.days), 2)

    return DEFAULT_PT_DAYS


@login_required
@api_view(['POST'])
def calculate_well_parameters(request):
    """
    Calculate DRL_DAYS, PT_DAYS, and Duration for a well
    """
    try:
        data = request.data
        field_name = data.get('field_name')
        well_profile = data.get('well_profile', 'DI')  # Default to Directional
        depth = data.get('depth')

        # Resolve location and well_type from request payload (compatibility keys)
        location_identifier = data.get('location') or data.get('location_id') or data.get('location_value')
        well_type = data.get('well_type') or data.get('wellType') or data.get('well_type_name')

        # Calculate drilling days (pass location if available so benchmark lookup can prefer location-specific rows)
        drl_days = calculate_drilling_days(field_name, well_profile, depth, location_identifier)

        # Calculate PT days using completion testing norms (location + depth + well_type)
        pt_days = calculate_pt_days(location_identifier, depth, well_type)
        
        # Calculate duration
        duration = None
        if drl_days is not None:
            duration = int(drl_days + pt_days)
        
        return Response({
            'success': True,
            'drl_days': drl_days,
            'pt_days': pt_days,
            'duration': duration,
            'message': 'Parameters calculated successfully' if drl_days else 'No benchmark found for this field'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error calculating well parameters: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# Daily Drilling Rate Management Views
# ============================================================================

@login_required
def daily_drilling_rate_management(request):
    """
    Page to manage daily drilling rates
    """
    return render(request, 'scheduler/daily_drilling_rate_management.html')


@login_required
@api_view(['GET'])
def get_daily_drilling_rates(request):
    """
    Get all daily drilling rates - filtered by user location if not admin
    """
    try:
        from .models import DailyDrillingRate
        rates = DailyDrillingRate.objects.all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            rates = rates.filter(location=user_location)
        
        data = []
        for rate in rates:
            data.append({
                'id': str(rate.id),
                'location': rate.location.location if rate.location else '',
                'depth_start': rate.depth_start,
                'depth_end': rate.depth_end,
                'field': rate.field,
                'per_day_depth': float(rate.per_day_depth),
                'loc_spec_factor': rate.loc_spec_factor,
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching daily drilling rates: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['POST'])
def create_daily_drilling_rate(request):
    """
    Create a new daily drilling rate
    """
    try:
        from .models import DailyDrillingRate, CompanyCode
        
        data = request.data
        
        # Get CompanyCode for location
        location_obj = None
        if data.get('location'):
            location_obj = CompanyCode.objects.filter(
                models.Q(location=data.get('location')) | 
                models.Q(name=data.get('location')) |
                models.Q(company_code=data.get('location'))
            ).first()
        
        # Normalize field name to title case for consistency
        field_value = data.get('field', '')
        if field_value:
            field_value = field_value.strip().title()
        
        rate = DailyDrillingRate.objects.create(
            location=location_obj,
            depth_start=int(data.get('depth_start')),
            depth_end=int(data.get('depth_end')),
            field=field_value,
            per_day_depth=Decimal(str(data.get('per_day_depth'))),
            loc_spec_factor=data.get('loc_spec_factor', 'Main Pool'),
        )
        
        return Response({
            'success': True,
            'id': str(rate.id),
            'message': 'Daily drilling rate created successfully'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating daily drilling rate: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_daily_drilling_rate(request, rate_id):
    """
    Update an existing daily drilling rate
    """
    try:
        from .models import DailyDrillingRate, CompanyCode
        
        rate = get_object_or_404(DailyDrillingRate, id=rate_id)
        data = request.data
        
        # Update location if provided
        if data.get('location'):
            location_obj = CompanyCode.objects.filter(
                models.Q(location=data.get('location')) | 
                models.Q(name=data.get('location')) |
                models.Q(company_code=data.get('location'))
            ).first()
            if location_obj:
                rate.location = location_obj
        
        # Update field if provided - normalize to title case
        if 'field' in data:
            field_value = data.get('field', '')
            if field_value:
                field_value = field_value.strip().title()
            rate.field = field_value
        
        rate.depth_start = int(data.get('depth_start', rate.depth_start))
        rate.depth_end = int(data.get('depth_end', rate.depth_end))
        rate.per_day_depth = Decimal(str(data.get('per_day_depth', rate.per_day_depth)))
        rate.loc_spec_factor = data.get('loc_spec_factor', rate.loc_spec_factor)
        rate.save()
        
        return Response({
            'success': True,
            'message': 'Daily drilling rate updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating daily drilling rate: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_daily_drilling_rate(request, rate_id):
    """
    Delete a daily drilling rate
    """
    try:
        from .models import DailyDrillingRate
        
        rate = get_object_or_404(DailyDrillingRate, id=rate_id)
        rate.delete()
        
        return Response({
            'success': True,
            'message': 'Daily drilling rate deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting daily drilling rate: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['GET'])
def get_location_field_combinations(request):
    """
    Get unique location-field combinations from DailyDrillingRate, DrillingBenchmark,
    Well, and StagedWell tables.
    Automatically filters by user's assigned location if they are not admin.
    """
    try:
        from .models import DailyDrillingRate, DrillingBenchmark, StagedWell, Well, CompanyCode
        from django.db.models import Q, Count
        
        # Apply location-based filtering for non-admin users
        location_filter = Q()
        if not user_can_view_all_locations(request.user):
            user_location_name = get_user_assigned_location(request.user)
            if user_location_name:
                # Get CompanyCode object for the user's location
                location_obj = CompanyCode.objects.filter(
                    Q(location=user_location_name) | Q(company_code=user_location_name)
                ).first()
                if location_obj:
                    location_filter = Q(location=location_obj)
        
        # Get combinations from DailyDrillingRate
        drilling_combos = DailyDrillingRate.objects.filter(
            location_filter,
            field__isnull=False
        ).exclude(field='').values(
            'location__location', 'field'
        ).annotate(count=Count('id'))
        
        # Get combinations from DrillingBenchmark
        benchmark_combos = DrillingBenchmark.objects.filter(
            location_filter,
            field__isnull=False
        ).exclude(field='').values(
            'location__location', 'field'
        ).annotate(count=Count('id'))
        
        # Get combinations from StagedWell (staging area)
        staged_combos = StagedWell.objects.filter(
            location_filter,
            field_name__isnull=False
        ).exclude(field_name='').values(
            'location__location', 'field_name'
        ).annotate(count=Count('id'))
        
        # Get combinations from Well (imported/uploaded wells ready for scheduling)
        well_combos = Well.objects.filter(
            location_filter,
            field_name__isnull=False,
            is_deleted=False
        ).exclude(field_name='').values(
            'location__location', 'field_name'
        ).annotate(count=Count('id'))
        
        # Combine and deduplicate
        combo_dict = {}
        
        for combo in drilling_combos:
            location = combo['location__location'] or ''
            field = combo['field']
            key = (location, field)
            if key not in combo_dict:
                combo_dict[key] = {
                    'location': location, 
                    'field': field, 
                    'count': 0,
                    'drilling_count': 0,
                    'benchmark_count': 0,
                    'staged_count': 0,
                    'well_count': 0
                }
            combo_dict[key]['drilling_count'] += combo['count']
            combo_dict[key]['count'] += combo['count']
        
        for combo in benchmark_combos:
            location = combo['location__location'] or ''
            field = combo['field']
            key = (location, field)
            if key not in combo_dict:
                combo_dict[key] = {
                    'location': location, 
                    'field': field, 
                    'count': 0,
                    'drilling_count': 0,
                    'benchmark_count': 0,
                    'staged_count': 0,
                    'well_count': 0
                }
            combo_dict[key]['benchmark_count'] += combo['count']
            combo_dict[key]['count'] += combo['count']
        
        for combo in staged_combos:
            location = combo['location__location'] or ''
            field = combo['field_name']
            key = (location, field)
            if key not in combo_dict:
                combo_dict[key] = {
                    'location': location, 
                    'field': field, 
                    'count': 0,
                    'drilling_count': 0,
                    'benchmark_count': 0,
                    'staged_count': 0,
                    'well_count': 0
                }
            combo_dict[key]['staged_count'] += combo['count']
            combo_dict[key]['count'] += combo['count']
        
        for combo in well_combos:
            location = combo['location__location'] or ''
            field = combo['field_name']
            key = (location, field)
            if key not in combo_dict:
                combo_dict[key] = {
                    'location': location, 
                    'field': field, 
                    'count': 0,
                    'drilling_count': 0,
                    'benchmark_count': 0,
                    'staged_count': 0,
                    'well_count': 0
                }
            combo_dict[key]['well_count'] += combo['count']
            combo_dict[key]['count'] += combo['count']
        
        # Convert to list and sort
        result = sorted(combo_dict.values(), key=lambda x: (x['location'], x['field']))
        
        return JsonResponse(result, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching location-field combinations: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['PUT'])
def update_location_field_combination(request):
    """
    Update field name across DailyDrillingRate, DrillingBenchmark, and StagedWell tables
    """
    try:
        from .models import DailyDrillingRate, DrillingBenchmark, StagedWell, CompanyCode
        
        data = request.data
        location_name = data.get('location', '')
        old_field = data.get('old_field', '')
        new_field = data.get('new_field', '')
        
        if not new_field:
            return Response({'error': 'New field name is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get the location object
        location_obj = None
        if location_name:
            location_obj = CompanyCode.objects.filter(location=location_name).first()
        
        updated_count = 0
        
        # Update DailyDrillingRate records
        if location_obj:
            drilling_updated = DailyDrillingRate.objects.filter(
                location=location_obj,
                field=old_field
            ).update(field=new_field)
            updated_count += drilling_updated
        else:
            # If no location specified, update all records with matching field
            drilling_updated = DailyDrillingRate.objects.filter(
                location__isnull=True,
                field=old_field
            ).update(field=new_field)
            updated_count += drilling_updated
        
        # Update DrillingBenchmark records
        if location_obj:
            benchmark_updated = DrillingBenchmark.objects.filter(
                location=location_obj,
                field=old_field
            ).update(field=new_field)
            updated_count += benchmark_updated
        else:
            benchmark_updated = DrillingBenchmark.objects.filter(
                location__isnull=True,
                field=old_field
            ).update(field=new_field)
            updated_count += benchmark_updated
        
        # Update StagedWell records
        if location_obj:
            staged_updated = StagedWell.objects.filter(
                location=location_obj,
                field_name=old_field
            ).update(field_name=new_field)
            updated_count += staged_updated
        else:
            staged_updated = StagedWell.objects.filter(
                location__isnull=True,
                field_name=old_field
            ).update(field_name=new_field)
            updated_count += staged_updated
        
        return Response({
            'success': True,
            'message': f'Updated {updated_count} records with new field name',
            'updated_count': updated_count
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating field name: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# Additional Ops Drilling Management Views
# ============================================================================

@login_required
def additional_ops_drilling_management(request):
    """
    Page to manage additional ops drilling norms
    """
    return render(request, 'scheduler/additional_ops_drilling_management.html')


@login_required
@api_view(['GET'])
def get_coring_norms(request):
    """
    Get all coring norms - filtered by user location if not admin
    """
    try:
        from .models import CoringNorm
        norms = CoringNorm.objects.select_related('location').all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            norms = norms.filter(location=user_location)
        
        data = []
        for norm in norms:
            data.append({
                'id': str(norm.id),
                'location_id': norm.location.id if norm.location else None,
                'location': norm.location.location if norm.location else None,
                'depth_start': norm.depth_start,
                'depth_end': norm.depth_end,
                'additional_days': float(norm.additional_days),
            })
        
        return Response(data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching coring norms: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_coring_norm(request):
    """
    Create a new coring norm
    """
    try:
        from .models import CoringNorm, CompanyCode
        
        data = request.data
        location_id = data.get('location_id')
        location = CompanyCode.objects.get(id=location_id) if location_id else None
        
        norm = CoringNorm.objects.create(
            location=location,
            depth_start=int(data.get('depth_start')),
            depth_end=int(data.get('depth_end')),
            additional_days=Decimal(str(data.get('additional_days'))),
        )
        
        return Response({
            'success': True,
            'id': str(norm.id),
            'message': 'Coring norm created successfully'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating coring norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_coring_norm(request, norm_id):
    """
    Update an existing coring norm
    """
    try:
        from .models import CoringNorm, CompanyCode
        
        norm = get_object_or_404(CoringNorm, id=norm_id)
        data = request.data
        
        location_id = data.get('location_id')
        if location_id:
            norm.location = CompanyCode.objects.get(id=location_id)
        elif 'location_id' in data and location_id is None:
            norm.location = None
            
        norm.depth_start = int(data.get('depth_start', norm.depth_start))
        norm.depth_end = int(data.get('depth_end', norm.depth_end))
        norm.additional_days = Decimal(str(data.get('additional_days', norm.additional_days)))
        norm.save()
        
        return Response({
            'success': True,
            'message': 'Coring norm updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating coring norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_coring_norm(request, norm_id):
    """
    Delete a coring norm
    """
    try:
        from .models import CoringNorm
        
        norm = get_object_or_404(CoringNorm, id=norm_id)
        norm.delete()
        
        return Response({
            'success': True,
            'message': 'Coring norm deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting coring norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['GET'])
def get_casing_norms(request):
    """
    Get all casing norms - filtered by user location if not admin
    """
    try:
        from .models import CasingNorm
        norms = CasingNorm.objects.select_related('location').all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            norms = norms.filter(location=user_location)
        
        data = []
        for norm in norms:
            data.append({
                'id': str(norm.id),
                'location_id': norm.location.id if norm.location else None,
                'location': norm.location.location if norm.location else None,
                'depth_start': norm.depth_start,
                'depth_end': norm.depth_end,
                'additional_days': float(norm.additional_days),
            })
        
        return Response(data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching casing norms: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_casing_norm(request):
    """
    Create a new casing norm
    """
    try:
        from .models import CasingNorm, CompanyCode
        
        data = request.data
        location_id = data.get('location_id')
        location = CompanyCode.objects.get(id=location_id) if location_id else None
        
        norm = CasingNorm.objects.create(
            location=location,
            depth_start=int(data.get('depth_start')),
            depth_end=int(data.get('depth_end')),
            additional_days=Decimal(str(data.get('additional_days'))),
        )
        
        return Response({
            'success': True,
            'id': str(norm.id),
            'message': 'Casing norm created successfully'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating casing norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_casing_norm(request, norm_id):
    """
    Update an existing casing norm
    """
    try:
        from .models import CasingNorm, CompanyCode
        
        norm = get_object_or_404(CasingNorm, id=norm_id)
        data = request.data
        
        location_id = data.get('location_id')
        if location_id:
            norm.location = CompanyCode.objects.get(id=location_id)
        elif 'location_id' in data and location_id is None:
            norm.location = None
            
        norm.depth_start = int(data.get('depth_start', norm.depth_start))
        norm.depth_end = int(data.get('depth_end', norm.depth_end))
        norm.additional_days = Decimal(str(data.get('additional_days', norm.additional_days)))
        norm.save()
        
        return Response({
            'success': True,
            'message': 'Casing norm updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating casing norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_casing_norm(request, norm_id):
    """
    Delete a casing norm
    """
    try:
        from .models import CasingNorm
        
        norm = get_object_or_404(CasingNorm, id=norm_id)
        norm.delete()
        
        return Response({
            'success': True,
            'message': 'Casing norm deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting casing norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['GET'])
def get_hermetical_testing_norms(request):
    """
    Get all hermetical testing norms - filtered by user location if not admin
    """
    try:
        from .models import HermeticalTestingNorm
        norms = HermeticalTestingNorm.objects.select_related('location').all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            norms = norms.filter(location=user_location)
        
        data = []
        for norm in norms:
            data.append({
                'id': str(norm.id),
                'location_id': norm.location.id if norm.location else None,
                'location': norm.location.location if norm.location else None,
                'depth_start': norm.depth_start,
                'depth_end': norm.depth_end,
                'norm_days': float(norm.norm_days),
            })
        
        return Response(data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching hermetical testing norms: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_hermetical_testing_norm(request):
    """
    Create a new hermetical testing norm
    """
    try:
        from .models import HermeticalTestingNorm, CompanyCode
        
        data = request.data
        location_id = data.get('location_id')
        location = CompanyCode.objects.get(id=location_id) if location_id else None
        
        norm = HermeticalTestingNorm.objects.create(
            location=location,
            depth_start=int(data.get('depth_start')),
            depth_end=int(data.get('depth_end')),
            norm_days=Decimal(str(data.get('norm_days'))),
        )
        
        return Response({
            'success': True,
            'id': str(norm.id),
            'message': 'Hermetical testing norm created successfully'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating hermetical testing norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_hermetical_testing_norm(request, norm_id):
    """
    Update an existing hermetical testing norm
    """
    try:
        from .models import HermeticalTestingNorm, CompanyCode
        
        norm = get_object_or_404(HermeticalTestingNorm, id=norm_id)
        data = request.data
        
        location_id = data.get('location_id')
        if location_id:
            norm.location = CompanyCode.objects.get(id=location_id)
        elif 'location_id' in data and location_id is None:
            norm.location = None
            
        norm.depth_start = int(data.get('depth_start', norm.depth_start))
        norm.depth_end = int(data.get('depth_end', norm.depth_end))
        norm.norm_days = Decimal(str(data.get('norm_days', norm.norm_days)))
        norm.save()
        
        return Response({
            'success': True,
            'message': 'Hermetical testing norm updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating hermetical testing norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_hermetical_testing_norm(request, norm_id):
    """
    Delete a hermetical testing norm
    """
    try:
        from .models import HermeticalTestingNorm
        
        norm = get_object_or_404(HermeticalTestingNorm, id=norm_id)
        norm.delete()
        
        return Response({
            'success': True,
            'message': 'Hermetical testing norm deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting hermetical testing norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['GET'])
def get_operation_norms(request):
    """
    Get all operation norms - filtered by user location if not admin
    """
    try:
        from .models import OperationNorm
        norms = OperationNorm.objects.select_related('location').all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            norms = norms.filter(location=user_location)
        
        data = []
        for norm in norms:
            data.append({
                'id': str(norm.id),
                'location_id': norm.location.id if norm.location else None,
                'location': norm.location.location if norm.location else None,
                'operation': norm.operation,
                'norm_rule': norm.norm_rule,
                'remarks': norm.remarks,
            })
        
        return Response(data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching operation norms: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_operation_norm(request):
    """
    Create a new operation norm
    """
    try:
        from .models import OperationNorm, CompanyCode
        
        data = request.data
        location_id = data.get('location_id')
        location = CompanyCode.objects.get(id=location_id) if location_id else None
        
        norm = OperationNorm.objects.create(
            location=location,
            operation=data.get('operation'),
            norm_rule=data.get('norm_rule'),
            remarks=data.get('remarks', ''),
        )
        
        return Response({
            'success': True,
            'id': str(norm.id),
            'message': 'Operation norm created successfully'
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating operation norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT'])
def update_operation_norm(request, norm_id):
    """
    Update an existing operation norm
    """
    try:
        from .models import OperationNorm, CompanyCode
        
        norm = get_object_or_404(OperationNorm, id=norm_id)
        data = request.data
        
        location_id = data.get('location_id')
        if location_id:
            norm.location = CompanyCode.objects.get(id=location_id)
        elif 'location_id' in data and location_id is None:
            norm.location = None
            
        norm.operation = data.get('operation', norm.operation)
        norm.norm_rule = data.get('norm_rule', norm.norm_rule)
        norm.remarks = data.get('remarks', norm.remarks)
        norm.save()
        
        return Response({
            'success': True,
            'message': 'Operation norm updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating operation norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_operation_norm(request, norm_id):
    """
    Delete an operation norm
    """
    try:
        from .models import OperationNorm
        
        norm = get_object_or_404(OperationNorm, id=norm_id)
        norm.delete()
        
        return Response({
            'success': True,
            'message': 'Operation norm deleted successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting operation norm: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =====================================================================
# Completion Testing Norms Management
# =====================================================================

@login_required
def completion_testing_management(request):
    """
    Page to manage completion testing norms
    """
    return render(request, 'scheduler/completion_testing_management.html')


@login_required
@api_view(['GET'])
def get_completion_testing_norms(request):
    """
    Get all completion testing norms - filtered by user location if not admin
    """
    try:
        from .models import CompletionTestingNorm
        norms = CompletionTestingNorm.objects.all()
        
        # Apply location-based filtering for non-admin users
        user_location = get_user_location(request.user)
        if user_location:
            norms = norms.filter(location=user_location)
        
        data = []
        for norm in norms:
            data.append({
                'id': str(norm.id),
                'location': norm.location.location if norm.location else '',
                'well_depth_start': norm.well_depth_start,
                'well_depth_end': norm.well_depth_end,
                'well_type': norm.well_type,
                'days': float(norm.days),
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching completion testing norms: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['POST'])
def create_completion_testing_norm(request):
    """
    Create a new completion testing norm
    """
    try:
        from .models import CompletionTestingNorm
        
        data = request.data
        norm = CompletionTestingNorm.objects.create(
            location=data.get('location'),
            well_depth_start=data.get('well_depth_start'),
            well_depth_end=data.get('well_depth_end'),
            well_type=data.get('well_type'),
            days=data.get('days')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Completion testing norm created successfully',
            'id': str(norm.id)
        })
        
    except Exception as e:
        logger.error(f"Error creating completion testing norm: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['PUT'])
def update_completion_testing_norm(request, norm_id):
    """
    Update a completion testing norm
    """
    try:
        from .models import CompletionTestingNorm
        
        norm = get_object_or_404(CompletionTestingNorm, id=norm_id)
        data = request.data
        
        norm.location = data.get('location', norm.location)
        norm.well_depth_start = data.get('well_depth_start', norm.well_depth_start)
        norm.well_depth_end = data.get('well_depth_end', norm.well_depth_end)
        norm.well_type = data.get('well_type', norm.well_type)
        norm.days = data.get('days', norm.days)
        norm.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Completion testing norm updated successfully'
        })
        
    except Exception as e:
        logger.error(f"Error updating completion testing norm: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['DELETE'])
def delete_completion_testing_norm(request, norm_id):
    """
    Delete a completion testing norm
    """
    try:
        from .models import CompletionTestingNorm
        
        norm = get_object_or_404(CompletionTestingNorm, id=norm_id)
        norm.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Completion testing norm deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting completion testing norm: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# Additional Tests Management Views
@login_required
def additional_tests_management(request):
    """
    Page to manage additional tests
    """
    return render(request, 'scheduler/additional_tests_management.html')


@login_required
@api_view(['GET'])
def get_additional_tests(request):
    """
    Get all additional tests
    """
    try:
        from .models import AdditionalTest
        tests = AdditionalTest.objects.select_related('location').all()
        
        # Filter by user's location if not admin/L1
        user_location = get_user_location(request.user)
        if user_location:
            tests = tests.filter(location=user_location)
        
        data = []
        for test in tests:
            data.append({
                'id': str(test.id),
                'location_id': test.location_id,
                'location': test.location.location if test.location else None,
                'job': test.job,
                'norm_time': test.norm_time,
                'notes': test.notes or '',
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching additional tests: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['POST'])
def create_additional_test(request):
    """
    Create a new additional test
    """
    try:
        from .models import AdditionalTest
        
        data = request.data
        test = AdditionalTest.objects.create(
            location_id=data.get('location_id') if data.get('location_id') else None,
            job=data.get('job'),
            norm_time=data.get('norm_time'),
            notes=data.get('notes', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Additional test created successfully',
            'id': str(test.id)
        })
        
    except Exception as e:
        logger.error(f"Error creating additional test: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['PUT'])
def update_additional_test(request, test_id):
    """
    Update an additional test
    """
    try:
        from .models import AdditionalTest
        
        test = get_object_or_404(AdditionalTest, id=test_id)
        data = request.data
        
        if 'location_id' in data:
            test.location_id = data.get('location_id') if data.get('location_id') else None
        test.job = data.get('job', test.job)
        test.norm_time = data.get('norm_time', test.norm_time)
        test.notes = data.get('notes', test.notes)
        test.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Additional test updated successfully'
        })
        
    except Exception as e:
        logger.error(f"Error updating additional test: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['DELETE'])
def delete_additional_test(request, test_id):
    """
    Delete an additional test
    """
    try:
        from .models import AdditionalTest
        
        test = get_object_or_404(AdditionalTest, id=test_id)
        test.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Additional test deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting additional test: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# LOCATION SPEC FACTOR MANAGEMENT VIEWS (Admin Only)
# =============================================================================

@login_required
@staff_member_required
def loc_spec_factors_management(request):
    """
    Page to manage location-specific factors (Admin only)
    """
    return render(request, 'scheduler/loc_spec_factors_management.html')


@login_required
@api_view(['GET'])
def get_loc_spec_factors(request):
    """
    Get all location spec factors
    """
    try:
        from .models import LocationSpecFactor
        factors = LocationSpecFactor.objects.select_related('location').all()
        
        # Filter by user's location if not admin/L1
        user_location = get_user_location(request.user)
        if user_location:
            factors = factors.filter(location=user_location)
        
        data = []
        for factor in factors:
            data.append({
                'id': str(factor.id),
                'location_id': str(factor.location_id) if factor.location_id else None,
                'location': factor.location.location if factor.location else None,
                'factor_value': factor.factor_value,
                'display_order': factor.display_order,
                'is_default': factor.is_default,
                'is_active': factor.is_active,
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching location spec factors: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['POST'])
def create_loc_spec_factor(request):
    """
    Create a new location spec factor (Admin only)
    """
    try:
        from .models import LocationSpecFactor, CompanyCode
        
        data = request.data
        
        # Get location by name or id
        location_id = data.get('location_id')
        location_name = data.get('location')
        
        if location_name and not location_id:
            # Look up location by name
            location_obj = CompanyCode.objects.filter(location__iexact=location_name).first()
            if location_obj:
                location_id = str(location_obj.id)
            else:
                return JsonResponse({'error': f'Location "{location_name}" not found'}, status=400)
        
        # If this is set as default, unset other defaults for this location
        if data.get('is_default'):
            LocationSpecFactor.objects.filter(
                location_id=location_id,
                is_default=True
            ).update(is_default=False)
        
        factor = LocationSpecFactor.objects.create(
            location_id=location_id,
            factor_value=data.get('factor_value'),
            display_order=data.get('display_order', 0),
            is_default=data.get('is_default', False),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Location spec factor created successfully',
            'id': str(factor.id)
        })
        
    except Exception as e:
        logger.error(f"Error creating location spec factor: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['PUT'])
def update_loc_spec_factor(request, factor_id):
    """
    Update a location spec factor (Admin only)
    """
    try:
        from .models import LocationSpecFactor, CompanyCode
        
        factor = get_object_or_404(LocationSpecFactor, id=factor_id)
        data = request.data
        
        # Get location by name or id
        location_id = data.get('location_id')
        location_name = data.get('location')
        
        if location_name and not location_id:
            # Look up location by name
            location_obj = CompanyCode.objects.filter(location__iexact=location_name).first()
            if location_obj:
                location_id = str(location_obj.id)
            else:
                return JsonResponse({'error': f'Location "{location_name}" not found'}, status=400)
        
        # If this is set as default, unset other defaults for this location
        if data.get('is_default') and not factor.is_default:
            target_location_id = location_id if location_id else factor.location_id
            LocationSpecFactor.objects.filter(
                location_id=target_location_id,
                is_default=True
            ).update(is_default=False)
        
        if location_id:
            factor.location_id = location_id
        if 'factor_value' in data:
            factor.factor_value = data.get('factor_value')
        if 'display_order' in data:
            factor.display_order = data.get('display_order', 0)
        if 'is_default' in data:
            factor.is_default = data.get('is_default', False)
        if 'is_active' in data:
            factor.is_active = data.get('is_active', True)
        
        factor.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Location spec factor updated successfully'
        })
        
    except Exception as e:
        logger.error(f"Error updating location spec factor: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['DELETE'])
def delete_loc_spec_factor(request, factor_id):
    """
    Delete a location spec factor (Admin only)
    """
    try:
        from .models import LocationSpecFactor
        
        factor = get_object_or_404(LocationSpecFactor, id=factor_id)
        factor.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Location spec factor deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting location spec factor: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['GET'])
def get_factors_for_location(request, location):
    """
    Get all active factors for a specific location
    """
    try:
        from .models import LocationSpecFactor
        
        factors = LocationSpecFactor.objects.filter(
            location__location__iexact=location,
            is_active=True
        ).order_by('display_order', 'factor_value')
        
        data = [{
            'value': factor.factor_value,
            'is_default': factor.is_default
        } for factor in factors]
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching factors for location: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# DATABASE VIEWER VIEWS
# =============================================================================

@login_required
@staff_member_required
def database_viewer(request):
    """
    Main database viewer page showing all tables categorized
    Auto-detects all database tables and categorizes them
    """
    from django.apps import apps
    from django.contrib.auth.models import User, Group, Permission
    from django.contrib.contenttypes.models import ContentType
    from django.contrib.sessions.models import Session
    from django.contrib.admin.models import LogEntry
    
    # Get all models from all apps
    all_models = apps.get_models()
    
    # Categorize models automatically
    categorized_models = {
        'Core Data': [],
        'Scheduling': [],
        'Staging & Import': [],
        'Benchmarks & Norms': [],
        'Personnel & Company': [],
        'User Management': [],
        'System': [],
    }
    
    # Define category rules
    for model in all_models:
        app_label = model._meta.app_label
        model_name = model.__name__
        
        # Skip Django internal models we don't want to show
        if app_label == 'contenttypes' and model_name != 'ContentType':
            continue
        
        # Categorize based on model name and app
        if model_name in ['Rig', 'Well']:
            categorized_models['Core Data'].append((app_label, model_name))
        elif model_name in ['Schedule', 'Assignment', 'ScheduleRig', 'ScheduleWell', 'UnassignedWell']:
            categorized_models['Scheduling'].append((app_label, model_name))
        elif 'Staged' in model_name or 'Import' in model_name:
            categorized_models['Staging & Import'].append((app_label, model_name))
        elif any(x in model_name for x in ['Benchmark', 'Norm', 'Rate', 'Test', 'Adjustment']):
            categorized_models['Benchmarks & Norms'].append((app_label, model_name))
        elif model_name in ['MasterPersonnelInfo', 'CompanyCode']:
            categorized_models['Personnel & Company'].append((app_label, model_name))
        elif app_label == 'auth' or model_name in ['UserProfile', 'User', 'Group', 'Permission']:
            categorized_models['User Management'].append((app_label, model_name))
        elif app_label in ['admin', 'sessions', 'contenttypes'] or model_name in ['ExternalAppSetting', 'LogEntry', 'Session', 'ContentType']:
            categorized_models['System'].append((app_label, model_name))
        else:
            # Uncategorized goes to Core Data by default
            categorized_models['Core Data'].append((app_label, model_name))
    
    # Define category metadata
    category_metadata = {
        'Core Data': {'icon': 'bi-database-fill', 'color': '#3b82f6'},
        'Scheduling': {'icon': 'bi-calendar-check-fill', 'color': '#10b981'},
        'Staging & Import': {'icon': 'bi-inbox-fill', 'color': '#f59e0b'},
        'Benchmarks & Norms': {'icon': 'bi-bar-chart-fill', 'color': '#8b5cf6'},
        'Personnel & Company': {'icon': 'bi-building', 'color': '#0ea5e9'},
        'User Management': {'icon': 'bi-people-fill', 'color': '#ec4899'},
        'System': {'icon': 'bi-gear-fill', 'color': '#6b7280'},
    }
    
    # Build category data with counts
    categories_data = []
    for category_name, model_list in categorized_models.items():
        if not model_list:  # Skip empty categories
            continue
            
        tables = []
        for app_label, model_name in model_list:
            try:
                model = apps.get_model(app_label, model_name)
                count = model.objects.count()
                
                # Get first 5 records
                preview_records = list(model.objects.all()[:5].values())
                
                tables.append({
                    'app_label': app_label,
                    'model_name': model_name,
                    'verbose_name': model._meta.verbose_name,
                    'verbose_name_plural': model._meta.verbose_name_plural,
                    'count': count,
                    'preview_records': preview_records,
                })
            except LookupError:
                # Model doesn't exist
                pass
        
        if tables:  # Only add category if it has tables
            metadata = category_metadata.get(category_name, {'icon': 'bi-table', 'color': '#6b7280'})
            categories_data.append({
                'name': category_name,
                'icon': metadata['icon'],
                'color': metadata['color'],
                'tables': tables,
            })
    
    return render(request, 'scheduler/database_viewer.html', {
        'categories': categories_data,
    })


@login_required
@staff_member_required
def database_table_detail(request, app_label, model_name):
    """
    Detailed view of a specific table with search and pagination
    """
    from django.apps import apps
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    try:
        model = apps.get_model(app_label, model_name)
    except LookupError:
        return JsonResponse({'error': 'Model not found'}, status=404)
    
    # Get search query
    search_query = request.GET.get('search', '').strip()
    
    # Get all records
    queryset = model.objects.all()
    
    # Apply search if provided
    if search_query:
        # Get all text fields
        text_fields = []
        for field in model._meta.get_fields():
            if hasattr(field, 'get_internal_type'):
                field_type = field.get_internal_type()
                if field_type in ['CharField', 'TextField', 'EmailField', 'URLField', 'SlugField']:
                    text_fields.append(field.name)
        
        # Build Q objects for search
        q_objects = Q()
        for field_name in text_fields:
            q_objects |= Q(**{f'{field_name}__icontains': search_query})
        
        if q_objects:
            queryset = queryset.filter(q_objects)
    
    # Get field information
    fields_info = []
    for field in model._meta.get_fields():
        if hasattr(field, 'get_internal_type'):
            fields_info.append({
                'name': field.name,
                'verbose_name': getattr(field, 'verbose_name', field.name),
                'type': field.get_internal_type(),
            })
    
    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(queryset, 50)  # 50 records per page
    page_obj = paginator.get_page(page_number)
    
    # Convert records to list of dicts
    records = []
    for obj in page_obj:
        record = {}
        for field_info in fields_info:
            field_name = field_info['name']
            try:
                value = getattr(obj, field_name)
                # Convert to string for display
                if value is None:
                    record[field_name] = ''
                elif isinstance(value, (datetime, date)):
                    record[field_name] = value.isoformat()
                elif isinstance(value, Decimal):
                    record[field_name] = str(value)
                else:
                    record[field_name] = str(value)
            except Exception:
                record[field_name] = ''
        records.append(record)
    
    context = {
        'app_label': app_label,
        'model_name': model_name,
        'verbose_name': model._meta.verbose_name,
        'verbose_name_plural': model._meta.verbose_name_plural,
        'total_count': queryset.count(),
        'fields': fields_info,
        'records': records,
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'scheduler/database_table_detail.html', context)


# =============================================================================
# COMPANY CODES MANAGEMENT VIEWS
# =============================================================================

@login_required
@staff_member_required
@api_view(['GET'])
def get_company_codes(request):
    """
    Get all company codes
    """
    try:
        from .models import CompanyCode
        codes = CompanyCode.objects.all()
        
        data = []
        for code in codes:
            data.append({
                'id': str(code.id),
                'fund_centre': code.fund_centre,
                'company_code': code.company_code,
                'cost_centre': code.cost_centre,
                'category': code.category,
                'name': code.name,
                'location': code.location,
                'city': code.city,
                'state': code.state,
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching company codes: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['POST'])
def create_company_code(request):
    """
    Create a new company code
    """
    try:
        from .models import CompanyCode
        
        data = request.data
        code = CompanyCode.objects.create(
            fund_centre=data['fund_centre'],
            company_code=data['company_code'],
            cost_centre=data['cost_centre'],
            category=data['category'],
            name=data['name'],
            location=data.get('location', ''),
            city=data['city'],
            state=data['state'],
        )
        
        return JsonResponse({
            'id': str(code.id),
            'fund_centre': code.fund_centre,
            'company_code': code.company_code,
            'cost_centre': code.cost_centre,
            'category': code.category,
            'name': code.name,
            'location': code.location,
            'city': code.city,
            'state': code.state,
            'message': 'Company code created successfully'
        }, status=201)
        
    except Exception as e:
        logger.error(f"Error creating company code: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['PUT'])
def update_company_code(request, code_id):
    """
    Update a company code
    """
    try:
        from .models import CompanyCode
        
        code = CompanyCode.objects.get(id=code_id)
        data = request.data
        
        code.fund_centre = data.get('fund_centre', code.fund_centre)
        code.company_code = data.get('company_code', code.company_code)
        code.cost_centre = data.get('cost_centre', code.cost_centre)
        code.category = data.get('category', code.category)
        code.name = data.get('name', code.name)
        code.location = data.get('location', code.location)
        code.city = data.get('city', code.city)
        code.state = data.get('state', code.state)
        code.save()
        
        return JsonResponse({
            'id': str(code.id),
            'fund_centre': code.fund_centre,
            'company_code': code.company_code,
            'cost_centre': code.cost_centre,
            'category': code.category,
            'name': code.name,
            'location': code.location,
            'city': code.city,
            'state': code.state,
            'message': 'Company code updated successfully'
        })
        
    except CompanyCode.DoesNotExist:
        return JsonResponse({'error': 'Company code not found'}, status=404)
    except Exception as e:
        logger.error(f"Error updating company code: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['DELETE'])
def delete_company_code(request, code_id):
    """
    Delete a company code
    """
    try:
        from .models import CompanyCode
        
        code = CompanyCode.objects.get(id=code_id)
        code.delete()
        
        return JsonResponse({'message': 'Company code deleted successfully'})
        
    except CompanyCode.DoesNotExist:
        return JsonResponse({'error': 'Company code not found'}, status=404)
    except Exception as e:
        logger.error(f"Error deleting company code: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@require_http_methods(["POST"])
def upload_company_codes(request):
    """
    Upload company codes from CSV file
    """
    import csv
    import io
    from .models import CompanyCode
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        csv_file = request.FILES['file']
        
        # Check file extension
        if not csv_file.name.endswith('.csv'):
            return JsonResponse({'error': 'File must be a CSV'}, status=400)
        
        # Read and decode file, handle BOM if present
        file_data = csv_file.read().decode('utf-8-sig')  # Handles BOM
        csv_reader = csv.DictReader(io.StringIO(file_data))
        
        created_count = 0
        updated_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):
            try:
                # Skip header row if present (defensive, in case DictReader misreads)
                if row['Fund Centre'].strip() == 'Fund Centre':
                    continue
                # Check if record exists
                code, created = CompanyCode.objects.update_or_create(
                    fund_centre=row['Fund Centre'].strip(),
                    cost_centre=row['Cost Centre'].strip(),
                    defaults={
                        'company_code': row['Company Code'].strip(),
                        'category': row['Category'].strip(),
                        'name': row['Name'].strip(),
                        'location': row.get('Location', '').strip().title() if row.get('Location', '').strip() else '',
                        'city': row['City'].strip(),
                        'state': row['State'].strip(),
                    }
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Successfully processed {created_count + updated_count} records'
        })
        
    except Exception as e:
        logger.error(f"Error uploading company codes: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['GET'])
def get_unique_mpi_locations(request):
    """
    Get unique locations from CompanyCode for dropdown
    """
    try:
        from .models import CompanyCode
        
        # Get unique locations from CompanyCode
        locations = CompanyCode.objects.values_list('location', flat=True).distinct().order_by('location')
        locations_list = [loc for loc in locations if loc]
        
        return JsonResponse({'locations': locations_list}, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching locations: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@api_view(['GET'])
def get_unique_mpi_fields(request):
    """
    Get unique fields/sectors from CompanyCode for dropdown
    """
    try:
        from .models import CompanyCode
        
        # Get unique locations from CompanyCode (fields are also stored as locations)
        fields = CompanyCode.objects.values_list('location', flat=True).distinct().order_by('location')
        fields_list = [field for field in fields if field]
        
        return JsonResponse({'fields': fields_list}, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching fields: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# MASTER PERSONNEL INFO (MPI) VIEWS
# =============================================================================

@login_required
@staff_member_required
@require_http_methods(["POST"])
def upload_mpi(request):
    """
    Upload Master Personnel Info from CSV file
    """
    import csv
    import io
    from datetime import datetime
    from django.db import transaction
    from .models import MasterPersonnelInfo
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        csv_file = request.FILES['file']
        
        # Check file extension
        if not csv_file.name.endswith('.csv'):
            return JsonResponse({'error': 'File must be a CSV'}, status=400)
        
        # Read and decode file, handle BOM if present
        file_data = csv_file.read().decode('utf-8-sig')
        csv_reader = csv.DictReader(io.StringIO(file_data))
        
        created_count = 0
        updated_count = 0
        errors = []
        
        def parse_date(date_str):
            """Parse date string in various formats"""
            if not date_str or date_str.strip() == '':
                return None
            try:
                date_str = date_str.strip()
                # Try different date formats (most common first)
                formats = [
                    '%d.%m.%Y',      # 31.12.2020
                    '%d-%m-%Y',      # 31-12-2020
                    '%d/%m/%Y',      # 31/12/2020
                    '%Y-%m-%d',      # 2020-12-31
                    '%m/%d/%Y',      # 12/31/2020
                    '%d %b %Y',      # 31 Dec 2020
                    '%d-%b-%Y',      # 31-Dec-2020
                    '%Y/%m/%d',      # 2020/12/31
                    '%d.%m.%y',      # 31.12.20
                    '%d-%m-%y',      # 31-12-20
                ]
                for fmt in formats:
                    try:
                        return datetime.strptime(date_str, fmt).date()
                    except ValueError:
                        continue
                # If no format matches, log and return None
                logger.warning(f"Could not parse date: {date_str}")
                return None
            except Exception as e:
                logger.error(f"Error parsing date {date_str}: {str(e)}")
                return None
        
        for row_num, row in enumerate(csv_reader, start=2):
            try:
                # Skip header row if present (duplicate headers in middle of file)
                if row.get('CPF NO', '').strip() == 'CPF NO':
                    continue
                
                # Get CPF NO and check if it's valid
                cpf_no = row.get('CPF NO', '').strip()
                
                # Skip completely empty rows
                if not cpf_no and not any(v.strip() for v in row.values() if v):
                    continue
                
                # Skip rows with missing CPF NO but log them only if they have other data
                if not cpf_no:
                    # Check if row has any other meaningful data
                    has_data = any(row.get(field, '').strip() for field in ['NAME', 'DESIGNATION', 'LOCATION'])
                    if has_data:
                        errors.append(f"Row {row_num}: Missing CPF NO (row has data but no CPF)")
                    continue
                
                # Prepare data
                mpi_data = {
                    'crc': row.get('CRC', '').strip() or None,
                    'duty_type': row.get('Duty Type', '').strip() or None,
                    'work_pattern': row.get('Work pattern', '').strip() or None,
                    'pwd': row.get('PwD', '').strip() or None,
                    'q_new': row.get('Q New', '').strip() or None,
                    'org_unit': row.get('ORG.UNIT', '').strip() or None,
                    'group_1': row.get('GROUP 1', '').strip() or None,
                    'group_2': row.get('GROUP 2', '').strip() or None,
                    'org_new': row.get('ORG NEW', '').strip() or None,
                    'org_unit_text': row.get('ORG.UNIT TEXT', '').strip() or None,
                    'position_text': row.get('POSITION TEXT', '').strip() or None,
                    'location': row.get('LOCATION', '').strip() or None,
                    'sector': row.get('SECTOR', '').strip() or None,
                    'name': row.get('NAME', '').strip() or '',
                    'designation': row.get('DESIGNATION', '').strip() or None,
                    'lvl': row.get('LVL', '').strip() or None,
                    'disp': row.get('DISP', '').strip() or None,
                    'subdisp': row.get('SUBDISP', '').strip() or None,
                    'gender_key': row.get('GENDER KEY', '').strip() or None,
                    'dob': parse_date(row.get('DOB', '')),
                    'dor': parse_date(row.get('DOR', '')),
                    'doj_ongc': parse_date(row.get('DOJ ONGC', '')),
                    'personal_area': row.get('PERSONAL AREA', '').strip() or None,
                    'state_deployed': row.get('STATE DEPLOYED', '').strip() or None,
                    'qual_text': row.get('QUAL_TEXT', '').strip() or None,
                    'home_state': row.get('HOME STATE', '').strip() or None,
                    'dl_designation_text': row.get('DL-DESIGNATION TEXT', '').strip() or None,
                    'dl_discipline_text': row.get('DL-DISCIPLINE TEXT', '').strip() or None,
                    'dl_sub_disp_text': row.get('DL-SUB DISP TEXT', '').strip() or None,
                    'date_of_join_post': parse_date(row.get('DATE OF JOIN POST', '')),
                    'eff_date_prom': parse_date(row.get('EFF DATE PROM', '')),
                    'date_of_join_per_area': parse_date(row.get('DATE OF JOIN PER AREA', '')),
                    'date_of_join_position': parse_date(row.get('DATE OF JOIN POSITION', '')),
                    'date_of_retirement': parse_date(row.get('DATE OF RETIREMENT', '')),
                    'type_i': row.get('Type-I', '').strip() or None,
                    'mobile_no': row.get('Mobile No', '').strip() or None,
                }
                
                # Update or create record with atomic transaction
                with transaction.atomic():
                    mpi, created = MasterPersonnelInfo.objects.update_or_create(
                        cpf_no=cpf_no,
                        defaults=mpi_data
                    )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Successfully processed {created_count + updated_count} records (Created: {created_count}, Updated: {updated_count})'
        })
        
    except Exception as e:
        logger.error(f"Error uploading MPI: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['GET'])
def search_mpi(request):
    """
    Search MPI by CPF, Name, or Location and include role information
    """
    try:
        from .models import MasterPersonnelInfo
        from django.db.models import Q
        
        query = request.GET.get('q', '').strip()
        
        if not query:
            return Response([], status=status.HTTP_200_OK)
        
        # Search in CPF, Name, Location, and Org Unit
        results = MasterPersonnelInfo.objects.filter(
            Q(cpf_no__icontains=query) |
            Q(name__icontains=query) |
            Q(location__icontains=query) |
            Q(org_unit__icontains=query) |
            Q(org_unit_text__icontains=query)
        )[:50]  # Limit to 50 results
        
        data = []
        for mpi in results:
            # Get role information from Django User if exists
            role_info = None
            try:
                from django.contrib.auth.models import User
                user = User.objects.get(username=mpi.cpf_no)
                if user.is_superuser:
                    role_info = {
                        'role': 'admin',
                        'role_display': 'Admin',
                        'assigned_location': user.profile.location.name if hasattr(user, 'profile') and user.profile and user.profile.location else mpi.location,
                    }
                elif user.is_staff:
                    role_info = {
                        'role': 'L1',
                        'role_display': 'L1',
                        'assigned_location': user.profile.location.name if hasattr(user, 'profile') and user.profile and user.profile.location else mpi.location,
                    }
                else:
                    role_info = {
                        'role': 'user',
                        'role_display': 'User',
                        'assigned_location': user.profile.location.name if hasattr(user, 'profile') and user.profile and user.profile.location else mpi.location,
                    }
            except User.DoesNotExist:
                pass
            except Exception:
                pass
            
            data.append({
                'id': str(mpi.id),
                'cpf_no': mpi.cpf_no or '',
                'name': mpi.name or '',
                'designation': mpi.designation or '',
                'location': mpi.location or '',
                'org_unit': mpi.org_unit or '',
                'org_unit_text': mpi.org_unit_text or '',
                'mobile_no': mpi.mobile_no or '',
                'position_text': mpi.position_text or '',
                'role': role_info,
            })
        
        return Response(data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error searching MPI: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return Response({'error': 'Failed to search MPI records'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@staff_member_required
def mpi_table_view(request):
    """
    Full MPI table page view
    """
    return render(request, 'scheduler/mpi_table.html')


@login_required
@staff_member_required
@api_view(['GET'])
def get_all_mpi(request):
    """
    Get all MPI records for the table view with server-side pagination and search
    """
    try:
        from .models import MasterPersonnelInfo
        from django.core.paginator import Paginator
        from django.db.models import Q
        
        # Get pagination and search parameters
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 100))  # Load 100 records at a time
        search_query = request.GET.get('search', '').strip()
        
        # Get all records
        all_records = MasterPersonnelInfo.objects.all()
        
        # Filter by user's location if not admin/L1
        user_location = get_user_location(request.user)
        if user_location:
            user_location_str = user_location.location
            all_records = all_records.filter(location__iexact=user_location_str)
        
        # Apply search filter if provided
        if search_query:
            all_records = all_records.filter(
                Q(cpf_no__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(designation__icontains=search_query) |
                Q(location__icontains=search_query) |
                Q(org_unit__icontains=search_query) |
                Q(org_unit_text__icontains=search_query) |
                Q(position_text__icontains=search_query) |
                Q(mobile_no__icontains=search_query)
            )
        
        # Order by name
        all_records = all_records.order_by('name')
        total_count = all_records.count()
        
        # Paginate
        paginator = Paginator(all_records, page_size)
        mpi_records = paginator.get_page(page)
        
        data = []
        for mpi in mpi_records:
            data.append({
                'id': str(mpi.id),
                'cpf_no': mpi.cpf_no,
                'crc': mpi.crc,
                'duty_type': mpi.duty_type,
                'work_pattern': mpi.work_pattern,
                'pwd': mpi.pwd,
                'q_new': mpi.q_new,
                'org_unit': mpi.org_unit,
                'group_1': mpi.group_1,
                'group_2': mpi.group_2,
                'org_new': mpi.org_new,
                'org_unit_text': mpi.org_unit_text,
                'position_text': mpi.position_text,
                'location': mpi.location,
                'sector': mpi.sector,
                'name': mpi.name,
                'designation': mpi.designation,
                'lvl': mpi.lvl,
                'disp': mpi.disp,
                'subdisp': mpi.subdisp,
                'gender_key': mpi.gender_key,
                'dob': mpi.dob.isoformat() if mpi.dob else None,
                'dor': mpi.dor.isoformat() if mpi.dor else None,
                'doj_ongc': mpi.doj_ongc.isoformat() if mpi.doj_ongc else None,
                'personal_area': mpi.personal_area,
                'state_deployed': mpi.state_deployed,
                'qual_text': mpi.qual_text,
                'home_state': mpi.home_state,
                'dl_designation_text': mpi.dl_designation_text,
                'dl_discipline_text': mpi.dl_discipline_text,
                'dl_sub_disp_text': mpi.dl_sub_disp_text,
                'date_of_join_post': mpi.date_of_join_post.isoformat() if mpi.date_of_join_post else None,
                'eff_date_prom': mpi.eff_date_prom.isoformat() if mpi.eff_date_prom else None,
                'date_of_join_per_area': mpi.date_of_join_per_area.isoformat() if mpi.date_of_join_per_area else None,
                'date_of_join_position': mpi.date_of_join_position.isoformat() if mpi.date_of_join_position else None,
                'date_of_retirement': mpi.date_of_retirement.isoformat() if mpi.date_of_retirement else None,
                'type_i': mpi.type_i,
                'mobile_no': mpi.mobile_no,
            })
        
        response_data = {
            'data': data,
            'total': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': paginator.num_pages,
        }
        
        return JsonResponse(response_data, safe=False)
        
    except Exception as e:
        logger.error(f"Error fetching MPI: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# USER ROLE MANAGEMENT APIs
# =============================================================================

@login_required
@staff_member_required
@api_view(['GET'])
def get_user_role(request):
    """
    Get role information for a specific user by CPF number or username
    """
    try:
        from .models import UserRole, MasterPersonnelInfo
        
        # Safely get data with None handling
        cpf_no = (request.GET.get('cpf_no') or '').strip()
        username = (request.GET.get('username') or '').strip()
        
        if cpf_no:
            try:
                role = UserRole.objects.get(cpf_no=cpf_no)
            except UserRole.DoesNotExist:
                return JsonResponse({'role': None, 'message': 'No role assigned'})
        elif username:
            try:
                role = UserRole.objects.get(user__username=username)
            except UserRole.DoesNotExist:
                return JsonResponse({'role': None, 'message': 'No role assigned'})
        else:
            return JsonResponse({'error': 'CPF number or username required'}, status=400)
        
        data = {
            'id': str(role.id),
            'username': role.user.username,
            'cpf_no': role.cpf_no,
            'role': role.role,
            'role_display': role.get_role_display(),
            'assigned_location': role.assigned_location,
            'can_view_all_locations': role.can_view_all_locations(),
            'is_admin': role.is_admin(),
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f"Error fetching user role: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['POST'])
def assign_user_role(request):
    """
    Assign or update role for a user
    """
    try:
        from .models import UserRole, MasterPersonnelInfo
        from django.contrib.auth.models import User
        
        # Safely get data with None handling
        cpf_no = (request.data.get('cpf_no') or '').strip()
        username = (request.data.get('username') or '').strip()
        role = (request.data.get('role') or '').strip()
        assigned_location = (request.data.get('assigned_location') or '').strip()
        
        if not role or role not in ['admin', 'L1', 'user']:
            return JsonResponse({'error': 'Invalid role. Must be admin, L1, or user'}, status=400)
        
        # Auto-populate location from MPI if not provided and role is 'user'
        if role == 'user' and not assigned_location and cpf_no:
            try:
                mpi = MasterPersonnelInfo.objects.get(cpf_no=cpf_no)
                if mpi.location:
                    assigned_location = mpi.location
                    logger.info(f"Auto-populated location '{assigned_location}' from MPI for CPF {cpf_no}")
            except MasterPersonnelInfo.DoesNotExist:
                pass
        
        # Find or create user account
        user_obj = None
        mpi_data = None
        
        if username:
            try:
                user_obj = User.objects.get(username=username)
            except User.DoesNotExist:
                return JsonResponse({'error': f'User {username} not found'}, status=404)
        elif cpf_no:
            # Try to find or create user
            try:
                user_obj = User.objects.get(username=cpf_no)
            except User.DoesNotExist:
                # Create new user from MPI data
                try:
                    mpi = MasterPersonnelInfo.objects.get(cpf_no=cpf_no)
                    mpi_data = mpi
                    user_obj = User.objects.create(
                        username=cpf_no,
                        first_name=mpi.name.split()[0] if mpi.name else '',
                        last_name=' '.join(mpi.name.split()[1:]) if mpi.name and len(mpi.name.split()) > 1 else '',
                        email=f"{cpf_no}@ongc.co.in"
                    )
                    user_obj.set_unusable_password()  # LDAP users don't have local passwords
                    user_obj.save()
                    logger.info(f"Created new user {cpf_no} from MPI data")
                except MasterPersonnelInfo.DoesNotExist:
                    return JsonResponse({'error': f'CPF {cpf_no} not found in personnel database'}, status=404)
        else:
            return JsonResponse({'error': 'CPF number or username required'}, status=400)
        
        # Update user permissions based on role
        if role == 'admin':
            user_obj.is_staff = True
            user_obj.is_superuser = True
        elif role == 'L1':
            user_obj.is_staff = True
            user_obj.is_superuser = False
        else:  # user
            user_obj.is_staff = False
            user_obj.is_superuser = False
        user_obj.save()
        
        # Ensure UserProfile exists (create if not) and update location/permissions
        from .models import UserProfile, CompanyCode
        profile, created = UserProfile.objects.get_or_create(user=user_obj)
        
        if role == 'user' and assigned_location:
            # Try to find matching CompanyCode object
            company_code_obj = CompanyCode.objects.filter(
                models.Q(location__icontains=assigned_location) |
                models.Q(company_code__icontains=assigned_location) |
                models.Q(name__icontains=assigned_location),
                is_active=True
            ).first()
            if company_code_obj:
                profile.location = company_code_obj
                profile.can_view_all_locations = False
            else:
                # Still set can_view_all_locations to False for 'user' role
                profile.location = None
                profile.can_view_all_locations = False
                logger.warning(f"Location '{assigned_location}' not found in CompanyCode table")
        elif role in ['admin', 'L1']:
            # Admin and L1 users can view all locations
            profile.location = None
            profile.can_view_all_locations = True
        else:
            # Default for 'user' role without location
            profile.location = None
            profile.can_view_all_locations = False
        profile.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User role assigned successfully',
            'user': {
                'username': user_obj.username,
                'cpf_no': cpf_no or user_obj.username,
                'role': role,
                'assigned_location': assigned_location,
                'is_staff': user_obj.is_staff,
                'is_superuser': user_obj.is_superuser,
            }
        })
        
    except Exception as e:
        logger.error(f"Error assigning user role: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def bulk_add_org_unit_users(request):
    """
    Bulk add all employees from an org_unit as users to Authorized Users table.
    Auto-populates location from MPI for each user.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        data = json.loads(request.body)
        org_unit = data.get('org_unit', '').strip()
        
        if not org_unit:
            return JsonResponse({'error': 'Org unit is required'}, status=400)
        
        # Find all employees in this org_unit
        employees = MasterPersonnelInfo.objects.filter(
            Q(org_unit__icontains=org_unit) | Q(org_unit_text__icontains=org_unit)
        )
        
        if not employees.exists():
            return JsonResponse({'error': f'No employees found in org unit: {org_unit}'}, status=404)
        
        success_count = 0
        error_count = 0
        errors = []
        
        for mpi in employees:
            try:
                if not mpi.cpf_no:
                    error_count += 1
                    errors.append(f'{mpi.name}: No CPF number')
                    continue
                
                # Create or get user
                user_obj, created = User.objects.get_or_create(
                    username=mpi.cpf_no,
                    defaults={
                        'first_name': mpi.name.split()[0] if mpi.name else '',
                        'last_name': ' '.join(mpi.name.split()[1:]) if mpi.name and len(mpi.name.split()) > 1 else '',
                        'email': f"{mpi.cpf_no}@ongc.co.in",
                        'is_staff': False,
                        'is_superuser': False,
                    }
                )
                
                if created:
                    user_obj.set_unusable_password()
                    user_obj.save()
                
                # Update UserProfile with location from MPI
                if hasattr(user_obj, 'profile') and user_obj.profile and mpi.location:
                    from .models import CompanyCode
                    company_code_obj = CompanyCode.objects.filter(
                        models.Q(location__icontains=mpi.location) |
                        models.Q(company_code__icontains=mpi.location) |
                        models.Q(name__icontains=mpi.location),
                        is_active=True
                    ).first()
                    
                    if company_code_obj:
                        user_obj.profile.location = company_code_obj
                        user_obj.profile.can_view_all_locations = False
                        user_obj.profile.save()
                
                success_count += 1
                logger.info(f"Bulk added user {mpi.cpf_no} ({mpi.name}) from org_unit {org_unit}")
                
            except Exception as e:
                error_count += 1
                errors.append(f'{mpi.cpf_no} ({mpi.name}): {str(e)}')
                logger.error(f"Error bulk adding user {mpi.cpf_no}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Bulk add completed: {success_count} users added, {error_count} errors',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10],  # Limit to first 10 errors
            'org_unit': org_unit,
        })
        
    except Exception as e:
        logger.error(f"Error in bulk add org unit users: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['DELETE'])
def remove_user_role(request):
    """
    Remove a user's role assignment
    """
    try:
        from .models import UserRole
        
        # Safely get data with None handling
        cpf_no = (request.data.get('cpf_no') or '').strip()
        username = (request.data.get('username') or '').strip()
        
        if cpf_no:
            try:
                role = UserRole.objects.get(cpf_no=cpf_no)
            except UserRole.DoesNotExist:
                return JsonResponse({'error': 'Role assignment not found'}, status=404)
        elif username:
            try:
                role = UserRole.objects.get(user__username=username)
            except UserRole.DoesNotExist:
                return JsonResponse({'error': 'Role assignment not found'}, status=404)
        else:
            return JsonResponse({'error': 'CPF number or username required'}, status=400)
        
        role.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Role assignment removed successfully'
        })
        
    except Exception as e:
        logger.error(f"Error removing user role: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['GET'])
def list_user_roles(request):
    """
    List all user role assignments (DEPRECATED - use list_authorized_users)
    """
    try:
        from .models import UserRole
        
        roles = UserRole.objects.select_related('user').all()
        
        data = []
        for role in roles:
            data.append({
                'id': str(role.id),
                'username': role.user.username,
                'cpf_no': role.cpf_no,
                'role': role.role,
                'role_display': role.get_role_display(),
                'assigned_location': role.assigned_location,
                'can_view_all_locations': role.can_view_all_locations(),
                'is_admin': role.is_admin(),
                'created_at': role.created_at.isoformat(),
            })
        
        return JsonResponse(data, safe=False)
        
    except Exception as e:
        logger.error(f"Error listing user roles: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# AUTHORIZED USERS API (NEW - LDAP Integration)
# =============================================================================

@login_required
@staff_member_required
@api_view(['GET'])
def list_authorized_users(request):
    """
    List all users from Django User table who can log in.
    Shows location from UserProfile (customizable) or MPI (default).
    Supports filtering and pagination
    """
    try:
        from django.contrib.auth.models import User
        from .models import MasterPersonnelInfo
        
        # Get filter parameters
        search = request.GET.get('search', '').strip()
        role_filter = request.GET.get('role', '').strip()
        status_filter = request.GET.get('status', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        
        # Base query - Show all Django User table entries
        users = User.objects.select_related('profile').all()
        
        # Apply filters
        if search:
            users = users.filter(
                models.Q(username__icontains=search) |
                models.Q(first_name__icontains=search) |
                models.Q(last_name__icontains=search) |
                models.Q(email__icontains=search)
            )
        
        # Role filter based on user permissions
        if role_filter:
            if role_filter == 'admin':
                users = users.filter(is_superuser=True)
            elif role_filter == 'L1':
                users = users.filter(is_staff=True, is_superuser=False)
            elif role_filter == 'user':
                users = users.filter(is_staff=False, is_superuser=False)
        
        # Status filter
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)
        
        # Count total
        total_count = users.count()
        
        # Apply pagination
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        users = users.order_by('username')[start_idx:end_idx]
        
        # Build response data
        data = []
        for user in users:
            # Determine role
            if user.is_superuser:
                role = 'admin'
                role_display = 'Admin'
            elif user.is_staff:
                role = 'L1'
                role_display = 'L1'
            else:
                role = 'user'
                role_display = 'User'
            
            # Get location - first try UserProfile, then MPI as default
            assigned_location = ''
            if hasattr(user, 'profile') and user.profile and user.profile.location:
                # Use the 'location' field from CompanyCode, not 'name'
                assigned_location = user.profile.location.location or user.profile.location.company_code
            else:
                # Get default location from MPI
                try:
                    mpi = MasterPersonnelInfo.objects.filter(cpf_no=user.username).first()
                    if mpi and mpi.location:
                        assigned_location = mpi.location
                except Exception:
                    pass
            
            # Get full name
            full_name = f"{user.first_name} {user.last_name}".strip()
            if not full_name:
                # Try to get from MPI
                try:
                    mpi = MasterPersonnelInfo.objects.filter(cpf_no=user.username).first()
                    if mpi:
                        full_name = mpi.name
                except Exception:
                    pass
            
            data.append({
                'id': str(user.id),
                'username': user.username,
                'cpf_no': user.username,
                'name': full_name or user.username,
                'email': user.email or '',
                'role': role,
                'role_display': role_display,
                'assigned_location': assigned_location,
                'is_active': user.is_active,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'created_at': user.date_joined.isoformat() if hasattr(user, 'date_joined') else None,
            })
        
        return JsonResponse({
            'users': data,
            'total_count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': (total_count + page_size - 1) // page_size
        })
        
    except Exception as e:
        logger.error(f"Error listing authorized users: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['POST'])
def update_authorized_user(request):
    """
    Update an authorized user's role and settings
    """
    try:
        from .models import AuthorizedUser
        
        cpf_no = (request.data.get('cpf_no') or '').strip()
        role = (request.data.get('role') or '').strip()
        assigned_location = (request.data.get('assigned_location') or '').strip()
        is_active = request.data.get('is_active', True)
        
        if not cpf_no:
            return JsonResponse({'error': 'CPF number is required'}, status=400)
        
        if role not in ['admin', 'L1', 'user']:
            return JsonResponse({'error': 'Invalid role'}, status=400)
        
        # Get or create authorized user
        user, created = AuthorizedUser.objects.get_or_create(
            cpf_no=cpf_no,
            defaults={'role': role, 'assigned_location': assigned_location, 'is_active': is_active}
        )
        
        if not created:
            user.role = role
            user.assigned_location = assigned_location
            user.is_active = is_active
            user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User {cpf_no} updated successfully',
            'user': {
                'cpf_no': user.cpf_no,
                'role': user.role,
                'assigned_location': user.assigned_location,
                'is_active': user.is_active,
            }
        })
        
    except Exception as e:
        logger.error(f"Error updating authorized user: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['POST'])
def toggle_user_active_status(request):
    """
    Activate or deactivate a user
    """
    try:
        from .models import AuthorizedUser
        
        cpf_no = (request.data.get('cpf_no') or '').strip()
        
        if not cpf_no:
            return JsonResponse({'error': 'CPF number is required'}, status=400)
        
        user = AuthorizedUser.objects.get(cpf_no=cpf_no)
        user.is_active = not user.is_active
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User {cpf_no} {"activated" if user.is_active else "deactivated"}',
            'is_active': user.is_active
        })
        
    except AuthorizedUser.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        logger.error(f"Error toggling user status: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['DELETE'])
def delete_authorized_user(request):
    """
    Revoke app access for an authorized user (deactivate and remove permissions)
    User remains in database and MPI, can authenticate via LDAP but cannot use app
    """
    try:
        from django.contrib.auth.models import User
        
        username = (request.data.get('username') or '').strip()
        cpf_no = (request.data.get('cpf_no') or '').strip()
        
        if not username and not cpf_no:
            return JsonResponse({'error': 'Username or CPF number is required'}, status=400)
        
        # Find user by username or CPF; if not present, create a shell user so we can flag as revoked
        if username:
            user, _ = User.objects.get_or_create(username=username)
        elif cpf_no:
            user, _ = User.objects.get_or_create(username=cpf_no)
        
        # Don't allow revoking access for current user
        if user == request.user:
            return JsonResponse({'error': 'You cannot revoke access for your own account'}, status=400)
        
        # Don't allow revoking access for superusers unless requester is also superuser
        if user.is_superuser and not request.user.is_superuser:
            return JsonResponse({'error': 'Only superusers can revoke access for other superusers'}, status=403)
        
        username_revoked = user.username
        
        # Revoke app access by deactivating user and removing permissions
        user.is_active = False
        user.is_staff = False
        user.is_superuser = False
        user.save()
        
        # Also update UserProfile if exists
        if hasattr(user, 'profile') and user.profile:
            user.profile.can_view_all_locations = False
            user.profile.location = None
            user.profile.save()
        
        logger.info(f"App access revoked for user {username_revoked} by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'App access revoked for {username_revoked}. User can authenticate via LDAP but cannot use the application.'
        })
        
    except User.DoesNotExist:
        # If user truly does not exist, treat as already revoked
        logger.warning(f"User {username or cpf_no} not found while revoking access; treating as already revoked")
        return JsonResponse({
            'success': True,
            'message': 'User not found in auth table; access already revoked.'
        })
    except Exception as e:
        logger.error(f"Error revoking user access: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_member_required
@api_view(['POST'])
def reactivate_authorized_user(request):
    """
    Restore app access for a previously revoked user
    User will be able to authenticate and use the app again
    """
    try:
        from django.contrib.auth.models import User
        
        username = (request.data.get('username') or '').strip()
        cpf_no = (request.data.get('cpf_no') or '').strip()
        
        if not username and not cpf_no:
            return JsonResponse({'error': 'Username or CPF number is required'}, status=400)
        
        # Find user by username or CPF; create shell user if absent so we can re-enable
        if username:
            user, _ = User.objects.get_or_create(username=username)
        elif cpf_no:
            user, _ = User.objects.get_or_create(username=cpf_no)
        
        username_restored = user.username
        
        # Restore app access
        user.is_active = True
        # Don't restore permissions automatically - admin should use Edit Role for that
        user.save()
        
        logger.info(f"App access restored for user {username_restored} by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': f'App access restored for {username_restored}. Use "Edit Role" to assign permissions.'
        })
        
    except User.DoesNotExist:
        logger.warning(f"User {username or cpf_no} not found while reactivating access")
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        logger.error(f"Error restoring user access: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================
# WELL BASKET MANAGEMENT VIEWS
# ============================================================

@login_required
def basket_creation_page(request):
    """
    Render the basket creation page where users can create well baskets
    """
    return render(request, 'scheduler/basket_creation.html')


@login_required
@api_view(['GET'])
def get_baskets(request):
    """
    Get all well baskets for the current user's location
    """
    try:
        from .models import WellBasket, CompanyCode
        
        user_location = get_user_location(request.user)
        
        # Filter by location if user has location restriction
        if user_location:
            baskets = WellBasket.objects.filter(location=user_location)
        else:
            # User can see all locations
            baskets = WellBasket.objects.all()
        
        # Optional location filter from query params (for users with all-locations access)
        location_filter = request.GET.get('location')
        if location_filter:
            # Try to find matching CompanyCode by location name or company_code
            from django.db.models import Q
            company_code = CompanyCode.objects.filter(
                Q(location__iexact=location_filter) | 
                Q(company_code__iexact=location_filter) |
                Q(name__icontains=location_filter)
            ).first()
            if company_code:
                baskets = baskets.filter(location=company_code)
        
        # Optional status filter
        status_filter = request.GET.get('status')
        if status_filter:
            baskets = baskets.filter(status=status_filter)
        
        # Build response data
        baskets_data = []
        for basket in baskets:
            wells_summary = basket.wells_summary
            baskets_data.append({
                'id': str(basket.id),
                'name': basket.name,
                'description': basket.description or '',
                'status': basket.status,
                'location': basket.location.code if basket.location else None,
                'location_name': basket.location.location if basket.location else None,
                'created_by': basket.created_by.username if basket.created_by else None,
                'created_at': basket.created_at.isoformat() if basket.created_at else None,
                'well_count': wells_summary['total'],
                'pending_count': wells_summary['pending'],
                'completed_count': wells_summary['completed'],
                'imported_count': wells_summary['imported'],
            })
        
        return Response({
            'success': True,
            'count': len(baskets_data),
            'baskets': baskets_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching baskets: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['GET'])
def get_basket_detail(request, basket_id):
    """
    Get details of a specific basket including its wells
    """
    try:
        from .models import WellBasket
        
        basket = get_object_or_404(WellBasket, id=basket_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and basket.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get wells in this basket
        wells = basket.staged_wells.all().order_by('name')
        wells_data = [{
            'id': str(well.id),
            'name': well.name,
            'asset_id': well.asset_id,
            'status': well.status,
            'field_name': well.field_name,
            'well_type': well.well_type,
            'depth': well.depth,
            'latitude': well.latitude,
            'longitude': well.longitude,
            'well_profile': well.well_profile,
            'rig_capacity_required_hp': well.rig_capacity_required_hp,
            'drl_days': float(well.drl_days) if well.drl_days else None,
            'pt_days': float(well.pt_days) if well.pt_days else None,
            'duration': well.duration,
            'bop_stack': well.bop_stack,
            'tds_requirement': well.tds_requirement,
            'footprint': well.footprint,
            'priority': well.priority,
            'preferred_rig': well.preferred_rig,
            'expected_potential': well.expected_potential,
            'rtd': well.rtd.isoformat() if well.rtd else None,
            'location': str(well.location.id) if well.location else None,
            'location_value': well.location.code if well.location else None,
            'imported_well_id': str(well.imported_well.id) if well.imported_well else None,
            'missing_fields': well.missing_fields,
            'is_ready_to_import': well.is_ready_to_import,
        } for well in wells]
        
        return Response({
            'success': True,
            'basket': {
                'id': str(basket.id),
                'name': basket.name,
                'description': basket.description or '',
                'status': basket.status,
                'location': basket.location.code if basket.location else None,
                'location_name': basket.location.location if basket.location else None,
                'created_by': basket.created_by.username if basket.created_by else None,
                'created_at': basket.created_at.isoformat() if basket.created_at else None,
                'wells': wells_data,
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error fetching basket detail: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def search_wells_for_basket(request):
    """
    Search for staged wells based on pasted well names.
    Returns found/not found summary without creating the basket.
    """
    try:
        from .models import StagedWell, CompanyCode
        
        user_location = get_user_location(request.user)
        well_names_input = request.data.get('well_names', '')
        location_filter = request.data.get('location', request.GET.get('location'))
        
        # Parse well names (comma or space/newline separated)
        # First replace commas with spaces, then split by whitespace
        well_names_input = well_names_input.replace(',', ' ')
        well_names = [name.strip() for name in well_names_input.split() if name.strip()]
        
        # Remove duplicates while preserving order
        seen = set()
        unique_well_names = []
        for name in well_names:
            name_upper = name.upper()
            if name_upper not in seen:
                seen.add(name_upper)
                unique_well_names.append(name)
        
        if not unique_well_names:
            return Response({
                'success': False,
                'error': 'No well names provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Query staged wells - location filter
        if user_location:
            # User has location restriction - use their location
            staged_wells = StagedWell.objects.filter(location=user_location)
        elif location_filter:
            # User has no restriction but provided a location filter
            from django.db.models import Q
            company_code = CompanyCode.objects.filter(
                Q(location__iexact=location_filter) | 
                Q(company_code__iexact=location_filter) |
                Q(name__icontains=location_filter)
            ).first()
            if company_code:
                staged_wells = StagedWell.objects.filter(location=company_code)
            else:
                staged_wells = StagedWell.objects.all()
        else:
            staged_wells = StagedWell.objects.all()
        
        # Don't exclude imported wells - we'll categorize them separately
        
        found_wells = []
        not_found_wells = []
        already_in_basket = []
        already_imported = []
        
        for name in unique_well_names:
            # Case-insensitive search
            matching_wells = staged_wells.filter(name__iexact=name)
            
            if matching_wells.exists():
                well = matching_wells.first()
                
                # Get list of baskets this well belongs to (for info purposes)
                existing_baskets = list(well.baskets.values_list('name', flat=True))
                existing_basket_names = ', '.join(existing_baskets) if existing_baskets else None
                
                if well.status == 'IMPORTED':
                    # Well has already been imported to scheduler
                    already_imported.append({
                        'id': str(well.id),
                        'staged_well_id': str(well.id),
                        'name': well.name,
                        'imported_well_id': str(well.imported_well_id) if well.imported_well_id else None,
                        'asset_id': well.asset_id,
                        'field_name': well.field_name,
                        'depth': well.depth,
                        'priority': well.priority,
                        'location_id': str(well.location.id) if well.location else None,
                        'location_code': well.location.code if well.location else None,
                        'existing_baskets': existing_basket_names,
                    })
                elif existing_baskets:
                    # Well is in other baskets but can still be added to this one
                    already_in_basket.append({
                        'id': str(well.id),
                        'staged_well_id': str(well.id),
                        'name': well.name,
                        'existing_basket': existing_basket_names,
                        'existing_baskets': existing_basket_names,
                        'asset_id': well.asset_id,
                        'field_name': well.field_name,
                        'depth': well.depth,
                        'priority': well.priority,
                        'status': well.status,
                        'location_id': str(well.location.id) if well.location else None,
                        'location_code': well.location.code if well.location else None,
                    })
                else:
                    found_wells.append({
                        'id': str(well.id),
                        'name': well.name,
                        'asset_id': well.asset_id,
                        'status': well.status,
                        'field_name': well.field_name,
                        'depth': well.depth,
                        'priority': well.priority,
                        'location_id': str(well.location.id) if well.location else None,
                        'location_code': well.location.code if well.location else None,
                    })
            else:
                not_found_wells.append(name)
        
        return Response({
            'success': True,
            'searched_count': len(unique_well_names),
            'found_wells': found_wells,
            'found_count': len(found_wells),
            'not_found_wells': not_found_wells,
            'not_found_count': len(not_found_wells),
            'already_in_basket': already_in_basket,
            'already_in_basket_count': len(already_in_basket),
            'already_imported': already_imported,
            'already_imported_count': len(already_imported),
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error searching wells for basket: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def create_basket(request):
    """
    Create a new basket with the specified wells.
    Wells from different locations are not allowed in the same basket.
    Wells that are already in another basket will be moved to this basket.
    Wells that are already imported will be added for reference/tracking.
    """
    try:
        from .models import WellBasket, StagedWell, CompanyCode
        
        user_location = get_user_location(request.user)
        
        basket_name = request.data.get('name', '').strip()
        description = request.data.get('description', '').strip()
        well_ids = request.data.get('well_ids', [])
        
        if not basket_name:
            return Response({
                'success': False,
                'error': 'Basket name is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not well_ids:
            return Response({
                'success': False,
                'error': 'At least one well must be selected'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get all wells and validate they are from the same location
        wells_to_add = []
        location_set = set()
        
        for well_id in well_ids:
            try:
                well = StagedWell.objects.get(id=well_id)
                wells_to_add.append(well)
                if well.location:
                    location_set.add(well.location.id)
            except StagedWell.DoesNotExist:
                continue
        
        if not wells_to_add:
            return Response({
                'success': False,
                'error': 'No valid wells found to add to basket'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate all wells are from the same location
        if len(location_set) > 1:
            return Response({
                'success': False,
                'error': 'Cannot create basket with wells from multiple locations. All wells must belong to the same location.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Determine location from the wells
        basket_location = wells_to_add[0].location if wells_to_add[0].location else None
        
        # If user has location restriction, ensure basket is for their location
        if user_location and basket_location and basket_location != user_location:
            return Response({
                'success': False,
                'error': 'You can only create baskets for wells in your assigned location'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if not basket_location:
            return Response({
                'success': False,
                'error': 'Could not determine location for basket'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        with transaction.atomic():
            # Create the basket
            basket = WellBasket.objects.create(
                name=basket_name,
                description=description,
                location=basket_location,
                created_by=request.user,
                status='ACTIVE'
            )
            
            # Assign wells to the basket using M2M relationship
            added_count = 0
            already_in_count = 0
            imported_count = 0
            skipped_count = 0
            
            for well in wells_to_add:
                try:
                    # Check location access for user
                    if user_location and well.location != user_location:
                        skipped_count += 1
                        continue
                    
                    # Check if well is already in this basket
                    if basket in well.baskets.all():
                        already_in_count += 1
                        continue
                    
                    # Add well to basket (wells can be in multiple baskets)
                    well.baskets.add(basket)
                    
                    # Track count by status
                    if well.status == 'IMPORTED':
                        imported_count += 1
                    else:
                        added_count += 1
                    
                except Exception as e:
                    logger.error(f"Error adding well {well.id} to basket: {str(e)}")
                    skipped_count += 1
        
        message_parts = []
        if added_count > 0:
            message_parts.append(f'{added_count} wells added')
        if imported_count > 0:
            message_parts.append(f'{imported_count} imported wells added for reference')
        if already_in_count > 0:
            message_parts.append(f'{already_in_count} wells already in basket')
        
        total_added = added_count + imported_count
        message = f'Basket "{basket_name}" created with ' + ', '.join(message_parts) if message_parts else f'Basket "{basket_name}" created'
        
        return Response({
            'success': True,
            'basket_id': str(basket.id),
            'basket_name': basket.name,
            'added_count': added_count,
            'imported_count': imported_count,
            'already_in_count': already_in_count,
            'skipped_count': skipped_count,
            'message': message
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error creating basket: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['PUT', 'PATCH'])
def update_basket(request, basket_id):
    """
    Update basket name or description
    """
    try:
        from .models import WellBasket
        
        basket = get_object_or_404(WellBasket, id=basket_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and basket.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Update allowed fields
        if 'name' in request.data:
            basket.name = request.data['name'].strip()
        if 'description' in request.data:
            basket.description = request.data['description'].strip()
        if 'status' in request.data and request.data['status'] in ['ACTIVE', 'FINALIZED', 'ARCHIVED']:
            basket.status = request.data['status']
        
        basket.save()
        
        return Response({
            'success': True,
            'message': f'Basket "{basket.name}" updated successfully'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error updating basket: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['DELETE'])
def delete_basket(request, basket_id):
    """
    Delete a basket and release all wells from it
    """
    try:
        from .models import WellBasket
        
        basket = get_object_or_404(WellBasket, id=basket_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and basket.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        basket_name = basket.name
        well_count = basket.well_count
        
        with transaction.atomic():
            # Clear all M2M relationships (Django will handle this automatically on delete, but explicit is better)
            basket.staged_wells.clear()
            
            # Delete the basket
            basket.delete()
        
        return Response({
            'success': True,
            'message': f'Basket "{basket_name}" deleted. {well_count} wells removed from basket.'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error deleting basket: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def remove_well_from_basket(request, basket_id, well_id):
    """
    Remove a specific well from a basket
    """
    try:
        from .models import WellBasket, StagedWell
        
        basket = get_object_or_404(WellBasket, id=basket_id)
        well = get_object_or_404(StagedWell, id=well_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and basket.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check well belongs to this basket (M2M relationship)
        if basket not in well.baskets.all():
            return Response({
                'success': False,
                'error': 'Well does not belong to this basket'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        well.baskets.remove(basket)
        
        return Response({
            'success': True,
            'message': f'Well "{well.name}" removed from basket'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error removing well from basket: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def add_wells_to_basket(request, basket_id):
    """
    Add additional wells to an existing basket
    """
    try:
        from .models import WellBasket, StagedWell
        
        basket = get_object_or_404(WellBasket, id=basket_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and basket.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        well_ids = request.data.get('well_ids', [])
        
        if not well_ids:
            return Response({
                'success': False,
                'error': 'No wells specified to add'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        added_count = 0
        skipped_count = 0
        already_in_count = 0
        
        with transaction.atomic():
            for well_id in well_ids:
                try:
                    well = StagedWell.objects.get(id=well_id)
                    
                    # Check location matches basket
                    if well.location != basket.location:
                        skipped_count += 1
                        continue
                    
                    # Check if well is already in this basket (M2M)
                    if basket in well.baskets.all():
                        already_in_count += 1
                        continue
                    
                    # Add well to basket (wells can be in multiple baskets)
                    well.baskets.add(basket)
                    added_count += 1
                    
                except StagedWell.DoesNotExist:
                    skipped_count += 1
        
        return Response({
            'success': True,
            'added_count': added_count,
            'already_in_count': already_in_count,
            'skipped_count': skipped_count,
            'message': f'{added_count} wells added to basket'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error adding wells to basket: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@login_required
@api_view(['POST'])
def finalize_basket_wells(request, basket_id):
    """
    Finalize all completed wells in a basket by moving them from StagedWell to Well table
    """
    try:
        from .models import WellBasket, StagedWell, Well
        
        basket = get_object_or_404(WellBasket, id=basket_id)
        
        # Check location access
        user_location = get_user_location(request.user)
        if user_location and basket.location != user_location:
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get all completed staged wells in this basket (M2M relationship)
        staged_wells = StagedWell.objects.filter(
            baskets=basket,
            status='COMPLETED'
        )
        
        if not staged_wells.exists():
            return Response({
                'success': False,
                'error': 'No completed wells found in this basket'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        imported_count = 0
        skipped_count = 0
        errors = []
        rtd_missing_wells = []
        
        # Use a single transaction so all on_commit ILM callbacks fire
        # only after the entire batch, preventing SQLite lock contention
        with transaction.atomic():
            for staged_well in staged_wells:
                try:
                    # Check if RTD date is set (required for main Well table)
                    if staged_well.rtd is None:
                        skipped_count += 1
                        rtd_missing_wells.append(staged_well.name)
                        errors.append(f'{staged_well.name}: Missing RTD Date')
                        continue
                    
                    # Check if well already exists in Well table
                    if Well.objects.filter(name=staged_well.name).exists():
                        skipped_count += 1
                        errors.append(f'{staged_well.name}: Well already exists in main database')
                        continue
                    
                    # Get next serial number
                    max_sn = Well.objects.aggregate(models.Max('sn'))['sn__max'] or 0
                    next_sn = max_sn + 1
                    
                    # Create the well in the main Well table with proper field mapping
                    well = Well.objects.create(
                        sn=next_sn,
                        location=staged_well.location,
                        asset_id=staged_well.asset_id,
                        name=staged_well.name,
                        well_type=staged_well.well_type,
                        well_profile=staged_well.well_profile,
                        depth=staged_well.depth,
                        rig_capacity_required_hp=staged_well.rig_capacity_required_hp,
                        drl_days=staged_well.drl_days,
                        pt_days=staged_well.pt_days,
                        duration=staged_well.duration,
                        latitude=staged_well.latitude,
                        longitude=staged_well.longitude,
                        rtd=staged_well.rtd,
                        bop_stack=staged_well.bop_stack,
                        tds_requirement=staged_well.tds_requirement,
                        footprint=staged_well.footprint,
                        preferred_rig=staged_well.preferred_rig or '',
                        expected_potential=staged_well.expected_potential or '',
                        priority=staged_well.priority,
                        field_name=staged_well.field_name or ''
                    )
                    
                    # ILM well pair distances are computed automatically
                    # by the post_save signal (well_saved_trigger_ilm) which uses
                    # transaction.on_commit() to start a background thread after
                    # this atomic block commits. No sync calculation needed here.
                    
                    # Update staged well status to IMPORTED
                    # Keep the well in the basket - don't set basket to None
                    staged_well.status = 'IMPORTED'
                    staged_well.imported_at = timezone.now()
                    staged_well.imported_well = well
                    staged_well.save()
                    
                    imported_count += 1
                    
                except Exception as e:
                    logger.error(f"Error importing staged well {staged_well.name}: {str(e)}")
                    errors.append(f'{staged_well.name}: {str(e)}')
                    skipped_count += 1
        
        return Response({
            'success': True,
            'imported_count': imported_count,
            'skipped_count': skipped_count,
            'rtd_missing_count': len(rtd_missing_wells),
            'rtd_missing_wells': rtd_missing_wells[:20],
            'errors': errors,
            'message': f'Successfully imported {imported_count} wells to main database'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error finalizing basket wells: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# ILM COST - WELL PAIR DISTANCE VIEWS
# =============================================================================

def calculate_ilm_days(rig, distance_m, location, norm_days, prefetched_adjustments=None):
    """
    Calculate ILM Days by applying adjustment rules to the base norm days.
    
    Logic:
    1. Start with base norm days
    2. Check for cluster movement rules (distance-based replacements)
    3. Apply equipment rules (TDS, DSA)
    4. Apply transportation rules (distance-based additions)
    
    Args:
        rig: Rig model instance
        distance_m: Distance between wells in meters
        location: CompanyCode model instance
        norm_days: Base norm days from RigBuildingNorm
        prefetched_adjustments: Optional pre-fetched list of RigBuildingAdjustment objects
                                 (pass this to avoid DB queries in bulk operations)
    
    Returns:
        Dictionary with ilm_days and applied_rules
    """
    if norm_days is None:
        return {'ilm_days': None, 'applied_rules': [], 'note': 'No norm mapping'}
    
    try:
        from .models import RigBuildingAdjustment
        
        # Use pre-fetched adjustments if provided (for bulk operations)
        if prefetched_adjustments is not None:
            adjustments = prefetched_adjustments
        else:
            # Get all active adjustment rules for this location
            adjustments = list(RigBuildingAdjustment.objects.filter(
                location=location,
                is_active=True
            ).order_by('-priority', 'category'))
        
        if not adjustments:
            return {'ilm_days': norm_days, 'applied_rules': [], 'note': 'No adjustment rules'}
        
        ilm_days = float(norm_days)
        applied_rules = []
        base_replaced = False
        
        # Get rig characteristics
        rig_type = rig.rig_type  # 'Mobile' or 'Fixed'
        tds_available = rig.tds_availability == 'Y'
        
        # Process each adjustment rule
        for adj in adjustments:
            rule_applies = True
            
            # Check rig type filter
            if adj.applies_to_rig_type:
                rig_type_filter = adj.applies_to_rig_type.lower()
                # Handle 'Type-I' which typically applies to Fixed rigs up to certain depth
                if rig_type_filter == 'type-i':
                    if rig_type != 'Fixed':
                        rule_applies = False
                elif rig_type_filter not in [rig_type.lower(), 'all', '']:
                    rule_applies = False
            
            # Check distance conditions
            if adj.min_distance is not None:
                if distance_m < float(adj.min_distance):
                    rule_applies = False
            
            if adj.max_distance is not None:
                if distance_m > float(adj.max_distance):
                    rule_applies = False
            
            if not rule_applies:
                continue
            
            # Apply the adjustment based on type
            if adj.adjustment_type == 'replace' and not base_replaced:
                # Replace base norm with this value (cluster movement rules)
                if adj.adjustment_value is not None:
                    ilm_days = float(adj.adjustment_value)
                    base_replaced = True
                    applied_rules.append({
                        'condition': adj.condition,
                        'action': f'Replaced base with {adj.adjustment_display}'
                    })
            
            elif adj.adjustment_type == 'add':
                # Add days to current value
                if adj.adjustment_value is not None:
                    # Special handling for TDS rule - only apply if rig has TDS
                    if 'tds' in adj.condition.lower() or 'top drive' in adj.condition.lower():
                        if tds_available:
                            ilm_days += float(adj.adjustment_value)
                            applied_rules.append({
                                'condition': adj.condition,
                                'action': adj.adjustment_display
                            })
                    else:
                        ilm_days += float(adj.adjustment_value)
                        applied_rules.append({
                            'condition': adj.condition,
                            'action': adj.adjustment_display
                        })
            
            elif adj.adjustment_type == 'per_unit':
                # Add days per unit (e.g., per 50 km)
                if adj.adjustment_value is not None and adj.unit:
                    # Parse unit to get base distance (e.g., "50 km" -> 50000 meters)
                    try:
                        unit_value = float(adj.unit.split()[0])
                        if 'km' in adj.unit.lower():
                            unit_value *= 1000  # Convert to meters
                        
                        # Calculate additional days based on distance beyond min_distance
                        min_dist = float(adj.min_distance) if adj.min_distance else 0
                        if distance_m > min_dist:
                            extra_distance = distance_m - min_dist
                            additional_units = extra_distance / unit_value
                            additional_days = additional_units * float(adj.adjustment_value)
                            ilm_days += additional_days
                            applied_rules.append({
                                'condition': adj.condition,
                                'action': f'+{additional_days:.1f} days ({additional_units:.1f} x {adj.unit})'
                            })
                    except (ValueError, IndexError):
                        pass
            
            elif adj.adjustment_type == 'included':
                # No extra days needed - this is informational
                pass
        
        # Round to 1 decimal place
        ilm_days = round(ilm_days, 1)
        
        return {
            'ilm_days': ilm_days,
            'applied_rules': applied_rules,
            'note': f'{len(applied_rules)} rules applied' if applied_rules else 'Base norm only'
        }
        
    except Exception as e:
        logger.error(f"Error calculating ILM days: {str(e)}")
        return {'ilm_days': norm_days, 'applied_rules': [], 'note': f'Error: {str(e)}'}


@login_required
def get_well_pair_distances(request):
    """
    Get well pair distances for a specific location and rig.
    Uses pre-computed ILM values stored on WellPairDistance for performance.
    Query params:
        - location: Required for data (Company code or location name)
        - rig_id: Optional UUID of specific rig (RECOMMENDED - returns only that rig's pairs)
        - rigs_only: If 'true', returns just the list of rigs with pair counts (fast)
        - page: Page number (1-indexed, default 1)
        - page_size: Items per page (default 100, max 500)
        - search: Optional search term to filter by well names
        - sort: Sort field (well_1, well_2, distance, ilm_days, norm_days) default: well_1
        - order: asc or desc (default asc)
    """
    try:
        from .models import WellPairDistance, CompanyCode, Rig
        from django.db.models import Count, F, Value, CharField
        from django.db.models.functions import Least, Greatest
        
        location_code = request.GET.get('location', '')
        rig_id = request.GET.get('rig_id', '')
        rigs_only = request.GET.get('rigs_only', '') == 'true'
        page = max(1, int(request.GET.get('page', 1)))
        page_size = min(500, max(10, int(request.GET.get('page_size', 100))))
        search_term = request.GET.get('search', '').strip()
        sort_field = request.GET.get('sort', 'well_1')
        sort_order = request.GET.get('order', 'asc')
        
        # If location not specified, apply user-based filtering
        if not location_code and not user_can_view_all_locations(request.user):
            location_code = get_user_assigned_location(request.user)
        
        # Require location
        if not location_code:
            if rigs_only:
                return JsonResponse({'success': True, 'rigs': [], 'total_pairs': 0})
            return JsonResponse({
                'success': True, 'location': 'All Locations',
                'count': 0, 'distances': [], 'total_count': 0,
                'page': 1, 'page_size': page_size, 'total_pages': 0,
                'message': 'Please select a location to view distances.'
            })
        
        # Resolve location
        try:
            location = CompanyCode.objects.get(
                models.Q(company_code=location_code) | models.Q(location=location_code)
            )
        except CompanyCode.DoesNotExist:
            return JsonResponse({'error': f'Location {location_code} not found'}, status=404)
        except CompanyCode.MultipleObjectsReturned:
            location = CompanyCode.objects.filter(
                models.Q(company_code=location_code) | models.Q(location=location_code)
            ).first()
        
        # Rigs-only mode: return list of rigs with pair counts (for dropdown)
        if rigs_only:
            rig_stats = (
                WellPairDistance.objects.filter(location=location)
                .values('rig_id', 'rig__name')
                .annotate(pair_count=Count('id'))
                .order_by('rig__name')
            )
            rigs = []
            for rs in rig_stats:
                rigs.append({
                    'rig_id': str(rs['rig_id']),
                    'rig_name': rs['rig__name'],
                    'pair_count': rs['pair_count'] // 2  # Stored bidirectionally
                })
            total = WellPairDistance.objects.filter(location=location).count() // 2
            return JsonResponse({'success': True, 'rigs': rigs, 'total_pairs': total})
        
        # Full data mode - require rig_id for performance
        if not rig_id:
            return JsonResponse({
                'success': True,
                'location': location_code,
                'count': 0, 'distances': [], 'total_count': 0,
                'page': 1, 'page_size': page_size, 'total_pages': 0,
                'message': 'Please select a rig to view its well pair distances.'
            })
        
        # Deduplicate bidirectional pairs using canonical ordering (well_1_id < well_2_id)
        # This avoids loading all rows into memory for dedup
        base_qs = WellPairDistance.objects.filter(
            location=location, rig_id=rig_id,
            well_1_id__lt=F('well_2_id')  # Only keep one direction
        ).select_related('rig', 'rig__rig_building_norm', 'well_1', 'well_2', 'location')
        
        # Apply search filter
        if search_term:
            base_qs = base_qs.filter(
                models.Q(well_1__name__icontains=search_term) |
                models.Q(well_2__name__icontains=search_term)
            )
        
        # Apply sorting
        sort_map = {
            'well_1': 'well_1__name',
            'well_2': 'well_2__name',
            'distance': 'distance_km',
            'ilm_days': 'ilm_days',
            'norm_days': 'rig__rig_building_norm__days',
            'location': 'location__location',
            'rig': 'rig__name',
        }
        order_by_field = sort_map.get(sort_field, 'well_1__name')
        if sort_order == 'desc':
            order_by_field = '-' + order_by_field
        base_qs = base_qs.order_by(order_by_field, 'well_1__name', 'well_2__name')
        
        # Get total count for pagination
        total_count = base_qs.count()
        total_pages = max(1, (total_count + page_size - 1) // page_size)
        page = min(page, total_pages)
        
        # Paginate
        offset = (page - 1) * page_size
        paginated_qs = base_qs[offset:offset + page_size]
        
        result = []
        for d in paginated_qs:
            norm_days = d.rig.rig_building_norm.days if d.rig.rig_building_norm else None
            distance_m = float(d.distance_km)
            
            # Use cached ILM values if available
            if d.ilm_days is not None:
                ilm_days = float(d.ilm_days)
                applied_rules = d.ilm_rules_applied or []
                ilm_note = d.ilm_note or ''
            else:
                ilm_result = calculate_ilm_days(d.rig, distance_m, d.location, norm_days)
                ilm_days = ilm_result['ilm_days']
                applied_rules = ilm_result['applied_rules']
                ilm_note = ilm_result['note']
            
            result.append({
                'id': str(d.id),
                'location': d.location.location if d.location else location_code,
                'rig_name': d.rig.name,
                'rig_id': str(d.rig_id),
                'rig_type': d.rig.rig_type,
                'rig_building_norm_days': norm_days,
                'ilm_days': ilm_days,
                'ilm_applied_rules': applied_rules,
                'ilm_note': ilm_note,
                'well_1_name': d.well_1.name,
                'well_1_id': str(d.well_1_id),
                'well_1_lat': float(d.well_1.latitude) if d.well_1.latitude else None,
                'well_1_lng': float(d.well_1.longitude) if d.well_1.longitude else None,
                'well_2_name': d.well_2.name,
                'well_2_id': str(d.well_2_id),
                'well_2_lat': float(d.well_2.latitude) if d.well_2.latitude else None,
                'well_2_lng': float(d.well_2.longitude) if d.well_2.longitude else None,
                'distance_m': distance_m,
            })
        
        return JsonResponse({
            'success': True,
            'location': location_code,
            'count': len(result),
            'total_count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'distances': result
        })
        
    except Exception as e:
        logger.error(f"Error getting well pair distances: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def refresh_ilm_cache_for_location(location, batch_size=500):
    """
    Recompute and store ILM days for all WellPairDistance records in a location.
    Called after distance recalculation or when adjustment rules change.
    Uses bulk_update with pre-fetched adjustments for high performance.
    """
    from .models import WellPairDistance, RigBuildingAdjustment
    
    # Pre-fetch adjustment rules once for the entire location (avoid N+1 queries)
    prefetched_adjustments = list(RigBuildingAdjustment.objects.filter(
        location=location,
        is_active=True
    ).order_by('-priority', 'category'))
    
    distances = WellPairDistance.objects.filter(location=location).select_related(
        'rig', 'rig__rig_building_norm'
    )
    
    total = distances.count()
    updated = 0
    
    # Process in batches using bulk_update
    offset = 0
    while offset < total:
        batch = list(distances[offset:offset + batch_size])
        if not batch:
            break
        
        to_update = []
        for d in batch:
            norm_days = d.rig.rig_building_norm.days if d.rig.rig_building_norm else None
            distance_m = float(d.distance_km)
            ilm_result = calculate_ilm_days(d.rig, distance_m, location, norm_days,
                                            prefetched_adjustments=prefetched_adjustments)
            d.ilm_days = ilm_result['ilm_days']
            d.ilm_note = ilm_result.get('note', '')
            d.ilm_rules_applied = ilm_result.get('applied_rules', [])
            to_update.append(d)
        
        if to_update:
            WellPairDistance.objects.bulk_update(
                to_update, ['ilm_days', 'ilm_note', 'ilm_rules_applied'],
                batch_size=batch_size
            )
            updated += len(to_update)
        offset += batch_size
    
    return updated


@login_required
@require_http_methods(["POST"])
def recalculate_well_pair_distances(request):
    """
    Recalculate all well pair distances for a location.
    POST body:
        - location: Company code (e.g., 'CAMBAY')
    """
    try:
        import json
        from .models import WellPairDistance, CompanyCode
        
        data = json.loads(request.body)
        location_code = data.get('location', '')
        
        if not location_code:
            return JsonResponse({'error': 'Location is required'}, status=400)
        
        # Get location - try both company_code and location fields
        try:
            location = CompanyCode.objects.get(
                models.Q(company_code=location_code) | models.Q(location=location_code)
            )
        except CompanyCode.DoesNotExist:
            return JsonResponse({'error': f'Location {location_code} not found'}, status=404)
        except CompanyCode.MultipleObjectsReturned:
            # If multiple matches, prefer exact company_code match, then location match
            location = CompanyCode.objects.filter(
                models.Q(company_code=location_code) | models.Q(location=location_code)
            ).first()
        
        # Recalculate all distances
        count = WellPairDistance.recalculate_all_for_location(location)
        
        # Also refresh the ILM cache for this location
        ilm_count = refresh_ilm_cache_for_location(location)
        
        return JsonResponse({
            'success': True,
            'message': f'Recalculated {count} well pair distances and {ilm_count} ILM values for {location_code}',
            'count': count
        })
        
    except Exception as e:
        logger.error(f"Error recalculating well pair distances: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_ilm_cost_summary(request):
    """
    Get ILM cost summary statistics for a location or all locations.
    Automatically filters by user's assigned location if they are not admin.
    Query params:
        - location: Optional - Company code (e.g., 'CAMBAY'). If not provided, uses user's location or all locations for admin.
    """
    try:
        from .models import WellPairDistance, CompanyCode, Rig, Well
        from django.db.models import Count, Avg, Min, Max
        
        location_code = request.GET.get('location', '')
        
        # If location not specified in query params, apply user-based filtering
        if not location_code and not user_can_view_all_locations(request.user):
            location_code = get_user_assigned_location(request.user)
        
        # Filter by location only if specified
        if location_code:
            # Get location - try both company_code and location fields
            try:
                location = CompanyCode.objects.get(
                    models.Q(company_code=location_code) | models.Q(location=location_code)
                )
            except CompanyCode.DoesNotExist:
                return JsonResponse({'error': f'Location {location_code} not found'}, status=404)
            except CompanyCode.MultipleObjectsReturned:
                # If multiple matches, prefer exact company_code match, then location match
                location = CompanyCode.objects.filter(
                    models.Q(company_code=location_code) | models.Q(location=location_code)
                ).first()
            
            # Get counts for specific location
            rigs_count = Rig.objects.filter(location=location, is_deleted=False).count()
            wells_count = Well.objects.filter(location=location, is_deleted=False).count()
            distances = WellPairDistance.objects.filter(location=location)
        else:
            # Get counts for all locations
            rigs_count = Rig.objects.filter(is_deleted=False).count()
            wells_count = Well.objects.filter(is_deleted=False).count()
            distances = WellPairDistance.objects.all()
        
        # Get unique pairs count (divide by 2 since we store both directions)
        unique_pairs = distances.count() // 2
        
        # Get statistics
        stats = distances.aggregate(
            avg_distance=Avg('distance_km'),
            min_distance=Min('distance_km'),
            max_distance=Max('distance_km')
        )
        
        return JsonResponse({
            'success': True,
            'location': location_code if location_code else 'All Locations',
            'rigs_count': rigs_count,
            'wells_count': wells_count,
            'unique_pairs': unique_pairs,
            'avg_distance_m': round(float(stats['avg_distance'] or 0), 2),
            'min_distance_m': round(float(stats['min_distance'] or 0), 2),
            'max_distance_m': round(float(stats['max_distance'] or 0), 2),
        })
        
    except Exception as e:
        logger.error(f"Error getting ILM cost summary: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)



@login_required
def video_tutorials(request):
    """View for displaying video tutorials"""
    from .models import VideoTutorial
    
    # Get all active tutorials grouped by category
    tutorials = VideoTutorial.objects.filter(is_active=True).order_by('category', 'order', 'title')
    
    # Group by category
    tutorials_by_category = {}
    for tutorial in tutorials:
        category = tutorial.get_category_display()
        if category not in tutorials_by_category:
            tutorials_by_category[category] = []
        tutorials_by_category[category].append(tutorial)
    
    context = {
        'tutorials_by_category': tutorials_by_category,
    }
    
    return render(request, 'scheduler/video_tutorials.html', context)


@login_required
def video_tutorial_detail(request, tutorial_id):
    """View for playing a specific video tutorial"""
    from .models import VideoTutorial
    
    tutorial = get_object_or_404(VideoTutorial, id=tutorial_id, is_active=True)
    
    # Increment view count
    tutorial.increment_view_count()
    
    # Get related tutorials from same category
    related_tutorials = VideoTutorial.objects.filter(
        category=tutorial.category,
        is_active=True
    ).exclude(id=tutorial_id).order_by('order', 'title')[:5]
    
    context = {
        'tutorial': tutorial,
        'related_tutorials': related_tutorials,
    }
    
    return render(request, 'scheduler/video_tutorial_detail.html', context)



@login_required
def stream_video_file(request, tutorial_id):
    """
    Stream video file with HTTP Range Request support for fast loading.
    Automatically uses the best available version (compressed > optimized > original).
    """
    from .models import VideoTutorial
    from .video_streaming import stream_video, get_video_content_type
    
    tutorial = get_object_or_404(VideoTutorial, id=tutorial_id, is_active=True)
    
    # Get the best available video file
    video_file = tutorial.get_best_video_file()
    
    if not video_file:
        raise Http404("Video file not found")
    
    # Get absolute path to video file
    video_path = video_file.path
    
    # Determine content type
    content_type = get_video_content_type(video_path)
    
    # Stream the video with range support
    return stream_video(request, video_path, content_type)


@login_required
def stream_hls_master(request, tutorial_id):
    """
    Stream HLS master playlist for adaptive bitrate streaming.
    """
    from .models import VideoTutorial
    from .video_streaming import has_hls_stream, get_hls_directory
    import os
    
    tutorial = get_object_or_404(VideoTutorial, id=tutorial_id, is_active=True)
    
    if not has_hls_stream(tutorial_id):
        raise Http404("HLS stream not available")
    
    hls_dir = get_hls_directory(tutorial_id)
    master_playlist_path = os.path.join(hls_dir, 'master.m3u8')
    
    if not os.path.exists(master_playlist_path):
        raise Http404("Master playlist not found")
    
    # Read and modify playlist to use correct URLs
    with open(master_playlist_path, 'r') as f:
        content = f.read()
    
    # Replace relative paths with absolute URLs
    from django.urls import reverse
    import re
    
    def replace_playlist_path(match):
        quality = match.group(1)
        return reverse('stream_hls_playlist', kwargs={'tutorial_id': tutorial_id, 'quality': quality})
    
    content = re.sub(r'(\w+)/playlist\.m3u8', replace_playlist_path, content)
    
    response = HttpResponse(content, content_type='application/vnd.apple.mpegurl')
    response['Cache-Control'] = 'public, max-age=3600'
    response['Access-Control-Allow-Origin'] = '*'
    
    return response


@login_required
def stream_hls_playlist(request, tutorial_id, quality):
    """
    Stream HLS quality-specific playlist.
    """
    from .models import VideoTutorial
    from .video_streaming import get_hls_directory
    import os
    
    tutorial = get_object_or_404(VideoTutorial, id=tutorial_id, is_active=True)
    
    hls_dir = get_hls_directory(tutorial_id)
    playlist_path = os.path.join(hls_dir, quality, 'playlist.m3u8')
    
    if not os.path.exists(playlist_path):
        raise Http404("Playlist not found")
    
    # Read and modify playlist to use correct segment URLs
    with open(playlist_path, 'r') as f:
        content = f.read()
    
    # Replace segment filenames with full URLs
    from django.urls import reverse
    import re
    
    def replace_segment_path(match):
        segment = match.group(1)
        return reverse('stream_hls_segment', kwargs={
            'tutorial_id': tutorial_id, 
            'quality': quality, 
            'segment': segment
        })
    
    content = re.sub(r'(segment\d+\.ts)', replace_segment_path, content)
    
    response = HttpResponse(content, content_type='application/vnd.apple.mpegurl')
    response['Cache-Control'] = 'public, max-age=3600'
    response['Access-Control-Allow-Origin'] = '*'
    
    return response


@login_required
def stream_hls_segment(request, tutorial_id, quality, segment):
    """
    Stream HLS segment file.
    """
    from .models import VideoTutorial
    from .video_streaming import get_hls_directory, stream_hls_segment as serve_segment
    import os
    
    tutorial = get_object_or_404(VideoTutorial, id=tutorial_id, is_active=True)
    
    hls_dir = get_hls_directory(tutorial_id)
    segment_path = os.path.join(hls_dir, quality, segment)
    
    if not os.path.exists(segment_path):
        raise Http404("Segment not found")
    
    return serve_segment(request, segment_path)
