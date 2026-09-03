from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.utils import timezone
from django.db import models
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, parser_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd
import csv
import io
import json
from datetime import datetime, date, timedelta
from decimal import Decimal
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import Location, UserProfile, Rig, Well, Schedule, Assignment, UnassignedWell
from .serializers import (
    RigSerializer, WellSerializer, ScheduleSerializer, AssignmentSerializer,
    UnassignedWellSerializer, ScheduleCreateSerializer, GanttDataSerializer,
    AssignmentUpdateSerializer, BulkDataUploadSerializer, ScheduleStatsSerializer
)
from .optimization import DrillingScheduler
from .well_rejection_analyzer import WellRejectionAnalyzer

logger = logging.getLogger(__name__)


def get_user_location(user):
    """
    Get the location for a user. Returns None if user can access all locations.
    
    Args:
        user: Django User object
    
    Returns:
        Location object or None if user has access to all locations
    """
    if not user or not user.is_authenticated:
        return None
    
    # Superusers and staff can see all locations
    if user.is_superuser or user.is_staff:
        return None
    
    # Check if user has a profile
    try:
        profile = user.profile
        if profile.can_view_all_locations:
            return None
        return profile.location
    except UserProfile.DoesNotExist:
        # Create profile if it doesn't exist
        UserProfile.objects.create(user=user)
        return None


def get_user_accessible_locations(user):
    """
    Get all locations a user can access.
    
    Args:
        user: Django User object
    
    Returns:
        QuerySet of Location objects
    """
    if not user or not user.is_authenticated:
        return Location.objects.none()
    
    # Superusers can see all locations
    if user.is_superuser:
        return Location.objects.filter(is_active=True)
    
    # Check if user has a profile
    try:
        profile = user.profile
        return profile.get_accessible_locations()
    except UserProfile.DoesNotExist:
        UserProfile.objects.create(user=user)
        return Location.objects.none()


def create_child_schedule(parent_schedule, branch_type="reschedule", custom_suffix=None):
    """
    Create a properly named child schedule with version management
    
    Args:
        parent_schedule: The parent Schedule instance
        branch_type: Type of branching (reschedule, add_well, delete_well, etc.)
        custom_suffix: Optional custom suffix instead of auto-generated version
    
    Returns:
        New Schedule instance with proper naming and relationships
    """
    # Get the base name from the root schedule
    base_name = parent_schedule.get_base_name()
    
    # Get next version number
    next_version = parent_schedule.get_next_version_number()
    
    # Create display name
    if custom_suffix:
        new_name = f"{base_name} {custom_suffix}"
    else:
        new_name = f"{base_name} v{next_version}"
    
    # Create the child schedule
    child_schedule = Schedule.objects.create(
        name=new_name,
        financial_year=parent_schedule.financial_year,
        parent_schedule=parent_schedule,
        branch_type=branch_type,
        version_number=next_version,
        status='RUNNING'
    )
    
    logger.info(f"Created child schedule: '{new_name}' (v{next_version}) from parent '{parent_schedule.name}'")
    return child_schedule


@login_required
def index(request):
    """Main dashboard view"""
    return render(request, 'scheduler/dashboard.html')


@login_required
def dashboard(request):
    """Dashboard page view"""
    return render(request, 'scheduler/dashboard.html')


@login_required
def about(request):
    """About page view"""
    return render(request, 'scheduler/about.html')


@login_required
def product_showcase(request):
    """Product showcase page for marketing/PR purposes"""
    return render(request, 'scheduler/product_showcase.html')


@login_required
def data_management(request):
    """Data management page view"""
    # Get unique asset_ids for location selector
    asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
    asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
    
    # Get user's location info
    user_location = None
    user_location_code = None
    can_view_all = True
    
    if request.user and request.user.is_authenticated:
        user_loc = get_user_location(request.user)
        if user_loc:
            user_location = user_loc.name
            user_location_code = user_loc.code
            can_view_all = False
    
    context = {
        'asset_ids': asset_ids,
        'user_location': user_location,
        'user_location_code': user_location_code,
        'can_view_all_locations': can_view_all,
    }
    return render(request, 'scheduler/data_management.html', context)


def scheduling(request):
    """Scheduling page view"""
    # Get unique asset_ids for location selector
    asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
    asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
    
    # Import financial year functions
    from .models import get_financial_year_choices, get_current_financial_year
    
    # Get user's location info
    user_location = None
    user_location_code = None
    can_view_all = True
    
    if request.user and request.user.is_authenticated:
        user_loc = get_user_location(request.user)
        if user_loc:
            user_location = user_loc.name
            user_location_code = user_loc.code
            can_view_all = False
    
    context = {
        'asset_ids': asset_ids,
        'financial_year_choices': get_financial_year_choices(),
        'current_financial_year': get_current_financial_year(),
        'user_location': user_location,
        'user_location_code': user_location_code,
        'can_view_all_locations': can_view_all,
    }
    return render(request, 'scheduler/scheduling.html', context)


def schedules_list(request):
    """Schedules management page view"""
    return render(request, 'scheduler/schedules.html')


def schedule_maps(request):
    """Generate and display movement maps for a schedule"""
    schedule_id = request.GET.get('schedule_id')

    
    if not schedule_id:
        # Get location filter if provided
        location_filter = request.GET.get('location', None)
        
        # Get available schedules for user to choose from
        available_schedules = Schedule.objects.all().order_by('-created_at')
        
        # Filter schedules by location if specified
        if location_filter:
            available_schedules = available_schedules.filter(
                assignments__well__asset_id=location_filter
            ).distinct()
        
        # Apply slice after filtering
        available_schedules = available_schedules[:10]
        
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        return render(request, 'scheduler/schedule_maps.html', {
            'error': 'No schedule ID provided',
            'available_schedules': available_schedules,
            'show_schedule_selector': True,
            'asset_ids': asset_ids,
            'selected_location': location_filter,
        })
    
    try:
        # Get location filter if provided
        location_filter = request.GET.get('location', None)
        
        # Get the schedule
        schedule = Schedule.objects.get(id=schedule_id)
        assignments = Assignment.objects.filter(schedule=schedule)
        
        # Filter assignments by location if specified
        if location_filter:
            assignments = assignments.filter(well__asset_id=location_filter)
        
        if not assignments.exists():
            error_message = f'No assignments found for this schedule'
            if location_filter:
                error_message += f' in location {location_filter}'
            # Get asset IDs for location filter
            asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
            asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
            
            return render(request, 'scheduler/schedule_maps.html', {
                'error': error_message,
                'schedule': schedule,
                'selected_location': location_filter,
                'asset_ids': asset_ids,
            })
        
        # Prepare data for the scheduler
        rigs_data = []
        wells_data = []
        
        # Get unique rigs and wells from assignments
        for assignment in assignments:
            rig = assignment.rig
            well = assignment.well
            
            # Add rig data
            rig_data = {
                'name': rig.name,
                'rig_type': rig.rig_type,
                'start_date': rig.start_date,
                'end_date': rig.end_date,
                'rig_capacity_hp': rig.rig_capacity_hp,
                'daily_cost_inr': rig.daily_cost_inr,
                'drilling_capacity_m': rig.drilling_capacity_m,
                'mobilization_time_days': rig.mobilization_time_days,
                'maintenance_schedule': rig.maintenance_schedule,
                'crew_availability': rig.crew_availability,
                'hpht_suitability': rig.hpht_suitability,
                'ilm_cost_fixed': rig.ilm_cost_fixed,
                'ilm_cost_per_km': rig.ilm_cost_per_km,
                'ilm_cost_cluster': rig.ilm_cost_cluster,
                'bop_stack': rig.bop_stack,
                'tds_availability': rig.tds_availability
            }
            
            # Add well data
            well_data = {
                'name': well.name,
                'sn': well.sn,
                'asset_id': well.asset_id,
                'well_type': well.well_type,
                'well_profile': well.well_profile,
                'depth': well.depth,
                'rig_capacity_required_hp': well.rig_capacity_required_hp,
                'drl_days': well.drl_days,
                'pt_days': well.pt_days,
                'duration': well.duration,
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'rtd': well.rtd,
                'bop_stack': well.bop_stack,
                'tds_requirement': well.tds_requirement,
                'footprint': well.footprint,
                'preferred_rig': well.preferred_rig,
                'expected_potential': well.expected_potential,
                'priority': well.priority
            }
            
            # Check if already added to avoid duplicates
            if not any(r['name'] == rig.name for r in rigs_data):
                rigs_data.append(rig_data)
            if not any(w['name'] == well.name for w in wells_data):
                wells_data.append(well_data)
        
        # Create scheduler instance with existing assignments
        from .optimization import DrillingScheduler
        scheduler = DrillingScheduler(rigs_data, wells_data)
        
        # Manually set results based on existing assignments
        scheduler.results = {
            'assignments': [],
            'unassigned_wells': [],
            'total_drilling_cost': 0,
            'total_ilm_cost': 0,
            'unassigned_wells_count': 0,
            'solver_status': 'OPTIMAL'
        }
        
        # Convert assignments to scheduler format
        for assignment in assignments:
            assignment_data = {
                'rig': assignment.rig.name,
                'well': assignment.well.name,
                'well_start_date': assignment.well_start_date,
                'well_end_date': assignment.well_end_date,
                'rig_start_date': assignment.rig.start_date,
                'rig_end_date': assignment.rig.end_date,
                'rtd': assignment.well.rtd,
                'priority': assignment.well.priority,
                'duration': assignment.well.duration,
                'depth': assignment.well.depth,
                'required_hp': assignment.well.rig_capacity_required_hp,
                'drilling_cost': float(assignment.drilling_cost or 0),
                'ilm_cost': float(assignment.ilm_cost or 0),
                'latitude': float(assignment.well.latitude),
                'longitude': float(assignment.well.longitude),
                'sequence_order': getattr(assignment, 'sequence_order', 1)
            }
            scheduler.results['assignments'].append(assignment_data)
        
        # Generate maps
        geographical_map_html = scheduler.generate_geographical_map()
        gantt_chart_html = scheduler.generate_gantt_chart()
        
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        context = {
            'schedule': schedule,
            'geographical_map_html': geographical_map_html,
            'gantt_chart_html': gantt_chart_html,
            'assignments_count': len(assignments),
            'asset_ids': asset_ids,
            'selected_location': location_filter,
        }
        
        return render(request, 'scheduler/schedule_maps.html', context)
        
    except Schedule.DoesNotExist:
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        return render(request, 'scheduler/schedule_maps.html', {
            'error': 'Schedule not found',
            'asset_ids': asset_ids,
        })
    except Exception as e:
        logger.error(f"Error generating schedule maps: {str(e)}")
        # Get asset IDs for location filter
        asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
        asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
        
        return render(request, 'scheduler/schedule_maps.html', {
            'error': f'Error generating maps: {str(e)}',
            'asset_ids': asset_ids,
        })


@login_required
@staff_member_required
def er_diagram(request):
    """ER Diagram and database schema visualization page - Admin only"""
    from django.apps import apps
    from django.db import connection
    
    # Get ALL models from ALL apps (not just scheduler app)
    all_models = []
    for app_config in apps.get_app_configs():
        all_models.extend(app_config.get_models())
    
    # Sort models by app name and model name for better organization
    app_models = sorted(all_models, key=lambda model: (model._meta.app_label, model.__name__))
    
    # Build model information
    models_info = {}
    apps_info = {}
    
    for model in app_models:
        model_name = model.__name__
        app_label = model._meta.app_label
        
        # Initialize app info if not exists
        if app_label not in apps_info:
            apps_info[app_label] = {
                'name': app_label,
                'models': [],
                'display_name': app_label.replace('_', ' ').title()
            }
        
        model_info = {
            'name': model_name,
            'app_label': app_label,
            'table_name': model._meta.db_table,
            'fields': [],
            'relationships': [],
            'doc_string': model.__doc__ or '',
            'verbose_name': str(model._meta.verbose_name),
            'verbose_name_plural': str(model._meta.verbose_name_plural),
        }
        
        # Add to app's models list
        apps_info[app_label]['models'].append(model_name)
        
        # Get field information
        for field in model._meta.get_fields():
            field_info = {
                'name': field.name,
                'type': field.__class__.__name__,
                'null': getattr(field, 'null', False),
                'blank': getattr(field, 'blank', False),
                'unique': getattr(field, 'unique', False),
                'primary_key': getattr(field, 'primary_key', False),
                'max_length': getattr(field, 'max_length', None),
                'help_text': getattr(field, 'help_text', ''),
                'choices': getattr(field, 'choices', None),
            }
            
            # Handle relationship fields
            if hasattr(field, 'related_model') and field.related_model:
                field_info['related_model'] = field.related_model.__name__
                field_info['related_table'] = field.related_model._meta.db_table
                
                # Add to relationships
                relationship = {
                    'from_model': model_name,
                    'to_model': field.related_model.__name__,
                    'field_name': field.name,
                    'relationship_type': field.__class__.__name__,
                }
                model_info['relationships'].append(relationship)
            
            model_info['fields'].append(field_info)
        
        models_info[model_name] = model_info
    
    # Get database statistics
    with connection.cursor() as cursor:
        stats = {}
        for model in app_models:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {model._meta.db_table}")
                result = cursor.fetchone()
                if result:
                    stats[model.__name__] = result[0]
                else:
                    stats[model.__name__] = 0
            except Exception as e:
                stats[model.__name__] = f"Error: {str(e)}"
    
    context = {
        'models_info': models_info,
        'apps_info': apps_info,
        'stats': stats,
        'total_models': len(app_models),
        'total_apps': len(apps_info),
    }
    
    return render(request, 'scheduler/er_diagram.html', context)


@login_required
def interactive_gantt(request):
    """Interactive Gantt chart view"""
    # Get asset IDs for location filter
    asset_ids = Well.objects.values_list('asset_id', flat=True).distinct().order_by('asset_id')
    asset_ids = [asset_id for asset_id in asset_ids if asset_id]  # Filter out empty/null values
    
    context = {
        'asset_ids': asset_ids,
    }
    return render(request, 'scheduler/interactive_gantt.html', context)


def schedule_detail(request, schedule_id):
    """Schedule detail view showing comprehensive schedule information"""
    schedule = get_object_or_404(Schedule, id=schedule_id)
    
    # Get all assignments for this schedule - using all_objects to include deleted items
    assignments = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well').order_by('rig__name', 'sequence_order')
    
    # Get unassigned wells
    unassigned_wells = UnassignedWell.objects.filter(schedule=schedule).select_related('well').order_by('well__sn')
    
    # Get input data - use all_objects to include any deleted wells/rigs that are in this schedule
    # This ensures historical schedules display correctly even if wells/rigs were deleted
    all_wells = Well.all_objects.all().order_by('sn')
    all_rigs = Rig.all_objects.all().order_by('name')
    
    # Identify assigned wells and rigs
    assigned_well_ids = assignments.values_list('well_id', flat=True)
    assigned_rig_ids = assignments.values_list('rig_id', flat=True)
    
    # Group assignments by rig for better display
    assignments_by_rig = {}
    for assignment in assignments:
        rig_name = assignment.rig.name
        if rig_name not in assignments_by_rig:
            assignments_by_rig[rig_name] = []
        assignments_by_rig[rig_name].append(assignment)
    
    # Calculate utilization statistics
    total_wells = all_wells.count()
    assigned_wells_count = len(assigned_well_ids)
    unassigned_wells_count = unassigned_wells.count()
    
    total_rigs = all_rigs.count()
    utilized_rigs_count = len(set(assigned_rig_ids))
    unused_rigs_count = total_rigs - utilized_rigs_count
    
    # Get unused rigs with reasons
    unused_rigs = all_rigs.exclude(id__in=assigned_rig_ids)
    
    # Calculate costs
    total_drilling_cost = assignments.aggregate(
        total=models.Sum('drilling_cost')
    )['total'] or Decimal('0.00')
    
    total_ilm_cost = assignments.aggregate(
        total=models.Sum('ilm_cost')
    )['total'] or Decimal('0.00')
    
    context = {
        'schedule': schedule,
        'assignments': assignments,
        'assignments_by_rig': assignments_by_rig,
        'unassigned_wells': unassigned_wells,
        'all_wells': all_wells,
        'all_rigs': all_rigs,
        'assigned_well_ids': assigned_well_ids,
        'assigned_rig_ids': assigned_rig_ids,
        'unused_rigs': unused_rigs,
        'stats': {
            'total_wells': total_wells,
            'assigned_wells_count': assigned_wells_count,
            'unassigned_wells_count': unassigned_wells_count,
            'total_rigs': total_rigs,
            'utilized_rigs_count': utilized_rigs_count,
            'unused_rigs_count': unused_rigs_count,
            'total_drilling_cost': total_drilling_cost,
            'total_ilm_cost': total_ilm_cost,
            'total_cost': total_drilling_cost + total_ilm_cost,
        }
    }
    
    return render(request, 'scheduler/schedule_detail.html', context)


def schedule_comparison(request):
    """Schedule comparison view for analyzing multiple schedules"""
    schedule_ids = request.GET.getlist('schedules')
    
    if not schedule_ids:
        # If no schedules specified, show empty comparison page
        return render(request, 'scheduler/schedule_comparison.html', {
            'schedules': [],
            'error_message': 'No schedules selected for comparison. Please select schedules from the scheduling page.'
        })
    
    if len(schedule_ids) > 3:
        # Limit to 3 schedules for better display
        schedule_ids = schedule_ids[:3]
    
    schedules_data = []
    
    for schedule_id in schedule_ids:
        try:
            schedule = Schedule.objects.get(id=schedule_id)
            
            # Get assignments and statistics for this schedule
            assignments = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
            unassigned_wells = UnassignedWell.objects.filter(schedule=schedule).select_related('well')
            
            # Calculate statistics
            total_drilling_cost = assignments.aggregate(
                total=models.Sum('drilling_cost')
            )['total'] or Decimal('0.00')
            
            total_ilm_cost = assignments.aggregate(
                total=models.Sum('ilm_cost')
            )['total'] or Decimal('0.00')
            
            total_cost = total_drilling_cost + total_ilm_cost
            
            # Get rig utilization - count unique rigs, not wells
            utilized_rigs = assignments.values_list('rig__name', flat=True).distinct()
            utilized_rigs_count = len(set(utilized_rigs))
            
            # Get detailed rig information
            rig_details = []
            for rig_name in set(utilized_rigs):
                rig_assignments = assignments.filter(rig__name=rig_name)
                rig_details.append({
                    'name': rig_name,
                    'wells_count': rig_assignments.count(),
                    'total_cost': rig_assignments.aggregate(
                        total=models.Sum('drilling_cost') + models.Sum('ilm_cost')
                    )['total'] or Decimal('0.00')
                })
            
            # Group assignments by rig
            assignments_by_rig = {}
            for assignment in assignments:
                rig_name = assignment.rig.name
                if rig_name not in assignments_by_rig:
                    assignments_by_rig[rig_name] = []
                assignments_by_rig[rig_name].append(assignment)
            
            # Calculate efficiency metrics
            assigned_wells_count = assignments.count()
            unassigned_wells_count = unassigned_wells.count()
            total_wells_count = assigned_wells_count + unassigned_wells_count
            success_rate = (assigned_wells_count / total_wells_count * 100) if total_wells_count > 0 else 0
            cost_per_well = (total_cost / assigned_wells_count) if assigned_wells_count > 0 else 0
            wells_per_rig = (assigned_wells_count / utilized_rigs_count) if utilized_rigs_count > 0 else 0
            
            schedule_data = {
                'schedule': schedule,
                'assignments': assignments,
                'assignments_by_rig': assignments_by_rig,
                'unassigned_wells': unassigned_wells,
                'rig_details': rig_details,
                'stats': {
                    'assigned_wells_count': assigned_wells_count,
                    'unassigned_wells_count': unassigned_wells_count,
                    'total_wells_count': total_wells_count,
                    'utilized_rigs_count': utilized_rigs_count,
                    'utilized_rigs': list(set(utilized_rigs)),
                    'total_drilling_cost': total_drilling_cost,
                    'total_ilm_cost': total_ilm_cost,
                    'total_cost': total_cost,
                    'success_rate': success_rate,
                    'cost_per_well': cost_per_well,
                    'wells_per_rig': wells_per_rig,
                }
            }
            
            schedules_data.append(schedule_data)
            
        except Schedule.DoesNotExist:
            continue
    
    if not schedules_data:
        return render(request, 'scheduler/schedule_comparison.html', {
            'schedules': [],
            'error_message': 'No valid schedules found for comparison.'
        })
    
    context = {
        'schedules': schedules_data,
        'comparison_count': len(schedules_data)
    }
    
    return render(request, 'scheduler/schedule_comparison.html', context)


class RigViewSet(viewsets.ModelViewSet):
    """ViewSet for managing drilling rigs"""
    queryset = Rig.all_objects.all()  # Include soft-deleted records
    serializer_class = RigSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Rig.all_objects.all()  # Include soft-deleted records
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                # Filter by location FK OR by asset_id matching location code (for legacy data)
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(location=user_location) | Q(asset_id__iexact=user_location.code)
                )
        
        rig_type = self.request.query_params.get('rig_type', None)
        available_now = self.request.query_params.get('available_now', None)
        asset_id = self.request.query_params.get('asset_id', None)
        location_id = self.request.query_params.get('location_id', None)
        
        if rig_type:
            queryset = queryset.filter(rig_type=rig_type)
        
        if available_now == 'true':
            today = timezone.now().date()
            queryset = queryset.filter(
                start_date__lte=today,
                end_date__gte=today
            )
        
        if asset_id:
            queryset = queryset.filter(asset_id=asset_id)
        
        # Additional location filter from query params (within user's accessible locations)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        return queryset

    def destroy(self, request, *args, **kwargs):
        """Override hard delete: perform soft-delete to avoid ProtectedError and preserve history"""
        try:
            rig = self.get_object()
            # If already deleted, return idempotent success
            if getattr(rig, 'is_deleted', False):
                return Response({'message': 'Rig already deleted'}, status=status.HTTP_200_OK)
            rig.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Rig soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete rig: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def soft_delete(self, request, pk=None):
        """Soft delete a rig without affecting existing schedules"""
        try:
            rig = self.get_object()
            if rig.is_deleted:
                return Response({'message': 'Rig already deleted'}, status=status.HTTP_200_OK)
            rig.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Rig soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete rig: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        """Bulk upload rigs from CSV file"""
        serializer = BulkDataUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        rigs_file = serializer.validated_data.get('rigs_file')
        if not rigs_file:
            return Response(
                {'error': 'No rigs file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Read CSV file
            df = pd.read_csv(rigs_file)
            created_rigs = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Map CSV columns to model fields (updated to match sample_rigs.csv format)
                    rig_data = {
                        'name': row.get('name') or row.get('Rig'),
                        'asset_id': row.get('asset_id') or row.get('Asset_ID'),
                        'rig_type': row.get('rig_type') or row.get('Rig Type'),
                        'start_date': pd.to_datetime(row.get('start_date') or row.get('Start Date'), dayfirst=True).date(),
                        'end_date': pd.to_datetime(row.get('end_date') or row.get('End Date'), dayfirst=True).date(),
                        'rig_capacity_hp': int(row.get('rig_capacity_hp') or row.get('Rig Capacity (HP)')),
                        'daily_cost_inr': Decimal(str(row.get('daily_cost_inr') or row.get('Daily Cost (INR)'))),
                        'drilling_capacity_m': int(row.get('drilling_capacity_m') or row.get('Drilling Capacity (m)')),
                        'mobilization_time_days': row.get('mobilization_time_days') or row.get('Mobilization Time (Days)'),
                        'maintenance_schedule': row.get('maintenance_schedule') or row.get('Maintenance Schedule'),
                        'crew_availability': row.get('crew_availability') or row.get('Crew Availability', 'OK'),
                        'hpht_suitability': row.get('hpht_suitability') or row.get('HPHT Suitability', 'N'),
                        'ilm_cost_fixed': Decimal(str(row.get('ilm_cost_fixed') or row.get('ILM COST FIXED'))),
                        'ilm_cost_per_km': Decimal(str(row.get('ilm_cost_per_km') or row.get('ILM COST per km'))),
                        'ilm_cost_cluster': Decimal(str(row.get('ilm_cost_cluster') or row.get('ILM COST CLUSTER'))),
                        'bop_stack': int(row.get('bop_stack') or row.get('BOP Stack')),
                        'tds_availability': row.get('tds_availability') or row.get('TDS Availability', 'Y'),
                    }
                    
                    # Clean the name
                    rig_name = str(rig_data['name']).strip()
                    rig_data['name'] = rig_name
                    
                    # Separate lookup key from data fields
                    defaults = {k: v for k, v in rig_data.items() if k != 'name'}
                    
                    # Use get_or_create with all_objects to handle both new and existing (including soft-deleted)
                    # This is atomic and works WITH the UNIQUE constraint, not against it
                    rig, created = Rig.all_objects.get_or_create(
                        name=rig_name,
                        defaults=defaults
                    )
                    
                    if not created:
                        # Record already exists (active or soft-deleted) - update all fields
                        was_deleted = rig.is_deleted
                        for key, value in rig_data.items():
                            setattr(rig, key, value)
                        
                        # Revive if soft-deleted
                        if was_deleted:
                            rig.is_deleted = False
                            rig.deleted_at = None
                            rig.deleted_by = None
                        
                        rig.save()
                        status = "revived" if was_deleted else "updated"
                        created_rigs.append(f"{rig.name} ({status})")
                    else:
                        # Truly new record
                        created_rigs.append(f"{rig.name} (new)")
                    
                except Exception as e:
                    errors.append(f"Row {index + 1} (Rig: {rig_data.get('name', 'Unknown')}): {str(e)}")
            
            return Response({
                'message': f'Successfully processed {len(created_rigs)} rigs',
                'created_rigs': created_rigs,
                'errors': errors
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export rigs data as CSV"""
        try:
            queryset = self.get_queryset()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="rigs_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = [
                'ID', 'Name', 'Asset ID', 'Rig Type', 'Start Date', 'End Date',
                'Rig Capacity (HP)', 'Daily Cost (INR)', 'Drilling Capacity (m)',
                'Mobilization Time (Days)', 'Maintenance Schedule', 'Crew Availability',
                'HPHT Suitability', 'ILM Cost Fixed', 'ILM Cost per km',
                'ILM Cost Cluster', 'BOP Stack', 'TDS Availability', 'Created At'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for rig in queryset:
                writer.writerow([
                    str(rig.id),
                    rig.name,
                    rig.asset_id or '',
                    rig.rig_type,
                    rig.start_date.strftime('%Y-%m-%d') if rig.start_date else '',
                    rig.end_date.strftime('%Y-%m-%d') if rig.end_date else '',
                    rig.rig_capacity_hp,
                    rig.daily_cost_inr,
                    rig.drilling_capacity_m,
                    rig.mobilization_time_days or '',
                    rig.maintenance_schedule or '',
                    rig.crew_availability or '',
                    rig.hpht_suitability or '',
                    rig.ilm_cost_fixed or '',
                    rig.ilm_cost_per_km or '',
                    rig.ilm_cost_cluster or '',
                    rig.bop_stack or '',
                    rig.tds_availability or '',
                    rig.created_at.strftime('%Y-%m-%d %H:%M:%S') if rig.created_at else ''
                ])
            
            return response
            
        except Exception as e:
            return JsonResponse(
                {'error': f'Failed to export CSV: {str(e)}'}, 
                status=500
            )


class WellViewSet(viewsets.ModelViewSet):
    """ViewSet for managing wells"""
    queryset = Well.all_objects.all()  # Include soft-deleted records
    serializer_class = WellSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Well.all_objects.all()  # Include soft-deleted records
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                # Filter by location FK OR by asset_id matching location code (for legacy data)
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(location=user_location) | Q(asset_id__iexact=user_location.code)
                )
        
        asset_id = self.request.query_params.get('asset_id', None)
        well_type = self.request.query_params.get('well_type', None)
        priority = self.request.query_params.get('priority', None)
        location_id = self.request.query_params.get('location_id', None)
        
        if asset_id:
            queryset = queryset.filter(asset_id=asset_id)
        if well_type:
            queryset = queryset.filter(well_type=well_type)
        if priority:
            queryset = queryset.filter(priority=priority)
        
        # Additional location filter from query params (within user's accessible locations)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        return queryset

    def destroy(self, request, *args, **kwargs):
        """Override hard delete: perform soft-delete to avoid ProtectedError and preserve history"""
        try:
            well = self.get_object()
            # If already deleted, return idempotent success
            if getattr(well, 'is_deleted', False):
                return Response({'message': 'Well already deleted'}, status=status.HTTP_200_OK)
            well.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Well soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete well: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def soft_delete(self, request, pk=None):
        """Soft delete a well without affecting existing schedules"""
        try:
            well = self.get_object()
            if well.is_deleted:
                return Response({'message': 'Well already deleted'}, status=status.HTTP_200_OK)
            well.soft_delete(user=request.user if request.user.is_authenticated else None)
            return Response({'message': 'Well soft-deleted successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete well: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        """Bulk upload wells from CSV file"""
        serializer = BulkDataUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        wells_file = serializer.validated_data.get('wells_file')
        if not wells_file:
            return Response(
                {'error': 'No wells file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Read CSV file
            df = pd.read_csv(wells_file)
            created_wells = []
            updated_wells = []
            errors = []
            
            # Get the next available SN
            from django.db.models import Max
            max_sn = Well.objects.aggregate(Max('sn'))['sn__max'] or 0
            next_sn = max_sn + 1
            
            for index, row in df.iterrows():
                try:
                    well_name = row.get('Well')
                    asset_id = row.get('Asset_ID')
                    
                    # Map CSV columns to model fields
                    well_data = {
                        'asset_id': asset_id,
                        'name': well_name,
                        'well_type': row.get('Type of well'),
                        'well_profile': row.get('Well Profile'),
                        'depth': int(row.get('Depth')),
                        'rig_capacity_required_hp': int(row.get('Rig Capacity Required (HP)')),
                        'drl_days': int(row.get('DRL_DAYS')),
                        'pt_days': int(row.get('PT_DAYS')),
                        'duration': int(row.get('Duration')),
                        'latitude': Decimal(str(row.get('Latitude'))),
                        'longitude': Decimal(str(row.get('Longitude'))),
                        'rtd': pd.to_datetime(row.get('RTD'), dayfirst=True).date(),
                        'bop_stack': int(row.get('BOP Stack')),
                        'tds_requirement': row.get('TDS Requirement', 'Y'),
                        'footprint': row.get('Footprint'),
                        'preferred_rig': row.get('Preferred Rig') if pd.notna(row.get('Preferred Rig')) else None,
                        'expected_potential': row.get('Expected_Potential') if pd.notna(row.get('Expected_Potential')) else None,
                        'priority': row.get('Priority', 'MEDIUM') if pd.notna(row.get('Priority')) else 'MEDIUM',
                    }
                    
                    # Clean the identifiers
                    well_name_clean = str(well_name).strip()
                    asset_id_clean = str(asset_id).strip()
                    well_data['name'] = well_name_clean
                    well_data['asset_id'] = asset_id_clean
                    
                    # Prepare defaults for get_or_create (exclude lookup fields)
                    defaults = {k: v for k, v in well_data.items() if k not in ['name', 'asset_id']}
                    defaults['sn'] = next_sn  # Will be used only if creating new
                    
                    # Use get_or_create with all_objects to handle both new and existing (including soft-deleted)
                    # This is atomic and works WITH the UNIQUE constraint, not against it
                    well, created = Well.all_objects.get_or_create(
                        name=well_name_clean,
                        asset_id=asset_id_clean,
                        defaults=defaults
                    )
                    
                    if not created:
                        # Record already exists (active or soft-deleted) - update all fields
                        was_deleted = well.is_deleted
                        for key, value in well_data.items():
                            if key not in ['name', 'asset_id']:  # Don't update lookup keys
                                setattr(well, key, value)
                        
                        # Revive if soft-deleted
                        if was_deleted:
                            well.is_deleted = False
                            well.deleted_at = None
                            well.deleted_by = None
                        
                        well.save()
                        status = "revived" if was_deleted else "updated"
                        updated_wells.append(f"{well.name} ({status})")
                    else:
                        # Truly new record
                        created_wells.append(f"{well.name} (new)")
                        next_sn += 1
                    
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
            
            return Response({
                'message': f'Successfully processed wells: {len(created_wells)} created, {len(updated_wells)} updated',
                'created_wells': created_wells,
                'updated_wells': updated_wells,
                'errors': errors
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export wells data as CSV"""
        try:
            queryset = self.get_queryset()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="wells_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = [
                'ID', 'SN', 'Asset ID', 'Name', 'Well Type', 'Well Profile', 'Depth',
                'Rig Capacity Required (HP)', 'DRL Days', 'PT Days', 'Duration',
                'Latitude', 'Longitude', 'RTD', 'BOP Stack', 'TDS Requirement',
                'Footprint', 'Preferred Rig', 'Expected Potential', 'Priority', 'Created At'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for well in queryset:
                writer.writerow([
                    str(well.id),
                    well.sn or '',
                    well.asset_id or '',
                    well.name,
                    well.well_type,
                    well.well_profile or '',
                    well.depth or '',
                    well.rig_capacity_required_hp or '',
                    well.drl_days or '',
                    well.pt_days or '',
                    well.duration or '',
                    well.latitude or '',
                    well.longitude or '',
                    well.rtd.strftime('%Y-%m-%d') if well.rtd else '',
                    well.bop_stack or '',
                    well.tds_requirement or '',
                    well.footprint or '',
                    well.preferred_rig or '',
                    well.expected_potential or '',
                    well.priority or '',
                    well.created_at.strftime('%Y-%m-%d %H:%M:%S') if well.created_at else ''
                ])
            
            return response
            
        except Exception as e:
            return JsonResponse(
                {'error': f'Failed to export CSV: {str(e)}'}, 
                status=500
            )


#new code piece
def run_full_optimization(rig_queryset=None, well_queryset=None, base_start_date=None, time_limit_seconds=60):
    """
    Runs the full optimizer and returns the result dictionary.
    """
    rigs = rig_queryset if rig_queryset is not None else Rig.objects.all()
    wells = well_queryset if well_queryset is not None else Well.objects.all()
    rigs_data = [rig.to_dict() for rig in rigs]
    wells_data = [well.to_dict() for well in wells]
    scheduler = DrillingScheduler(rigs_data, wells_data, base_start_date)
    scheduler.preprocess_data()
    scheduler.setup_variables()
    scheduler.add_constraints()
    scheduler.add_ilm_constraints()
    scheduler.set_objective()
    result = scheduler.solve(time_limit_seconds=time_limit_seconds)
    return result
#-- new code piece
class ScheduleViewSet(viewsets.ModelViewSet):
    """ViewSet for managing schedules"""
    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Schedule.objects.all().order_by('-created_at')
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                queryset = queryset.filter(location=user_location)
        
        # Check if request has query_params (DRF Request) or GET (Django Request)
        if hasattr(self.request, 'query_params'):
            status_filter = self.request.query_params.get('status', None)
            asset_id_filter = self.request.query_params.get('asset_id', None)
            location_id = self.request.query_params.get('location_id', None)
        else:
            status_filter = self.request.GET.get('status', None)
            asset_id_filter = self.request.GET.get('asset_id', None)
            location_id = self.request.GET.get('location_id', None)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter schedules by asset_id (location) - legacy support
        if asset_id_filter:
            # Get schedules that have assignments with wells from the specified asset_id
            queryset = queryset.filter(
                assignments__well__asset_id=asset_id_filter
            ).distinct()
        
        # Additional location filter from query params (within user's accessible locations)
        if location_id:
            queryset = queryset.filter(location_id=location_id)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def create_schedule(self, request):
        """Create and run a new schedule optimization"""
        serializer = ScheduleCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        
        try:
            with transaction.atomic():
                # Determine location for the schedule
                schedule_location = None
                
                # If user has a specific location, use that
                if request.user and request.user.is_authenticated:
                    user_location = get_user_location(request.user)
                    if user_location:
                        schedule_location = user_location
                
                # Create schedule instance
                schedule = Schedule.objects.create(
                    name=validated_data['name'],
                    financial_year=validated_data['financial_year'],
                    location=schedule_location,
                    status='RUNNING'
                )
                
                # Get rigs and wells
                rigs = Rig.objects.filter(id__in=validated_data['rig_ids'])
                wells = Well.objects.filter(id__in=validated_data['well_ids'])
                
                # If no location set yet, try to get from first well
                if not schedule_location and wells.exists():
                    first_well = wells.first()
                    if first_well.location:
                        schedule.location = first_well.location
                        schedule.save()
                
                if not rigs.exists():
                    return Response(
                        {'error': 'No valid rigs found'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                if not wells.exists():
                    return Response(
                        {'error': 'No valid wells found'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Run optimization
                try:
                    # Convert QuerySets to dictionaries for the optimizer
                    rigs_data = list(rigs.values())
                    wells_data = list(wells.values())
                    
                    scheduler = DrillingScheduler(rigs_data, wells_data)
                    scheduler.preprocess_data()
                    scheduler.setup_variables()
                    scheduler.add_constraints()
                    scheduler.add_ilm_constraints()
                    scheduler.set_objective()
                    
                    results = scheduler.solve(validated_data['time_limit_seconds'])
                    
                    if results and results.get('status') in ['OPTIMAL', 'FEASIBLE']:
                        success = True
                        
                        # Update schedule with results
                        schedule.status = 'COMPLETED'
                        schedule.completed_at = timezone.now()
                        schedule.total_drilling_cost = Decimal(str(results.get('total_drilling_cost', 0)))
                        schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
                        schedule.project_end_date = results.get('project_end_date')
                        schedule.unassigned_wells_count = results.get('unassigned_wells_count', 0)
                        schedule.solver_status = results.get('solver_status')
                        schedule.solve_time_seconds = results.get('solve_time_seconds')
                        schedule.save()
                        
                        # Store the selected rigs and wells for this schedule
                        from .models import ScheduleRig, ScheduleWell
                        
                        # Create ScheduleRig entries
                        schedule_rigs = []
                        for rig in rigs:
                            schedule_rigs.append(ScheduleRig(schedule=schedule, rig=rig))
                        ScheduleRig.objects.bulk_create(schedule_rigs, ignore_conflicts=True)
                        
                        # Create ScheduleWell entries
                        schedule_wells = []
                        for well in wells:
                            schedule_wells.append(ScheduleWell(schedule=schedule, well=well))
                        ScheduleWell.objects.bulk_create(schedule_wells, ignore_conflicts=True)
                        
                        # Create assignments with proper sequence order
                        # First, group assignments by rig and calculate proper sequence order
                        assignments_by_rig = {}
                        for assignment_data in results.get('assignments', []):
                            rig_name = assignment_data['rig']
                            if rig_name not in assignments_by_rig:
                                assignments_by_rig[rig_name] = []
                            assignments_by_rig[rig_name].append(assignment_data)
                        
                        # Sort each rig's assignments by start date and assign sequence numbers
                        for rig_name, rig_assignments in assignments_by_rig.items():
                            # Sort by start date
                            rig_assignments.sort(key=lambda x: x['well_start_date'])
                            # Assign sequence numbers starting from 1
                            for i, assignment_data in enumerate(rig_assignments, 1):
                                assignment_data['calculated_sequence_order'] = i
                        
                        # Now create the assignments with proper sequence numbers
                        for assignment_data in results.get('assignments', []):
                            try:
                                rig = rigs.get(name=assignment_data['rig'])
                                well = wells.get(name=assignment_data['well'])
                            except (Rig.DoesNotExist, Well.DoesNotExist) as e:
                                logger.warning(f"Assignment creation failed - {str(e)}: rig={assignment_data.get('rig')}, well={assignment_data.get('well')}")
                                continue
                                
                            Assignment.objects.create(
                                schedule=schedule,
                                rig=rig,
                                well=well,
                                well_start_date=assignment_data['well_start_date'],
                                well_end_date=assignment_data['well_end_date'],
                                rtd_check=assignment_data.get('rtd_check', 'OK'),
                                well_start_check=assignment_data.get('well_start_check', 'OK'),
                                well_end_check=assignment_data.get('well_end_check', 'OK'),
                                depth_check=assignment_data.get('depth_check', 'OK'),
                                hp_check=assignment_data.get('hp_check', 'OK'),
                                bop_check=assignment_data.get('bop_check', 'OK'),
                                tds_check=assignment_data.get('tds_check', 'OK'),
                                rig_type_check=assignment_data.get('rig_type_check', 'OK'),
                                drilling_cost=Decimal(str(assignment_data.get('drilling_cost_inr', assignment_data.get('drilling_cost', 0)))),
                                ilm_cost=Decimal(str(assignment_data.get('ilm_cost', 0))),
                                sequence_order=assignment_data.get('calculated_sequence_order', 1)
                            )
                        
                        # Create unassigned wells with detailed rejection analysis
                        assigned_well_names = [assignment['well'] for assignment in results.get('assignments', [])]
                        
                        # Initialize rejection analyzer
                        wells_data = []
                        rigs_data = []
                        
                        for well in wells:
                            wells_data.append({
                                'name': well.name,
                                'depth': well.depth,
                                'duration': well.duration,
                                'rtd': well.rtd,
                                'rig_capacity_required_hp': well.rig_capacity_required_hp,
                                'bop_stack': well.bop_stack,
                                'tds_requirement': well.tds_requirement,
                                'priority': well.priority,
                                'latitude': float(well.latitude) if well.latitude else 0.0,
                                'longitude': float(well.longitude) if well.longitude else 0.0,
                                'footprint': well.footprint,
                            })
                        
                        for rig in rigs:
                            rigs_data.append({
                                'name': rig.name,
                                'start_date': rig.start_date,
                                'end_date': rig.end_date,
                                'rig_capacity_hp': rig.rig_capacity_hp,
                                'drilling_capacity_m': rig.drilling_capacity_m,
                                'bop_stack': rig.bop_stack,
                                'tds_availability': rig.tds_availability,
                                'daily_cost_inr': float(rig.daily_cost_inr),
                                'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                                'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                                'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                                'rig_type': rig.rig_type,
                            })
                        
                        # Convert to DataFrames for analysis
                        wells_df = pd.DataFrame(wells_data)
                        rigs_df = pd.DataFrame(rigs_data)
                        
                        # Initialize analyzer
                        analyzer = WellRejectionAnalyzer(wells_df, rigs_df, timezone.now().date())
                        
                        for well_name in results.get('unassigned_wells', []):
                            try:
                                well = wells.get(name=well_name)
                            except Well.DoesNotExist:
                                logger.warning(f"Unassigned well creation failed - Well not found: {well_name}")
                                continue
                            
                            # Get detailed rejection reason
                            detailed_reason = analyzer.analyze_well_rejection(well_name, assigned_well_names)
                            
                            UnassignedWell.objects.create(
                                schedule=schedule,
                                well=well,
                                reason=detailed_reason
                            )
                        
                        return Response(ScheduleSerializer(schedule).data)
                    
                    else:
                        schedule.status = 'FAILED'
                        schedule.solver_status = results.get('status', 'NO_SOLUTION') if results else 'NO_SOLUTION'
                        schedule.save()
                        
                        return Response(
                            {'error': 'No feasible solution found'}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                        
                except Exception as e:
                    schedule.status = 'FAILED'
                    schedule.save()
                    logger.error(f"Optimization failed: {str(e)}")
                    return Response(
                        {'error': f'Optimization failed: {str(e)}'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                    
        except Exception as e:
            logger.error(f"Schedule creation failed: {str(e)}")
            return Response(
                {'error': f'Schedule creation failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def reschedule_with_actuals(self, request, pk=None):
        """Re-optimize an existing schedule using actual start/end dates.

        Expected payload:
        {
          "actuals": [
             {"well": "Well-1", "rig": "Rig-1", "actual_start_date": "2025-05-10", "actual_end_date": "2025-06-05"},
             {"assignment_id": "<uuid>", "actual_start_date": "2025-07-01"}
          ],
          "time_limit_seconds": 60
        }
        - If rig is omitted, it will be inferred from the current schedule's assignment for that well/assignment.
        - If only one of start/end is provided, only that boundary is pinned.
        - A new Schedule will be created to store results (original remains unchanged).
        """
        schedule = self.get_object()

        payload = request.data if isinstance(request.data, dict) else {}
        actuals = payload.get('actuals', []) or []
        time_limit_seconds = int(payload.get('time_limit_seconds', 60))

        if not isinstance(actuals, list) or len(actuals) == 0:
            return Response({'error': 'Provide a non-empty "actuals" list'}, status=status.HTTP_400_BAD_REQUEST)

        # Build current rigs/wells dataset from the ORIGINAL SCOPE of the schedule
        from .models import ScheduleRig, ScheduleWell
        
        # Get the originally selected rigs and wells for this schedule
        schedule_rigs = ScheduleRig.objects.filter(schedule=schedule).select_related('rig')
        schedule_wells = ScheduleWell.objects.filter(schedule=schedule).select_related('well')
        
        if not schedule_rigs.exists() or not schedule_wells.exists():
            # Fallback to assignment-based detection if no scope tracking available
            assignments_qs = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
            if not assignments_qs.exists():
                return Response({'error': 'No assignments or scope found for the provided schedule'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Use rigs and wells from assignments as fallback
            rigs = [a.rig for a in assignments_qs]
            wells = [a.well for a in assignments_qs]
            rigs_seen = set([r.name for r in rigs])
            wells_seen = set([w.name for w in wells])
        else:
            # Use the originally selected scope
            rigs = [sr.rig for sr in schedule_rigs]
            wells = [sw.well for sw in schedule_wells]
            rigs_seen = set([r.name for r in rigs])
            wells_seen = set([w.name for w in wells])

        rigs_data: list[dict] = []
        wells_data: list[dict] = []

        # Build the data structures for rigs and wells
        for rig in rigs:
            if rig.name not in [r['name'] for r in rigs_data]:  # Avoid duplicates
                rigs_data.append({
                    'name': rig.name,
                    'rig_type': rig.rig_type,
                    'start_date': rig.start_date,
                    'end_date': rig.end_date,
                    'rig_capacity_hp': rig.rig_capacity_hp,
                    'daily_cost_inr': float(rig.daily_cost_inr),
                    'drilling_capacity_m': rig.drilling_capacity_m,
                    'mobilization_time_days': rig.mobilization_time_days,
                    'maintenance_schedule': rig.maintenance_schedule,
                    'crew_availability': rig.crew_availability,
                    'hpht_suitability': rig.hpht_suitability,
                    'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                    'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                    'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                    'bop_stack': rig.bop_stack,
                    'tds_availability': rig.tds_availability,
                })
        
        for well in wells:
            if well.name not in [w['name'] for w in wells_data]:  # Avoid duplicates
                wells_data.append({
                    'name': well.name,
                    'sn': well.sn,
                    'asset_id': well.asset_id,
                    'well_type': well.well_type,
                    'well_profile': well.well_profile,
                    'depth': well.depth,
                    'rig_capacity_required_hp': well.rig_capacity_required_hp,
                    'drl_days': well.drl_days,
                    'pt_days': well.pt_days,
                    'duration': well.duration,
                    'latitude': float(well.latitude),
                    'longitude': float(well.longitude),
                    'rtd': well.rtd,
                    'bop_stack': well.bop_stack,
                    'tds_requirement': well.tds_requirement,
                    'footprint': well.footprint,
                    'preferred_rig': well.preferred_rig,
                    'expected_potential': well.expected_potential,
                    'priority': well.priority,
                })
        
        # Build mapping from existing assignments
        assignments_qs = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
        well_to_rig: dict[str, str] = {}
        assignment_map: dict[str, dict] = {}
        
        for a in assignments_qs:
            well_to_rig[a.well.name] = a.rig.name
            assignment_map[str(a.id)] = {
                'well': a.well.name,
                'rig': a.rig.name,
            }

        # STEP 1: Build comprehensive actuals list from EXISTING assignments with actual dates
        # This ensures ALL locked dates are preserved with absolute priority
        normalized_actuals: list[dict] = []
        
        # Add ALL existing actual dates from original schedule as FIXED constraints
        for assignment in assignments_qs:
            if assignment.actual_start_date or assignment.actual_end_date:
                normalized_actuals.append({
                    'well': assignment.well.name,
                    'rig': assignment.rig.name,
                    'actual_start_date': assignment.actual_start_date,
                    'actual_end_date': assignment.actual_end_date,
                })
        
        # STEP 2: Process new actuals provided in the API call (these override existing if same well-rig)
        actuals_map = {}  # Track well-rig combinations to avoid duplicates
        for existing in normalized_actuals:
            key = (existing['well'], existing['rig'])
            actuals_map[key] = existing
            
        for rec in actuals:
            if not isinstance(rec, dict):
                continue
            well = rec.get('well')
            rig = rec.get('rig')
            assignment_id = rec.get('assignment_id')
            
            # Infer well/rig from assignment_id if not provided
            if not well and assignment_id and str(assignment_id) in assignment_map:
                well = assignment_map[str(assignment_id)]['well']
                rig = rig or assignment_map[str(assignment_id)]['rig']
            if well and not rig:
                # infer rig from current schedule
                rig = well_to_rig.get(well)
            if not well or not rig:
                # skip records we cannot map to current schedule
                continue
            
            # Filter out None values and only include records with actual dates
            actual_start_date = rec.get('actual_start_date')
            actual_end_date = rec.get('actual_end_date')
            
            if actual_start_date or actual_end_date:
                key = (well, rig)
                # Override existing or add new actual dates
                actuals_map[key] = {
                    'well': well,
                    'rig': rig,
                    'actual_start_date': actual_start_date,
                    'actual_end_date': actual_end_date,
                }

        # Convert back to list with only valid actual dates
        normalized_actuals = []
        for actual_data in actuals_map.values():
            # Only include if at least one actual date is provided and not None
            if actual_data.get('actual_start_date') or actual_data.get('actual_end_date'):
                normalized_actuals.append(actual_data)

        logger.info(f"Fixed actuals for optimization: {len(normalized_actuals)} well-rig combinations with locked dates")
        
        if not normalized_actuals:
            logger.warning("No actual dates found in existing assignments or provided actuals - running normal optimization")
            # If no actual dates to lock, run normal optimization instead
            scheduler = DrillingScheduler(rigs_data, wells_data)
            results = scheduler.solve(time_limit_seconds=time_limit_seconds)
        else:
            # Run rescheduling optimizer with locked actual dates as FIXED constraints
            scheduler = DrillingScheduler(rigs_data, wells_data)
            results = scheduler.solve_with_actuals(normalized_actuals, time_limit_seconds=time_limit_seconds)

        if not results or results.get('status') not in ['OPTIMAL', 'FEASIBLE']:
            # Provide detailed analysis when solution is infeasible
            failure_analysis = scheduler.analyze_infeasible_solution(normalized_actuals)
            solver_status = results.get('status') if results else 'NO_RESULT'
            
            return Response({
                'error': f"No feasible solution: {solver_status}",
                'error_type': 'INFEASIBLE_SOLUTION',
                'detailed_analysis': failure_analysis,
                'solver_status': solver_status,
                'actuals_provided': len(normalized_actuals)
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create a properly named child schedule
        new_schedule = create_child_schedule(schedule, branch_type="reschedule")
        
        # Update with optimization results
        new_schedule.status = 'COMPLETED'
        new_schedule.completed_at = timezone.now()
        new_schedule.total_drilling_cost = Decimal(str(results.get('total_drilling_cost', 0)))
        new_schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
        new_schedule.project_end_date = results.get('project_end_date')
        new_schedule.unassigned_wells_count = len(results.get('unassigned_wells', []))
        new_schedule.solver_status = results.get('solver_status')
        new_schedule.solve_time_seconds = results.get('solve_time_seconds')
        new_schedule.save()
        
        # Copy the scope tracking from the original schedule
        # Create ScheduleRig entries for the new schedule
        schedule_rigs = []
        for rig in rigs:
            schedule_rigs.append(ScheduleRig(schedule=new_schedule, rig=rig))
        ScheduleRig.objects.bulk_create(schedule_rigs, ignore_conflicts=True)
        
        # Create ScheduleWell entries for the new schedule
        schedule_wells = []
        for well in wells:
            schedule_wells.append(ScheduleWell(schedule=new_schedule, well=well))
        ScheduleWell.objects.bulk_create(schedule_wells, ignore_conflicts=True)

        # Create assignments with proper sequence order
        assignments_by_rig: dict[str, list] = {}
        for ad in results.get('assignments', []):
            assignments_by_rig.setdefault(ad['rig'], []).append(ad)
        for rig_name, arr in assignments_by_rig.items():
            arr.sort(key=lambda x: x['well_start_date'])
            for i, ad in enumerate(arr, 1):
                ad['calculated_sequence_order'] = i

        # Persist assignments
        # Build quick lookup
        rigs_by_name = {r.name: r for r in Rig.objects.filter(name__in=rigs_seen)}
        wells_by_name = {w.name: w for w in Well.objects.filter(name__in=wells_seen)}
        
        # Create a lookup for actual dates from the provided actuals list AND existing assignments
        actuals_lookup = {}
        
        # First, get existing actual dates from the original schedule
        for assignment in assignments_qs:
            if assignment.actual_start_date or assignment.actual_end_date:
                key = (assignment.well.name, assignment.rig.name)
                actuals_lookup[key] = {
                    'actual_start_date': assignment.actual_start_date,
                    'actual_end_date': assignment.actual_end_date
                }
        
        # Then, override/add with provided actuals (new actuals take precedence)
        for actual in normalized_actuals:
            key = (actual['well'], actual['rig'])
            actuals_lookup[key] = {
                'actual_start_date': actual.get('actual_start_date'),
                'actual_end_date': actual.get('actual_end_date')
            }

        # Build a lookup for original planned dates from the parent schedule
        original_planned_lookup = {}
        for assignment in assignments_qs:
            key = (assignment.well.name, assignment.rig.name)
            # If the assignment already has original planned dates, use those
            # Otherwise, use the current well_start/end_date as the original
            original_planned_lookup[key] = {
                'original_planned_start': assignment.original_planned_start or assignment.well_start_date,
                'original_planned_end': assignment.original_planned_end or assignment.well_end_date
            }
        
        for ad in results.get('assignments', []):
            rig_obj = rigs_by_name.get(ad['rig'])
            well_obj = wells_by_name.get(ad['well'])
            if not rig_obj or not well_obj:
                continue
            
            # Check if this well-rig combination has actual dates
            actual_dates = actuals_lookup.get((ad['well'], ad['rig']), {})
            
            # Get original planned dates (preserve them from parent schedule)
            original_planned = original_planned_lookup.get((ad['well'], ad['rig']), {})
            
            Assignment.objects.create(
                schedule=new_schedule,
                rig=rig_obj,
                well=well_obj,
                well_start_date=ad['well_start_date'],
                well_end_date=ad['well_end_date'],
                original_planned_start=original_planned.get('original_planned_start'),
                original_planned_end=original_planned.get('original_planned_end'),
                actual_start_date=actual_dates.get('actual_start_date'),
                actual_end_date=actual_dates.get('actual_end_date'),
                rtd_check=ad.get('rtd_check', 'OK'),
                well_start_check=ad.get('well_start_check', 'OK'),
                well_end_check=ad.get('well_end_check', 'OK'),
                depth_check=ad.get('depth_check', 'OK'),
                hp_check=ad.get('hp_check', 'OK'),
                bop_check=ad.get('bop_check', 'OK'),
                tds_check=ad.get('tds_check', 'OK'),
                rig_type_check=ad.get('rig_type_check', 'OK'),
                drilling_cost=Decimal(str(ad.get('drilling_cost_inr', ad.get('drilling_cost', 0)))),
                ilm_cost=Decimal(str(ad.get('ilm_cost', 0))),
                sequence_order=ad.get('calculated_sequence_order', 1),
            )

        # Save unassigned wells info
        assigned_well_names = [a['well'] for a in results.get('assignments', [])]
        wells_df = pd.DataFrame([{
            'name': w.name,
            'depth': w.depth,
            'duration': w.duration,
            'rtd': w.rtd,
            'rig_capacity_required_hp': w.rig_capacity_required_hp,
            'bop_stack': w.bop_stack,
            'tds_requirement': w.tds_requirement,
            'priority': w.priority,
            'latitude': float(w.latitude),
            'longitude': float(w.longitude),
            'footprint': w.footprint,
        } for w in Well.objects.filter(name__in=wells_seen)])
        rigs_df = pd.DataFrame([{
            'name': r.name,
            'start_date': r.start_date,
            'end_date': r.end_date,
            'rig_capacity_hp': r.rig_capacity_hp,
            'drilling_capacity_m': r.drilling_capacity_m,
            'bop_stack': r.bop_stack,
            'tds_availability': r.tds_availability,
            'daily_cost_inr': float(r.daily_cost_inr),
            'ilm_cost_fixed': float(r.ilm_cost_fixed),
            'ilm_cost_per_km': float(r.ilm_cost_per_km),
            'ilm_cost_cluster': float(r.ilm_cost_cluster),
            'rig_type': r.rig_type,
        } for r in Rig.objects.filter(name__in=rigs_seen)])

        analyzer = WellRejectionAnalyzer(wells_df, rigs_df, timezone.now().date())
        for well_name in results.get('unassigned_wells', []):
            try:
                well_obj = wells_by_name.get(well_name) or Well.objects.get(name=well_name)
            except Well.DoesNotExist:
                continue
            UnassignedWell.objects.create(
                schedule=new_schedule,
                well=well_obj,
                reason=analyzer.analyze_well_rejection(well_name, assigned_well_names)
            )

        return Response(ScheduleSerializer(new_schedule).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def gantt_data(self, request, pk=None):
        """Get Gantt chart data for a schedule"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        assignments = Assignment.objects.filter(schedule=schedule).order_by('rig__name', 'sequence_order')
        
        tasks = []
        rigs = set()
        wells = set()
        
        for assignment in assignments:
            task = {
                'id': str(assignment.id),
                'name': assignment.well.name,
                'rig': assignment.rig.name,
                'start': assignment.well_start_date.isoformat(),
                'end': assignment.well_end_date.isoformat(),
                'duration': assignment.well.duration,
                'priority': assignment.well.priority,
                'well_type': assignment.well.well_type,
                'depth': assignment.well.depth,
                'asset': assignment.well.asset_id,
                'latitude': float(assignment.well.latitude),
                'longitude': float(assignment.well.longitude),
                'drilling_cost': float(assignment.drilling_cost),
                'ilm_cost': float(assignment.ilm_cost),
                'sequence_order': assignment.sequence_order,
                'original_planned_start': assignment.original_planned_start.isoformat() if assignment.original_planned_start else None,
                'original_planned_end': assignment.original_planned_end.isoformat() if assignment.original_planned_end else None,
                'actual_start_date': assignment.actual_start_date.isoformat() if assignment.actual_start_date else None,
                'actual_end_date': assignment.actual_end_date.isoformat() if assignment.actual_end_date else None,
                'has_actuals': bool(assignment.actual_start_date or assignment.actual_end_date),
                'is_well_deleted': assignment.well.is_deleted,
                'is_rig_deleted': assignment.rig.is_deleted,
                'checks': {
                    'rtd': assignment.rtd_check,
                    'hp': assignment.hp_check,
                    'depth': assignment.depth_check,
                    'bop': assignment.bop_check,
                    'tds': assignment.tds_check,
                    'rig_type': assignment.rig_type_check,
                }
            }
            tasks.append(task)
            rigs.add(assignment.rig.name)
            wells.add(assignment.well.name)
        
        # Calculate date range
        if tasks:
            start_dates = [task['start'] for task in tasks]
            end_dates = [task['end'] for task in tasks]
            date_range = {
                'start': min(start_dates),
                'end': max(end_dates)
            }
        else:
            date_range = {'start': None, 'end': None}
        
        data = {
            'schedule_id': str(schedule.id),
            'tasks': tasks,
            'rigs': sorted(list(rigs)),
            'wells': sorted(list(wells)),
            'date_range': date_range
        }
        
        serializer = GanttDataSerializer(data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reschedule_task(self, request, pk=None):
        """Reschedule a specific task and re-optimize the entire schedule"""
        schedule = self.get_object()
        



        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task_id = request.data.get('task_id')
        new_start_date = request.data.get('new_start_date')
        
        if not task_id or not new_start_date:
            return Response(
                {'error': 'task_id and new_start_date are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            from decimal import Decimal
            
            # Parse the new start date
            if isinstance(new_start_date, str):
                new_start_date = datetime.strptime(new_start_date, '%Y-%m-%d').date()
            
            # Get the assignment to reschedule
            assignment = Assignment.objects.get(id=task_id, schedule=schedule)
            
            # Create a properly named child schedule for the rescheduling
            new_schedule = create_child_schedule(schedule, branch_type="reschedule")
            
            # Get wells and rigs from the ORIGINAL SCOPE of the schedule
            from .models import ScheduleRig, ScheduleWell
            
            # Get the originally selected rigs and wells for this schedule
            schedule_rigs = ScheduleRig.objects.filter(schedule=schedule).select_related('rig')
            schedule_wells = ScheduleWell.objects.filter(schedule=schedule).select_related('well')
            
            if not schedule_rigs.exists() or not schedule_wells.exists():
                # Fallback to assignment-based detection if no scope tracking available
                original_assignments = Assignment.objects.filter(schedule=schedule)
                wells_in_schedule = [a.well for a in original_assignments]
                rigs_in_schedule = list(set([a.rig for a in original_assignments]))
            else:
                # Use the originally selected scope
                rigs_in_schedule = [sr.rig for sr in schedule_rigs]
                wells_in_schedule = [sw.well for sw in schedule_wells]
            
            # Collect ALL existing actual dates as FIXED constraints  
            original_assignments = Assignment.objects.filter(schedule=schedule)
            actuals_for_optimizer = []
            
            # Add ALL existing locked dates as fixed constraints
            for orig_assignment in original_assignments:
                if orig_assignment.actual_start_date or orig_assignment.actual_end_date:
                    actuals_for_optimizer.append({
                        'well': orig_assignment.well.name,
                        'rig': orig_assignment.rig.name,
                        'actual_start_date': orig_assignment.actual_start_date,
                        'actual_end_date': orig_assignment.actual_end_date,
                    })
            
            # Add the NEW reschedule constraint (this overrides any existing actual dates for this well-rig combo)
            reschedule_actual = {
                'well': assignment.well.name,
                'rig': assignment.rig.name,
                'actual_start_date': new_start_date,
                'actual_end_date': None  # Let optimizer calculate end date based on duration
            }
            
            # Remove any existing actual for this well-rig combination and add the new one
            actuals_for_optimizer = [a for a in actuals_for_optimizer 
                                   if not (a['well'] == assignment.well.name and a['rig'] == assignment.rig.name)]
            actuals_for_optimizer.append(reschedule_actual)
            
            logger.info(f"Rescheduling with {len(actuals_for_optimizer)} fixed actual constraints")
            
            # Convert rigs and wells to data format for optimizer
            rigs_data = []
            wells_data = []
            
            for rig in rigs_in_schedule:
                rigs_data.append({
                    'name': rig.name,
                    'rig_type': rig.rig_type,
                    'start_date': rig.start_date,
                    'end_date': rig.end_date,
                    'rig_capacity_hp': rig.rig_capacity_hp,
                    'daily_cost_inr': float(rig.daily_cost_inr),
                    'drilling_capacity_m': rig.drilling_capacity_m,
                    'mobilization_time_days': rig.mobilization_time_days,
                    'maintenance_schedule': rig.maintenance_schedule,
                    'crew_availability': rig.crew_availability,
                    'hpht_suitability': rig.hpht_suitability,
                    'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                    'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                    'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                    'bop_stack': rig.bop_stack,
                    'tds_availability': rig.tds_availability,
                })

            for well in wells_in_schedule:
                wells_data.append({
                    'name': well.name,
                    'sn': well.sn,
                    'asset_id': well.asset_id,
                    'well_type': well.well_type,
                    'well_profile': well.well_profile,
                    'depth': well.depth,
                    'rig_capacity_required_hp': well.rig_capacity_required_hp,
                    'drl_days': well.drl_days,
                    'pt_days': well.pt_days,
                    'duration': well.duration,
                    'latitude': float(well.latitude),
                    'longitude': float(well.longitude),
                    'rtd': well.rtd,
                    'bop_stack': well.bop_stack,
                    'tds_requirement': well.tds_requirement,
                    'footprint': well.footprint,
                    'preferred_rig': well.preferred_rig,
                    'expected_potential': well.expected_potential,
                    'priority': well.priority,
                })
            
            # Run optimization with locked actual dates as FIXED constraints
            from .optimization import DrillingScheduler
            scheduler = DrillingScheduler(rigs_data, wells_data)
            
            if actuals_for_optimizer:
                results = scheduler.solve_with_actuals(actuals_for_optimizer, time_limit_seconds=60)
            else:
                results = scheduler.solve(time_limit_seconds=60)
                
            success = results and results.get('status') in ['OPTIMAL', 'FEASIBLE']
            
            if success:
                # Update schedule with results
                new_schedule.status = 'COMPLETED'
                new_schedule.completed_at = timezone.now()
                new_schedule.total_drilling_cost = Decimal(str(results.get('total_drilling_cost', 0)))
                new_schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
                new_schedule.project_end_date = results.get('project_end_date')
                new_schedule.unassigned_wells_count = len(results.get('unassigned_wells', []))
                new_schedule.solver_status = results.get('solver_status')
                new_schedule.solve_time_seconds = results.get('solve_time_seconds')
                new_schedule.save()
                
                # Copy scope tracking from original schedule
                from .models import ScheduleRig, ScheduleWell
                
                # Create ScheduleRig entries for the new schedule
                schedule_rigs = []
                for rig in rigs_in_schedule:
                    schedule_rigs.append(ScheduleRig(schedule=new_schedule, rig=rig))
                ScheduleRig.objects.bulk_create(schedule_rigs, ignore_conflicts=True)
                
                # Create ScheduleWell entries for the new schedule  
                schedule_wells = []
                for well in wells_in_schedule:
                    schedule_wells.append(ScheduleWell(schedule=new_schedule, well=well))
                ScheduleWell.objects.bulk_create(schedule_wells, ignore_conflicts=True)
                
                # Create assignments with sequence order and preserve actual dates
                assignments_by_rig = {}
                for ad in results.get('assignments', []):
                    assignments_by_rig.setdefault(ad['rig'], []).append(ad)
                for rig_name, arr in assignments_by_rig.items():
                    arr.sort(key=lambda x: x['well_start_date'])
                    for i, ad in enumerate(arr, 1):
                        ad['calculated_sequence_order'] = i
                
                # Create assignment lookup for actual dates
                actuals_lookup = {}
                for actual in actuals_for_optimizer:
                    key = (actual['well'], actual['rig'])
                    actuals_lookup[key] = {
                        'actual_start_date': actual.get('actual_start_date'),
                        'actual_end_date': actual.get('actual_end_date')
                    }
                
                # Create assignments
                for ad in results.get('assignments', []):
                    try:
                        rig = next(r for r in rigs_in_schedule if r.name == ad['rig'])
                        well = next(w for w in wells_in_schedule if w.name == ad['well'])
                        
                        # Get actual dates for this well-rig combination
                        actual_dates = actuals_lookup.get((ad['well'], ad['rig']), {})
                        
                        Assignment.objects.create(
                            schedule=new_schedule,
                            rig=rig,
                            well=well,
                            well_start_date=ad['well_start_date'],
                            well_end_date=ad['well_end_date'],
                            actual_start_date=actual_dates.get('actual_start_date'),
                            actual_end_date=actual_dates.get('actual_end_date'),
                            rtd_check=ad.get('rtd_check', 'OK'),
                            well_start_check=ad.get('well_start_check', 'OK'),
                            well_end_check=ad.get('well_end_check', 'OK'),
                            depth_check=ad.get('depth_check', 'OK'),
                            hp_check=ad.get('hp_check', 'OK'),
                            bop_check=ad.get('bop_check', 'OK'),
                            tds_check=ad.get('tds_check', 'OK'),
                            rig_type_check=ad.get('rig_type_check', 'OK'),
                            drilling_cost=Decimal(str(ad.get('drilling_cost_inr', ad.get('drilling_cost', 0)))),
                            ilm_cost=Decimal(str(ad.get('ilm_cost', 0))),
                            sequence_order=ad.get('calculated_sequence_order', 1),
                        )
                    except (StopIteration, Exception) as e:
                        logger.warning(f"Could not create assignment for {ad.get('well')} -> {ad.get('rig')}: {e}")
                        continue
                
                # Return the new schedule data
                return Response({
                    'success': True,
                    'new_schedule_id': str(new_schedule.id),
                    'message': 'Schedule rescheduled successfully with locked dates preserved'
                })
            else:
                new_schedule.status = 'FAILED'
                new_schedule.save()
                return Response(
                    {'error': 'Rescheduling optimization failed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Assignment.DoesNotExist:
            return Response(
                {'error': 'Assignment not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Rescheduling error: {e}")
            return Response(
                {'error': f'Rescheduling failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def debug_assignments(self, request, pk=None):
        """Debug endpoint to see assignment data structure"""
        schedule = self.get_object()
        assignments = Assignment.objects.filter(schedule=schedule)
        
        debug_data = []
        for assignment in assignments:
            debug_data.append({
                'id': str(assignment.id),
                'id_type': type(assignment.id).__name__,
                'well_name': assignment.well.name,
                'rig_name': assignment.rig.name,
                'sequence_order': assignment.sequence_order,
            })
        
        return Response({
            'schedule_id': str(schedule.id),
            'assignment_count': len(debug_data),
            'assignments': debug_data
        })
    
    @action(detail=True, methods=['post'])
    def delete_assignment(self, request, pk=None):
        """Delete an assignment and re-optimize with remaining wells/rigs from the original schedule scope"""
        schedule = self.get_object()
        
        logger.info(f"Delete assignment request received for schedule: {schedule.id}")
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request data: {request.data}")
        logger.info(f"Content type: {request.content_type}")
        
        if schedule.status != 'COMPLETED':
            logger.warning(f"Schedule {schedule.id} is not completed, status: {schedule.status}")
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task_id = request.data.get('task_id')
        logger.info(f"Task ID to delete: {task_id}")
        
        if not task_id:
            logger.error("No task_id provided in request")
            return Response(
                {'error': 'task_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            
            # Get the assignment to delete
            logger.info(f"Looking for assignment with ID: {task_id} in schedule: {schedule.id}")
            assignment_to_delete = Assignment.objects.get(id=task_id, schedule=schedule)
            logger.info(f"Found assignment to delete: {assignment_to_delete.well.name} on {assignment_to_delete.rig.name}")
            
            # Create a properly named child schedule for deleting a well
            new_schedule = create_child_schedule(
                schedule, 
                branch_type="delete_well", 
                custom_suffix=f"- {assignment_to_delete.well.name}"
            )
            logger.info(f"Created new schedule: {new_schedule.id}")
            
            # REVISED APPROACH: Use ORIGINAL SCHEDULE SCOPE (wells and rigs that were selected for the parent schedule)
            # Get the well that was deleted
            deleted_well = assignment_to_delete.well
            logger.info(f"Deleted well: {deleted_well.name}")
            
            # Get wells and rigs from the ORIGINAL SCOPE of the schedule
            from .models import ScheduleRig, ScheduleWell
            
            # Get the originally selected rigs and wells for this schedule
            schedule_rigs = ScheduleRig.objects.filter(schedule=schedule).select_related('rig')
            schedule_wells = ScheduleWell.objects.filter(schedule=schedule).select_related('well')
            
            if not schedule_rigs.exists() or not schedule_wells.exists():
                # Fallback to assignment-based detection if no scope tracking available
                logger.info("No ScheduleRig/ScheduleWell records found, using assignment-based detection")
                original_assignments = Assignment.objects.filter(schedule=schedule)
                wells_in_schedule = [a.well for a in original_assignments]
                rigs_in_schedule = list(set([a.rig for a in original_assignments]))
                
                # Also include unassigned wells from the original schedule
                from .models import UnassignedWell
                unassigned_wells = UnassignedWell.objects.filter(schedule=schedule).select_related('well')
                unassigned_well_objects = [uw.well for uw in unassigned_wells]
                
                # Combine assigned and unassigned wells
                all_wells_in_scope = set(wells_in_schedule + unassigned_well_objects)
            else:
                # Use the originally selected scope
                logger.info(f"Found {schedule_rigs.count()} rigs and {schedule_wells.count()} wells in original scope")
                rigs_in_schedule = [sr.rig for sr in schedule_rigs]
                all_wells_in_scope = set([sw.well for sw in schedule_wells])
            
            # Remove the deleted well from the scope
            all_wells_in_scope.discard(deleted_well)
            
            # Convert to queryset for optimization
            well_ids = [w.id for w in all_wells_in_scope]
            rig_ids = [r.id for r in rigs_in_schedule]
            
            all_wells_except_deleted = Well.objects.filter(id__in=well_ids)
            all_rigs = Rig.objects.filter(id__in=rig_ids)
            
            logger.info(f"Re-optimizing with {all_wells_except_deleted.count()} wells (from original scope minus deleted) and {all_rigs.count()} rigs (from original scope)")
            
            # Let the optimizer determine the base date automatically based on rig availability
            # This matches the behavior of normal optimization runs
            base_start_date = None  # Let DrillingScheduler use earliest rig start date
            logger.info(f"Using automatic base start date (will be determined by optimizer based on rig availability)")
            
            # Run optimization with fresh database state (all wells except deleted + all rigs)
            success = self._run_optimization_with_constraints(
                new_schedule, all_rigs, all_wells_except_deleted, {}, base_start_date, original_schedule=schedule
            )
            
            if success:
                # Refresh the schedule data and return it for frontend update
                new_schedule.refresh_from_db()
                serializer = ScheduleSerializer(new_schedule)
                return Response({
                    'success': True,
                    'schedule': serializer.data,
                    'new_schedule_id': str(new_schedule.id),
                    'deleted_well_id': str(deleted_well.id),
                    'deleted_well_name': deleted_well.name,
                    'message': f'Assignment for {assignment_to_delete.well.name} deleted and remaining wells from original scope re-optimized'
                })
            else:
                new_schedule.status = 'FAILED'
                new_schedule.save()
                return Response(
                    {'error': 'Re-optimization after deletion failed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Assignment.DoesNotExist:
            logger.error(f"Assignment with ID {task_id} not found in schedule {schedule.id}")
            return Response(
                {'error': 'Assignment not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError as e:
            logger.error(f"Invalid task_id format: {task_id}, error: {e}")
            return Response(
                {'error': f'Invalid task ID format: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Deletion error: {e}")
            logger.error(f"Error type: {type(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Deletion failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def add_assignment(self, request, pk=None):
        """Add an unscheduled well to the current schedule and re-optimize"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        well_id = request.data.get('well_id')
        
        if not well_id:
            return Response(
                {'error': 'well_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from datetime import datetime
            
            # Get the well to add
            well_to_add = Well.objects.get(id=well_id)
            
            # Check if the well is already scheduled in this schedule
            existing_assignment = Assignment.objects.filter(schedule=schedule, well=well_to_add).first()
            if existing_assignment:
                return Response(
                    {'error': f'Well {well_to_add.name} is already scheduled in this schedule'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create a properly named child schedule for adding a well
            new_schedule = create_child_schedule(
                schedule, 
                branch_type="add_well", 
                custom_suffix=f"+ {well_to_add.name}"
            )
            
            # Copy scope tracking from original schedule and add the new well
            from .models import ScheduleRig, ScheduleWell
            
            # Copy original schedule's rigs
            original_schedule_rigs = ScheduleRig.objects.filter(schedule=schedule)
            for orig_rig in original_schedule_rigs:
                ScheduleRig.objects.create(schedule=new_schedule, rig=orig_rig.rig)
                
            # Copy original schedule's wells and add the new one
            original_schedule_wells = ScheduleWell.objects.filter(schedule=schedule)
            for orig_well in original_schedule_wells:
                ScheduleWell.objects.create(schedule=new_schedule, well=orig_well.well)
            # Add the new well
            ScheduleWell.objects.create(schedule=new_schedule, well=well_to_add)
            
            # Get all wells from current schedule plus the new well
            original_assignments = Assignment.objects.filter(schedule=schedule)
            
            # Extract locked wells (those with actual dates) to preserve their positions
            locked_wells_data = []
            for assignment in original_assignments:
                if assignment.actual_start_date or assignment.actual_end_date:
                    locked_wells_data.append({
                        'well': assignment.well.name,
                        'rig': assignment.rig.name,
                        'actual_start_date': assignment.actual_start_date.isoformat() if assignment.actual_start_date else None,
                        'actual_end_date': assignment.actual_end_date.isoformat() if assignment.actual_end_date else None,
                    })
            
            logger.info(f"Found {len(locked_wells_data)} locked wells with actual dates that will be preserved")
            for locked in locked_wells_data:
                logger.info(f"  - Locked: {locked['well']} on {locked['rig']} ({locked['actual_start_date']} to {locked['actual_end_date']})")
            
            # Get well IDs from current assignments plus the new well
            well_ids_current = [a.well.id for a in original_assignments]
            well_ids_current.append(well_to_add.id)
            rig_ids_in_schedule = list(set([a.rig.id for a in original_assignments]))
            
            # Create QuerySets for all wells (current + new) and rigs
            wells_in_schedule = Well.objects.filter(id__in=well_ids_current)
            rigs_in_schedule = Rig.objects.filter(id__in=rig_ids_in_schedule)
            
            logger.info(f"Re-optimizing with {wells_in_schedule.count()} wells (added {well_to_add.name}) and {rigs_in_schedule.count()} rigs")
            
            # Use the earliest start date from the original schedule
            if original_assignments.exists():
                earliest_start_date = min(a.well_start_date for a in original_assignments)
                base_start_date = earliest_start_date
                logger.info(f"Using base start date from original schedule: {base_start_date}")
            else:
                # Let the optimizer determine base date from rig availability
                base_start_date = None
                logger.info(f"Using automatic base start date determination")
            
            # Run optimization with the added well and locked constraints
            success = self._run_optimization_with_constraints(
                new_schedule, rigs_in_schedule, wells_in_schedule, locked_wells_data, base_start_date, original_schedule=schedule
            )
            
            if success:
                return Response({
                    'success': True,
                    'new_schedule_id': str(new_schedule.id),
                    'message': f'Well {well_to_add.name} added and schedule re-optimized'
                })
            else:
                new_schedule.status = 'FAILED'
                new_schedule.save()
                return Response(
                    {'error': 'Re-optimization after adding well failed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Well.DoesNotExist:
            return Response(
                {'error': 'Well not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Add assignment error: {e}")
            return Response(
                {'error': f'Adding well failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _run_optimization_with_constraints(self, schedule, rigs, wells, fixed_assignments=None, base_start_date=None, original_schedule=None):
        """Run optimization with optional fixed assignments and custom base start date"""
        try:
            from .optimization import DrillingScheduler
            from datetime import date, datetime
            from decimal import Decimal
            
            # Use provided base_start_date or let optimizer determine it automatically
            # The DrillingScheduler will automatically use earliest rig start date when None is passed
            
            logger.info(f"Starting optimization with {wells.count() if hasattr(wells, 'count') else len(wells)} wells and {rigs.count() if hasattr(rigs, 'count') else len(rigs)} rigs")
            if base_start_date:
                logger.info(f"Using provided base start date: {base_start_date}")
            else:
                logger.info(f"Using automatic base start date determination from rig availability")
            
            # Convert QuerySets to dictionaries for the optimizer
            rigs_data = list(rigs.values()) if hasattr(rigs, 'values') else rigs
            wells_data = list(wells.values()) if hasattr(wells, 'values') else wells
            
            # DEBUGGING: Log exact wells being optimized
            well_names = [w['name'] for w in wells_data]
            logger.info(f"Wells being optimized: {sorted(well_names)}")
            
            # DEBUGGING: Check for any wells that might have problematic constraints
            problematic_wells = []
            for well in wells_data:
                if not well.get('rtd') or not well.get('duration') or well.get('duration', 0) <= 0:
                    problematic_wells.append(f"{well['name']}: rtd={well.get('rtd')}, duration={well.get('duration')}")
            
            if problematic_wells:
                logger.warning(f"Wells with potential constraint issues: {problematic_wells}")
            
            # Initialize scheduler with the correct base start date
            scheduler = DrillingScheduler(rigs_data, wells_data, base_start_date=base_start_date)
            
            try:
                scheduler.preprocess_data()
            except Exception as e:
                logger.error(f"Data preprocessing error: {e}")
                return False
                
            try:
                scheduler.setup_variables()
                scheduler.add_constraints()
            except Exception as e:
                logger.error(f"Constraint setup error: {e}")
                return False
            
            try:
                scheduler.set_objective()
            except Exception as e:
                logger.error(f"Objective setting error: {e}")
                return False
            
            # If we have locked wells (fixed_assignments), use solve_with_actuals to respect them
            if fixed_assignments and len(fixed_assignments) > 0:
                logger.info(f"Running optimization WITH {len(fixed_assignments)} locked wells as fixed constraints")
                results = scheduler.solve_with_actuals(fixed_assignments, time_limit_seconds=60)
            else:
                # Use standard solver for unconstrained optimization
                logger.info("Running standard optimization with 60s time limit")
                results = scheduler.solve(time_limit_seconds=60)
            
            if results and results.get('status') in ['OPTIMAL', 'FEASIBLE']:
                logger.info(f"Optimization successful: {len(results['assignments'])} assignments created")
                
                # Clear any existing assignments and unscheduled wells for this schedule to prevent duplicates
                Assignment.objects.filter(schedule=schedule).delete()
                UnassignedWell.objects.filter(schedule=schedule).delete()
                
                # Validate assignments for overlaps before creating them
                assignment_data_list = results['assignments']
                rig_schedules = {}  # Track each rig's schedule to prevent overlaps
                
                # Sort assignments by rig and start date to check for overlaps
                assignment_data_list.sort(key=lambda x: (x['rig'], x['well_start_date']))
                
                valid_assignments = []
                for assignment_data in assignment_data_list:
                    rig_name = assignment_data['rig']
                    start_date = assignment_data['well_start_date']
                    end_date = assignment_data['well_end_date']
                    
                    # Check for overlaps with existing assignments for this rig
                    has_overlap = False
                    if rig_name in rig_schedules:
                        for existing_start, existing_end in rig_schedules[rig_name]:
                            # Check if dates overlap
                            if not (end_date <= existing_start or start_date >= existing_end):
                                logger.warning(f"Overlap detected for rig {rig_name}: {assignment_data['well']} ({start_date}-{end_date}) overlaps with existing assignment ({existing_start}-{existing_end})")
                                has_overlap = True
                                break
                    
                    if not has_overlap:
                        valid_assignments.append(assignment_data)
                        # Track this rig's schedule
                        if rig_name not in rig_schedules:
                            rig_schedules[rig_name] = []
                        rig_schedules[rig_name].append((start_date, end_date))
                
                logger.info(f"Creating {len(valid_assignments)} valid assignments (removed {len(assignment_data_list) - len(valid_assignments)} overlapping assignments)")
                
                # Calculate proper sequence order for valid assignments
                assignments_by_rig = {}
                for assignment_data in valid_assignments:
                    rig_name = assignment_data['rig']
                    if rig_name not in assignments_by_rig:
                        assignments_by_rig[rig_name] = []
                    assignments_by_rig[rig_name].append(assignment_data)
                
                # Sort each rig's assignments by start date and assign sequence numbers
                for rig_name, rig_assignments in assignments_by_rig.items():
                    # Sort by start date
                    rig_assignments.sort(key=lambda x: x['well_start_date'])
                    # Assign sequence numbers starting from 1
                    for i, assignment_data in enumerate(rig_assignments, 1):
                        assignment_data['calculated_sequence_order'] = i
                
                # Prepare actual dates and original planned dates lookup if original schedule is provided
                actual_dates_lookup = {}
                original_planned_lookup = {}
                if original_schedule:
                    original_assignments = Assignment.objects.filter(schedule=original_schedule)
                    for orig_assignment in original_assignments:
                        key = (orig_assignment.well.name, orig_assignment.rig.name)
                        
                        # Preserve actual dates if they exist
                        if orig_assignment.actual_start_date or orig_assignment.actual_end_date:
                            actual_dates_lookup[key] = {
                                'actual_start_date': orig_assignment.actual_start_date,
                                'actual_end_date': orig_assignment.actual_end_date
                            }
                        
                        # Preserve original planned dates if they exist, otherwise use current planned dates
                        original_planned_lookup[key] = {
                            'original_planned_start': orig_assignment.original_planned_start or orig_assignment.well_start_date,
                            'original_planned_end': orig_assignment.original_planned_end or orig_assignment.well_end_date
                        }
                
                # Create assignments for the new schedule with proper sequence order
                for assignment_data in valid_assignments:
                    well = Well.objects.get(name=assignment_data['well'])
                    rig = Rig.objects.get(name=assignment_data['rig'])
                    
                    # Get actual dates if they exist for this well-rig combination
                    actual_dates = actual_dates_lookup.get((well.name, rig.name), {})
                    
                    # Get original planned dates if they exist
                    original_planned = original_planned_lookup.get((well.name, rig.name), {})
                    
                    Assignment.objects.create(
                        schedule=schedule,
                        rig=rig,
                        well=well,
                        well_start_date=assignment_data['well_start_date'],
                        well_end_date=assignment_data['well_end_date'],
                        original_planned_start=original_planned.get('original_planned_start'),
                        original_planned_end=original_planned.get('original_planned_end'),
                        actual_start_date=actual_dates.get('actual_start_date'),
                        actual_end_date=actual_dates.get('actual_end_date'),
                        rtd_check=assignment_data.get('rtd_check', 'OK'),
                        well_start_check=assignment_data.get('well_start_check', 'OK'),
                        well_end_check=assignment_data.get('well_end_check', 'OK'),
                        depth_check=assignment_data.get('depth_check', 'OK'),
                        hp_check=assignment_data.get('hp_check', 'OK'),
                        bop_check=assignment_data.get('bop_check', 'OK'),
                        tds_check=assignment_data.get('tds_check', 'OK'),
                        rig_type_check=assignment_data.get('rig_type_check', 'OK'),
                        drilling_cost=Decimal(str(assignment_data.get('drilling_cost_inr', assignment_data.get('drilling_cost', 0)))),
                        ilm_cost=Decimal(str(assignment_data.get('ilm_cost', 0))),
                        sequence_order=assignment_data.get('calculated_sequence_order', 1)
                    )
                
                # Create unassigned wells with detailed rejection analysis
                assigned_well_names = [assignment['well'] for assignment in valid_assignments]
                
                # Get all wells and rigs for analysis
                all_wells = Well.objects.all()
                all_rigs = Rig.objects.all()
                
                # Prepare data for analyzer
                wells_data = []
                rigs_data = []
                
                for well in all_wells:
                    wells_data.append({
                        'name': well.name,
                        'depth': well.depth,
                        'duration': well.duration,
                        'rtd': well.rtd,
                        'rig_capacity_required_hp': well.rig_capacity_required_hp,
                        'bop_stack': well.bop_stack,
                        'tds_requirement': well.tds_requirement,
                        'priority': well.priority,
                        'latitude': float(well.latitude) if well.latitude else 0.0,
                        'longitude': float(well.longitude) if well.longitude else 0.0,
                        'footprint': well.footprint,
                    })
                
                for rig in all_rigs:
                    rigs_data.append({
                        'name': rig.name,
                        'start_date': rig.start_date,
                        'end_date': rig.end_date,
                        'rig_capacity_hp': rig.rig_capacity_hp,
                        'drilling_capacity_m': rig.drilling_capacity_m,
                        'bop_stack': rig.bop_stack,
                        'tds_availability': rig.tds_availability,
                        'daily_cost_inr': float(rig.daily_cost_inr),
                        'ilm_cost_fixed': float(rig.ilm_cost_fixed),
                        'ilm_cost_per_km': float(rig.ilm_cost_per_km),
                        'ilm_cost_cluster': float(rig.ilm_cost_cluster),
                        'rig_type': rig.rig_type,
                    })
                
                # Convert to DataFrames for analysis
                wells_df = pd.DataFrame(wells_data)
                rigs_df = pd.DataFrame(rigs_data)
                
                # Initialize analyzer
                analyzer = WellRejectionAnalyzer(wells_df, rigs_df, timezone.now().date())
                
                for unassigned_well_name in results.get('unassigned_wells', []):
                    # Handle both simple string format and dict format
                    if isinstance(unassigned_well_name, dict):
                        well_name = unassigned_well_name.get('well', unassigned_well_name.get('name'))
                        # If reason is already provided in dict, use it, otherwise analyze
                        existing_reason = unassigned_well_name.get('reason', '')
                        if existing_reason and existing_reason != 'Not assigned':
                            reason = existing_reason
                        else:
                            reason = analyzer.analyze_well_rejection(well_name, assigned_well_names)
                    else:
                        well_name = str(unassigned_well_name)
                        reason = analyzer.analyze_well_rejection(well_name, assigned_well_names)
                    
                    well = Well.objects.get(name=well_name)
                    UnassignedWell.objects.create(
                        schedule=schedule,
                        well=well,
                        reason=reason
                    )
                
                # Update schedule
                schedule.status = 'COMPLETED'
                schedule.total_drilling_cost = Decimal(str(results['total_drilling_cost']))
                schedule.total_ilm_cost = Decimal(str(results.get('total_ilm_cost', 0)))
                schedule.project_end_date = results.get('project_end_date')
                schedule.unassigned_wells_count = results.get('unassigned_wells_count', 0)
                schedule.solver_status = results.get('solver_status', 'UNKNOWN')
                schedule.solve_time_seconds = results.get('solve_time_seconds', 0)
                schedule.completed_at = datetime.now()
                schedule.save()
                
                return True
            else:
                return False
                
        except Exception as e:
            logger.error(f"Optimization error: {e}")
            return False
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Get schedule statistics"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        assignments = Assignment.objects.filter(schedule=schedule)
        unassigned = UnassignedWell.objects.filter(schedule=schedule)
        
        # Calculate statistics
        total_drilling_cost_cr = float(schedule.total_drilling_cost or 0) / 10000000
        total_ilm_cost_cr = float(schedule.total_ilm_cost or 0) / 10000000
        total_cost_cr = total_drilling_cost_cr + total_ilm_cost_cr
        
        # Project duration
        if assignments.exists():
            project_start = assignments.aggregate(min_start=models.Min('well_start_date'))['min_start']
            project_end = assignments.aggregate(max_end=models.Max('well_end_date'))['max_end']
            project_duration_days = (project_end - project_start).days + 1 if project_start and project_end else 0
        else:
            project_duration_days = 0
        
        # Priority breakdown
        priority_counts = {}
        for priority in ['HIGH', 'MEDIUM', 'LOW']:
            assigned = assignments.filter(well__priority=priority).count()
            unassigned_count = unassigned.filter(well__priority=priority).count()
            priority_counts[priority] = {
                'assigned': assigned,
                'unassigned': unassigned_count,
                'total': assigned + unassigned_count
            }
        
        # Rig utilization
        rig_utilizations = []
        for rig in Rig.objects.filter(assignments__schedule=schedule).distinct():
            rig_assignments = assignments.filter(rig=rig)
            total_assigned_days = sum(a.well.duration for a in rig_assignments)
            total_available_days = rig.duration_days
            utilization_pct = (total_assigned_days / total_available_days * 100) if total_available_days > 0 else 0
            
            rig_utilizations.append({
                'rig_name': rig.name,
                'total_available_days': total_available_days,
                'total_assigned_days': total_assigned_days,
                'utilization_percentage': round(utilization_pct, 2),
                'wells_assigned': rig_assignments.count(),
                'idle_days': max(0, total_available_days - total_assigned_days),
                'drilling_cost': float(sum(a.drilling_cost for a in rig_assignments)),
                'ilm_cost': float(sum(a.ilm_cost for a in rig_assignments))
            })
        
        stats_data = {
            'total_assignments': assignments.count(),
            'total_unassigned': unassigned.count(),
            'total_drilling_cost_cr': total_drilling_cost_cr,
            'total_ilm_cost_cr': total_ilm_cost_cr,
            'total_cost_cr': total_cost_cr,
            'project_duration_days': project_duration_days,
            'rig_utilization': rig_utilizations,
            'priority_breakdown': priority_counts
        }
        
        serializer = ScheduleStatsSerializer(stats_data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def unscheduled_resources(self, request, pk=None):
        """Get wells and rigs that were selected for the schedule but not assigned/used"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get wells that were SELECTED for this schedule (input)
        selected_well_ids = schedule.selected_wells.values_list('well_id', flat=True)
        
        # Get wells that were ASSIGNED in this schedule (output)
        assigned_well_ids = Assignment.objects.filter(schedule=schedule).values_list('well_id', flat=True)
        
        # Get unassigned wells WITH rejection reasons
        unassigned_wells_qs = schedule.unassigned_wells.select_related('well').order_by('well__name')
        
        # Get rigs that were SELECTED for this schedule (input)
        selected_rig_ids = schedule.selected_rigs.values_list('rig_id', flat=True)
        
        # Get rigs that were USED in this schedule (output)
        used_rig_ids = Assignment.objects.filter(schedule=schedule).values_list('rig_id', flat=True).distinct()
        
        # Get unused rigs (selected but not used)
        unused_rig_ids = set(selected_rig_ids) - set(used_rig_ids)
        unused_rigs = Rig.objects.filter(id__in=unused_rig_ids).order_by('name')
        
        # Serialize unassigned wells with rejection reasons
        unscheduled_wells_data = []
        for unassigned in unassigned_wells_qs:
            well = unassigned.well
            unscheduled_wells_data.append({
                'id': str(well.id),
                'name': well.name,
                'asset_id': well.asset_id,
                'priority': well.priority,
                'depth': well.depth,
                'duration': well.duration,
                'well_type': well.well_type,
                'rtd': well.rtd.strftime('%Y-%m-%d'),
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'reason': unassigned.reason  # Include rejection reason
            })
        
        # Serialize unused rigs with reason
        unused_rigs_data = []
        for rig in unused_rigs:
            unused_rigs_data.append({
                'id': str(rig.id),
                'name': rig.name,
                'rig_type': rig.rig_type,
                'capacity_hp': rig.rig_capacity_hp,
                'daily_cost': float(rig.daily_cost_inr),
                'drilling_capacity': rig.drilling_capacity_m,
                'start_date': rig.start_date.strftime('%Y-%m-%d'),
                'end_date': rig.end_date.strftime('%Y-%m-%d'),
                'duration_days': rig.duration_days,
                'reason': 'Not selected by optimization algorithm (higher cost or lower efficiency)'  # Default reason for rigs
            })
        
        return Response({
            'unscheduled_wells': unscheduled_wells_data,
            'unused_rigs': unused_rigs_data,
            'summary': {
                'total_wells_selected': len(selected_well_ids),
                'wells_assigned': len(assigned_well_ids),
                'unscheduled_wells_count': len(unscheduled_wells_data),
                'total_rigs_selected': len(selected_rig_ids),
                'rigs_used': len(set(used_rig_ids)),
                'unused_rigs_count': len(unused_rigs_data)
            }
        })
    
    @action(detail=True, methods=['get'])
    def rig_statistics(self, request, pk=None):
        """Get detailed rig statistics including rig-months, well type split, and meterage"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        assignments = Assignment.objects.filter(schedule=schedule).select_related('rig', 'well')
        
        # Group assignments by rig
        rig_stats = {}
        for assignment in assignments:
            rig_name = assignment.rig.name
            if rig_name not in rig_stats:
                rig_stats[rig_name] = {
                    'rig_name': rig_name,
                    'rig_type': assignment.rig.rig_type,
                    'total_days': 0,
                    'dev_wells': 0,
                    'exp_wells': 0,
                    'dev_days': 0,
                    'exp_days': 0,
                    'dev_meterage': 0,
                    'exp_meterage': 0,
                }
            
            # Calculate operating days for this assignment
            if assignment.well_start_date and assignment.well_end_date:
                days = (assignment.well_end_date - assignment.well_start_date).days + 1
            else:
                days = assignment.well.duration
            
            rig_stats[rig_name]['total_days'] += days
            
            # Classify by well type
            is_exploration = assignment.well.well_type.upper() in ['EXP', 'EXPLORATION']
            
            if is_exploration:
                rig_stats[rig_name]['exp_wells'] += 1
                rig_stats[rig_name]['exp_days'] += days
                rig_stats[rig_name]['exp_meterage'] += assignment.well.depth
            else:
                rig_stats[rig_name]['dev_wells'] += 1
                rig_stats[rig_name]['dev_days'] += days
                rig_stats[rig_name]['dev_meterage'] += assignment.well.depth
        
        # Calculate rig-months and format data
        rig_statistics = []
        for rig_name, stats in rig_stats.items():
            # Calculate rig-months (total_days / 30.41)
            total_rig_months = stats['total_days'] / 30.41
            dev_rig_months = stats['dev_days'] / 30.41
            exp_rig_months = stats['exp_days'] / 30.41
            
            rig_statistics.append({
                'rig_name': stats['rig_name'],
                'rig_type': stats['rig_type'],
                'total_rig_months': round(total_rig_months, 2),
                'total_days': stats['total_days'],
                'development': {
                    'well_count': stats['dev_wells'],
                    'rig_months': round(dev_rig_months, 2),
                    'days': stats['dev_days'],
                    'meterage': stats['dev_meterage']
                },
                'exploration': {
                    'well_count': stats['exp_wells'],
                    'rig_months': round(exp_rig_months, 2),
                    'days': stats['exp_days'],
                    'meterage': stats['exp_meterage']
                },
                'total_meterage': stats['dev_meterage'] + stats['exp_meterage']
            })
        
        # Sort by rig name
        rig_statistics.sort(key=lambda x: x['rig_name'])
        
        return Response({
            'rig_statistics': rig_statistics,
            'summary': {
                'total_rigs': len(rig_statistics),
                'total_rig_months': round(sum(r['total_rig_months'] for r in rig_statistics), 2),
                'total_meterage': sum(r['total_meterage'] for r in rig_statistics)
            }
        })
    
    @action(detail=True, methods=['get'])
    def allocated_resources(self, request, pk=None):
        """Get wells and rigs allocated to the selected schedule"""
        schedule = self.get_object()
        
        if schedule.status != 'COMPLETED':
            return Response(
                {'error': 'Schedule is not completed'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get wells assigned to this schedule
        assigned_wells = Assignment.objects.filter(schedule=schedule).select_related('well').order_by('well__name')
        
        # Get rigs used in this schedule
        used_rigs = Assignment.objects.filter(schedule=schedule).select_related('rig').values_list('rig', flat=True).distinct()
        allocated_rigs = Rig.objects.filter(id__in=used_rigs).order_by('name')
        
        # Serialize allocated wells
        allocated_wells_data = []
        for assignment in assigned_wells:
            well = assignment.well
            allocated_wells_data.append({
                'id': str(well.id),
                'name': well.name,
                'asset_id': well.asset_id,
                'priority': well.priority,
                'depth': well.depth,
                'duration': well.duration,
                'well_type': well.well_type,
                'rtd': well.rtd.strftime('%Y-%m-%d'),
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'assigned_start_date': assignment.well_start_date.strftime('%Y-%m-%d'),
                'assigned_end_date': assignment.well_end_date.strftime('%Y-%m-%d'),
                'assigned_rig': assignment.rig.name
            })
        
        # Serialize allocated rigs
        allocated_rigs_data = []
        for rig in allocated_rigs:
            allocated_rigs_data.append({
                'id': str(rig.id),
                'name': rig.name,
                'rig_type': rig.rig_type,
                'capacity_hp': rig.rig_capacity_hp,
                'daily_cost': float(rig.daily_cost_inr),
                'drilling_capacity': rig.drilling_capacity_m,
                'start_date': rig.start_date.strftime('%Y-%m-%d'),
                'end_date': rig.end_date.strftime('%Y-%m-%d'),
                'duration_days': rig.duration_days
            })
        
        return Response({
            'allocated_wells': allocated_wells_data,
            'allocated_rigs': allocated_rigs_data,
            'summary': {
                'total_wells_in_schedule': len(allocated_wells_data),
                'total_rigs_in_schedule': len(allocated_rigs_data),
                'schedule_name': schedule.name,
                'schedule_status': schedule.status
            }
        })
    
    @action(detail=True, methods=['post'])
    def update_assignment(self, request, pk=None):
        """Update assignment dates (for drag-and-drop functionality)"""
        schedule = self.get_object()
        serializer = AssignmentUpdateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        
        try:
            assignment = Assignment.objects.get(
                id=validated_data['assignment_id'],
                schedule=schedule
            )
            
            # Update start date and calculate end date
            new_start_date = validated_data['new_start_date']
            new_end_date = new_start_date + timedelta(days=assignment.well.duration - 1)
            
            assignment.well_start_date = new_start_date
            assignment.well_end_date = new_end_date
            
            # Update rig if provided
            if 'new_rig_id' in validated_data:
                new_rig = get_object_or_404(Rig, id=validated_data['new_rig_id'])
                assignment.rig = new_rig
            
            # Re-validate constraints
            checks = self._validate_assignment(assignment)
            assignment.rtd_check = checks.get('rtd_check', 'OK')
            assignment.well_start_check = checks.get('well_start_check', 'OK')
            assignment.well_end_check = checks.get('well_end_check', 'OK')
            assignment.depth_check = checks.get('depth_check', 'OK')
            assignment.hp_check = checks.get('hp_check', 'OK')
            assignment.bop_check = checks.get('bop_check', 'OK')
            assignment.tds_check = checks.get('tds_check', 'OK')
            assignment.rig_type_check = checks.get('rig_type_check', 'OK')
            
            assignment.save()
            
            return Response(AssignmentSerializer(assignment).data)
            
        except Assignment.DoesNotExist:
            return Response(
                {'error': 'Assignment not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Update failed: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export schedules data as CSV"""
        try:
            queryset = self.get_queryset()
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="schedules_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = [
                'ID', 'Name', 'Status', 'Total Drilling Cost', 'Total ILM Cost',
                'Project End Date', 'Unassigned Wells Count', 'Solver Status',
                'Solve Time (seconds)', 'Created At', 'Updated At', 'Completed At'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for schedule in queryset:
                writer.writerow([
                    str(schedule.id),
                    schedule.name,
                    schedule.status,
                    schedule.total_drilling_cost or 0,
                    schedule.total_ilm_cost or 0,
                    schedule.project_end_date.strftime('%Y-%m-%d') if schedule.project_end_date else '',
                    schedule.unassigned_wells_count or 0,
                    schedule.solver_status or '',
                    schedule.solve_time_seconds or 0,
                    schedule.created_at.strftime('%Y-%m-%d %H:%M:%S') if schedule.created_at else '',
                    schedule.updated_at.strftime('%Y-%m-%d %H:%M:%S') if schedule.updated_at else '',
                    schedule.completed_at.strftime('%Y-%m-%d %H:%M:%S') if schedule.completed_at else ''
                ])
            
            return response
            
        except Exception as e:
            return JsonResponse(
                {'error': f'Failed to export CSV: {str(e)}'}, 
                status=500
            )
    
    def _validate_assignment(self, assignment):
        """Validate an assignment against constraints"""
        well = assignment.well
        rig = assignment.rig
        well_start_date = assignment.well_start_date
        well_end_date = assignment.well_end_date
        
        checks = {}
        
        # RTD check
        checks['rtd_check'] = 'OK' if well_start_date >= well.rtd else 'NOK'
        
        # Well start date check
        checks['well_start_check'] = 'OK' if (well_start_date - well.rtd).days >= 0 else 'NOK'
        
        # Well end date check
        checks['well_end_check'] = 'OK' if well_end_date <= rig.end_date else 'NOK'
        
        # Depth check
        checks['depth_check'] = 'OK' if rig.drilling_capacity_m >= well.depth else 'NOK'
        
        # HP check
        checks['hp_check'] = 'OK' if rig.rig_capacity_hp >= well.rig_capacity_required_hp else 'NOK'
        
        # BOP check
        checks['bop_check'] = 'OK' if rig.bop_stack >= well.bop_stack else 'NOK'
        
        # TDS check
        if well.tds_requirement == 'Y':
            checks['tds_check'] = 'OK' if rig.tds_availability == 'Y' else 'NOK'
        else:
            checks['tds_check'] = 'OK'
        
        # Rig type check
        checks['rig_type_check'] = 'OK' if well.footprint == rig.rig_type else 'NOK'
        
        return checks


class AssignmentViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing assignments"""
    queryset = Assignment.objects.all()
    serializer_class = AssignmentSerializer
    permission_classes = [permissions.AllowAny]
    
    def get_queryset(self):
        queryset = Assignment.objects.all()
        
        # Apply location-based filtering for authenticated users
        if self.request.user and self.request.user.is_authenticated:
            user_location = get_user_location(self.request.user)
            if user_location:
                # Filter assignments by schedule location or well location
                queryset = queryset.filter(
                    models.Q(schedule__location=user_location) |
                    models.Q(well__location=user_location)
                )
        
        # Check if request has query_params (DRF Request) or GET (Django Request)
        if hasattr(self.request, 'query_params'):
            schedule_id = self.request.query_params.get('schedule_id', None)
            rig_id = self.request.query_params.get('rig_id', None)
        else:
            schedule_id = self.request.GET.get('schedule_id', None)
            rig_id = self.request.GET.get('rig_id', None)
        
        if schedule_id:
            queryset = queryset.filter(schedule_id=schedule_id)
        if rig_id:
            queryset = queryset.filter(rig_id=rig_id)
        
        return queryset.order_by('rig__name', 'sequence_order')


@csrf_exempt
def export_schedule_csv(request, schedule_id):
    """Export schedule to CSV format"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        assignments = Assignment.objects.filter(schedule=schedule).order_by('rig__name', 'sequence_order')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="schedule_{schedule.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Rig', 'Well', 'RTD', 'RTD Check', 'Well Start Date', 'Well Start Date Check (+ve)',
            'Well End Date', 'Well End Date Check (-ve)', 'Rig Start Date', 'Rig End Date',
            'Well Priority', 'Duration (days)', 'required_depth', 'rig depth capacity',
            'depth check', 'required_hp', 'rig hp capacity', 'HP check', 'required_bop',
            'rig bop capacity', 'BOP check', 'required_tds', 'rig tds availability',
            'TDS check', 'required_rig_type', 'rig type', 'Rig Type check',
            'Latitude', 'Longitude'
        ])
        
        # Write data
        for assignment in assignments:
            writer.writerow([
                assignment.rig.name,
                assignment.well.name,
                assignment.well.rtd,
                assignment.rtd_check,
                assignment.well_start_date,
                assignment.well_start_check,
                assignment.well_end_date,
                assignment.well_end_check,
                assignment.rig.start_date,
                assignment.rig.end_date,
                assignment.well.priority,
                assignment.well.duration,
                assignment.well.depth,
                assignment.rig.drilling_capacity_m,
                assignment.depth_check,
                assignment.well.rig_capacity_required_hp,
                assignment.rig.rig_capacity_hp,
                assignment.hp_check,
                assignment.well.bop_stack,
                assignment.rig.bop_stack,
                assignment.bop_check,
                assignment.well.tds_requirement,
                assignment.rig.tds_availability,
                assignment.tds_check,
                assignment.well.footprint,
                assignment.rig.rig_type,
                assignment.rig_type_check,
                assignment.well.latitude,
                assignment.well.longitude
            ])
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def export_schedule_excel(request, schedule_id):
    """Export schedule to Excel format with multiple sheets and formatting"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        assignments = Assignment.objects.filter(schedule=schedule).order_by('rig__name', 'sequence_order')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet if it exists
        if wb.active:
            wb.remove(wb.active)
        
        # Create schedule overview sheet
        ws_overview = wb.create_sheet("Schedule Overview")
        
        # Add title and basic info
        ws_overview['A1'] = f'Schedule: {schedule.name}'
        ws_overview['A1'].font = Font(size=16, bold=True)
        ws_overview['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_overview['A3'] = f'Total Wells: {assignments.count()}'
        ws_overview['A4'] = f'Total Rigs: {assignments.values("rig").distinct().count()}'
        
        # Create detailed assignments sheet
        ws_details = wb.create_sheet("Detailed Assignments")
        
        # Headers
        headers = [
            'Rig', 'Well', 'Asset ID', 'Well Type', 'Priority', 'RTD', 'RTD Check', 
            'Well Start Date', 'Well Start Check', 'Well End Date', 'Well End Check',
            'Rig Start Date', 'Rig End Date', 'Duration (days)', 'Depth (m)', 
            'Required HP', 'Rig HP', 'HP Check', 'Required BOP', 'Rig BOP', 'BOP Check',
            'Required TDS', 'Rig TDS', 'TDS Check', 'Required Type', 'Rig Type', 'Type Check',
            'Latitude', 'Longitude', 'Sequence Order'
        ]
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_num, assignment in enumerate(assignments, 2):
            data = [
                assignment.rig.name,
                assignment.well.name,
                assignment.well.asset_id,
                assignment.well.well_type,
                assignment.well.priority,
                assignment.well.rtd,
                assignment.rtd_check,
                assignment.well_start_date,
                assignment.well_start_check,
                assignment.well_end_date,
                assignment.well_end_check,
                assignment.rig.start_date,
                assignment.rig.end_date,
                assignment.well.duration,
                assignment.well.depth,
                assignment.well.rig_capacity_required_hp,
                assignment.rig.rig_capacity_hp,
                assignment.hp_check,
                assignment.well.bop_stack,
                assignment.rig.bop_stack,
                assignment.bop_check,
                assignment.well.tds_requirement,
                assignment.rig.tds_availability,
                assignment.tds_check,
                assignment.well.footprint,
                assignment.rig.rig_type,
                assignment.rig_type_check,
                assignment.well.latitude,
                assignment.well.longitude,
                assignment.sequence_order
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_details.cell(row=row_num, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Color code based on checks
                if col in [7, 9, 11, 18, 21, 24, 27]:  # Check columns
                    if value == 'PASS':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == 'FAIL':
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Create rig summary sheet
        ws_rigs = wb.create_sheet("Rig Summary")
        
        # Get rig summary data
        rig_data = []
        for rig in assignments.values('rig__name', 'rig__rig_type', 'rig__rig_capacity_hp', 'rig__start_date', 'rig__end_date').distinct():
            rig_assignments = assignments.filter(rig__name=rig['rig__name'])
            rig_data.append([
                rig['rig__name'],
                rig['rig__rig_type'],
                rig['rig__rig_capacity_hp'],
                rig['rig__start_date'],
                rig['rig__end_date'],
                rig_assignments.count(),
                sum([a.well.duration for a in rig_assignments])
            ])
        
        # Rig headers
        rig_headers = ['Rig Name', 'Type', 'Capacity (HP)', 'Start Date', 'End Date', 'Wells Assigned', 'Total Duration (days)']
        
        for col, header in enumerate(rig_headers, 1):
            cell = ws_rigs.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, rig_row in enumerate(rig_data, 2):
            for col, value in enumerate(rig_row, 1):
                ws_rigs.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for ws in [ws_overview, ws_details, ws_rigs]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="schedule_{schedule.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload_unified(request):
    """Unified bulk upload endpoint for both rigs and wells"""
    try:
        data_type = request.data.get('data_type')
        file_obj = request.data.get('file')
        
        if not data_type:
            return Response(
                {'error': 'data_type parameter is required (rigs or wells)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not file_obj:
            return Response(
                {'error': 'file parameter is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if data_type == 'rigs':
            return handle_rigs_upload(file_obj)
        elif data_type == 'wells':
            return handle_wells_upload(file_obj)
        else:
            return Response(
                {'error': 'data_type must be either "rigs" or "wells"'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
    except Exception as e:
        return Response(
            {'error': f'Upload failed: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def handle_rigs_upload(file_obj):
    """Handle rigs CSV upload"""
    try:
        df = pd.read_csv(file_obj)
        created_count = 0
        errors = []
        
        # Log the actual CSV columns for debugging
        print(f"CSV Columns found: {list(df.columns)}")
        
        # Column mapping from CSV headers to model fields
        column_mapping = {
            'Rig': 'name',
            'Name': 'name',  # Handle both "Rig" and "Name" columns
            'Asset ID': 'asset_id',
            'Rig Type': 'rig_type',
            'Start Date': 'start_date', 
            'End Date': 'end_date',
            'Rig Capacity (HP)': 'rig_capacity_hp',
            'Daily Cost (INR)': 'daily_cost_inr',
            'Drilling Capacity (m)': 'drilling_capacity_m',
            'Mobilization Time (Days)': 'mobilization_time_days',
            'Maintenance Schedule': 'maintenance_schedule',
            'Crew Availability': 'crew_availability',
            'HPHT Suitability': 'hpht_suitability',
            'ILM COST FIXED': 'ilm_cost_fixed',
            'ILM Cost Fixed': 'ilm_cost_fixed',  # Handle different capitalizations
            'ILM COST per km': 'ilm_cost_per_km',
            'ILM Cost per km': 'ilm_cost_per_km', 
            'ILM COST CLUSTER': 'ilm_cost_cluster',
            'ILM Cost Cluster': 'ilm_cost_cluster',
            'BOP Stack': 'bop_stack',
            'TDS Availability': 'tds_availability'
        }
        
        # Check if required columns are present
        if 'Rig' not in df.columns and 'Name' not in df.columns and 'name' not in df.columns:
            return Response(
                {'error': f'Required column "Rig" or "Name" not found. CSV columns are: {list(df.columns)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Rename columns to match model fields
        df = df.rename(columns=column_mapping)
        
        for idx, (index, row) in enumerate(df.iterrows()):
            try:
                # Convert row to dictionary and handle data types
                rig_data = row.to_dict()
                
                # Convert date strings to proper date format (handle multiple formats)
                if 'start_date' in rig_data and pd.notna(rig_data['start_date']):
                    date_str = str(rig_data['start_date'])
                    try:
                        # Try DD-MM-YYYY format first
                        rig_data['start_date'] = pd.to_datetime(date_str, format='%d-%m-%Y').date()
                    except ValueError:
                        try:
                            # Try DD/MM/YY format (e.g., 01/05/24)
                            rig_data['start_date'] = pd.to_datetime(date_str, format='%d/%m/%y').date()
                        except ValueError:
                            try:
                                # Try DD/MM/YYYY format
                                rig_data['start_date'] = pd.to_datetime(date_str, format='%d/%m/%Y').date()
                            except ValueError:
                                # Use general date parsing as last resort
                                rig_data['start_date'] = pd.to_datetime(date_str).date()
                
                if 'end_date' in rig_data and pd.notna(rig_data['end_date']):
                    date_str = str(rig_data['end_date'])
                    try:
                        # Try DD-MM-YYYY format first
                        rig_data['end_date'] = pd.to_datetime(date_str, format='%d-%m-%Y').date()
                    except ValueError:
                        try:
                            # Try DD/MM/YY format (e.g., 01/05/24)
                            rig_data['end_date'] = pd.to_datetime(date_str, format='%d/%m/%y').date()
                        except ValueError:
                            try:
                                # Try DD/MM/YYYY format
                                rig_data['end_date'] = pd.to_datetime(date_str, format='%d/%m/%Y').date()
                            except ValueError:
                                # Use general date parsing as last resort
                                rig_data['end_date'] = pd.to_datetime(date_str).date()
                
                # Handle empty/null values and special strings like "Nil"
                for key, value in rig_data.items():
                    if pd.isna(value) or value == '' or str(value).upper() == 'NIL':
                        if key in ['mobilization_time_days', 'maintenance_schedule']:
                            rig_data[key] = None
                        elif key == 'hpht_suitability':
                            rig_data[key] = 'N'  # Default value
                        elif key == 'crew_availability':
                            rig_data[key] = 'OK'  # Default value
                        elif key == 'tds_availability':
                            rig_data[key] = 'Y'  # Default value
                
                # Convert YES/NO to Y/N
                if 'hpht_suitability' in rig_data:
                    if str(rig_data['hpht_suitability']).upper() in ['YES', 'NO']:
                        rig_data['hpht_suitability'] = 'Y' if str(rig_data['hpht_suitability']).upper() == 'YES' else 'N'
                
                # Create or update rig - USE all_objects to find soft-deleted records too!
                rig, created = Rig.all_objects.get_or_create(
                    name=rig_data['name'],
                    defaults=rig_data
                )
                
                if created:
                    created_count += 1
                else:
                    # Update existing rig (may be soft-deleted)
                    was_deleted = rig.is_deleted
                    for key, value in rig_data.items():
                        if key != 'name':  # Don't update the name (it's the identifier)
                            setattr(rig, key, value)
                    
                    # Revive if soft-deleted
                    if was_deleted:
                        rig.is_deleted = False
                        rig.deleted_at = None
                        rig.deleted_by = None
                    
                    rig.save()
                    
            except Exception as e:
                error_msg = f"Row {idx + 1} (Rig: {row.get('name', 'Unknown')}): {str(e)}"
                errors.append(error_msg)
                print(f"Error processing rig row {idx + 1}: {e}")  # Also log to console
                continue
        
        return Response({
            'message': f'Successfully processed {len(df)} rows',
            'created': created_count,
            'updated': len(df) - created_count - len(errors),
            'errors': errors,
            'error_count': len(errors)
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process rigs file: {str(e)}'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


def handle_wells_upload(file_obj):
    """Handle wells CSV upload"""
    try:
        df = pd.read_csv(file_obj)
        created_count = 0
        errors = []
        
        # Log the actual CSV columns for debugging
        print(f"Wells CSV Columns found: {list(df.columns)}")
        
        # Column mapping from CSV headers to model fields
        column_mapping = {
            'SN': 'sn',
            'Asset_ID': 'asset_id',
            'Asset ID': 'asset_id',  # Handle both formats
            'Well': 'name',
            'Name': 'name',  # Handle both "Well" and "Name" columns
            'Type of well': 'well_type',
            'Well Type': 'well_type',  # Handle both formats
            'Well Profile': 'well_profile',
            'Depth': 'depth',
            'Rig Capacity Required (HP)': 'rig_capacity_required_hp',
            'DRL_DAYS': 'drl_days',
            'DRL Days': 'drl_days',  # Handle both formats
            'PT_DAYS': 'pt_days',
            'PT Days': 'pt_days',  # Handle both formats
            'Duration': 'duration',
            'Latitude': 'latitude',
            'Longitude': 'longitude',
            'RTD': 'rtd',
            'BOP Stack': 'bop_stack',
            'TDS Requirement': 'tds_requirement',
            'Footprint': 'footprint',
            'Preferred Rig': 'preferred_rig',
            'Expected_Potential': 'expected_potential',
            'Expected Potential': 'expected_potential',  # Handle both formats
            'Priority': 'priority'
        }
        
        # Check if required columns are present (SN is now optional - auto-generated)
        if 'Well' not in df.columns and 'Name' not in df.columns and 'name' not in df.columns:
            return Response(
                {'error': f'Required column "Well" or "Name" not found. CSV columns are: {list(df.columns)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if 'Asset_ID' not in df.columns and 'Asset ID' not in df.columns and 'asset_id' not in df.columns:
            return Response(
                {'error': f'Required column "Asset_ID" or "Asset ID" not found. CSV columns are: {list(df.columns)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Rename columns to match model fields
        df = df.rename(columns=column_mapping)
        
        # Filter out completely empty rows
        df = df.dropna(how='all')
        
        # Get the next available SN for auto-generation
        from django.db.models import Max
        max_sn = Well.objects.aggregate(Max('sn'))['sn__max'] or 0
        next_sn = max_sn + 1
        
        # Define valid model fields (exclude auto-generated fields and extra columns)
        valid_fields = {
            'sn', 'asset_id', 'name', 'well_type', 'well_profile', 'depth',
            'rig_capacity_required_hp', 'drl_days', 'pt_days', 'duration',
            'latitude', 'longitude', 'rtd', 'bop_stack', 'tds_requirement',
            'footprint', 'preferred_rig', 'expected_potential', 'priority'
        }
        
        for idx, (index, row) in enumerate(df.iterrows()):
            try:
                # Convert row to dictionary and handle data types
                well_data = row.to_dict()
                
                # Remove columns that aren't valid model fields (like ID, Created At)
                well_data = {k: v for k, v in well_data.items() if k in valid_fields}
                
                # Convert date strings to proper format (handle multiple formats)
                if 'rtd' in well_data and pd.notna(well_data['rtd']):
                    rtd_value = str(well_data['rtd'])
                    try:
                        # Try DD-MM-YYYY format first
                        well_data['rtd'] = pd.to_datetime(rtd_value, format='%d-%m-%Y').date()
                    except ValueError:
                        try:
                            # Try DD/MM/YY format (e.g., 05/01/24)
                            well_data['rtd'] = pd.to_datetime(rtd_value, format='%d/%m/%y').date()
                        except ValueError:
                            try:
                                # Try DD/MM/YYYY format
                                well_data['rtd'] = pd.to_datetime(rtd_value, format='%d/%m/%Y').date()
                            except ValueError:
                                try:
                                    # Try YYYY-MM-DD format
                                    well_data['rtd'] = pd.to_datetime(rtd_value, format='%Y-%m-%d').date()
                                except ValueError:
                                    # Use general date parsing as last resort
                                    well_data['rtd'] = pd.to_datetime(rtd_value).date()
                
                # Handle empty/null values and special strings like "Nil"
                for key, value in well_data.items():
                    if pd.isna(value) or value == '' or str(value).upper() == 'NIL':
                        if key in ['preferred_rig', 'expected_potential']:
                            well_data[key] = None
                        elif key == 'priority':
                            well_data[key] = 'MEDIUM'  # Default priority
                        elif key == 'tds_requirement':
                            well_data[key] = 'Y'  # Default value
                
                # Convert YES/NO to Y/N for TDS requirement
                if 'tds_requirement' in well_data:
                    if str(well_data['tds_requirement']).upper() in ['YES', 'NO']:
                        well_data['tds_requirement'] = 'Y' if str(well_data['tds_requirement']).upper() == 'YES' else 'N'
                
                # Handle priority field properly
                if 'priority' in well_data:
                    priority_val = str(well_data['priority']).upper().strip()
                    if priority_val in ['HIGH', 'MEDIUM', 'LOW']:
                        well_data['priority'] = priority_val
                    elif priority_val == '' or pd.isna(well_data['priority']):
                        well_data['priority'] = 'MEDIUM'
                    else:
                        well_data['priority'] = 'MEDIUM'  # Default for unrecognized values
                
                # Get well name and asset_id for unique lookup
                well_name = well_data.get('name')
                asset_id = well_data.get('asset_id')
                
                # Auto-generate SN if not provided
                if 'sn' not in well_data or pd.isna(well_data.get('sn')):
                    well_data['sn'] = next_sn
                
                # Create or update well using name + asset_id as unique key
                # USE all_objects to find soft-deleted records too!
                well, created = Well.all_objects.get_or_create(
                    name=well_name,
                    asset_id=asset_id,
                    defaults=well_data
                )
                
                if created:
                    created_count += 1
                    next_sn += 1  # Increment for next auto-generated SN
                else:
                    # Update existing well (may be soft-deleted)
                    was_deleted = well.is_deleted
                    for key, value in well_data.items():
                        if key != 'sn':  # Don't update the serial number
                            setattr(well, key, value)
                    
                    # Revive if soft-deleted
                    if was_deleted:
                        well.is_deleted = False
                        well.deleted_at = None
                        well.deleted_by = None
                    
                    well.save()
                    
            except Exception as e:
                error_msg = f"Row {idx + 1} (Well: {row.get('name', 'Unknown')}): {str(e)}"
                errors.append(error_msg)
                print(f"Error processing well row {idx + 1}: {e}")  # Also log to console
                continue
        
        return Response({
            'message': f'Successfully processed {len(df)} rows',
            'created': created_count,
            'updated': len(df) - created_count - len(errors),
            'errors': errors,
            'error_count': len(errors)
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process wells file: {str(e)}'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
def wells_by_location(request):
    """API endpoint to get wells filtered by asset_id/location"""
    asset_id = request.GET.get('asset_id')
    
    try:
        if asset_id:
            wells = Well.objects.filter(asset_id=asset_id)
        else:
            wells = Well.objects.all()
            
        wells_data = []
        for well in wells:
            wells_data.append({
                'id': str(well.id),
                'name': well.name,
                'sn': well.sn,
                'asset_id': well.asset_id,
                'well_type': well.well_type,
                'well_profile': well.well_profile,
                'depth': well.depth,
                'rig_capacity_required_hp': well.rig_capacity_required_hp,
                'drl_days': well.drl_days,
                'pt_days': well.pt_days,
                'duration': well.duration,
                'latitude': float(well.latitude),
                'longitude': float(well.longitude),
                'rtd': well.rtd.isoformat() if well.rtd else None,
                'bop_stack': well.bop_stack,
                'tds_requirement': well.tds_requirement,
                'footprint': well.footprint,
                'preferred_rig': well.preferred_rig,
                'expected_potential': float(well.expected_potential) if well.expected_potential else None,
                'priority': well.priority
            })
            
        return JsonResponse({
            'success': True,
            'wells': wells_data,
            'count': len(wells_data),
            'asset_id': asset_id
        })
        
    except Exception as e:
        logger.error(f"Error fetching wells by location: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET']) 
def rigs_by_location(request):
    """API endpoint to get rigs filtered by asset_id/location"""
    asset_id = request.GET.get('asset_id')
    
    try:
        if asset_id:
            rigs = Rig.objects.filter(asset_id=asset_id)
        else:
            rigs = Rig.objects.all()
            
        rigs_data = []
        for rig in rigs:
            rigs_data.append({
                'id': str(rig.id),
                'name': rig.name,
                'asset_id': rig.asset_id,
                'rig_type': rig.rig_type,
                'start_date': rig.start_date.isoformat() if rig.start_date else None,
                'end_date': rig.end_date.isoformat() if rig.end_date else None,
                'rig_capacity_hp': rig.rig_capacity_hp,
                'daily_cost_inr': float(rig.daily_cost_inr) if rig.daily_cost_inr else None,
                'drilling_capacity_m': rig.drilling_capacity_m,
                'mobilization_time_days': rig.mobilization_time_days,
                'maintenance_schedule': rig.maintenance_schedule,
                'crew_availability': rig.crew_availability,
                'hpht_suitability': rig.hpht_suitability,
                'ilm_cost_fixed': float(rig.ilm_cost_fixed) if rig.ilm_cost_fixed else None,
                'ilm_cost_per_km': float(rig.ilm_cost_per_km) if rig.ilm_cost_per_km else None,
                'ilm_cost_cluster': float(rig.ilm_cost_cluster) if rig.ilm_cost_cluster else None,
                'bop_stack': rig.bop_stack,
                'tds_availability': rig.tds_availability
            })
            
        return JsonResponse({
            'success': True,
            'rigs': rigs_data,
            'count': len(rigs_data),
            'asset_id': asset_id
        })
        
    except Exception as e:
        logger.error(f"Error fetching rigs by location: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@login_required
def get_user_location_info(request):
    """
    Get the current user's location information and accessible locations.
    
    Returns:
        - user_location: The location assigned to the user (or null if admin)
        - can_view_all: Boolean indicating if user can view all locations
        - accessible_locations: List of locations the user can access
    """
    try:
        user = request.user
        user_location = get_user_location(user)
        accessible_locations = get_user_accessible_locations(user)
        
        # Format the response
        response_data = {
            'user': {
                'username': user.username,
                'email': user.email,
                'is_superuser': user.is_superuser,
                'is_staff': user.is_staff,
            },
            'user_location': None,
            'can_view_all': user.is_superuser,
            'accessible_locations': []
        }
        
        # Check user's profile
        if hasattr(user, 'profile') and user.profile:
            response_data['can_view_all'] = user.profile.can_view_all_locations or user.is_superuser
            if user.profile.location:
                response_data['user_location'] = {
                    'id': str(user.profile.location.id),
                    'code': user.profile.location.code,
                    'name': user.profile.location.name,
                }
        
        # Format accessible locations
        for location in accessible_locations:
            response_data['accessible_locations'].append({
                'id': str(location.id),
                'code': location.code,
                'name': location.name,
                'rig_count': location.rigs.count(),
                'well_count': location.wells.count(),
            })
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Error getting user location info: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_all_locations(request):
    """
    Get all active locations. Available to all authenticated users.
    """
    try:
        locations = Location.objects.filter(is_active=True).order_by('code')
        
        locations_data = []
        for location in locations:
            locations_data.append({
                'id': str(location.id),
                'code': location.code,
                'name': location.name,
                'description': location.description,
                'rig_count': location.rigs.count(),
                'well_count': location.wells.count(),
            })
        
        return Response({
            'locations': locations_data,
            'count': len(locations_data)
        })
        
    except Exception as e:
        logger.error(f"Error getting all locations: {str(e)}")
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
